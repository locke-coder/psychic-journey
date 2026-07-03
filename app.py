"""Streamlit entry point for the input-driven sales closing forecast tool."""

from __future__ import annotations

import hashlib
import hmac
import io
import math
import os
import tempfile
from html import escape
from pathlib import Path
from typing import Any, Mapping

import altair as alt
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

try:
    import streamlit as st
except ModuleNotFoundError:  # pragma: no cover - local test runtime may omit Streamlit.
    st = None

from src import history_schema
from src.backtest_engine import build_backtest_dataset, summarize_by_forecast_model
from src.close_cycle_engine import _coerce_is_close_day, build_close_cycle_summary
from src.display_labels import (
    get_forecast_model_label,
    get_metric_label,
    get_operation_mode,
    get_status_label,
    get_strategy_code,
    get_strategy_group,
    get_strategy_label,
    get_strategy_short_description,
)
from src.excel_exporter import export_daily_report

try:
    from src.excel_exporter import (
        SCENARIO_GRID_REQUIRED_COLUMNS as _EXCEL_SCENARIO_GRID_REQUIRED_COLUMNS,
        prepare_scenario_grid_export_frame as _excel_prepare_scenario_grid_export_frame,
    )
except ImportError:
    _EXCEL_SCENARIO_GRID_REQUIRED_COLUMNS = None
    _excel_prepare_scenario_grid_export_frame = None
from src.final_actual_store import load_final_actuals
from src.forecast_models import (
    F1_CUMULATIVE_RATE,
    F2_LAST_TWO_CLOSES,
    F3_DAY_CLOSE_WEIGHTED,
    run_forecast_model,
)
from src.loader import load_input
from src.operator_sample_store import (
    get_operator_sample_location,
    get_operator_sample_path,
    get_packaged_sample_path,
    load_sample_with_source,
    read_operator_metadata,
    save_operator_sample,
)
from src.history_store import append_forecast_history, build_forecast_history_rows
from src.next_close import calculate_next_close_required
from src.overachievement_models import (
    N1_MAINTAIN_TARGET,
    N2_MONITOR_BUFFER,
    N3_QUALITY_CHECK,
    NEUTRAL,
    O1_TARGET_HOLD_BUFFER,
    O2_STRETCH_TARGET_CAPTURE,
    O3_QUALITY_GUARD_RELIEF,
    OVERACHIEVEMENT,
    PROVISION,
    run_neutral_strategy,
    run_overachievement_strategy,
)
from src.provision_models import (
    P1_ALL_REMAINING,
    P2_CLOSE_DAY_FOCUSED,
    P3_NON_CLOSE_DAY_FOCUSED,
    run_provision_model,
)
from src.report_builder import (
    FORECAST_MODEL_DEFINITIONS as REPORT_FORECAST_MODEL_DEFINITIONS,
    NEUTRAL_STRATEGY_DEFINITIONS as REPORT_NEUTRAL_STRATEGY_DEFINITIONS,
    OVERACHIEVEMENT_STRATEGY_DEFINITIONS as REPORT_OVERACHIEVEMENT_STRATEGY_DEFINITIONS,
    PROVISION_STRATEGY_DEFINITIONS as REPORT_PROVISION_STRATEGY_DEFINITIONS,
    RISK_LEVEL_DEFINITIONS as REPORT_RISK_LEVEL_DEFINITIONS,
    build_daily_report_text,
)
from src.scenario_runner import run_scenario_grid
from src.schema import get_metric_columns, load_model_config
from src.ui_components import (
    format_krw,
    render_download_card,
    render_history_card,
    render_kpi_card,
    render_operation_mode_card,
    render_pace_header,
    render_report_card,
    render_scenario_card,
    render_section_header,
    render_status_badge,
    scenario_description,
    scenario_display_name,
    status_label,
)
from src.ui_navigation import (
    PAGE_DEFINITIONS,
    get_current_page,
    page_title,
    validate_page_key,
)
from src.ui_pages import render_page
from src.ui_styles import inject_global_styles
from src.ui_theme import get_pace_check_css
from src.validator import validate_input
from src.visualization_builder import (
    build_close_day_markers,
    build_close_cycle_cumulative_source,
    build_forecast_model_mini_chart_source,
    build_pace_projection_chart_data,
    build_strategy_arrival_compare_source,
)


REPO_ROOT = Path(__file__).resolve().parent
SAMPLE_INPUT_PATH = REPO_ROOT / "data" / "sample" / "input_sample.csv"
HISTORICAL_SAMPLE_INPUT_PATH = REPO_ROOT / "data" / "sample" / "historical_input_sample.csv"
OUTPUT_DIR = REPO_ROOT / "outputs"
SAVED_ACTUALS_PATH = OUTPUT_DIR / "saved_actuals.csv"
AUDIT_READONLY_QUERY_PARAM = "audit_readonly"
AUDIT_READONLY_TRUE_VALUES = {"1", "true", "yes", "on"}
INPUT_TEMPLATE_FILENAME = "month_close_forecast_input_template.xlsx"
HISTORICAL_INPUT_TEMPLATE_FILENAME = "historical_month_close_forecast_input_template.xlsx"
SAMPLE_INPUT_SOURCE_LABEL = "샘플 데이터"
HISTORICAL_SAMPLE_INPUT_SOURCE_LABEL = "과거 샘플 데이터"
OPERATOR_SAMPLE_SOURCE_LABEL = "운영 저장본"
PACKAGED_SAMPLE_DISPLAY_LABEL = "내장 샘플"
SAVED_ACTUALS_SOURCE_LABEL = "내장 샘플 + 저장 실적"
UPLOAD_SAMPLE_DISPLAY_LABEL = "업로드 파일"
INPUT_TEMPLATE_HEADERS = (
    "date",
    "day_name",
    "business_day_no",
    "is_close_day",
    "close_type",
    "sales_target_daily",
    "recognized_target_daily",
    "sales_actual_cum",
    "recognized_actual_cum",
    "memo",
)
INPUT_TEMPLATE_SAMPLE_ROWS = (
    (
        "YYYY-MM-DD",
        "display_only",
        1,
        False,
        "",
        None,
        None,
        None,
        None,
        "",
    ),
)
FALLBACK_SCENARIO_GRID_REQUIRED_COLUMNS = (
    "scenario",
    "forecast_model",
    "model_name",
    "expected_month_end_amount",
    "target_status",
    "target_variance",
    "surplus_to_target",
    "strategy_type",
    "strategy_code",
    "overachievement_strategy",
    "strategy_label",
    "strategy_group",
    "stretch_uplift",
    "revised_monthly_target",
    "remaining_surplus_buffer",
    "minimum_remaining_to_hit_target",
    "relief_amount",
    "recommended_action",
    "risk_note",
)
SCENARIO_GRID_REQUIRED_COLUMNS = (
    _EXCEL_SCENARIO_GRID_REQUIRED_COLUMNS or FALLBACK_SCENARIO_GRID_REQUIRED_COLUMNS
)
HISTORICAL_INPUT_TEMPLATE_SAMPLE_ROWS = (
    ("2026-03-02", "월", 1, True, "월초특수", 35.0, 32.0, 33.0, 30.0, "과거 월 예시"),
    ("2026-03-03", "화", 2, False, "일반", 2.5, 2.3, 38.0, 34.5, "과거 월 예시"),
    ("2026-03-05", "목", 3, True, "목마감", 14.0, 12.8, 51.0, 46.3, "과거 월 예시"),
    ("2026-04-01", "수", 1, True, "월초특수", 34.0, 31.0, 31.0, 28.3, "다음 월은 1부터 다시 시작"),
    ("2026-04-02", "목", 2, False, "일반", 2.6, 2.4, 35.0, 31.8, "다음 월은 1부터 다시 시작"),
    ("2026-04-06", "월", 3, True, "월마감", 14.5, 13.3, 47.0, 42.7, "다음 월은 1부터 다시 시작"),
)
HISTORY_TAB_LABEL = "예측 이력 / Backtest"
CURRENT_INPUT_DF_SESSION_KEY = "pace_current_input_df"
CURRENT_INPUT_SOURCE_SESSION_KEY = "pace_current_input_source"
CURRENT_INPUT_SOURCE_OVERRIDE_SESSION_KEY = "pace_current_input_source_override"
HISTORICAL_INPUT_DF_SESSION_KEY = "pace_historical_input_df"
HISTORICAL_INPUT_SOURCE_SESSION_KEY = "pace_historical_input_source"
HISTORICAL_SAMPLE_DISABLED_SESSION_KEY = "pace_historical_sample_disabled"
OPERATOR_SAMPLE_NOTICE_SESSION_KEY = "pace_operator_sample_notice"
PACE_METRIC_SESSION_KEY = "pace_metric"
PACE_AS_OF_DATE_SESSION_KEY = "pace_as_of_date"
PACE_AS_OF_DATE_DEFAULT_SESSION_KEY = "pace_as_of_date_default_token"
PACE_FORECAST_CHOICE_SESSION_KEY = "pace_forecast_choice"
PACE_STRATEGY_CHOICE_SESSION_KEY = "pace_strategy_choice"
PACE_CLOSE_CAP_SESSION_KEY = "pace_close_day_cap_rate"
PACE_NON_CLOSE_CAP_SESSION_KEY = "pace_non_close_day_cap_rate"
PACE_SELECTED_SCENARIO_SESSION_KEY = "pace_selected_scenario_id"
ACCESS_SESSION_STATE_KEY = "limited_distribution_access_granted"
ACCESS_PASSWORD_SETTING_KEYS = (
    "APP_ACCESS_PASSWORD",
    "app_access_password",
    "access_password",
)
ACCESS_PASSWORD_HASH_SETTING_KEYS = (
    "APP_ACCESS_PASSWORD_SHA256",
    "app_access_password_sha256",
    "access_password_sha256",
)
COMPARE_LABEL = "전체 비교"
APP_TIMEZONE = "Asia/Seoul"
CHART_COLOR_RANGE = (
    "#2563EB",
    "#059669",
    "#D97706",
    "#7C3AED",
    "#DC2626",
    "#0F766E",
    "#9333EA",
    "#4B5563",
)
WEIGHTED_FORECAST_COLUMN_TOKENS = ("weighted_forecast", "weighted_forecast_amount")
CONFIDENCE_BAND_COLUMN_PAIRS = (
    ("confidence_lower", "confidence_upper"),
    ("forecast_lower", "forecast_upper"),
    ("lower_bound", "upper_bound"),
)

FORECAST_MODEL_OPTIONS = {
    "F1": F1_CUMULATIVE_RATE,
    "F2": F2_LAST_TWO_CLOSES,
    "F3": F3_DAY_CLOSE_WEIGHTED,
}
PROVISION_STRATEGY_OPTIONS = {
    "P1": P1_ALL_REMAINING,
    "P2": P2_CLOSE_DAY_FOCUSED,
    "P3": P3_NON_CLOSE_DAY_FOCUSED,
}
OVERACHIEVEMENT_STRATEGY_OPTIONS = {
    "O1": O1_TARGET_HOLD_BUFFER,
    "O2": O2_STRETCH_TARGET_CAPTURE,
    "O3": O3_QUALITY_GUARD_RELIEF,
}
NEUTRAL_STRATEGY_OPTIONS = {
    "N1": N1_MAINTAIN_TARGET,
    "N2": N2_MONITOR_BUFFER,
    "N3": N3_QUALITY_CHECK,
}
SCENARIO_STRATEGY_OPTIONS = {
    **PROVISION_STRATEGY_OPTIONS,
    **OVERACHIEVEMENT_STRATEGY_OPTIONS,
    **NEUTRAL_STRATEGY_OPTIONS,
}
FORECAST_MODEL_DEFINITIONS = {
    "F1": {
        "name": "누적 달성률 모델",
        "description": "기준일까지의 누적 실적/누적 목표 달성률을 남은 모든 일자 목표에 동일하게 적용합니다.",
        "formula": "forecast = current_actual_cum + remaining_target * r_cum",
    },
    "F2": {
        "name": "직전 2개 완료 마감차수 모델",
        "description": "기준일까지 완료된 최근 2개 마감차수의 실적/목표 비율을 남은 일자에 적용합니다.",
        "formula": "forecast = current_actual_cum + remaining_target * r_last2",
    },
    "F3": {
        "name": "마감일/비마감일 가중 모델",
        "description": "마감일과 비마감일의 과거 달성률을 분리해 남은 일자의 마감 여부별로 적용합니다.",
        "formula": "forecast = current_actual_cum + close_target * r_close + non_close_target * r_non_close",
    },
}
PROVISION_STRATEGY_DEFINITIONS = {
    "P1": {
        "name": "전체 잔여일 배분",
        "description": "예상 부족분을 기준일 이후 모든 잔여 일자에 기존 일 목표 비중대로 배분합니다.",
    },
    "P2": {
        "name": "마감일 우선 배분",
        "description": "예상 부족분을 기준일 이후 마감일에 우선 배분하고, 상한 초과분은 설정에 따라 재배분합니다.",
    },
    "P3": {
        "name": "비마감일 우선 배분",
        "description": "예상 부족분을 기준일 이후 비마감일에 우선 배분하고, 상한 초과분은 설정에 따라 재배분합니다.",
    },
}
OVERACHIEVEMENT_STRATEGY_DEFINITIONS = {
    "O1": {
        "name": "목표 유지 안전버퍼",
        "description": "월 목표는 유지하고 초과 예상분을 취소, 철회, 미결제, 실적 조정 리스크 방어용 버퍼로 둡니다.",
    },
    "O2": {
        "name": "상향 목표 전환",
        "description": "초과 예상분 일부를 추가 성장 목표로 전환하고 남은 초과분은 안전버퍼로 유지합니다.",
    },
    "O3": {
        "name": "계약 품질 방어",
        "description": "공식 목표는 유지하고 결제완료율, 순계약, 상담 품질, 취소/철회/미결제 리스크 관리에 집중합니다.",
    },
}
NEUTRAL_STRATEGY_DEFINITIONS = {
    "N1": {
        "name": "목표 유지",
        "description": "목표와 거의 같은 상태로 보고 공식 월 목표와 잔여 실행 계획을 유지합니다.",
    },
    "N2": {
        "name": "버퍼 모니터링",
        "description": "목표선 부근의 작은 변동이 미달로 바뀌지 않도록 취소, 철회, 미결제 변동을 점검합니다.",
    },
    "N3": {
        "name": "품질 점검",
        "description": "무리한 추가 상향보다 계약 품질, 결제완료율, 순계약 상태를 확인합니다.",
    },
}
SCENARIO_STRATEGY_DEFINITIONS = {
    **PROVISION_STRATEGY_DEFINITIONS,
    **OVERACHIEVEMENT_STRATEGY_DEFINITIONS,
    **NEUTRAL_STRATEGY_DEFINITIONS,
}
RISK_LEVEL_DEFINITIONS = {
    "Green": "예상 달성률이 100% 이상이고 보정 상태가 정상이거나 부족분이 없는 상태입니다.",
    "Yellow": "예상 달성률이 95% 이상이거나 필요 상향분이 잔여 목표의 5% 이하인 상태입니다.",
    "Red": "예상 달성률이 90% 이상이거나 필요 상향분이 잔여 목표의 15% 이하인 상태입니다.",
    "Black": "계산 오류, 상한 부족, 또는 예상 달성률/필요 상향 부담이 높은 상태입니다.",
    "N/A": "해당 보정 전략을 적용할 수 없는 상태입니다.",
}

AMOUNT_COLUMNS = {
    "monthly_target",
    "current_target_cum",
    "current_actual_cum",
    "remaining_target",
    "forecast_amount",
    "expected_month_end_amount",
    "gap_to_target",
    "target_variance",
    "target_variance_after_strategy",
    "surplus_to_target",
    "base_forecast_amount",
    "strategy_expected_amount",
    "surplus_buffer",
    "compare_value",
    "forecast_reference_value",
    "required_uplift",
    "allocated_uplift",
    "unallocated_uplift",
    "revised_remaining_target",
    "operating_target_reference",
    "stretch_uplift",
    "revised_monthly_target",
    "remaining_surplus_buffer",
    "minimum_remaining_to_hit_target",
    "relief_amount",
    "forecast_after_provision",
    "gap_after_provision",
    "next_close_required",
    "original_target",
    "uplift",
    "revised_target",
    "cap_target",
    "expected_after_revision",
    "final_actual",
    "target_cum",
    "actual_cum",
    "cancellation_amount",
    "net_actual",
    "forecast_error",
    "abs_error",
    "mean_abs_error",
    "bias",
    "weighted_forecast",
    "weighted_forecast_amount",
    "confidence_lower",
    "confidence_upper",
    "forecast_lower",
    "forecast_upper",
    "lower_bound",
    "upper_bound",
}
RATE_COLUMNS = {
    "forecast_rate",
    "expected_rate",
    "allocation_weight",
    "final_achievement_rate",
    "cumulative_achievement_rate",
    "error_rate",
    "signed_error_rate",
    "mean_error_rate",
    "median_error_rate",
}
TECHNICAL_CODE_COLUMNS = {
    "scenario",
    "scenario_id",
    "strategy_key",
    "strategy_code",
    "forecast_key",
}
DIRECT_EDITABLE_COLUMNS = (
    "sales_target_daily",
    "recognized_target_daily",
    "sales_actual_cum",
    "recognized_actual_cum",
)
ACTUAL_CUM_COLUMNS = (
    "sales_actual_cum",
    "recognized_actual_cum",
)
SAVED_ACTUAL_COLUMNS = (
    "date",
    "business_day_no",
    *ACTUAL_CUM_COLUMNS,
)
REPORT_GLOSSARY_GROUPS = (
    ("예측모델(F)", REPORT_FORECAST_MODEL_DEFINITIONS),
    ("목표 보정 전략(P)", REPORT_PROVISION_STRATEGY_DEFINITIONS),
    ("초과달성 운영전략(O)", REPORT_OVERACHIEVEMENT_STRATEGY_DEFINITIONS),
    ("유지/모니터링 전략(N)", REPORT_NEUTRAL_STRATEGY_DEFINITIONS),
    ("위험등급", REPORT_RISK_LEVEL_DEFINITIONS),
)
STRATEGY_LEVEL_COLUMNS = (
    "scenario_id",
    "forecast_model",
    "target_status",
    "provision_strategy",
    "strategy_difference_summary",
    "stretch_uplift",
    "revised_monthly_target",
    "remaining_surplus_buffer",
    "minimum_remaining_to_hit_target",
    "relief_amount",
    "revised_remaining_target",
    "strategy_type",
    "monthly_target",
    "forecast_amount",
    "target_variance",
    "gap_to_target",
    "surplus_to_target",
    "required_uplift",
    "forecast_after_provision",
    "gap_after_provision",
    "risk_level",
    "status",
    "recommended_action",
)
STRATEGY_LEVEL_CHART_COLUMNS = (
    "stretch_uplift",
    "revised_monthly_target",
    "remaining_surplus_buffer",
    "minimum_remaining_to_hit_target",
    "relief_amount",
    "revised_remaining_target",
)
SCENARIO_DAILY_FORECAST_COLUMNS = (
    "date",
    "date_label",
    "week_start",
    "week_end",
    "week_label",
    "week_no",
    "business_day_no",
    "is_close_day",
    "day_type",
    "close_type",
    "scenario_id",
    "series_type",
    "line_group",
    "daily_expected",
    "forecast_cum",
    "target_cum",
    "monthly_target",
    "achievement_rate",
    "target_achievement_rate",
    "achievement_label",
    "target_achievement_label",
    "forecast_label",
    "target_cum_label",
    "risk_level_label",
    "is_selected",
    "is_as_of_date",
)
SCENARIO_DAILY_DETAIL_COLUMNS = (
    "date",
    "scenario_id",
    "series_type",
    "day_type",
    "close_type",
    "daily_expected",
    "forecast_cum",
    "target_cum",
    "achievement_rate",
    "target_achievement_rate",
)
REMAINING_OPERATION_DIRECTION_COLUMNS = (
    "date",
    "date_label",
    "scenario_id",
    "strategy_type",
    "operation_mode",
    "day_type",
    "close_type",
    "original_target",
    "uplift",
    "revised_target",
    "expected_daily",
    "expected_rate",
    "direction",
    "direction_detail",
)
VALIDATION_COLUMN_LABELS = {
    "date": "날짜",
    "business_day_no": "영업일 번호",
    "is_close_day": "마감일 여부",
    "close_type": "마감 유형",
    "sales_target_daily": "판매실적 일 목표",
    "recognized_target_daily": "인정실적 일 목표",
    "sales_actual_cum": "판매실적 누적 실적",
    "recognized_actual_cum": "인정실적 누적 실적",
    "target_daily": "선택 지표의 일 목표",
    "actual_cum": "선택 지표의 누적 실적",
    "actual_daily": "일 실적",
    "monthly_target": "월 전체 목표",
    "current_target_cum": "기준일까지의 누적 목표",
    "current_actual_cum": "기준일까지의 누적 실적",
    "as_of_date": "기준일",
    "metric": "지표",
    "uplift_effective_rate": "목표 상향 반영률",
}
DISPLAY_COLUMN_LABELS = {
    **VALIDATION_COLUMN_LABELS,
    "scenario": "시나리오",
    "scenario_id": "시나리오",
    "forecast_model": "예측모델",
    "forecast_basis": "고정 예측모델",
    "strategy_key": "전략 코드",
    "strategy_code": "전략 코드",
    "strategy_label": "전략명",
    "strategy_group": "전략군",
    "provision_strategy": "운영전략",
    "strategy_effect_type": "O전략 차이",
    "strategy_difference_summary": "전략 차이 요약",
    "model_name": "모델명",
    "compare_value": "비교 기준값",
    "forecast_reference_value": "F예측 월말 예상",
    "forecast_rate": "예상 달성률",
    "remaining_target": "잔여 목표",
    "forecast_amount": get_metric_label("forecast_amount"),
    "expected_month_end_amount": get_metric_label("expected_month_end_amount"),
    "gap_to_target": "목표 미달 예상분",
    "target_variance": get_metric_label("target_variance"),
    "target_variance_after_strategy": "전략 적용 후 목표 대비 차이",
    "surplus_to_target": get_metric_label("surplus_to_target"),
    "base_forecast_amount": "기준 F예측값",
    "strategy_expected_amount": "전략 운영 기준값",
    "surplus_buffer": "안전버퍼",
    "target_status": get_metric_label("target_status"),
    "strategy_type": "전략 구분",
    "overachievement_strategy": "초과달성 전략",
    "required_uplift": "필요 상향",
    "allocated_uplift": "배분된 상향",
    "unallocated_uplift": "미배분 상향",
    "revised_remaining_target": "수정 잔여 목표",
    "operating_target_reference": "운영 기준 목표",
    "stretch_uplift": "상향 목표 전환분",
    "revised_monthly_target": "운영전략 월 목표",
    "remaining_surplus_buffer": "잔여 안전버퍼",
    "minimum_remaining_to_hit_target": "목표 달성 최소 잔여 실적",
    "relief_amount": "품질관리 여유분",
    "forecast_after_provision": "전략 반영 후 예상",
    "gap_after_provision": "전략 반영 후 부족분",
    "next_close_date": "다음 마감일",
    "next_close_required": "다음 마감 누적선 필요실적",
    "risk_level": "위험등급",
    "status": "계산 상태",
    "recommended_action": "권장 조치",
    "recommended": "추천 여부",
    "risk_note": "리스크 메모",
    "comment": "설명",
    "warnings": "확인 사항",
    "date": "날짜",
    "day_name": "요일",
    "cycle_id": "마감차수",
    "cycle_start_date": "마감차수 시작일",
    "cycle_end_date": "마감차수 종료일",
    "is_completed": "완료 여부",
    "target_sum": "마감차수 목표 합계",
    "actual_sum": "마감차수 실적 합계",
    "achievement_rate": "마감차수 달성률",
    "target_cum": "누적 목표선",
    "actual_cum": "누적 실적",
    "cumulative_achievement_rate": "누적 달성률",
    "row_count": "입력 행 수",
    "original_target": "기존 일 목표",
    "uplift": "추가 배분 목표",
    "revised_target": "수정 후 일 목표",
    "cap_target": "일별 허용 상한",
    "expected_after_revision": "수정 후 예상 일 실적",
    "expected_rate": "예상 달성률",
    "allocation_weight": "배분 비중",
    "run_id": "저장 ID",
    "run_datetime": "저장 시각",
    "target_month": "대상 월",
    "as_of_date": "기준일",
    "metric": "지표",
    "strategy_id": "전략 ID",
    "final_actual": "월마감 확정 실적",
    "final_achievement_rate": "확정 달성률",
    "final_status": "확정 목표 상태",
    "cancellation_amount": "취소/철회 금액",
    "net_actual": "순 확정 실적",
    "updated_at": "확정 저장 시각",
    "forecast_error": "예측 오차",
    "abs_error": "절대 오차",
    "error_rate": "오차율",
    "signed_error_rate": "방향성 오차율",
    "over_forecast_flag": "과대 예측 여부",
    "under_forecast_flag": "과소 예측 여부",
    "sample_count": "표본 수",
    "mean_abs_error": "평균 절대 오차",
    "mean_error_rate": "평균 오차율",
    "median_error_rate": "중앙 오차율",
    "bias": "bias",
    "best_model_by_error_rate": "최저 오차 모델",
    "weighted_forecast": "Weighted Forecast",
    "weighted_forecast_amount": "Weighted Forecast",
    "confidence_lower": "Confidence Band 하단",
    "confidence_upper": "Confidence Band 상단",
    "forecast_lower": "Confidence Band 하단",
    "forecast_upper": "Confidence Band 상단",
    "lower_bound": "Confidence Band 하단",
    "upper_bound": "Confidence Band 상단",
}
METRIC_DISPLAY_LABELS = {
    "sales": "판매실적",
    "recognized": "인정실적",
}
DISPLAY_VALUE_LABELS = {
    **METRIC_DISPLAY_LABELS,
    F1_CUMULATIVE_RATE: "F1 누적 달성률 모델",
    F2_LAST_TWO_CLOSES: "F2 직전 2개 완료 마감차수 모델",
    F3_DAY_CLOSE_WEIGHTED: "F3 마감일/비마감일 가중 모델",
    P1_ALL_REMAINING: f"P1 {get_strategy_label('P1')}",
    P2_CLOSE_DAY_FOCUSED: f"P2 {get_strategy_label('P2')}",
    P3_NON_CLOSE_DAY_FOCUSED: f"P3 {get_strategy_label('P3')}",
    O1_TARGET_HOLD_BUFFER: f"O1 {get_strategy_label('O1')}",
    O2_STRETCH_TARGET_CAPTURE: f"O2 {get_strategy_label('O2')}",
    O3_QUALITY_GUARD_RELIEF: f"O3 {get_strategy_label('O3')}",
    N1_MAINTAIN_TARGET: "N1 목표 유지",
    N2_MONITOR_BUFFER: "N2 버퍼 모니터링",
    N3_QUALITY_CHECK: "N3 품질 점검",
    "UNDER_TARGET": get_status_label("UNDER_TARGET"),
    "ON_TARGET": get_status_label("ON_TARGET"),
    "OVER_TARGET": get_status_label("OVER_TARGET"),
    "UNKNOWN_TARGET_STATUS": "계산 불가",
    "PROVISION": "목표 보정",
    "OVERACHIEVEMENT": "초과달성 운영",
    "NEUTRAL": "유지/모니터링",
    "Green": "낮음",
    "Yellow": "주의",
    "Red": "높음",
    "Black": "매우 높음",
    "N/A": "해당 없음",
    "OK": "정상",
    "NO_GAP": "부족분 없음",
    "CAPACITY_LIMITED": "배분 한도 초과",
    "NOT_APPLICABLE": "적용 불가",
    "CALCULATION_ERROR": "계산 오류",
    "OVER_TARGET_MANAGED": "초과달성 관리",
    "ON_TARGET_MAINTAIN": "목표선 유지",
}
NEXT_CLOSE_REQUIRED_LABEL = "다음 마감 누적선 필요실적"
NEXT_CLOSE_REQUIRED_EXPLANATION = (
    "다음 마감 필요실적은 월 목표 부족분이 아니라, "
    "다음 마감일까지의 누적 계획선을 맞추기 위해 필요한 실적입니다."
)
SECURITY_WARNING_TEXT = (
    "운영 주의: Public Streamlit 또는 외부 공개 URL에는 실제 영업실적을 업로드하지 마세요. "
    "실데이터는 Private/사내망/권한 통제 환경에서만 사용하세요."
)
KPI_HELP_TEXTS = {
    NEXT_CLOSE_REQUIRED_LABEL: NEXT_CLOSE_REQUIRED_EXPLANATION,
}
TARGET_STATUS_OPERATION_MODE_LABELS = {
    "UNDER_TARGET": get_operation_mode("UNDER_TARGET"),
    "ON_TARGET": get_operation_mode("ON_TARGET"),
    "OVER_TARGET": get_operation_mode("OVER_TARGET"),
}
OVERACHIEVEMENT_MATRIX_LABELS = {
    "O1": get_strategy_label("O1"),
    "O2": get_strategy_label("O2"),
    "O3": get_strategy_label("O3"),
    O1_TARGET_HOLD_BUFFER: get_strategy_label("O1"),
    O2_STRETCH_TARGET_CAPTURE: get_strategy_label("O2"),
    O3_QUALITY_GUARD_RELIEF: get_strategy_label("O3"),
}
VALIDATION_MESSAGE_TRANSLATIONS = {
    "business_day_no must be in ascending order.": (
        "영업일 번호가 순서대로 정렬되어 있어야 합니다. 1, 2, 3처럼 증가하는지 확인해 주세요."
    ),
    "as_of_date must exist in the input table.": (
        "선택한 기준일이 입력표 날짜에 없습니다. 입력표에 있는 날짜로 기준일을 선택해 주세요."
    ),
    "target_daily must not contain missing or invalid values.": (
        "선택한 지표의 일 목표에 빈값 또는 숫자가 아닌 값이 있습니다. 일 목표 칸을 숫자로 입력해 주세요."
    ),
    "target_daily must not be negative.": (
        "선택한 지표의 일 목표에 음수가 있습니다. 목표는 0 이상으로 입력해 주세요."
    ),
    "monthly_target must be greater than 0.": (
        "월 전체 목표 합계가 0입니다. 일 목표를 하나 이상 입력해 주세요."
    ),
    "actual_cum must be populated through as_of_date.": (
        "기준일까지의 누적 실적에 빈값이 있습니다. 기준일 이전과 기준일의 누적 실적을 모두 입력해 주세요."
    ),
    "actual_cum after as_of_date is populated.": (
        "기준일 이후 날짜에도 누적 실적이 입력되어 있습니다. 미래 날짜의 실적이 의도한 입력인지 확인해 주세요."
    ),
    "At least one is_close_day=True row is required.": (
        "마감일로 표시된 행이 없습니다. 마감일 여부 칸에 TRUE, Y, 1 중 하나로 표시된 행을 하나 이상 입력해 주세요."
    ),
    "F2 fallback warning: fewer than two completed close days exist through as_of_date.": (
        "기준일까지 완료된 마감일이 2개 미만이라 F2 방식은 F1 방식으로 대신 계산될 수 있습니다."
    ),
    "actual_cum decreases in the input table.": (
        "누적 실적이 이전 날짜보다 줄어드는 구간이 있습니다. 차감이나 환입이 의도된 값인지 확인해 주세요."
    ),
    "actual_daily calculated from actual_cum is negative.": (
        "누적 실적 차이로 계산한 일 실적이 음수입니다. 누적 실적이 중간에 감소했는지 확인해 주세요."
    ),
    "close_type is missing; blank close_type is allowed.": (
        "마감 유형 열이 없습니다. 계산은 가능하지만 마감 구분 설명은 비어 있을 수 있습니다."
    ),
    "close_type is blank for one or more close days.": (
        "마감일로 표시된 행 중 마감 유형이 비어 있습니다. 중간마감, 월마감 같은 구분값을 입력하면 해석이 쉬워집니다."
    ),
    "date contains missing or invalid values.": (
        "날짜 칸에 비어 있거나 읽을 수 없는 값이 있습니다. YYYY-MM-DD 형식으로 입력해 주세요."
    ),
    "as_of_date must be a valid date.": (
        "기준일을 날짜로 읽을 수 없습니다. 기준일을 다시 선택하거나 YYYY-MM-DD 형식으로 입력해 주세요."
    ),
    "is_close_day contains unsupported values.": (
        "마감일 여부 칸에 읽을 수 없는 값이 있습니다. TRUE/FALSE, Y/N, 1/0 중 하나로 입력해 주세요."
    ),
    "Calculation unavailable: as_of_date is not present in the input rows.": (
        "선택한 기준일을 입력표에서 찾을 수 없어 계산할 수 없습니다. 입력표에 있는 날짜로 기준일을 선택해 주세요."
    ),
    "Calculation unavailable: as_of_date is not present in input rows.": (
        "선택한 기준일을 입력표에서 찾을 수 없어 계산할 수 없습니다. 입력표에 있는 날짜로 기준일을 선택해 주세요."
    ),
    "Calculation unavailable: actual cumulative value is missing at as_of_date.": (
        "기준일의 누적 실적이 비어 있어 계산할 수 없습니다. 기준일 행의 누적 실적을 입력해 주세요."
    ),
    "Calculation unavailable: monthly_target is zero.": (
        "월 전체 목표가 0이라 계산할 수 없습니다. 일 목표를 입력해 주세요."
    ),
    "Calculation unavailable: current_target_cum is zero.": (
        "기준일까지의 누적 목표가 0이라 달성률 계산이 어렵습니다. 기준일 이전 일 목표를 확인해 주세요."
    ),
    "Calculation unavailable: uplift_effective_rate is zero.": (
        "목표 상향을 반영할 수 있는 비율이 0이라 보정 계산을 할 수 없습니다. 보정 전략이나 상한 설정을 확인해 주세요."
    ),
    "No next close day is present after as_of_date.": (
        "기준일 이후에 다음 마감일이 없습니다. 다음 마감 누적선 필요실적은 계산하지 않습니다."
    ),
    "Negative actual_daily values are present before or at as_of_date.": (
        "기준일까지의 일 실적 중 음수로 계산되는 날이 있습니다. 누적 실적 감소가 의도된 값인지 확인해 주세요."
    ),
    "F2_LAST_TWO_CLOSES fallback to F1_CUMULATIVE_RATE: fewer than two completed close cycles are available.": (
        "F2는 기준일까지 완료된 마감차수가 2개 미만이라 F1 방식으로 대신 계산했습니다."
    ),
    "F2_LAST_TWO_CLOSES fallback to F1_CUMULATIVE_RATE: last two completed close cycles have zero target.": (
        "F2는 최근 2개 완료 마감차수의 목표 합계가 0이라 F1 방식으로 대신 계산했습니다."
    ),
    "F3_DAY_CLOSE_WEIGHTED fallback: close-day and non-close-day historical targets are both required.": (
        "F3는 마감일과 비마감일의 과거 목표가 모두 필요하지만 일부가 부족해 F2 방식으로 대신 계산했습니다."
    ),
}
VALIDATION_TERM_REPLACEMENTS = {
    "F1_CUMULATIVE_RATE": "F1 누적 달성률 방식",
    "F2_LAST_TWO_CLOSES": "F2 직전 2개 마감차수 방식",
    "F3_DAY_CLOSE_WEIGHTED": "F3 마감일/비마감일 가중 방식",
    "business_day_no": "영업일 번호",
    "is_close_day": "마감일 여부",
    "close_type": "마감 유형",
    "target_daily": "일 목표",
    "actual_daily": "일 실적",
    "actual_cum": "누적 실적",
    "monthly_target": "월 전체 목표",
    "current_target_cum": "기준일까지의 누적 목표",
    "current_actual_cum": "기준일까지의 누적 실적",
    "as_of_date": "기준일",
    "uplift_effective_rate": "목표 상향 반영률",
    "fallback": "대체 계산",
}
VISUAL_METRIC_DEFINITIONS = {
    "daily_forecast_cum": {
        "label": "주간 누적 예상",
        "unit": "억원",
        "definition": "기준일까지는 확정 누적 실적, 기준일 이후는 각 시나리오의 예상 실적을 주간 말 기준으로 누적한 값입니다.",
    },
    "daily_target_cum": {
        "label": "주간 누적 목표선",
        "unit": "억원",
        "definition": "입력표의 일 목표를 날짜 순서대로 누적한 뒤 주간 말 기준으로 표시한 공식 계획선입니다.",
    },
    "daily_achievement_rate": {
        "label": "주간 월 목표 달성률",
        "unit": "%",
        "definition": "각 주의 누적 실적 또는 누적 예상 실적을 공식 월 목표로 나눈 비율입니다.",
    },
    "forecast_amount": {
        "label": "월말 예상 실적(보정 전)",
        "unit": "억원",
        "definition": "선택한 F 방식만 적용해 계산한 월말 예상 실적입니다. 잔여 목표를 다시 배분하기 전 숫자입니다.",
    },
    "forecast_after_provision": {
        "label": "월말 예상 실적(보정 후)",
        "unit": "억원",
        "definition": "F 방식의 예상에 P 보정 또는 O/N 운영전략을 반영한 뒤의 월말 예상 실적입니다.",
    },
    "gap_to_target": {
        "label": "목표 미달 예상분",
        "unit": "억원",
        "definition": "월 목표보다 예상 실적이 부족한 금액입니다. 목표 이상이면 0으로 표시합니다.",
    },
    "target_variance": {
        "label": "목표 대비 차이",
        "unit": "억원",
        "definition": "월말 예상 실적에서 월 목표를 뺀 값입니다. 양수면 초과 예상, 음수면 미달 예상입니다.",
    },
    "surplus_to_target": {
        "label": "초과 예상분",
        "unit": "억원",
        "definition": "월 목표를 넘길 것으로 예상되는 금액입니다. 미달이면 0으로 표시합니다.",
    },
    "target_cum": {
        "label": "누적 목표선",
        "unit": "억원",
        "definition": "마감차수별 목표 합계를 입력 row 순서대로 누적한 목표선입니다.",
    },
    "actual_cum": {
        "label": "누적 실적",
        "unit": "억원",
        "definition": "마감차수별 실적 합계를 입력 row 순서대로 누적한 확정/입력 실적 흐름입니다.",
    },
    "cumulative_achievement_rate": {
        "label": "누적 달성률",
        "unit": "%",
        "definition": "누적 실적을 누적 목표선으로 나눈 비율입니다. 마감일 판단은 입력의 is_close_day 기반 차수만 사용합니다.",
    },
    "required_uplift": {
        "label": "목표 달성에 필요한 추가 실적",
        "unit": "억원",
        "definition": "목표 미달 시 월 목표를 맞추기 위해 기준일 이후 잔여 기간에 추가로 만들어야 하는 실적 규모입니다.",
    },
    "gap_after_provision": {
        "label": "보정 후 목표 차이(+부족/-초과)",
        "unit": "억원",
        "definition": "보정 후에도 월 목표와 예상 실적 사이에 남은 부족분입니다. 초과달성 전략에서는 부족분 대신 초과 예상분과 버퍼를 함께 확인합니다.",
    },
    "monthly_target": {
        "label": "공식 월 목표",
        "unit": "억원",
        "definition": "입력표의 일 목표를 모두 더한 공식 월 목표입니다.",
    },
    "revised_monthly_target": {
        "label": "운영전략상 월 목표 수준",
        "unit": "억원",
        "definition": "선택한 운영전략에서 관리할 월 목표 수준입니다. O2는 초과 예상분 일부를 반영해 상향될 수 있습니다.",
    },
    "stretch_uplift": {
        "label": "Stretch 전환분",
        "unit": "억원",
        "definition": "O2 전략에서 초과 예상분 중 추가 성장 목표로 전환한 금액입니다.",
    },
    "remaining_surplus_buffer": {
        "label": "잔여 안전버퍼",
        "unit": "억원",
        "definition": "초과 예상분 중 취소, 철회, 미결제, 실적 조정 리스크에 대비해 남겨 두는 금액입니다.",
    },
    "minimum_remaining_to_hit_target": {
        "label": "목표 달성까지 필요한 최소 잔여 실적",
        "unit": "억원",
        "definition": "현재 누적 실적 기준으로 공식 월 목표를 맞추는 데 필요한 최소 잔여 실적입니다.",
    },
    "relief_amount": {
        "label": "품질관리 여유분",
        "unit": "억원",
        "definition": "O3 전략에서 무리한 추가 압박 대신 계약 품질과 결제완료율 관리에 활용할 수 있는 여유분입니다.",
    },
    "original_target": {
        "label": "기존 일 목표",
        "unit": "억원",
        "definition": "입력표에 원래 들어 있던 해당 날짜의 목표입니다.",
    },
    "uplift": {
        "label": "추가 배분 목표",
        "unit": "억원",
        "definition": "목표 달성을 위해 해당 날짜에 새로 더 배분한 목표입니다.",
    },
    "revised_target": {
        "label": "수정 후 일 목표",
        "unit": "억원",
        "definition": "기존 일 목표에 추가 배분 목표를 더한 값입니다.",
    },
    "cap_target": {
        "label": "일별 허용 상한",
        "unit": "억원",
        "definition": "설정한 cap rate를 반영했을 때 해당 날짜에 배분할 수 있는 최대 목표입니다.",
    },
    "expected_after_revision": {
        "label": "수정 후 예상 일 실적",
        "unit": "억원",
        "definition": "수정 후 일 목표와 예상 달성률을 바탕으로 계산한 해당 날짜의 예상 실적입니다.",
    },
    "target_sum": {
        "label": "마감차수 목표 합계",
        "unit": "억원",
        "definition": "해당 마감차수 기간에 포함된 일 목표를 모두 더한 값입니다.",
    },
    "actual_sum": {
        "label": "마감차수 실적 합계",
        "unit": "억원",
        "definition": "해당 마감차수 기간에 실제로 발생한 실적을 모두 더한 값입니다.",
    },
    "achievement_rate": {
        "label": "마감차수 달성률",
        "unit": "%",
        "definition": "마감차수 실적 합계를 마감차수 목표 합계로 나눈 비율입니다.",
    },
}
VISUAL_READING_GUIDES = {
    "scenario_daily_progress": {
        "title": "주간 목표 달성률 전망",
        "steps": (
            "100% 기준선을 중심으로 기준일까지의 확정 달성률과 이후 주간 시나리오 달성률을 비교합니다.",
            "마우스 휠과 드래그로 특정 주간 구간을 확대/축소해 마감 전후 변화를 봅니다.",
            "선 끝 라벨의 최종 달성률과 목표 대비 차이를 보고 우위 시나리오를 고릅니다.",
        ),
        "decision": "100% 위로 안정적으로 올라서고, 같은 주차에서 더 높은 달성률을 보이는 조합이 시각적으로 우위입니다.",
    },
    "scenario_amount": {
        "title": "월말 예상 실적 비교",
        "steps": (
            "먼저 공식 월 목표를 기준선으로 잡습니다.",
            "월말 예상 실적(보정 전)이 목표보다 낮은지, 높은지 확인합니다.",
            "전략 반영 후 예상이 목표선을 회복하는지 보고 선택 전략의 효과를 판단합니다.",
        ),
        "decision": "전략 반영 후 예상이 목표보다 낮으면 보정 강도가 부족하고, 목표보다 높으면 초과분 관리 전략을 함께 봅니다.",
    },
    "scenario_target_position": {
        "title": "목표선 대비 예상 실적",
        "steps": (
            "빨간 기준선은 공식 월 목표입니다.",
            "막대 끝이 기준선 오른쪽이면 목표 초과, 왼쪽이면 목표 미달입니다.",
            "선택 시나리오는 진한 테두리로 표시해 다른 조합과 바로 비교합니다.",
        ),
        "decision": "기준선을 넘기는 조합 중 위험등급과 운영 부담이 낮은 조합을 우선 검토합니다.",
    },
    "scenario_gap_position": {
        "title": "부족/초과 금액",
        "steps": (
            "0선을 기준으로 오른쪽은 초과 예상, 왼쪽은 미달 예상입니다.",
            "막대 길이가 길수록 목표선에서 더 멀리 떨어진 조합입니다.",
            "미달 막대는 보정 필요 규모, 초과 막대는 버퍼 또는 Stretch 후보 규모로 봅니다.",
        ),
        "decision": "미달 폭이 큰 조합은 실행 부담이 크고, 초과 폭이 큰 조합은 초과분 관리 전략이 중요합니다.",
    },
    "scenario_heatmap": {
        "title": "시나리오 조합 지도",
        "steps": (
            "행은 예측모델(F1~F3), 열은 운영전략(P/O/N)을 뜻합니다.",
            "붉은 칸은 목표 미달, 초록 칸은 목표 초과 방향입니다.",
            "칸 안 숫자는 월 목표 대비 부족/초과 금액입니다.",
        ),
        "decision": "같은 색이 반복되면 모델보다 시장 흐름 영향이 크고, 행마다 색이 달라지면 예측모델 선택 민감도가 큽니다.",
    },
    "scenario_status": {
        "title": "목표 상태 확인",
        "steps": (
            "목표 대비 차이는 양수면 초과 예상, 음수면 미달 예상으로 읽습니다.",
            "미달 구간은 목표 미달 예상분과 목표 달성에 필요한 추가 실적을 함께 봅니다.",
            "초과 구간은 초과 예상분이 운영상 버퍼인지, 상향 목표로 전환할 수 있는지 확인합니다.",
        ),
        "decision": "부족분이 큰 시나리오는 실행 부담이 높고, 초과분이 큰 시나리오는 품질 리스크와 Stretch 전환 가능성을 검토합니다.",
    },
    "scenario_matrix": {
        "title": "시나리오 숫자표",
        "steps": (
            "행은 예측모델(F1~F3), 열은 운영전략(P/O/N)을 뜻합니다.",
            "먼저 같은 행 안에서 전략별 월말 예상 실적 차이를 비교합니다.",
            "그다음 같은 열 안에서 예측모델별 민감도를 비교해 보수적/공격적 전망 차이를 봅니다.",
        ),
        "decision": "숫자가 가장 큰 조합보다, KPI 위험등급과 실행 가능한 보정 전략까지 같이 맞는 조합을 선택합니다.",
    },
    "target_allocation": {
        "title": "일자별 목표 변화",
        "steps": (
            "기존 일 목표를 해당 날짜의 원래 계획선으로 봅니다.",
            "추가 배분 목표는 부족분을 메우기 위해 새로 얹은 부담입니다.",
            "수정 후 일 목표는 기존 목표와 추가 배분이 반영된 최종 관리 목표입니다.",
        ),
        "decision": "수정 후 일 목표가 특정 날짜에 과도하게 몰리면 배분 전략이나 상한 설정을 재검토합니다.",
    },
    "target_uplift": {
        "title": "일자별 추가 부담",
        "steps": (
            "막대는 목표 달성을 위해 해당 날짜에 추가로 얹힌 목표입니다.",
            "마감일에 막대가 몰리면 마감 당일 의존도가 높은 계획입니다.",
            "일반일에도 막대가 넓게 퍼지면 남은 기간 전체에 부담을 분산한 계획입니다.",
        ),
        "decision": "특정 날짜의 추가 부담이 과도하면 P2/P3 같은 배분 전략을 바꿔 비교합니다.",
    },
    "target_stack": {
        "title": "기존 목표와 추가 배분",
        "steps": (
            "회색은 기존 일 목표, 주황색은 새로 더한 추가 배분 목표입니다.",
            "막대 전체 높이가 수정 후 일 목표입니다.",
            "상한선에 가까운 날짜가 많을수록 실행 여력이 부족하다는 신호입니다.",
        ),
        "decision": "추가 배분이 상한선 근처에 반복되면 목표 상향 강도나 잔여 기간 운영 계획을 재검토합니다.",
    },
    "target_cap": {
        "title": "상한과 예상 실적",
        "steps": (
            "일별 허용 상한은 해당 날짜에 배분 가능한 최대 목표선입니다.",
            "수정 후 예상 일 실적이 상한보다 낮은지 확인해 목표가 현실적인지 봅니다.",
            "상한에 계속 붙어 있으면 잔여 기간 전체의 실행 여력이 부족하다는 신호입니다.",
        ),
        "decision": "상한을 자주 넘거나 근접하면 마감일/비마감일 배분 방식 또는 월 목표 상향 부담을 다시 봅니다.",
    },
    "close_cycle_amount": {
        "title": "마감차수별 목표와 실적",
        "steps": (
            "각 마감차수의 목표 합계를 기준으로 잡습니다.",
            "실적 합계가 목표 합계보다 높으면 해당 차수는 초과 달성, 낮으면 미달입니다.",
            "미달 차수가 반복되는지, 특정 마감차수에서만 흔들리는지 흐름을 봅니다.",
        ),
        "decision": "최근 마감차수에서 미달이 커질수록 남은 기간 보정 필요성이 높아집니다.",
    },
    "close_cycle_rate": {
        "title": "마감차수별 달성률",
        "steps": (
            "100%를 기준선으로 두고 각 마감차수가 목표를 넘겼는지 확인합니다.",
            "달성률이 상승 중인지 하락 중인지 방향을 봅니다.",
            "최근 2개 마감차수의 흐름은 F2 예측모델의 근거가 되므로 특히 따로 봅니다.",
        ),
        "decision": "최근 달성률이 100% 아래로 내려가면 월말 예상보다 실제 마감 리스크를 더 보수적으로 봅니다.",
    },
    "strategy_amount": {
        "title": "전략별 월 목표 수준",
        "steps": (
            "공식 월 목표와 월말 예상 실적의 간격을 먼저 봅니다.",
            "운영전략상 월 목표 수준이 공식 목표와 같은지, 상향됐는지 확인합니다.",
            "상향된 목표가 예상 실적 안에서 감당되는지 비교합니다.",
        ),
        "decision": "상향 목표가 예상 실적보다 높으면 공격적인 전략이고, 공식 목표를 유지하면 방어적인 전략입니다.",
    },
    "strategy_buffer": {
        "title": "전략별 초과/완화 수준",
        "steps": (
            "초과 예상분이 있다면 안전버퍼, Stretch 전환분, 품질관리 여유분으로 나뉘는 구조를 봅니다.",
            "목표 미달이면 목표 달성에 필요한 추가 실적이 얼마나 남는지 봅니다.",
            "각 전략이 숫자를 키우는 전략인지, 리스크를 줄이는 전략인지 구분합니다.",
        ),
        "decision": "초과분이 크면 성장 전환과 버퍼 유지의 균형을, 부족분이 크면 실제 추가 실행 가능성을 우선 판단합니다.",
    },
}


def main() -> None:
    """Render the Streamlit app."""
    if st is None:
        raise RuntimeError("Streamlit is required to run app.py. Install requirements.txt first.")

    st.set_page_config(page_title="마감 페이스 체크", layout="wide")
    _inject_app_styles()
    base_config = load_model_config()
    active_page = get_current_page(st)
    audit_readonly = _is_audit_readonly_mode(st)
    page_context = _build_page_context(base_config)
    page_context["audit_readonly"] = audit_readonly
    page_context = _with_page_callbacks(page_context, base_config)

    _render_same_window_side_nav(st, active_page)
    _render_same_window_top_status(active_page, _top_nav_meta(page_context))
    if audit_readonly:
        _render_audit_readonly_banner()
    st.markdown('<main class="page-shell">', unsafe_allow_html=True)
    if active_page == "home":
        _render_home_workbench_page(page_context)
    else:
        render_page(active_page, page_context)
    st.markdown("</main>", unsafe_allow_html=True)


def _with_page_callbacks(
    context: dict[str, Any],
    base_config: dict[str, Any],
) -> dict[str, Any]:
    context = dict(context)
    context.update(
        {
            "render_input_page": lambda: _render_input_data_page(
                base_config,
                audit_readonly=bool(context.get("audit_readonly", False)),
            ),
            "render_forecast_strategy_page": lambda: _render_forecast_strategy_detail_page(context),
            "render_forecast_page": lambda: _render_forecast_detail_page(context),
            "render_scenarios_page": lambda: _render_scenarios_detail_page(context),
            "render_report_page": lambda: _render_report_detail_page(context),
            "render_history_page": lambda: _render_history_detail_page(context),
            "render_excel_page": lambda: _render_excel_detail_page(context),
            "render_audit_page": lambda: _render_audit_detail_page(context),
        }
    )
    return context


def _build_page_context(base_config: dict[str, Any]) -> dict[str, Any]:
    df, source_label = _get_current_input_state()
    historical_df, historical_source_label = _get_historical_input_state()
    metric, as_of_date, forecast_choice, provision_choice, config = _normalize_app_settings(
        df,
        base_config,
    )

    results = calculate_validated_results(df, as_of_date, metric, config)
    validation_result = results["validation"]
    scenario_df = _as_dataframe(results.get("scenario_df"))
    next_close_result = dict(results.get("next_close_result") or {})
    close_cycle_df = _as_dataframe(results.get("close_cycle_df"))

    selected_scenario_id = ""
    selected_row = pd.Series(dtype=object)
    forecast_result: dict[str, object] = {}
    provision_result: dict[str, object] = {}
    revised_targets_df = pd.DataFrame()
    historical_context: dict[str, object] = {"has_data": False}
    report_text = "입력 후 계산됩니다."
    summary_dict: dict[str, Any] = {}

    if not validation_result.get("errors") and not scenario_df.empty:
        selected_scenario_id = _default_selected_scenario_id(
            scenario_df,
            forecast_choice,
            provision_choice,
        )
        selected_row = _selected_scenario_row(scenario_df, selected_scenario_id)
        forecast_result, provision_result = run_selected_scenario_detail(
            df,
            as_of_date,
            metric,
            selected_scenario_id,
            config,
        )
        revised_targets_df = _as_dataframe(provision_result.get("allocation_by_day"))
        historical_context = build_historical_context(
            historical_df,
            df,
            as_of_date,
            metric,
            validation_result,
            historical_source_label,
        )
        report_text = build_daily_report_text(
            scenario_df,
            next_close_result,
            selected_scenario_id=selected_scenario_id,
        )
        summary_dict = build_summary_dict(
            validation_result,
            selected_row,
            next_close_result,
            metric,
            as_of_date,
        )

    return {
        "df": df,
        "source_label": source_label,
        "historical_df": historical_df,
        "historical_source_label": historical_source_label,
        "metric": metric,
        "as_of_date": as_of_date,
        "forecast_choice": forecast_choice,
        "provision_choice": provision_choice,
        "config": config,
        "validation_result": validation_result,
        "scenario_df": scenario_df,
        "next_close_result": next_close_result,
        "close_cycle_df": close_cycle_df,
        "selected_scenario_id": selected_scenario_id,
        "selected_row": selected_row,
        "forecast_result": forecast_result,
        "provision_result": provision_result,
        "revised_targets_df": revised_targets_df,
        "historical_context": historical_context,
        "report_text": report_text,
        "summary_dict": summary_dict,
        "report_name": _expected_report_name(metric, as_of_date),
    }


def _get_current_input_state() -> tuple[pd.DataFrame, str]:
    stored_df = st.session_state.get(CURRENT_INPUT_DF_SESSION_KEY)
    stored_source = st.session_state.get(CURRENT_INPUT_SOURCE_SESSION_KEY)
    if isinstance(stored_df, pd.DataFrame):
        return stored_df.copy(), str(stored_source or SAMPLE_INPUT_SOURCE_LABEL)

    df, source_info = load_sample_with_source("current_input")
    _warn_operator_sample_fallback("현재 입력 샘플", source_info)
    if source_info.get("source") in {"operator", "github"}:
        source_label = OPERATOR_SAMPLE_SOURCE_LABEL
        prepared_df = df.copy()
    else:
        source_label = SAMPLE_INPUT_SOURCE_LABEL
        saved_actuals = _load_saved_actuals_for_ui()
        prepared_df, default_source = apply_latest_upload_policy(df, source_label, saved_actuals)
        if default_source == "saved":
            source_label = SAVED_ACTUALS_SOURCE_LABEL
    _store_current_input_state(prepared_df, source_label)
    return prepared_df, source_label


def _store_current_input_state(df: pd.DataFrame, source_label: str) -> None:
    st.session_state[CURRENT_INPUT_DF_SESSION_KEY] = df.copy()
    st.session_state[CURRENT_INPUT_SOURCE_SESSION_KEY] = source_label


def _get_historical_input_state() -> tuple[pd.DataFrame, str]:
    stored_df = st.session_state.get(HISTORICAL_INPUT_DF_SESSION_KEY)
    stored_source = st.session_state.get(HISTORICAL_INPUT_SOURCE_SESSION_KEY)
    sample_disabled = bool(st.session_state.get(HISTORICAL_SAMPLE_DISABLED_SESSION_KEY, False))
    if isinstance(stored_df, pd.DataFrame) and (not stored_df.empty or stored_source or sample_disabled):
        return stored_df.copy(), str(stored_source or "")
    if sample_disabled:
        return pd.DataFrame(), ""
    try:
        historical_df, source_info = load_sample_with_source("historical_input")
    except Exception:  # noqa: BLE001 - keep the page usable if the bundled sample is missing.
        return pd.DataFrame(), ""
    _warn_operator_sample_fallback("과거 샘플", source_info)
    source_label = (
        OPERATOR_SAMPLE_SOURCE_LABEL
        if source_info.get("source") in {"operator", "github"}
        else HISTORICAL_SAMPLE_INPUT_SOURCE_LABEL
    )
    _store_historical_input_state(historical_df, source_label)
    return historical_df, source_label


def _store_historical_input_state(df: pd.DataFrame, source_label: str) -> None:
    st.session_state[HISTORICAL_INPUT_DF_SESSION_KEY] = df.copy()
    st.session_state[HISTORICAL_INPUT_SOURCE_SESSION_KEY] = source_label


def _normalize_app_settings(
    df: pd.DataFrame,
    base_config: dict[str, Any],
) -> tuple[str, pd.Timestamp, str, str, dict[str, Any]]:
    metric = str(st.session_state.get(PACE_METRIC_SESSION_KEY, "sales"))
    if metric not in {"sales", "recognized"}:
        metric = "sales"

    dates = pd.to_datetime(df["date"], errors="raise")
    date_values = [timestamp.date() for timestamp in dates]
    default_date = default_as_of_date(df, metric).date()
    default_token = f"{metric}:{default_date.isoformat()}"
    if st.session_state.get(PACE_AS_OF_DATE_DEFAULT_SESSION_KEY) != default_token:
        st.session_state[PACE_AS_OF_DATE_DEFAULT_SESSION_KEY] = default_token
        st.session_state[PACE_AS_OF_DATE_SESSION_KEY] = default_date
    requested_date = st.session_state.get(PACE_AS_OF_DATE_SESSION_KEY, default_date)
    try:
        as_of_date = pd.Timestamp(requested_date).date()
    except Exception:  # noqa: BLE001 - UI state fallback only.
        as_of_date = default_date
    if as_of_date not in date_values:
        as_of_date = default_date if default_date in date_values else date_values[0]

    forecast_choice = str(st.session_state.get(PACE_FORECAST_CHOICE_SESSION_KEY, COMPARE_LABEL))
    if forecast_choice not in {"F1", "F2", "F3", COMPARE_LABEL}:
        forecast_choice = COMPARE_LABEL

    provision_choice = str(st.session_state.get(PACE_STRATEGY_CHOICE_SESSION_KEY, COMPARE_LABEL))
    if provision_choice not in {"P1", "P2", "P3", "O1", "O2", "O3", "N1", "N2", "N3", COMPARE_LABEL}:
        provision_choice = COMPARE_LABEL

    close_day_cap_rate = _session_float(
        PACE_CLOSE_CAP_SESSION_KEY,
        base_config.get("close_day_cap_rate", 1.30),
    )
    non_close_day_cap_rate = _session_float(
        PACE_NON_CLOSE_CAP_SESSION_KEY,
        base_config.get("non_close_day_cap_rate", 1.50),
    )
    config = build_runtime_config(
        base_config,
        close_day_cap_rate,
        non_close_day_cap_rate,
    )

    st.session_state[PACE_METRIC_SESSION_KEY] = metric
    st.session_state[PACE_AS_OF_DATE_SESSION_KEY] = as_of_date
    st.session_state[PACE_FORECAST_CHOICE_SESSION_KEY] = forecast_choice
    st.session_state[PACE_STRATEGY_CHOICE_SESSION_KEY] = provision_choice
    st.session_state[PACE_CLOSE_CAP_SESSION_KEY] = close_day_cap_rate
    st.session_state[PACE_NON_CLOSE_CAP_SESSION_KEY] = non_close_day_cap_rate
    return metric, pd.Timestamp(as_of_date), forecast_choice, provision_choice, config


def _session_float(key: str, default: object) -> float:
    try:
        return float(st.session_state.get(key, default))
    except (TypeError, ValueError):
        return float(default)


def _default_selected_scenario_id(
    scenario_df: pd.DataFrame,
    forecast_choice: str,
    provision_choice: str,
) -> str:
    if scenario_df.empty or "scenario_id" not in scenario_df.columns:
        return ""

    candidates = _filter_scenarios(scenario_df, forecast_choice, provision_choice)
    if candidates.empty:
        candidates = scenario_df
    scenario_ids = candidates["scenario_id"].astype(str).tolist()
    stored = str(st.session_state.get(PACE_SELECTED_SCENARIO_SESSION_KEY, ""))
    if stored in scenario_ids:
        return stored
    selected = scenario_ids[0]
    st.session_state[PACE_SELECTED_SCENARIO_SESSION_KEY] = selected
    return selected


def _top_nav_meta(context: Mapping[str, Any]) -> dict[str, object]:
    df = _as_dataframe(context.get("df"))
    as_of_date = context.get("as_of_date")
    selected_row = _as_series(context.get("selected_row"))
    audit_readonly = bool(context.get("audit_readonly", False))
    header_context = _pace_header_context(df, as_of_date)
    target_month = "입력 후 계산됩니다"
    try:
        target_month = pd.Timestamp(as_of_date).strftime("%Y-%m")
    except Exception:  # noqa: BLE001 - display fallback only.
        pass
    return {
        "target_month": target_month,
        "business_day": (
            f"{header_context.get('current_business_day_no')} / "
            f"{header_context.get('total_business_days')}"
        ),
        "close_day": header_context.get("close_day_label", "입력 후 계산됩니다"),
        "operation_mode": (
            "읽기 전용 감리 모드"
            if audit_readonly
            else status_label(selected_row.get("target_status")) or "로컬 운영"
        ),
    }


def _render_same_window_side_nav(st_module: Any, active_page: str) -> None:
    """Render primary navigation with Streamlit buttons and query-param routing."""
    with st_module.sidebar:
        st_module.markdown(
            '<div class="nav-rail nav-rail--buttons">'
            '<div class="nav-rail__title">페이지 이동</div>'
            "</div>",
            unsafe_allow_html=True,
        )
        for page_key, definition in PAGE_DEFINITIONS.items():
            is_active = page_key == active_page
            label = definition["title"]
            button_label = f"현재 · {label}" if is_active else str(label)
            if st_module.button(
                button_label,
                key=f"same_window_nav_{page_key}",
                use_container_width=True,
                type="primary" if is_active else "secondary",
            ):
                _navigate_same_window(st_module, page_key)
        st_module.markdown(
            '<section class="security-warning-block">'
            '<div class="security-warning-block__label">운영 보안</div>'
            f"<strong>{escape(SECURITY_WARNING_TEXT)}</strong>"
            "</section>",
            unsafe_allow_html=True,
        )


def _render_same_window_top_status(active_page: str, meta: Mapping[str, object]) -> None:
    """Render status-only top navigation chrome without internal links."""
    meta_items = (
        ("기준월", meta.get("target_month", "입력 후 계산됩니다.")),
        ("영업일", meta.get("business_day", "입력 후 계산됩니다.")),
        ("마감일", meta.get("close_day", "입력 후 계산됩니다.")),
        ("운영모드", meta.get("operation_mode", "로컬 운영")),
    )
    pills = "".join(
        f'<span class="pace-pill{primary}">{escape(label)}: {escape(str(value))}</span>'
        for index, (label, value) in enumerate(meta_items)
        for primary in (" is-primary" if index == 0 else "",)
    )
    st.markdown(
        '<section class="same-window-top-status">'
        '<div class="same-window-top-status__brand">'
        '<span class="pace-brand-mark"></span>'
        '<span><strong>마감 페이스 체크</strong><small>한 창 안에서 페이지 이동</small></span>'
        "</div>"
        f'<div class="same-window-top-status__page">{escape(page_title(active_page))}</div>'
        f'<div class="same-window-top-status__meta">{pills}</div>'
        "</section>",
        unsafe_allow_html=True,
    )


def _navigate_same_window(st_module: Any, page_key: object) -> None:
    safe_page = validate_page_key(page_key)
    st_module.session_state["pace_current_page"] = safe_page
    audit_readonly = _is_audit_readonly_mode(st_module)
    try:
        st_module.query_params["page"] = safe_page
        if audit_readonly:
            st_module.query_params[AUDIT_READONLY_QUERY_PARAM] = "1"
    except Exception:  # noqa: BLE001 - Streamlit compatibility only.
        params = {"page": safe_page}
        if audit_readonly:
            params[AUDIT_READONLY_QUERY_PARAM] = "1"
        st_module.experimental_set_query_params(**params)
    _rerun_streamlit(st_module)


def _rerun_streamlit(st_module: Any) -> None:
    if hasattr(st_module, "rerun"):
        st_module.rerun()
    else:  # pragma: no cover - compatibility for older Streamlit runtimes.
        st_module.experimental_rerun()


def _is_audit_readonly_mode(st_module: Any) -> bool:
    """Return True when the current request asks for read-only audit mode."""
    try:
        raw_value = st_module.query_params.get(AUDIT_READONLY_QUERY_PARAM, "")
    except Exception:  # noqa: BLE001 - Streamlit compatibility only.
        try:
            raw_value = st_module.experimental_get_query_params().get(
                AUDIT_READONLY_QUERY_PARAM,
                "",
            )
        except Exception:  # noqa: BLE001 - missing query API in tests/fakes.
            raw_value = ""

    if isinstance(raw_value, (list, tuple)):
        raw_value = raw_value[-1] if raw_value else ""
    return str(raw_value).strip().lower() in AUDIT_READONLY_TRUE_VALUES


def _render_audit_readonly_banner() -> None:
    st.info(
        "읽기 전용 감리 모드: 화면 조회, 페이지 이동, 캡처 중에는 "
        "saved_actuals 저장, 예측 이력 저장, Excel 재생성을 비활성화합니다."
    )


def _expected_report_name(metric: str, as_of_date: object) -> str:
    date_token = pd.Timestamp(as_of_date).strftime("%Y%m%d")
    return f"daily_report_{metric}_{date_token}.xlsx"


def _render_home_workbench_page(context: Mapping[str, Any]) -> None:
    """Render the A Workbench Compact home page with visual projection first."""
    df = _as_dataframe(context.get("df"))
    selected_row = _as_series(context.get("selected_row"))
    validation_result = dict(context.get("validation_result") or {})
    scenario_df = _as_dataframe(context.get("scenario_df"))
    next_close_result = dict(context.get("next_close_result") or {})
    report_text = str(context.get("report_text") or "입력 후 계산합니다.")
    target_status = selected_row.get("target_status")

    st.markdown(
        """
        <section class="workbench-shell top-status-bar">
          <div class="page-header-compact">
            <div>
              <div class="page-header-compact__eyebrow">오늘의 마감 보드</div>
              <h1>달성 추이 및 월말 예측 구간</h1>
              <p>현재 누적 실적, 목표선, 예상 도착 구간을 첫 화면에서 바로 확인합니다.</p>
            </div>
          </div>
        </section>
        """,
        unsafe_allow_html=True,
    )
    _render_month_close_status_panel(
        context,
        selected_row,
        validation_result,
        next_close_result,
    )

    chart_col, decision_col = st.columns([0.68, 0.32], gap="large")
    with chart_col:
        _render_projection_chart_card(context)
    with decision_col:
        _render_home_decision_panel(selected_row, validation_result, next_close_result)

    _render_home_status_facts(validation_result, selected_row, next_close_result)

    _render_home_scenario_summary(
        scenario_df,
        str(context.get("selected_scenario_id") or ""),
        target_status,
    )
    _render_home_overachievement_strategy_summary(
        scenario_df,
        str(context.get("selected_scenario_id") or ""),
        target_status,
    )

    lower_cols = st.columns([0.62, 0.38], gap="large")
    with lower_cols[0]:
        preview = report_text.strip()
        if len(preview) > 620:
            preview = preview[:620].rstrip() + "\n..."
        st.markdown(
            render_section_header(
                "보고 메모 preview",
                "보고문 원문을 짧게 확인하고 상세 페이지에서 전체 문안을 검토합니다.",
            ),
            unsafe_allow_html=True,
        )
        st.markdown(render_report_card(preview or "예측 계산 후 보고 메모가 표시됩니다."), unsafe_allow_html=True)
    with lower_cols[1]:
        st.markdown(
            render_section_header(
                "Excel 공유 readiness",
                "outputs/latest 기준 최신 공유본 안내와 archive_invalid 제외 원칙을 확인합니다.",
            ),
            unsafe_allow_html=True,
        )
        st.markdown(
            render_download_card(str(context.get("report_name") or "예측 계산 후 생성됩니다.")),
            unsafe_allow_html=True,
        )


def _render_month_close_status_panel(
    context: Mapping[str, Any],
    selected_row: pd.Series,
    validation_result: Mapping[str, Any],
    next_close_result: Mapping[str, Any],
) -> None:
    df = _as_dataframe(context.get("df"))
    as_of_date = context.get("as_of_date")
    target_status = selected_row.get("target_status")
    scenario_id = str(selected_row.get("scenario_id") or "")
    strategy_code = get_strategy_code(scenario_id or selected_row.get("provision_strategy"))
    strategy_label = (
        f"{strategy_code} {get_strategy_label(strategy_code)}"
        if strategy_code
        else "예측 계산 후 표시"
    )
    current_day_no = _projection_current_day_no(df, as_of_date)
    total_business_days = _as_float(df["business_day_no"].max()) if "business_day_no" in df else float("nan")
    progress_value = (
        f"{int(current_day_no)} / {int(total_business_days)}"
        if math.isfinite(current_day_no) and math.isfinite(total_business_days)
        else "입력 후 계산"
    )
    basis_month = (
        pd.Timestamp(as_of_date).strftime("%Y-%m")
        if not _is_missing(as_of_date)
        else "입력 후 계산"
    )
    status_text = get_status_label(target_status)
    status_tone = {
        "UNDER_TARGET": "under",
        "ON_TARGET": "on",
        "OVER_TARGET": "over",
    }.get(str(target_status), "on")
    items = (
        ("기준월", basis_month),
        ("영업일 진행", progress_value),
        (get_metric_label("target_status"), status_text),
        (get_metric_label("expected_month_end_amount"), format_amount(selected_row.get("forecast_after_provision"))),
        (get_metric_label("target_variance"), _format_signed_amount(selected_row.get("target_variance"))),
        (
            get_metric_label("next_close_required_amount"),
            format_amount(next_close_result.get("required_to_recover_next_close_cum")),
        ),
        ("운영모드", get_operation_mode(target_status)),
        ("추천 전략", strategy_label),
    )
    item_html = "".join(
        '<div class="month-close-hero__item">'
        f"<span>{escape(str(label))}</span>"
        f"<strong>{escape(str(value))}</strong>"
        "</div>"
        for label, value in items
    )
    st.markdown(
        '<section class="month-close-hero">'
        '<div class="month-close-hero__head">'
        '<div>'
        '<div class="month-close-hero__eyebrow">월마감 상태판</div>'
        '<h1>월마감 상태판</h1>'
        '<p>현재 상태, 다음 마감 누적선, 추천 운영전략을 한 번에 확인합니다.</p>'
        '</div>'
        f'<span class="strategy-badge strategy-badge--{escape(status_tone)}">{escape(strategy_label)}</span>'
        '</div>'
        f'<div class="month-close-hero__grid">{item_html}</div>'
        '</section>',
        unsafe_allow_html=True,
    )


def _render_home_status_facts(
    validation_result: Mapping[str, Any],
    selected_row: pd.Series,
    next_close_result: Mapping[str, Any],
) -> None:
    scenario_id = str(selected_row.get("scenario_id") or "")
    strategy_code = get_strategy_code(scenario_id or selected_row.get("provision_strategy"))
    strategy_label = (
        f"{strategy_code} {get_strategy_label(strategy_code)}"
        if strategy_code
        else "예측 계산 후 표시"
    )
    facts = (
        (get_metric_label("current_cumulative_actual"), format_amount(validation_result.get("current_actual_cum"))),
        (get_metric_label("expected_month_end_amount"), format_amount(selected_row.get("forecast_after_provision"))),
        (get_metric_label("target_variance"), _format_signed_amount(selected_row.get("target_variance"))),
        (
            get_metric_label("next_close_required_amount"),
            format_amount(next_close_result.get("required_to_recover_next_close_cum")),
        ),
        ("추천 전략", strategy_label),
    )
    fact_html = "".join(
        '<div class="metric-card-compact">'
        f'<span>{escape(str(label))}</span>'
        f'<strong>{escape(str(value))}</strong>'
        "</div>"
        for label, value in facts
    )
    st.markdown(f'<div class="workbench-fact-row">{fact_html}</div>', unsafe_allow_html=True)


def _render_projection_chart_card(context: Mapping[str, Any]) -> None:
    df = _as_dataframe(context.get("df"))
    metric = str(context.get("metric") or "sales")
    selected_row = _as_series(context.get("selected_row"))
    scenario_df = _as_dataframe(context.get("scenario_df"))
    as_of_date = context.get("as_of_date")
    target_status = selected_row.get("target_status")
    columns = get_metric_columns(metric)
    current_day_no = _projection_current_day_no(df, as_of_date)
    forecast_payload = _build_projection_forecast_payload(
        scenario_df,
        selected_row,
        dict(context.get("forecast_result") or {}),
    )
    close_markers = build_close_day_markers(df, current_day_no=current_day_no)
    projection_df = build_pace_projection_chart_data(
        df,
        forecast_payload,
        target_status,
        current_day_no=current_day_no,
        target_daily_column=columns["target_daily"],
        actual_cum_column=columns["actual_cum"],
    )
    if not projection_df.empty and not close_markers.empty:
        next_days = close_markers.loc[
            close_markers["is_next_close_day"],
            "business_day_no",
        ].astype(str)
        if not next_days.empty:
            projection_df["is_next_close_day"] = projection_df["business_day_no"].astype(str).isin(set(next_days))

    with st.container(border=True):
        st.markdown(
            """
            <div class="projection-chart-card">
              <div class="projection-chart-card__head">
                <div>
                  <div class="projection-chart-card__label">달성 추이 및 월말 예측 구간</div>
                  <div class="projection-chart-card__copy">현재까지의 누적 실적 흐름과 기존 예측 모델을 기준으로 잔여 영업일의 예상 도착 구간을 표시합니다.</div>
                </div>
              </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        if projection_df.empty:
            st.info(str(projection_df.attrs.get("empty_state") or "예측 계산 후 Projection 차트를 표시합니다."))
        else:
            _render_pace_projection_chart(projection_df)
            _render_projection_interpretation(projection_df, selected_row)


def _build_projection_forecast_payload(
    scenario_df: pd.DataFrame,
    selected_row: pd.Series,
    forecast_result: Mapping[str, Any],
) -> dict[str, object]:
    summary = _forecast_summary(scenario_df)
    return {
        "F1": summary.get("F1"),
        "F2": summary.get("F2"),
        "F3": summary.get("F3"),
        "forecast_mid": forecast_result.get(
            "forecast_amount",
            selected_row.get("forecast_amount", selected_row.get("forecast_after_provision")),
        ),
    }


def _projection_current_day_no(df: pd.DataFrame, as_of_date: object) -> object | None:
    if df.empty or "date" not in df.columns or "business_day_no" not in df.columns:
        return None
    try:
        dates = pd.to_datetime(df["date"], errors="coerce").dt.normalize()
        as_of_timestamp = pd.Timestamp(as_of_date).normalize()
    except Exception:  # noqa: BLE001 - display fallback only.
        return None
    rows = df.loc[dates == as_of_timestamp]
    if rows.empty:
        return None
    return rows.iloc[-1].get("business_day_no")


def _render_pace_projection_chart(source: pd.DataFrame) -> None:
    chart_source = source.copy()
    chart_font = '-apple-system, BlinkMacSystemFont, "Segoe UI", "Apple SD Gothic Neo", "Malgun Gothic", "Noto Sans KR", sans-serif'
    chart_source["date"] = pd.to_datetime(chart_source["date"], errors="coerce")
    chart_source["date_label"] = chart_source["date"].dt.strftime("%Y-%m-%d")
    for column in (
        "business_day_no",
        "actual_cum",
        "target_cum",
        "forecast_low",
        "forecast_mid",
        "forecast_high",
        "projection_mid",
    ):
        chart_source[column] = pd.to_numeric(chart_source[column], errors="coerce")

    chart_source = chart_source.dropna(subset=["business_day_no", "date"])
    if chart_source.empty:
        st.info("현재 누적 실적이 입력되면 실제 추이선이 표시됩니다.")
        return

    tooltip = [
        alt.Tooltip("business_day_no:Q", title="영업일차"),
        alt.Tooltip("date_label:N", title="날짜"),
        alt.Tooltip("target_cum:Q", title="누적 목표선", format=chart_value_format("억원")),
        alt.Tooltip("actual_cum:Q", title="현재까지 확정 실적", format=chart_value_format("억원")),
        alt.Tooltip("projection_mid:Q", title="향후 예측 중심선", format=chart_value_format("억원")),
        alt.Tooltip("forecast_low:Q", title="예상 도착 구간 하단", format=chart_value_format("억원")),
        alt.Tooltip("forecast_high:Q", title="예상 도착 구간 상단", format=chart_value_format("억원")),
        alt.Tooltip("is_close_day:N", title="is_close_day"),
        alt.Tooltip("zone:N", title="구간"),
    ]
    business_days = chart_source["business_day_no"].dropna()
    min_business_day = max(1, int(math.floor(float(business_days.min()))))
    max_business_day = max(min_business_day, int(math.ceil(float(business_days.max()))))
    x_axis = alt.Axis(
        title=None,
        tickMinStep=1,
        values=list(range(min_business_day, max_business_day + 1)),
        labelAngle=0,
        labelExpr="format(datum.value, 'd') + 'WD'",
        grid=False,
    )
    x_left_padding = 0.35
    x_scale = alt.Scale(
        domain=[max(0.5, min_business_day - x_left_padding), max_business_day],
        nice=False,
        zero=False,
    )
    x_encoding = alt.X(
        "business_day_no:Q",
        title=None,
        axis=x_axis,
        scale=x_scale,
    )
    y_axis_hidden = alt.Axis(
        title=None,
        labels=False,
        ticks=False,
        domain=False,
        grid=False,
    )
    y_encoding = alt.Y(
        "value:Q",
        title=None,
        axis=y_axis_hidden,
        scale=_projection_value_scale(chart_source),
    )
    projection_zoom = alt.selection_interval(
        bind="scales",
        encodings=["x", "y"],
        name="projection_zoom",
    )
    projection_hover = alt.selection_point(
        fields=["business_day_no"],
        nearest=True,
        on="pointerover",
        clear="pointerout",
        empty=False,
        name="projection_hover",
    )
    projection_click = alt.selection_point(
        fields=["business_day_no"],
        nearest=True,
        on="click",
        clear="dblclick",
        empty=False,
        name="projection_click",
    )

    band_source = chart_source.loc[chart_source["is_projection_period"]].dropna(
        subset=["forecast_low", "forecast_high"]
    )
    actual_source = chart_source.loc[chart_source["actual_cum"].notna()]
    projection_source = chart_source.loc[chart_source["projection_mid"].notna()]
    current_source = chart_source.loc[chart_source["is_current_point"]]
    next_close_source = chart_source.loc[chart_source["is_next_close_day"]]

    target_source = chart_source.copy()
    target_source["value"] = target_source["target_cum"]
    actual_line_source = actual_source.copy()
    actual_line_source["value"] = actual_line_source["actual_cum"]
    projection_line_source = projection_source.copy()
    projection_line_source["value"] = projection_line_source["projection_mid"]
    guide_source = chart_source.copy()
    guide_source["value"] = guide_source["actual_cum"].combine_first(
        guide_source["projection_mid"]
    ).combine_first(guide_source["target_cum"])

    band = (
        alt.Chart(band_source)
        .mark_area(color="#b7791f", opacity=0.16, interpolate="linear")
        .encode(
            x=x_encoding,
            y=alt.Y(
                "forecast_low:Q",
                title=None,
                axis=y_axis_hidden,
                scale=_projection_value_scale(chart_source),
            ),
            y2="forecast_high:Q",
            tooltip=tooltip,
        )
    )
    target_line = (
        alt.Chart(target_source)
        .mark_line(color="#8a94a1", strokeDash=[7, 4], strokeWidth=2.1)
        .encode(x=x_encoding, y=y_encoding, tooltip=tooltip)
    )
    actual_line = (
        alt.Chart(actual_line_source)
        .mark_line(color="#14756f", strokeWidth=3.3, interpolate="linear")
        .encode(x=x_encoding, y=y_encoding, tooltip=tooltip)
    )
    actual_points = (
        alt.Chart(actual_line_source)
        .mark_point(color="#14756f", filled=True, size=38, opacity=0.92)
        .encode(x=x_encoding, y=y_encoding, tooltip=tooltip)
    )
    projection_line = (
        alt.Chart(projection_line_source)
        .mark_line(color="#b7791f", strokeDash=[6, 4], strokeWidth=2.8, interpolate="linear")
        .encode(x=x_encoding, y=y_encoding, tooltip=tooltip)
    )
    projection_points = (
        alt.Chart(projection_line_source)
        .mark_point(color="#b7791f", filled=True, size=28, opacity=0.82)
        .encode(x=x_encoding, y=y_encoding, tooltip=tooltip)
    )
    current_point_source = current_source.copy()
    current_point_source["value"] = current_point_source["actual_cum"]
    current_point_source["marker_label"] = "현재 위치"
    current_point = (
        alt.Chart(current_point_source)
        .mark_point(color="#202833", filled=True, size=125)
        .encode(x=x_encoding, y=y_encoding, tooltip=tooltip)
    )
    current_label = (
        alt.Chart(current_point_source)
        .mark_text(
            align="left",
            baseline="middle",
            dx=8,
            dy=-16,
            font=chart_font,
            fontSize=12,
            fontWeight=700,
            color="#202833",
        )
        .encode(x=x_encoding, y=y_encoding, text="marker_label:N", tooltip=tooltip)
    )
    next_close_rule = (
        alt.Chart(next_close_source)
        .mark_rule(color="#536170", strokeDash=[3, 3], strokeWidth=2.1)
        .encode(x=x_encoding, tooltip=tooltip)
    )
    next_close_point_source = next_close_source.copy()
    next_close_point_source["value"] = next_close_point_source["target_cum"]
    next_close_point_source["marker_label"] = "다음 마감"
    next_close_point = (
        alt.Chart(next_close_point_source)
        .mark_point(color="#536170", filled=True, size=108, shape="triangle-up")
        .encode(x=x_encoding, y=y_encoding, tooltip=tooltip)
    )
    next_close_label = (
        alt.Chart(next_close_point_source)
        .mark_text(
            align="left",
            baseline="middle",
            dx=8,
            dy=16,
            font=chart_font,
            fontSize=12,
            fontWeight=700,
            color="#536170",
        )
        .encode(x=x_encoding, y=y_encoding, text="marker_label:N", tooltip=tooltip)
    )
    guide_selector = (
        alt.Chart(guide_source)
        .mark_rule(opacity=0.001, strokeWidth=24)
        .encode(x=x_encoding, tooltip=tooltip)
        .add_params(projection_hover, projection_click)
    )
    hover_guide = (
        alt.Chart(guide_source)
        .mark_rule(color="#1f2937", strokeWidth=1.5, opacity=0.45)
        .encode(x=x_encoding, tooltip=tooltip)
        .transform_filter(projection_hover)
    )
    click_guide = (
        alt.Chart(guide_source)
        .mark_rule(color="#0f766e", strokeWidth=2.2, opacity=0.72)
        .encode(x=x_encoding, tooltip=tooltip)
        .transform_filter(projection_click)
    )
    hover_point = (
        alt.Chart(guide_source)
        .mark_point(color="#1f2937", filled=True, size=70, opacity=0.78)
        .encode(x=x_encoding, y=y_encoding, tooltip=tooltip)
        .transform_filter(projection_hover)
    )
    click_point = (
        alt.Chart(guide_source)
        .mark_point(color="#0f766e", filled=True, size=90, opacity=0.92)
        .encode(x=x_encoding, y=y_encoding, tooltip=tooltip)
        .transform_filter(projection_click)
    )

    chart = (
        (
            band
            + target_line
            + actual_line
            + actual_points
            + projection_line
            + projection_points
            + next_close_rule
            + next_close_point
            + next_close_label
            + current_point
            + current_label
            + hover_guide
            + click_guide
            + hover_point
            + click_point
            + guide_selector
        )
        .add_params(projection_zoom)
        .properties(height=320, background="#f7f9fb")
        .configure(background="#f7f9fb")
        .configure_axis(
            labelFont=chart_font,
            titleFont=chart_font,
            labelFontSize=11,
            titleFontSize=12,
            labelColor="#65717f",
            titleColor="#65717f",
            grid=False,
            domain=False,
        )
        .configure_legend(
            labelFont=chart_font,
            titleFont=chart_font,
            labelFontSize=12,
            titleFontSize=12,
        )
        .configure_title(font=chart_font, fontSize=14, fontWeight=700)
        .configure_view(strokeWidth=0, fill="#f7f9fb")
    )
    st.altair_chart(chart, use_container_width=True)
    st.markdown(
        """
        <div class="chart-legend-row">
          <span><i class="legend-target"></i>누적 목표선</span>
          <span><i class="legend-actual"></i>현재까지 확정 실적</span>
          <span><i class="legend-projection"></i>향후 예측 중심선</span>
          <span><i class="legend-band"></i>예상 도착 구간</span>
          <span><i class="legend-current"></i>현재 위치</span>
          <span><i class="legend-close"></i>다음 마감</span>
        </div>
        """,
        unsafe_allow_html=True,
    )


def _projection_value_scale(chart_source: pd.DataFrame) -> alt.Scale:
    values = pd.concat(
        [
            pd.to_numeric(chart_source[column], errors="coerce")
            for column in (
                "actual_cum",
                "target_cum",
                "forecast_low",
                "forecast_mid",
                "forecast_high",
                "projection_mid",
            )
            if column in chart_source.columns
        ],
        ignore_index=True,
    ).dropna()
    if values.empty:
        return alt.Scale(zero=True)
    lower = max(0.0, float(values.min()) * 0.94)
    upper = float(values.max()) * 1.06
    if math.isclose(lower, upper):
        upper = lower + 1.0
    return alt.Scale(domain=[lower, upper], nice=True)


def _render_projection_interpretation(
    projection_df: pd.DataFrame,
    selected_row: pd.Series,
) -> None:
    status = str(selected_row.get("target_status") or "UNKNOWN_TARGET_STATUS")
    variance = selected_row.get("target_variance")
    low = projection_df.attrs.get("forecast_low_final")
    mid = projection_df.attrs.get("forecast_mid_final")
    high = projection_df.attrs.get("forecast_high_final")
    caption = (
        "현재 흐름은 목표선 대비 "
        f"{_format_signed_amount(variance)} 수준이며, "
        f"월말 예상 도착 구간은 {_target_status_arrival_label(status)}입니다. "
        f"예상 도착 구간: {format_amount(low)} ~ {format_amount(high)}, 중심선 {format_amount(mid)}."
    )
    st.markdown(
        f'<div class="projection-chart-caption">{escape(caption)}</div>',
        unsafe_allow_html=True,
    )


def _target_status_arrival_label(target_status: object) -> str:
    labels = {
        "UNDER_TARGET": "UNDER_TARGET 목표선 미달 구간",
        "ON_TARGET": "ON_TARGET 계획선 근접 구간",
        "OVER_TARGET": "OVER_TARGET 초과달성 관리 구간",
    }
    return labels.get(str(target_status), "계산 확인 구간")


def _render_home_decision_panel(
    selected_row: pd.Series,
    validation_result: Mapping[str, Any],
    next_close_result: Mapping[str, Any],
) -> None:
    target_status = selected_row.get("target_status")
    scenario_id = str(selected_row.get("scenario_id") or "")
    _forecast_key, strategy_key = _split_scenario_id(scenario_id)
    strategy_name = get_strategy_label(strategy_key or selected_row.get("provision_strategy"))
    risk_level = _localize_display_value(selected_row.get("risk_level", "N/A"))
    status = _localize_display_value(selected_row.get("status", ""))
    decision_rows = (
        ("목표 상태", _localize_display_value(target_status)),
        ("월마감 예상", format_amount(selected_row.get("forecast_after_provision"))),
        ("목표 대비 차이", _format_signed_amount(selected_row.get("target_variance"))),
        (
            "다음 마감 누적선 필요실적",
            format_amount(next_close_result.get("required_to_recover_next_close_cum")),
        ),
        ("운영모드", _operation_mode_label(target_status)),
        ("권장 전략", f"{strategy_key or '-'} {strategy_name}"),
        ("리스크 메모", f"{risk_level} / {status}"),
        ("다음 액션", _home_next_action_text(target_status, validation_result, next_close_result)),
    )
    rows_html = "".join(
        '<div class="decision-panel__row">'
        f'<span>{escape(str(label))}</span>'
        f'<strong>{escape(str(value))}</strong>'
        "</div>"
        for label, value in decision_rows
    )
    st.markdown(
        '<aside class="decision-panel">'
        '<div class="decision-panel__label">Next Action Panel</div>'
        '<h2>오늘 판단 카드</h2>'
        f"{rows_html}"
        "</aside>",
        unsafe_allow_html=True,
    )


def _home_next_action_text(
    target_status: object,
    validation_result: Mapping[str, Any],
    next_close_result: Mapping[str, Any],
) -> str:
    _ = validation_result
    required = _as_float(next_close_result.get("required_to_recover_next_close_cum"))
    if str(target_status) == "UNDER_TARGET":
        if math.isfinite(required) and required > 0:
            return f"다음 마감일까지 누적선 기준 {format_amount(required)} 확보 계획을 먼저 확인합니다."
        return "잔여 목표 보정전략 P1/P2/P3 중 실행 부담이 낮은 안을 확인합니다."
    if str(target_status) == "OVER_TARGET":
        return "O1 버퍼 유지, O2 Stretch 전환, O3 품질 방어 중 보고 기준을 선택합니다."
    if str(target_status) == "ON_TARGET":
        return "현재 계획선을 유지하면서 다음 마감일 입력 누락 여부를 점검합니다."
    return "입력 검증 후 예측 계산을 다시 실행합니다."


def render_next_action_panel(page_key: str, context: Mapping[str, Any]) -> None:
    """Render the compact next action panel shared by all workbench pages."""
    validation_result = dict(context.get("validation_result") or {})
    selected_row = _as_series(context.get("selected_row"))
    next_close_result = dict(context.get("next_close_result") or {})
    target_status = selected_row.get("target_status")
    page_label = {
        "home": "홈",
        "input": "입력 · 데이터",
        "forecast_strategy": "예측 · 전략 통합",
        "report": "보고 메모",
        "history": "예측 이력",
        "excel": "Excel 공유",
        "audit": "검증 · 운영관리",
    }.get(page_key, page_key)
    action_items = {
        "home": (
            _home_next_action_text(target_status, validation_result, next_close_result),
            "예측 · 전략 통합에서 target_status와 모델별 월마감 예상값을 확인합니다.",
        ),
        "input": (
            "누락 행, 빈 누적 실적, is_close_day 입력 상태를 먼저 확인합니다.",
            "입력 검증 오류가 없으면 예측 · 전략 통합으로 이동합니다.",
        ),
        "forecast_strategy": (
            "상단 결론 요약에서 목표 상태, 목표 대비 차이, 다음 마감 누적선을 먼저 확인합니다.",
            "선택 시나리오를 바꿔 운영 판단판과 상세 표를 같은 화면에서 비교합니다.",
        ),
        "report": (
            "보고 메모 원문을 복사용 영역에서 확인합니다.",
            "Excel 공유 페이지에서 기존 outputs/latest 산출물만 다운로드합니다.",
        ),
        "history": (
            "완료월 비교와 현재 예측 신뢰도 흐름을 확인합니다.",
            "필요할 때만 예측 이력 저장 버튼을 사용합니다.",
        ),
        "excel": (
            "outputs/latest 파일 목록과 수정시각을 읽기 전용으로 확인합니다.",
            "최신 리포트 재생성은 명시적 버튼 클릭 시에만 실행합니다.",
        ),
        "audit": (
            "pytest, Gate Runner, forbidden scan의 U03-A1.1 로그 상태를 확인합니다.",
            "화면 캡처 감리와 Auth Gate 후속 복원 항목을 분리해 관리합니다.",
        ),
    }.get(page_key, ("현재 페이지의 입력과 계산 상태를 확인합니다.",))
    items_html = "".join(f"<li>{escape(str(item))}</li>" for item in action_items)
    st.markdown(
        '<aside class="next-action-panel">'
        f'<div class="next-action-panel__label">다음 액션 · next action</div>'
        f"<h3>{escape(page_label)} 작업 흐름</h3>"
        f"<ul>{items_html}</ul>"
        "</aside>",
        unsafe_allow_html=True,
    )


def _render_home_scenario_summary(
    scenario_df: pd.DataFrame,
    selected_scenario_id: str,
    target_status: object,
) -> None:
    _ = selected_scenario_id, target_status
    st.markdown(
        render_section_header(
            "시나리오 요약 3개",
            "F1/F2/F3 예측모델별 최종 실적 가능성과 목표 대비 차이를 비교합니다.",
        ),
        unsafe_allow_html=True,
    )
    if scenario_df.empty:
        st.info("예측 계산 후 시나리오 요약이 표시됩니다.")
        return

    summary_rows = build_home_forecast_model_summary(scenario_df)
    if summary_rows.empty:
        st.info("F1/F2/F3 예측모델별 요약 데이터가 없습니다.")
        return

    display_rows = _dedupe_converged_forecast_summary_rows(summary_rows)
    cards = "".join(
        _render_forecast_model_summary_card(row)
        for row in display_rows.to_dict("records")
    )
    st.markdown(f'<div class="strategy-card-row">{cards}</div>', unsafe_allow_html=True)


def _render_home_overachievement_strategy_summary(
    scenario_df: pd.DataFrame,
    selected_scenario_id: str,
    target_status: object,
) -> None:
    over_rows = build_home_overachievement_strategy_summary(
        scenario_df,
        selected_scenario_id,
    )
    if over_rows.empty:
        return

    state_note = (
        "현재 관리 대상"
        if str(target_status) == "OVER_TARGET"
        else "현재 상태에서는 참고용"
    )
    cards = "".join(
        _render_overachievement_strategy_summary_card(row)
        for row in over_rows.to_dict("records")
    )
    st.markdown(
        render_section_header(
            "초과달성 운영전략",
            "O1/O2/O3는 같은 base forecast를 반복하지 않고 버퍼, Stretch, 품질 방어 지표로 분리합니다.",
        ),
        unsafe_allow_html=True,
    )
    st.markdown(
        f'<section class="strategy-section is-active-management">'
        '<div class="strategy-section__head">'
        '<div><div class="strategy-section__status">OVER_TARGET</div>'
        '<p>기준 F예측값은 참고값으로 두고, 전략별 운영 기준값과 여유분을 비교합니다.</p></div>'
        f'<span>{escape(state_note)}</span>'
        "</div>"
        f'<div class="strategy-section__cards">{cards}</div>'
        "</section>",
        unsafe_allow_html=True,
    )


FORECAST_MODEL_SUMMARY_COLUMNS = (
    "forecast_key",
    "forecast_model",
    "model_name",
    "expected_month_end_amount",
    "target_variance",
    "target_status",
    "recommended_action",
    "convergence_notice",
)


OVERACHIEVEMENT_SUMMARY_COLUMNS = (
    "scenario_id",
    "forecast_key",
    "strategy_key",
    "base_forecast_amount",
    "strategy_expected_amount",
    "target_variance_after_strategy",
    "surplus_buffer",
    "stretch_uplift",
    "revised_monthly_target",
    "relief_amount",
    "recommended_action",
    "target_status",
)


def build_home_forecast_model_summary(scenario_df: pd.DataFrame) -> pd.DataFrame:
    """Return one F1/F2/F3 forecast row per model for the home summary cards."""
    if scenario_df.empty or "scenario_id" not in scenario_df.columns:
        return pd.DataFrame(columns=FORECAST_MODEL_SUMMARY_COLUMNS)

    rows: list[dict[str, object]] = []
    for forecast_key in ("F1", "F2", "F3"):
        model_df = _scenario_df_for_forecast_key(scenario_df, forecast_key)
        if model_df.empty:
            continue
        row = model_df.iloc[0]
        target_status_value = row.get("target_status")
        rows.append(
            {
                "forecast_key": forecast_key,
                "forecast_model": FORECAST_MODEL_OPTIONS.get(forecast_key, forecast_key),
                "model_name": FORECAST_MODEL_DEFINITIONS.get(forecast_key, {}).get(
                    "name",
                    forecast_key,
                ),
                "expected_month_end_amount": row.get("forecast_amount"),
                "target_variance": row.get("target_variance"),
                "target_status": target_status_value,
                "recommended_action": _recommended_strategy_for_target_status(
                    target_status_value
                ),
                "convergence_notice": "",
            }
        )
    return pd.DataFrame(rows, columns=FORECAST_MODEL_SUMMARY_COLUMNS)


def build_home_overachievement_strategy_summary(
    scenario_df: pd.DataFrame,
    selected_scenario_id: str | None = None,
) -> pd.DataFrame:
    """Return O1/O2/O3 display rows with strategy-specific operating values."""
    if scenario_df.empty or "scenario_id" not in scenario_df.columns:
        return pd.DataFrame(columns=OVERACHIEVEMENT_SUMMARY_COLUMNS)

    forecast_key = _selected_forecast_key(selected_scenario_id) or "F1"
    rows = scenario_df.loc[
        scenario_df["scenario_id"].astype(str).map(
            lambda value: _split_scenario_id(value)[1] in {"O1", "O2", "O3"}
        )
    ].copy()
    if rows.empty:
        return pd.DataFrame(columns=OVERACHIEVEMENT_SUMMARY_COLUMNS)

    focused = rows.loc[rows["scenario_id"].astype(str).str.startswith(f"{forecast_key}_")]
    if focused.empty:
        forecast_key = _split_scenario_id(str(rows.iloc[0].get("scenario_id") or ""))[0]
        focused = rows.loc[rows["scenario_id"].astype(str).str.startswith(f"{forecast_key}_")]
    if focused.empty:
        focused = rows

    result_rows: list[dict[str, object]] = []
    for strategy_key in ("O1", "O2", "O3"):
        matches = focused.loc[
            focused["scenario_id"].astype(str).map(
                lambda value: _split_scenario_id(value)[1] == strategy_key
            )
        ]
        if matches.empty:
            continue
        row = matches.iloc[0]
        base_forecast = _first_finite_amount(
            row.get("forecast_amount"),
            row.get("forecast_after_provision"),
        )
        strategy_expected = _strategy_expected_amount_for_overachievement(row)
        revised_target = _first_finite_amount(
            row.get("revised_monthly_target"),
            row.get("monthly_target"),
        )
        target_variance_after_strategy = _amount_difference(
            strategy_expected,
            revised_target,
        )
        result_rows.append(
            {
                "scenario_id": row.get("scenario_id"),
                "forecast_key": _split_scenario_id(str(row.get("scenario_id") or ""))[0],
                "strategy_key": strategy_key,
                "base_forecast_amount": base_forecast,
                "strategy_expected_amount": strategy_expected,
                "target_variance_after_strategy": target_variance_after_strategy,
                "surplus_buffer": row.get("remaining_surplus_buffer"),
                "stretch_uplift": row.get("stretch_uplift"),
                "revised_monthly_target": row.get("revised_monthly_target"),
                "relief_amount": row.get("relief_amount"),
                "recommended_action": row.get("recommended_action"),
                "target_status": row.get("target_status"),
            }
        )

    return pd.DataFrame(result_rows, columns=OVERACHIEVEMENT_SUMMARY_COLUMNS)


def _dedupe_converged_forecast_summary_rows(summary_rows: pd.DataFrame) -> pd.DataFrame:
    if summary_rows.empty or "expected_month_end_amount" not in summary_rows.columns:
        return summary_rows

    working = summary_rows.copy()
    working["_convergence_key"] = working["expected_month_end_amount"].map(
        _forecast_convergence_key
    )
    display_rows: list[dict[str, object]] = []
    for _, group in working.groupby("_convergence_key", sort=False, dropna=False):
        group = group.drop(columns=["_convergence_key"], errors="ignore")
        first = group.iloc[0].to_dict()
        if len(group) > 1:
            forecast_keys = "/".join(group["forecast_key"].astype(str).tolist())
            model_ids = " / ".join(group["forecast_model"].astype(str).tolist())
            first.update(
                {
                    "forecast_key": forecast_keys,
                    "forecast_model": model_ids,
                    "model_name": "예측값 수렴",
                    "recommended_action": (
                        f"{forecast_keys} 예측값이 같은 구간입니다. "
                        "중복 카드 대신 하나의 수렴 카드로 보고 운영전략을 선택합니다."
                    ),
                    "convergence_notice": "예측값 수렴",
                }
            )
        display_rows.append(first)
    return pd.DataFrame(display_rows, columns=FORECAST_MODEL_SUMMARY_COLUMNS)


def _render_forecast_model_summary_card(row: Mapping[str, object]) -> str:
    target_status_value = row.get("target_status")
    notice = str(row.get("convergence_notice") or "")
    notice_html = (
        f'<div class="scenario-card__group">{escape(notice)}</div>' if notice else ""
    )
    metrics = (
        ("forecast_model", row.get("forecast_model")),
        ("model_name", row.get("model_name")),
        ("expected_month_end_amount", format_amount(row.get("expected_month_end_amount"))),
        ("target_variance", format_amount(row.get("target_variance"))),
        ("target_status", row.get("target_status")),
        ("추천 운영전략", row.get("recommended_action")),
    )
    metrics_html = "".join(
        '<div class="scenario-card__metric">'
        f'<div class="scenario-card__metric-label">{escape(str(label))}</div>'
        f'<div class="scenario-card__metric-value">{escape(str(value))}</div>'
        "</div>"
        for label, value in metrics
    )
    return (
        '<article class="scenario-card scenario-card--neutral forecast-summary-card">'
        '<div class="scenario-card__topline">'
        f'<div><div class="scenario-card__id">{escape(str(row.get("forecast_key") or ""))}</div>'
        f"{notice_html}</div>"
        f"{render_status_badge(target_status_value)}"
        "</div>"
        f'<div class="scenario-card__name">{escape(str(row.get("model_name") or ""))}</div>'
        f'<div class="scenario-card__description">{escape(str(row.get("recommended_action") or ""))}</div>'
        f'<div class="scenario-card__metrics">{metrics_html}</div>'
        "</article>"
    )


def _render_overachievement_strategy_summary_card(row: Mapping[str, object]) -> str:
    strategy_key = str(row.get("strategy_key") or "")
    definition = OVERACHIEVEMENT_STRATEGY_DEFINITIONS.get(strategy_key, {})
    detail_label, detail_value = _overachievement_summary_detail_metric(row, strategy_key)
    metrics = (
        ("운영 기준 목표", format_amount(row.get("revised_monthly_target"))),
        (detail_label, format_amount(detail_value)),
        ("기준 F예측값", format_amount(row.get("base_forecast_amount"))),
        (
            "전략 적용 후 차이",
            format_amount(row.get("target_variance_after_strategy")),
        ),
    )
    metrics_html = "".join(
        '<div class="scenario-card__metric">'
        f'<div class="scenario-card__metric-label">{escape(str(label))}</div>'
        f'<div class="scenario-card__metric-value">{escape(str(value))}</div>'
        "</div>"
        for label, value in metrics
    )
    kind_class = strategy_key.lower() if strategy_key in {"O1", "O2", "O3"} else "neutral"
    return (
        f'<article class="scenario-card scenario-card--{escape(kind_class)} is-emphasis">'
        '<div class="scenario-card__topline">'
        f'<div><div class="scenario-card__id">{escape(str(row.get("scenario_id") or strategy_key))}</div>'
        '<div class="scenario-card__group">초과달성 운영전략</div></div>'
        f"{render_status_badge(row.get('target_status'))}"
        "</div>"
        f'<div class="scenario-card__name">{escape(strategy_key)} {escape(str(definition.get("name", strategy_key)))}</div>'
        f'<div class="scenario-card__description">{escape(str(definition.get("description", "")))}</div>'
        f'<div class="scenario-card__metrics">{metrics_html}</div>'
        "</article>"
    )


def _overachievement_summary_detail_metric(
    row: Mapping[str, object],
    strategy_key: str,
) -> tuple[str, object]:
    if strategy_key == "O2":
        return "Stretch 전환분", row.get("stretch_uplift")
    if strategy_key == "O3":
        return "품질관리 여유분", row.get("relief_amount")
    return "잔여 안전버퍼", row.get("surplus_buffer")


def _recommended_strategy_for_target_status(target_status: object) -> str:
    status = "" if _is_missing(target_status) else str(target_status)
    if status == "OVER_TARGET":
        return "O1 버퍼 유지, O2 Stretch 전환, O3 품질 방어 중 운영 기준을 선택합니다."
    if status == "ON_TARGET":
        return "N1 목표 유지, N2 버퍼 모니터링, N3 품질 점검으로 변동 리스크를 봅니다."
    if status == "UNDER_TARGET":
        return "P1 전체 잔여일, P2 마감일 우선, P3 비마감일 우선 보정 전략을 비교합니다."
    return "입력값과 예측 상태를 확인한 뒤 운영전략을 선택합니다."


def _strategy_expected_amount_for_overachievement(row: Mapping[str, object]) -> float:
    scenario_id = str(row.get("scenario_id") or "")
    strategy_key = _split_scenario_id(scenario_id)[1]
    base_forecast = _first_finite_amount(
        row.get("forecast_amount"),
        row.get("forecast_after_provision"),
    )
    monthly_target = _as_float(row.get("monthly_target"))
    revised_target = _as_float(row.get("revised_monthly_target"))
    current_actual = _as_float(row.get("current_actual_cum"))
    minimum_remaining = _as_float(row.get("minimum_remaining_to_hit_target"))

    if strategy_key == "O2" and math.isfinite(revised_target):
        return revised_target
    if strategy_key == "O3":
        if math.isfinite(current_actual) and math.isfinite(minimum_remaining):
            return current_actual + max(0.0, minimum_remaining)
        if math.isfinite(monthly_target):
            return monthly_target
    return base_forecast


def _first_finite_amount(*values: object) -> float:
    for value in values:
        number = _as_float(value)
        if math.isfinite(number):
            return number
    return float("nan")


def _amount_difference(left: object, right: object) -> float:
    left_number = _as_float(left)
    right_number = _as_float(right)
    if math.isfinite(left_number) and math.isfinite(right_number):
        return left_number - right_number
    return float("nan")


def _forecast_convergence_key(value: object) -> object:
    number = _as_float(value)
    if math.isfinite(number):
        return round(number, 6)
    return str(value)


def _render_strategy_reference_sections(
    scenario_df: pd.DataFrame,
    selected_scenario_id: str,
    target_status: object,
) -> None:
    """Render visible strategy cards for under, over, and neutral states."""
    st.markdown(
        render_section_header(
            "전략 카드",
            "현재 관리 대상 전략을 먼저 보고, 다른 상태 전략은 참고용으로 접어 둡니다.",
        ),
        unsafe_allow_html=True,
    )
    if scenario_df.empty:
        st.info("시나리오 계산 후 전략 카드가 표시됩니다.")
        return

    forecast_key = _selected_forecast_key(selected_scenario_id) or "F1"
    _, selected_strategy_key = _split_scenario_id(selected_scenario_id)
    sections = (
        {
            "status": "UNDER_TARGET",
            "title": "UNDER_TARGET",
            "caption": "목표 미달 시 잔여 목표를 보정하는 P1/P2/P3 전략입니다.",
            "suffixes": ("P1", "P2", "P3"),
            "full_ids": {
                "P1": P1_ALL_REMAINING,
                "P2": P2_CLOSE_DAY_FOCUSED,
                "P3": P3_NON_CLOSE_DAY_FOCUSED,
            },
        },
        {
            "status": "OVER_TARGET",
            "title": "OVER_TARGET",
            "caption": "초과달성 구간에서 버퍼, Stretch 전환, 품질 방어를 분리해 봅니다.",
            "suffixes": ("O1", "O2", "O3"),
            "full_ids": {
                "O1": O1_TARGET_HOLD_BUFFER,
                "O2": O2_STRETCH_TARGET_CAPTURE,
                "O3": O3_QUALITY_GUARD_RELIEF,
            },
        },
        {
            "status": "ON_TARGET",
            "title": "ON_TARGET / Neutral",
            "caption": "목표 근접 구간에서는 유지, 모니터링, 품질 점검 전략을 참고합니다.",
            "suffixes": ("N1", "N2", "N3"),
            "full_ids": {
                "N1": N1_MAINTAIN_TARGET,
                "N2": N2_MONITOR_BUFFER,
                "N3": N3_QUALITY_CHECK,
            },
        },
    )

    def render_strategy_section(section: Mapping[str, object]) -> None:
        is_active = str(target_status) == section["status"]
        section_class = " is-active-management" if is_active else ""
        state_label = "현재 관리 대상" if is_active else "현재 상태에서는 참고용"
        cards_html = _strategy_section_cards_html(
            scenario_df,
            forecast_key,
            tuple(section["suffixes"]),
            dict(section["full_ids"]),
            is_active=is_active,
            selected_strategy_key=selected_strategy_key,
        )
        st.markdown(
            f'<section class="strategy-section{section_class}">'
            '<div class="strategy-section__head">'
            f'<div><div class="strategy-section__status">{escape(str(section["title"]))}</div>'
            f'<p>{escape(str(section["caption"]))}</p></div>'
            f'<span>{escape(state_label)}</span>'
            "</div>"
            f'<div class="strategy-section__cards">{cards_html}</div>'
            "</section>",
            unsafe_allow_html=True,
        )

    active_sections = [
        section for section in sections if str(target_status) == str(section["status"])
    ]
    reference_sections = [
        section for section in sections if str(target_status) != str(section["status"])
    ]
    for section in active_sections or sections[:1]:
        render_strategy_section(section)

    if reference_sections:
        with st.expander("다른 상태 전략 참고", expanded=False):
            for section in reference_sections:
                render_strategy_section(section)


def _render_strategy_arrival_inline_summary(
    scenario_df: pd.DataFrame,
    selected_scenario_id: str,
) -> None:
    source = build_strategy_arrival_compare_source(scenario_df, selected_scenario_id)
    if source.empty:
        st.info("전략별 월말 도착점 compact chart는 시나리오 계산 후 표시됩니다.")
        return

    source = source.copy()
    source["strategy_label"] = source["strategy_key"].map(_compact_strategy_label)
    source["selected_label"] = source["is_selected"].map(lambda value: "선택" if value else "비교")
    source = source.sort_values(["strategy_key", "scenario_id"], kind="mergesort")
    target_value = (
        _as_float(source["monthly_target"].dropna().iloc[0])
        if source["monthly_target"].notna().any()
        else float("nan")
    )
    if bool(source.attrs.get("fallback_used")):
        st.markdown(
            '<div class="scenario-inline-chart-title">'
            '<strong>전략별 운영 기준 compact summary</strong>'
            '<span>월말 예상은 동일, 운영 기준 차이를 표로 확인합니다.</span>'
            "</div>",
            unsafe_allow_html=True,
        )
        _render_strategy_compare_fallback(source)
        return

    st.markdown(
        '<div class="scenario-inline-chart-title">'
        '<strong>전략별 월말 도착점 compact chart</strong>'
        '<span>ScenarioGrid의 월말 예상 실적 차이가 있을 때만 그래프로 비교합니다.</span>'
        "</div>",
        unsafe_allow_html=True,
    )
    _render_strategy_arrival_html_bars(source, target_value)


def _strategy_section_cards_html(
    scenario_df: pd.DataFrame,
    forecast_key: str,
    suffixes: tuple[str, ...],
    full_ids: dict[str, str],
    *,
    is_active: bool,
    selected_strategy_key: str = "",
) -> str:
    cards: list[str] = []
    for suffix in suffixes:
        row = _strategy_reference_row(scenario_df, forecast_key, suffix)
        full_id = full_ids.get(suffix, suffix)
        inactive_class = "" if is_active else " strategy-card-inactive"
        active_class = " strategy-card-active" if is_active else ""
        selected_class = " is-recommended" if is_active and suffix == selected_strategy_key else ""
        if is_active and suffix == selected_strategy_key:
            badge_html = '<span class="strategy-card-shell__badge is-recommended-badge">권장</span>'
        elif is_active:
            badge_html = ""
        else:
            badge_html = '<span class="strategy-card-shell__badge is-reference">참고용</span>'
        if row is None:
            definition = SCENARIO_STRATEGY_DEFINITIONS.get(suffix, {})
            card_html = (
                '<article class="scenario-card scenario-card--neutral">'
                '<div class="scenario-card__topline">'
                f'<div><div class="scenario-card__id">{escape(suffix)}</div>'
                '<div class="scenario-card__group">현재 상태에서는 참고용</div></div>'
                "</div>"
                f'<div class="scenario-card__name">{escape(str(definition.get("name", suffix)))}</div>'
                f'<div class="scenario-card__description">{escape(str(definition.get("description", "전략 정의를 참고합니다.")))}</div>'
                '<div class="scenario-card__metrics">'
                '<div class="scenario-card__metric">'
                '<div class="scenario-card__metric-label">적용 상태</div>'
                '<div class="scenario-card__metric-value">참고용</div>'
                "</div>"
                '<div class="scenario-card__metric">'
                '<div class="scenario-card__metric-label">계산 기준</div>'
                '<div class="scenario-card__metric-value">기존 산식 유지</div>'
                "</div>"
                "</div>"
                "</article>"
            )
        elif suffix in {"O1", "O2", "O3"}:
            card_html = _render_overachievement_operation_basis_card(row, suffix)
        else:
            card_html = render_scenario_card(row)
        cards.append(
            f'<div class="strategy-card-shell{inactive_class}{active_class}{selected_class}">'
            '<div class="strategy-card-shell__head">'
            f'<div class="strategy-card-shell__code">{escape(suffix)} · {escape(full_id)}</div>'
            f"{badge_html}"
            "</div>"
            f"{card_html}"
            "</div>"
        )
    return "".join(cards)


def _render_overachievement_operation_basis_card(
    row: Mapping[str, object],
    suffix: str,
) -> str:
    strategy_id = str(
        row.get("overachievement_strategy")
        or row.get("provision_strategy")
        or suffix
    )
    display_id = strategy_id if strategy_id else suffix
    scenario_id = str(row.get("scenario_id") or suffix)
    revised_target = _first_finite_amount(
        row.get("revised_monthly_target"),
        row.get("monthly_target"),
    )
    base_forecast = _first_finite_amount(
        row.get("forecast_amount"),
        row.get("forecast_after_provision"),
    )
    strategy_expected = _strategy_expected_amount_for_overachievement(row)
    target_variance_after_strategy = _amount_difference(strategy_expected, revised_target)
    detail_label, detail_value = _overachievement_basis_detail_metric(row, suffix)
    metrics = (
        ("운영 기준 목표", format_amount(revised_target)),
        (detail_label, format_amount(detail_value)),
        ("기준 F예측값", format_amount(base_forecast)),
        ("전략 적용 후 차이", format_amount(target_variance_after_strategy)),
    )
    metrics_html = "".join(
        '<div class="scenario-card__metric">'
        f'<div class="scenario-card__metric-label">{escape(str(label))}</div>'
        f'<div class="scenario-card__metric-value">{escape(str(value))}</div>'
        "</div>"
        for label, value in metrics
    )
    kind_class = suffix.lower() if suffix in {"O1", "O2", "O3"} else "neutral"
    return (
        f'<article class="scenario-card scenario-card--{escape(kind_class)} is-emphasis">'
        '<div class="scenario-card__topline">'
        f'<div><div class="scenario-card__id">{escape(scenario_id)}</div>'
        '<div class="scenario-card__group">초과달성 운영</div></div>'
        f"{render_status_badge(row.get('target_status'))}"
        "</div>"
        f'<div class="scenario-card__name">{escape(scenario_display_name(display_id))}</div>'
        f'<div class="scenario-card__description">{escape(scenario_description(display_id))}</div>'
        f'<div class="scenario-card__metrics">{metrics_html}</div>'
        "</article>"
    )


def _overachievement_basis_detail_metric(
    row: Mapping[str, object],
    suffix: str,
) -> tuple[str, object]:
    if suffix == "O2":
        return "Stretch 전환분", row.get("stretch_uplift")
    if suffix == "O3":
        return "품질관리 여유분", row.get("relief_amount")
    return "잔여 안전버퍼", row.get("remaining_surplus_buffer")


def _strategy_reference_row(
    scenario_df: pd.DataFrame,
    forecast_key: str,
    suffix: str,
) -> Mapping[str, object] | None:
    if scenario_df.empty or "scenario_id" not in scenario_df.columns:
        return None
    rows = scenario_df.loc[
        scenario_df["scenario_id"].astype(str).map(lambda value: _split_scenario_id(value)[1] == suffix)
    ]
    if rows.empty:
        return None
    focused = rows.loc[rows["scenario_id"].astype(str).str.startswith(f"{forecast_key}_")]
    if not focused.empty:
        return focused.iloc[0].to_dict()
    return rows.iloc[0].to_dict()


def _render_forecast_model_mini_chart(
    model_rows_df: pd.DataFrame,
    selected_row: pd.Series,
    validation_result: Mapping[str, Any],
) -> None:
    st.markdown(
        render_section_header(
            "F1/F2/F3 비교 mini chart",
            "월말 예상값을 선으로 잇지 않고, 목표선 대비 차이를 막대로 비교합니다.",
        ),
        unsafe_allow_html=True,
    )
    if model_rows_df.empty:
        st.info("예측 계산 후 F1/F2/F3 비교 차트를 표시합니다.")
        return

    source = build_forecast_model_mini_chart_source(
        model_rows_df,
        selected_row,
        monthly_target=validation_result.get("monthly_target"),
    )
    if source.empty:
        st.info(str(source.attrs.get("empty_state") or "예측 계산 후 F1/F2/F3 비교 차트를 표시합니다."))
        return

    source = source.copy()
    source["target_status_label"] = source["target_status"].map(_localize_display_value)
    target_value = _as_float(source.attrs.get("target_line_value"))
    representative_value = _as_float(source.attrs.get("representative_value"))
    representative_model_key = str(source.attrs.get("representative_model_key") or "")
    target_source = (
        pd.DataFrame({"value": [target_value], "label": ["목표선"]})
        if math.isfinite(target_value)
        else pd.DataFrame(columns=["value", "label"])
    )
    representative_source = (
        pd.DataFrame(
            {
                "label": [representative_model_key],
                "value": [representative_value],
                "marker_label": ["대표 월마감 예상값"],
            }
        )
        if representative_model_key in set(source["label"]) and math.isfinite(representative_value)
        else pd.DataFrame(columns=["label", "value", "marker_label"])
    )
    scale_source = pd.concat(
        [
            source.loc[:, ["value"]],
            target_source.loc[:, ["value"]],
            representative_source.loc[:, ["value"]],
        ],
        ignore_index=True,
    )
    order = ["F1", "F2", "F3"]
    chart = _build_forecast_model_comparison_chart(
        source,
        target_source,
        representative_source,
        scale_source,
    )
    st.altair_chart(chart, use_container_width=True)
    st.caption(
        f"F1/F2/F3 mini chart data rows: {len(source)} | "
        f"target_status: {_localize_display_value(selected_row.get('target_status'))}"
    )
    return

    chart = (
        (
            alt.Chart(source)
            .mark_bar(cornerRadiusTopLeft=3, cornerRadiusTopRight=3, opacity=0.94, stroke="#25312f", strokeWidth=0.5)
            .encode(
                x=alt.X("label:N", title=None, sort=order, axis=alt.Axis(labelAngle=0)),
                y=alt.Y(
                    "value:Q",
                    title="월말 예상 / 목표",
                    axis=alt.Axis(format=chart_value_format("억원")),
                    scale=_auto_value_scale(scale_source),
                ),
                color=alt.value("#2f6f68"),
                opacity=alt.condition("datum.is_selected_model", alt.value(1.0), alt.value(0.78)),
                tooltip=[
                    alt.Tooltip("forecast_model:N", title="예측모델"),
                    alt.Tooltip("value:Q", title="월말 예상", format=chart_value_format("억원")),
                    alt.Tooltip("target_status_label:N", title="target_status"),
                ],
            )
            + alt.Chart(source)
            .mark_point(color="#25312f", filled=True, size=92, opacity=0.9)
            .encode(
                x=alt.X("label:N", title=None, sort=order),
                y=alt.Y("value:Q", scale=_auto_value_scale(scale_source)),
                tooltip=[
                    alt.Tooltip("forecast_model:N", title="예측모델"),
                    alt.Tooltip("value:Q", title="월말 예상", format=chart_value_format("억원")),
                ],
            )
            + alt.Chart(representative_source)
            .mark_point(color="#b7791f", filled=True, size=150, shape="diamond", opacity=0.95)
            .encode(
                x=alt.X("label:N", title=None, sort=order),
                y=alt.Y("value:Q", scale=_auto_value_scale(scale_source)),
                tooltip=[
                    alt.Tooltip("marker_label:N", title="구분"),
                    alt.Tooltip("value:Q", title="금액", format=chart_value_format("억원")),
                ],
            )
            + alt.Chart(target_source)
            .mark_rule(color="#7f8785", strokeDash=[6, 4], strokeWidth=2.4)
            .encode(
                y="value:Q",
                tooltip=[alt.Tooltip("value:Q", title="목표선", format=chart_value_format("억원"))],
            )
            + alt.Chart(target_source)
            .mark_text(align="left", dx=4, dy=-6, fontSize=12, fontWeight=700, color="#636967")
            .encode(x=alt.value(5), y="value:Q", text=alt.value("목표선"))
        )
        .properties(height=220)
        .configure_axis(labelFontSize=11, titleFontSize=12)
    )
    st.altair_chart(chart, use_container_width=True)
    st.caption(
        f"F1/F2/F3 mini chart data rows: {len(source)} | "
        f"target_status: {_localize_display_value(selected_row.get('target_status'))}"
    )


def _build_forecast_model_comparison_chart(
    source: pd.DataFrame,
    target_source: pd.DataFrame,
    representative_source: pd.DataFrame,
    scale_source: pd.DataFrame,
) -> alt.Chart:
    """Build an intuitive model comparison chart around the target delta baseline."""
    order = ["F1", "F2", "F3"]
    value_format = chart_value_format("억원")
    chart_source = source.copy()
    target_value = (
        _as_float(target_source["value"].iloc[0])
        if not target_source.empty and "value" in target_source
        else float("nan")
    )
    chart_source["value_label"] = chart_source["value"].map(lambda value: f"{_as_float(value):,.1f}")
    chart_source["target_delta"] = (
        pd.to_numeric(chart_source["value"], errors="coerce") - target_value
        if math.isfinite(target_value)
        else float("nan")
    )
    finite_deltas = pd.to_numeric(chart_source["target_delta"], errors="coerce").dropna()
    max_delta = max(0.0, float(finite_deltas.max())) if not finite_deltas.empty else 1.0
    min_delta = min(0.0, float(finite_deltas.min())) if not finite_deltas.empty else 0.0
    spread = max(max_delta - min_delta, abs(max_delta), abs(min_delta), 1.0)
    delta_padding = spread * 0.18
    delta_scale = alt.Scale(domain=[min_delta - delta_padding, max_delta + delta_padding], nice=True)
    chart_source["target_delta_label"] = chart_source["target_delta"].map(
        lambda value: f"{_as_float(value):+,.1f}억" if math.isfinite(_as_float(value)) else ""
    )
    chart_source["forecast_label"] = chart_source["value"].map(
        lambda value: f"{_as_float(value):,.1f}억 예상" if math.isfinite(_as_float(value)) else ""
    )
    chart_source["bar_color"] = chart_source.apply(
        lambda row: "#b7791f"
        if bool(row.get("is_selected_model"))
        else ("#14756f" if _as_float(row.get("target_delta")) >= 0 else "#c2410c"),
        axis=1,
    )
    selected_source = chart_source.loc[chart_source["is_selected_model"].astype(bool)].copy()

    base = alt.Chart(chart_source).encode(
        x=alt.X(
            "label:N",
            title=None,
            sort=order,
            axis=alt.Axis(labelAngle=0, labelFontSize=12, labelFontWeight=700),
        ),
        y=alt.Y(
            "target_delta:Q",
            title="목표 대비 차이",
            axis=alt.Axis(format="+,.1f", grid=False, labelFontSize=11),
            scale=delta_scale,
        ),
        tooltip=[
            alt.Tooltip("forecast_model:N", title="예측모델"),
            alt.Tooltip("value:Q", title="월말 예상", format=value_format),
            alt.Tooltip("target_delta:Q", title="목표 대비 차이", format="+,.1f"),
            alt.Tooltip("target_status_label:N", title="target_status"),
        ],
    )
    zero_source = pd.DataFrame({"zero": [0.0], "label": ["목표선"]})
    target_baseline = (
        alt.Chart(zero_source)
        .mark_rule(color="#7f8785", strokeDash=[6, 4], strokeWidth=2.1)
        .encode(
            y=alt.Y("zero:Q", scale=delta_scale),
            tooltip=[alt.Tooltip("label:N", title="기준")],
        )
    )
    baseline_label = (
        alt.Chart(zero_source)
        .mark_text(align="left", baseline="bottom", dx=6, dy=-4, fontSize=12, fontWeight=700, color="#636967")
        .encode(x=alt.value(8), y=alt.Y("zero:Q", scale=delta_scale), text=alt.value("목표선"))
    )
    delta_bars = base.mark_bar(cornerRadiusTopLeft=5, cornerRadiusTopRight=5, opacity=0.9).encode(
        y=alt.Y(
            "target_delta:Q",
            title="목표 대비 차이",
            axis=alt.Axis(format="+,.1f", grid=False, labelFontSize=11),
            scale=delta_scale,
        ),
        y2=alt.Y2(datum=0),
        color=alt.Color("bar_color:N", scale=None, legend=None),
    )
    model_points = base.mark_point(filled=True, size=120, stroke="#ffffff", strokeWidth=1.4).encode(
        color=alt.Color("bar_color:N", scale=None, legend=None),
    )
    selected_point = (
        alt.Chart(selected_source)
        .mark_point(color="#b7791f", filled=True, size=220, shape="diamond", stroke="#ffffff", strokeWidth=1.5)
        .encode(
            x=alt.X("label:N", title=None, sort=order),
            y=alt.Y("target_delta:Q", scale=delta_scale),
            tooltip=[
                alt.Tooltip("forecast_model:N", title="선택 모델"),
                alt.Tooltip("value:Q", title="대표 월말 예상", format=value_format),
                alt.Tooltip("target_delta:Q", title="목표 대비 차이", format="+,.1f"),
            ],
        )
    )
    delta_labels = base.mark_text(
        align="center",
        baseline="bottom",
        dy=-7,
        fontSize=13,
        fontWeight=700,
        color="#202833",
    ).encode(text="target_delta_label:N")
    forecast_labels = base.mark_text(
        align="center",
        baseline="top",
        dy=12,
        fontSize=11,
        color="#536170",
    ).encode(y=alt.Y("zero:Q", scale=delta_scale), text="forecast_label:N")
    return (
        (
            target_baseline
            + baseline_label
            + delta_bars
            + model_points
            + selected_point
            + delta_labels
            + forecast_labels
        )
        .properties(height=220)
        .configure_axis(labelFontSize=11, titleFontSize=12)
        .configure_view(stroke=None)
    )


def _render_strategy_arrival_compact_chart(
    scenario_df: pd.DataFrame,
    selected_scenario_id: str,
) -> None:
    st.markdown(
        render_section_header(
            "전략별 월말 도착점 비교",
            "P1/P2/P3와 O1/O2/O3의 월말 예상이 같으면 운영 기준 비교로 전환합니다.",
        ),
        unsafe_allow_html=True,
    )
    source = build_strategy_arrival_compare_source(scenario_df, selected_scenario_id)
    if source.empty:
        st.info("시나리오 계산 후 전략별 월말 도착점 비교 차트를 표시합니다.")
        return

    source = source.copy()
    source["strategy_label"] = source["strategy_key"].map(_compact_strategy_label)
    source["selected_label"] = source["is_selected"].map(lambda value: "선택" if value else "비교")
    source = source.sort_values(["strategy_key", "scenario_id"], kind="mergesort")

    scenario_order = source["scenario_id"].tolist()
    target_value = _as_float(source["monthly_target"].dropna().iloc[0]) if source["monthly_target"].notna().any() else float("nan")
    target_source = pd.DataFrame({"monthly_target": [target_value]}).dropna()

    if bool(source.attrs.get("fallback_used")):
        _render_strategy_compare_fallback(source)
        return

    _render_strategy_arrival_html_bars(source, target_value)
    bars = (
        alt.Chart(source)
        .mark_bar(cornerRadiusTopRight=4, cornerRadiusBottomRight=4, opacity=0.95, stroke="#25312f", strokeWidth=0.35)
        .encode(
            y=alt.Y("scenario_id:N", title="시나리오", sort=scenario_order),
            x=alt.X(
                "compare_value:Q",
                title=str(source.attrs.get("compare_label") or "월말 도착점"),
                axis=alt.Axis(format=chart_value_format("억원")),
                scale=_auto_value_scale(source.rename(columns={"compare_value": "value"})),
            ),
            color=alt.Color(
                "strategy_key:N",
                title="전략",
                scale=alt.Scale(
                    domain=["P1", "P2", "P3", "O1", "O2", "O3", "N1", "N2", "N3"],
                    range=[
                        "#b48632",
                        "#c49a4d",
                        "#d2b06e",
                        "#567c5d",
                        "#4c6f8f",
                        "#b48632",
                        "#51758c",
                        "#6f8795",
                        "#8aa0aa",
                    ],
                ),
            ),
            opacity=alt.condition("datum.is_selected", alt.value(1.0), alt.value(0.74)),
            tooltip=[
                alt.Tooltip("scenario_id:N", title="시나리오"),
                alt.Tooltip("strategy_label:N", title="전략 해석"),
                alt.Tooltip("compare_label:N", title="비교 기준"),
                alt.Tooltip("compare_value:Q", title="비교값", format=chart_value_format("억원")),
                alt.Tooltip("forecast_after_provision:Q", title="월말 예상 실적", format=chart_value_format("억원")),
                alt.Tooltip("monthly_target:Q", title="목표선", format=chart_value_format("억원")),
                alt.Tooltip("target_variance:Q", title="목표 대비 차이", format=chart_value_format("억원")),
                alt.Tooltip("selected_label:N", title="선택 여부"),
            ],
        )
    )
    target_rule = (
        alt.Chart(target_source)
        .mark_rule(color="#8a8f98", strokeDash=[5, 4], strokeWidth=2)
        .encode(
            x="monthly_target:Q",
            tooltip=[alt.Tooltip("monthly_target:Q", title="목표선", format=chart_value_format("억원"))],
        )
    )
    text = (
        alt.Chart(source)
        .mark_text(align="left", dx=6, fontSize=12, fontWeight=700, color="#25312f")
        .encode(
            y=alt.Y("scenario_id:N", sort=scenario_order),
            x="compare_value:Q",
            text=alt.Text("compare_value:Q", format=chart_value_format("억원")),
        )
    )
    chart = (
        (bars + target_rule + text)
        .properties(height=max(220, len(source) * 34))
        .configure_axis(labelFontSize=11, titleFontSize=12)
        .configure_legend(labelFontSize=11, titleFontSize=12)
    )
    st.altair_chart(chart, use_container_width=True)
    st.caption("O1: 버퍼 유지 / O2: Stretch 전환 / O3: 품질 방어")


def _compact_strategy_label(strategy_key: object) -> str:
    code = get_strategy_code(strategy_key)
    if code in {"P1", "P2", "P3", "O1", "O2", "O3", "N1", "N2", "N3"}:
        return f"{code} {get_strategy_label(code)}"
    return str(strategy_key)


def active_strategy_suffixes_for_status(target_status: object) -> tuple[str, str, str]:
    status = str(target_status or "")
    if status == "OVER_TARGET":
        return ("O1", "O2", "O3")
    if status == "ON_TARGET":
        return ("N1", "N2", "N3")
    return ("P1", "P2", "P3")


def _render_strategy_compare_fallback(source: pd.DataFrame) -> None:
    compare_label = str(source.attrs.get("compare_label") or "운영 기준")
    classification = str(source.attrs.get("classification") or "")
    st.info(
        "O전략은 월말 예상 실적을 다시 예측하지 않습니다. 같은 F예측값 안에서 운영 기준 목표, Stretch 전환분, 안전버퍼, 품질관리 여유분을 나눕니다."
    )
    st.caption(
        f"동일값 분류: {classification} | 최종 표시 방식: table | 비교 기준: {compare_label}"
    )
    working_source = source.copy()
    if "strategy_key" in working_source.columns:
        working_source["strategy_effect_type"] = working_source["strategy_key"].map(
            _strategy_effect_type_label
        )
    table_columns = [
        column
        for column in (
            "scenario_id",
            "strategy_key",
            "strategy_effect_type",
            "strategy_type",
            "forecast_after_provision",
            "revised_monthly_target",
            "remaining_surplus_buffer",
            "stretch_uplift",
            "relief_amount",
            "minimum_remaining_to_hit_target",
            "compare_value",
            "recommended_action",
        )
        if column in working_source.columns
    ]
    display = working_source.loc[:, table_columns].copy()
    display = display.rename(
        columns={
            "forecast_after_provision": "forecast_reference_value",
            "revised_monthly_target": "operating_target_reference",
        }
    )
    formatted = _format_display_df(display)
    if "compare_value" in display.columns:
        formatted = formatted.rename(columns={"비교 기준값": f"{compare_label} 비교 기준값"})
    st.dataframe(formatted, hide_index=True, use_container_width=True)


def _render_strategy_arrival_html_bars(source: pd.DataFrame, target_value: float) -> None:
    numeric = source.copy()
    value_column = "compare_value" if "compare_value" in numeric.columns else "forecast_after_provision"
    value_label = str(numeric.attrs.get("compare_label") or "월말 도착점")
    numeric[value_column] = pd.to_numeric(
        numeric[value_column],
        errors="coerce",
    )
    numeric = numeric.dropna(subset=[value_column])
    if numeric.empty:
        st.info("전략별 월말 도착점 표시 데이터가 부족합니다.")
        return

    values = numeric[value_column].astype(float)
    max_value = max(float(values.max()), float(target_value) if math.isfinite(target_value) else 0.0, 1.0)
    rows: list[str] = []
    for _, row in numeric.iterrows():
        scenario_id = str(row.get("scenario_id", ""))
        strategy_key = str(row.get("strategy_key", ""))
        value = _as_float(row.get(value_column))
        width = max(8.0, min(100.0, value / max_value * 100.0)) if math.isfinite(value) else 8.0
        selected_class = " is-selected" if bool(row.get("is_selected")) else ""
        rows.append(
            '<div class="compact-arrival-row">'
            f'<div class="compact-arrival-row__label">{escape(scenario_id)}<span>{escape(_compact_strategy_label(strategy_key))}</span></div>'
            '<div class="compact-arrival-row__track">'
            f'<div class="compact-arrival-row__bar compact-arrival-row__bar--{escape(strategy_key[:1].lower())}{selected_class}" '
            f'style="width:{width:.1f}%"></div>'
            "</div>"
            f'<div class="compact-arrival-row__value">{escape(format_amount(value))}</div>'
            "</div>"
        )
    target_html = (
        f'<div class="compact-arrival-target">목표선 {escape(format_amount(target_value))}</div>'
        if math.isfinite(target_value)
        else ""
    )
    st.markdown(
        '<div class="compact-arrival-chart">'
        f'<div class="compact-arrival-target">{escape(value_label)}</div>'
        f"{target_html}"
        f"{''.join(rows)}"
        "</div>",
        unsafe_allow_html=True,
    )


def _render_input_data_page(
    base_config: dict[str, Any],
    audit_readonly: bool = False,
) -> None:
    st.markdown(
        render_section_header(
            "입력 · 데이터",
            "영업일정, is_close_day, 일별 목표, 누적 실적 입력 상태를 확인합니다.",
        ),
        unsafe_allow_html=True,
    )
    df, source_label = _render_file_upload()
    if df is None:
        st.stop()
    historical_df, historical_source_label = _render_historical_upload()

    df, source_label, historical_df, historical_source_label = _render_operator_sample_management(
        df,
        source_label,
        historical_df,
        historical_source_label,
        audit_readonly=audit_readonly,
    )
    st.caption(f"입력 소스: {source_label}")
    df = _render_input_editor(df, source_label, audit_readonly=audit_readonly)
    source_override = st.session_state.pop(CURRENT_INPUT_SOURCE_OVERRIDE_SESSION_KEY, None)
    if isinstance(source_override, str) and source_override:
        source_label = source_override
    _store_current_input_state(df, source_label)
    _store_historical_input_state(historical_df, historical_source_label)

    metric, as_of_date, _forecast_choice, _provision_choice, config = _render_settings(
        df,
        base_config,
    )
    results = calculate_validated_results(df, as_of_date, metric, config)
    validation_result = results["validation"]

    st.header("입력 검증")
    st.caption("마감일 판정은 is_close_day 컬럼만 사용하며, day_name은 표시용입니다.")
    _render_validation(validation_result)
    _render_input_state_summary(df, validation_result)
    render_next_action_panel(
        "input",
        {
            "validation_result": validation_result,
        },
    )


def _render_forecast_detail_page(context: Mapping[str, Any]) -> None:
    _render_forecast_strategy_detail_page(context)


def _render_scenarios_detail_page(context: Mapping[str, Any]) -> None:
    _render_forecast_strategy_detail_page(context)


def _render_forecast_strategy_detail_page(context: Mapping[str, Any]) -> None:
    if _render_validation_guard(context):
        return

    validation_result = dict(context["validation_result"])
    scenario_df = _as_dataframe(context["scenario_df"])
    next_close_result = dict(context["next_close_result"])
    close_cycle_df = _as_dataframe(context.get("close_cycle_df"))
    forecast_choice = str(context.get("forecast_choice") or COMPARE_LABEL)
    provision_choice = str(context.get("provision_choice") or COMPARE_LABEL)
    selected_scenario_id = _render_selected_scenario_picker(
        scenario_df,
        forecast_choice,
        provision_choice,
    )
    st.session_state[PACE_SELECTED_SCENARIO_SESSION_KEY] = selected_scenario_id
    selected_row = _selected_scenario_row(scenario_df, selected_scenario_id)
    _forecast_result, provision_result = run_selected_scenario_detail(
        _as_dataframe(context["df"]),
        context["as_of_date"],
        str(context["metric"]),
        selected_scenario_id,
        dict(context["config"]),
    )
    revised_targets_df = _as_dataframe(provision_result.get("allocation_by_day"))

    st.markdown(
        '<section class="forecast-strategy-board">'
        '<div class="forecast-strategy-board__eyebrow">Unified forecast strategy board</div>'
        "<h2>예측 · 전략 통합 보드</h2>"
        "<p>F1/F2/F3 예측과 P/O/N 운영전략을 같은 화면에서 비교합니다. "
        "상단은 현재 선택 시나리오 기준 결론, 하단은 전체 시나리오 비교입니다.</p>"
        "</section>",
        unsafe_allow_html=True,
    )
    _render_forecast_strategy_summary_board(
        validation_result,
        next_close_result,
        selected_scenario_id,
        selected_row,
    )

    st.markdown(
        render_section_header(
            "F1/F2/F3 모델 비교",
            "예측모델 간 월말 예상 차이만 compact하게 비교합니다.",
        ),
        unsafe_allow_html=True,
    )
    model_rows_df = _build_forecast_strategy_model_rows(scenario_df)
    if model_rows_df.empty:
        st.info("F1/F2/F3 모델 비교 데이터가 없습니다.")
    else:
        display_columns = [
            "forecast_model",
            "model_name",
            "expected_month_end_amount",
            "target_status",
            "target_variance",
            "risk_level",
        ]
        st.dataframe(
            _format_display_df(model_rows_df.loc[:, display_columns]),
            hide_index=True,
            use_container_width=True,
        )
        _render_forecast_model_mini_chart(model_rows_df, selected_row, validation_result)

    st.markdown(
        render_section_header(
            "선택 시나리오와 운영 판단판",
            "선택값을 바꾸면 추천 행과 상세 분석이 같은 화면에서 함께 갱신됩니다.",
        ),
        unsafe_allow_html=True,
    )
    _render_strategy_reference_sections(
        scenario_df,
        selected_scenario_id,
        selected_row.get("target_status"),
    )
    _render_scenario_operation_matrix(scenario_df, selected_scenario_id)

    with st.expander("상세 차트와 잔여목표", expanded=False):
        _render_strategy_arrival_inline_summary(scenario_df, selected_scenario_id)
        st.markdown("**CloseCycle / Daily revised target**")
        if close_cycle_df.empty:
            st.info("마감차수 요약 데이터가 없습니다.")
        else:
            st.dataframe(_format_display_df(close_cycle_df), hide_index=True, use_container_width=True)
        _render_target_or_strategy_table(
            scenario_df,
            selected_scenario_id,
            revised_targets_df,
            show_all_forecast_models=True,
        )
        _render_forecast_strategy_chart_tabs(
            _as_dataframe(context["df"]),
            scenario_df,
            selected_scenario_id,
            selected_row,
            revised_targets_df,
            close_cycle_df,
            next_close_result,
            validation_result,
            str(context["metric"]),
            context["as_of_date"],
            dict(context["config"]),
        )
    with st.expander("원본 ScenarioGrid", expanded=False):
        _render_selected_scenario_summary(selected_scenario_id, selected_row)
        st.dataframe(_format_display_df(scenario_df), use_container_width=True)
    with st.expander("과거 이력 / Backtest 참고", expanded=False):
        _render_historical_context_panel(
            dict(context.get("historical_context") or {}),
            scenario_df,
            selected_scenario_id,
        )
        _render_forecast_history_backtest_tab(
            scenario_df,
            str(context["metric"]),
            context["as_of_date"],
            audit_readonly=bool(context.get("audit_readonly", False)),
        )

    next_action_context = dict(context)
    next_action_context.update(
        {
            "selected_scenario_id": selected_scenario_id,
            "selected_row": selected_row,
            "revised_targets_df": revised_targets_df,
        }
    )
    render_next_action_panel("forecast_strategy", next_action_context)


def _render_forecast_strategy_summary_board(
    validation_result: Mapping[str, Any],
    next_close_result: Mapping[str, Any],
    selected_scenario_id: str,
    selected_row: pd.Series,
) -> None:
    target_status = selected_row.get("target_status")
    strategy_code = get_strategy_code(
        _scenario_strategy_source(selected_row.to_dict()) or selected_scenario_id
    )
    strategy_label = get_strategy_label(strategy_code)
    strategy_group = get_strategy_group(strategy_code)
    summary_items = (
        ("목표 상태", _localize_display_value(target_status), "현재 선택 시나리오 기준"),
        ("월마감 예상 실적", format_amount(selected_row.get("forecast_after_provision")), selected_scenario_id),
        ("목표 대비 차이", _format_signed_amount(selected_row.get("target_variance")), "공식 월 목표 대비"),
        ("초과 예상분", format_amount(selected_row.get("surplus_to_target")), "초과달성 운영 버퍼"),
        (
            "다음 마감 누적선 필요실적",
            format_amount(next_close_result.get("required_to_recover_next_close_cum")),
            _format_date(next_close_result.get("next_close_date")),
        ),
        ("운영모드", _operation_mode_label(target_status), "P/O/N 전략군 판단"),
        ("권장 전략", f"{strategy_code} {strategy_label}", strategy_group),
        (
            "다음 액션",
            _home_next_action_text(target_status, validation_result, next_close_result),
            "보고 전 확인",
        ),
    )
    cards_html = "".join(
        '<div class="unified-decision-strip__item">'
        f"<span>{escape(str(label))}</span>"
        f"<strong>{escape(str(value))}</strong>"
        f"<small>{escape(str(note))}</small>"
        "</div>"
        for label, value, note in summary_items
    )
    st.markdown(
        '<section class="unified-decision-strip strategy-recommendation-pulse">'
        f"{cards_html}"
        "</section>",
        unsafe_allow_html=True,
    )


def _build_forecast_strategy_model_rows(scenario_df: pd.DataFrame) -> pd.DataFrame:
    model_rows = []
    for forecast_key in ("F1", "F2", "F3"):
        model_df = _scenario_df_for_forecast_key(scenario_df, forecast_key)
        if model_df.empty:
            continue
        row = model_df.iloc[0]
        expected_amount = _first_finite_amount(
            row.get("forecast_amount"),
            row.get("forecast_after_provision"),
        )
        model_rows.append(
            {
                "forecast_model": forecast_key,
                "model_name": FORECAST_MODEL_DEFINITIONS.get(forecast_key, {}).get(
                    "name",
                    forecast_key,
                ),
                "expected_month_end_amount": expected_amount,
                "forecast_amount": expected_amount,
                "target_status": row.get("target_status"),
                "target_variance": row.get("target_variance"),
                "risk_level": row.get("risk_level"),
            }
        )
    return pd.DataFrame(model_rows)


SCENARIO_OPERATION_COLUMNS = (
    "scenario",
    "forecast_model",
    "target_status",
    "strategy_code",
    "strategy_label",
    "strategy_group",
    "expected_month_end_amount",
    "target_variance",
    "surplus_to_target",
    "recommended_action",
    "risk_note",
    "recommended",
)


def build_scenario_operation_matrix(
    scenario_df: pd.DataFrame,
    selected_scenario_id: str | None = None,
) -> pd.DataFrame:
    """Return display-ready ScenarioGrid rows without changing calculations."""
    if scenario_df.empty:
        return pd.DataFrame(columns=SCENARIO_OPERATION_COLUMNS)

    rows: list[dict[str, object]] = []
    selected_id = str(selected_scenario_id or "")
    for _, row in scenario_df.iterrows():
        scenario_id = str(row.get("scenario_id") or row.get("scenario") or "")
        strategy_source = _scenario_strategy_source(row)
        strategy_code = get_strategy_code(strategy_source or scenario_id)
        rows.append(
            {
                "scenario": scenario_id,
                "forecast_model": row.get("forecast_model"),
                "target_status": row.get("target_status"),
                "strategy_code": strategy_code,
                "strategy_label": get_strategy_label(strategy_code),
                "strategy_group": get_strategy_group(strategy_code),
                "expected_month_end_amount": _first_finite_amount(
                    row.get("forecast_after_provision"),
                    row.get("forecast_amount"),
                ),
                "target_variance": row.get("target_variance"),
                "surplus_to_target": row.get("surplus_to_target"),
                "recommended_action": row.get("recommended_action"),
                "risk_note": _scenario_risk_note(row),
                "recommended": "추천" if scenario_id == selected_id else "",
            }
        )
    return pd.DataFrame(rows, columns=SCENARIO_OPERATION_COLUMNS)


def _render_scenario_operation_matrix(
    scenario_df: pd.DataFrame,
    selected_scenario_id: str,
) -> None:
    st.markdown(
        render_section_header(
            "운영 판단판",
            "전략 코드는 유지하되, 기본 표는 한국어 전략명과 전략군을 우선 표시합니다.",
        ),
        unsafe_allow_html=True,
    )
    matrix = build_scenario_operation_matrix(scenario_df, selected_scenario_id)
    if matrix.empty:
        st.info("시나리오 운영 판단판에 표시할 데이터가 없습니다.")
        return

    st.markdown(
        '<div class="scenario-operation-note">'
        "추천 행은 현재 선택된 시나리오 기준입니다. 원본 ScenarioGrid는 아래 접힌 영역에서 확인합니다."
        "</div>",
        unsafe_allow_html=True,
    )
    display = _format_display_df(matrix)
    recommended_label = _display_column_label("recommended")

    def highlight_recommended(row: pd.Series) -> list[str]:
        is_recommended = str(row.get(recommended_label, "")) == "추천"
        style = "background-color: #eef6f4; font-weight: 700;" if is_recommended else ""
        return [style] * len(row)

    st.dataframe(
        display.style.apply(highlight_recommended, axis=1),
        hide_index=True,
        use_container_width=True,
    )


def _scenario_strategy_source(row: Mapping[str, object]) -> object:
    for key in (
        "overachievement_strategy",
        "provision_strategy",
        "neutral_strategy",
        "strategy_id",
        "strategy_code",
        "scenario_id",
    ):
        value = row.get(key)
        if not _is_missing(value) and str(value):
            return value
    return ""


def _scenario_risk_note(row: Mapping[str, object]) -> str:
    for key in ("risk_note", "comment", "warnings", "recommended_action"):
        value = row.get(key)
        if _is_missing(value):
            continue
        if isinstance(value, (list, tuple, set)):
            text = ", ".join(str(item) for item in value if str(item))
        else:
            text = str(value)
        if text.strip():
            return text.strip()
    return "특이 리스크 없음"


def _render_report_detail_page(context: Mapping[str, Any]) -> None:
    if _render_validation_guard(context):
        return

    report_text = str(context.get("report_text") or "입력 후 계산됩니다.")
    selected_row = _as_series(context.get("selected_row"))
    validation_result = dict(context.get("validation_result") or {})
    next_close_result = dict(context.get("next_close_result") or {})
    report_basis_date = _format_date(context.get("as_of_date"))
    report_name = str(context.get("report_name") or "기존 Excel 리포트 조회 전")
    target_status = selected_row.get("target_status")
    st.markdown(
        render_section_header(
            "보고 메모",
            "판단 메모는 내부 검토용 구조화 요약이고, 복사용 보고문은 공유 채널에 붙여넣는 최종 문안입니다.",
        ),
        unsafe_allow_html=True,
    )
    st.markdown(
        '<div class="report-ia-note">'
        "동일한 report_builder 결과를 기준으로 하되, 위쪽은 의사결정에 필요한 근거를 카드로 나누고 "
        "아래쪽은 전체 선택 후 복사하기 쉬운 원문 형태로 제공합니다."
        "</div>",
        unsafe_allow_html=True,
    )
    st.markdown(
        '<div class="report-meta-row">'
        f'<span><strong>보고 기준일</strong><br>{escape(report_basis_date)}</span>'
        f'<span><strong>Excel 포함 여부</strong><br>{escape(report_name)} 기준, 이 화면에서는 새 파일을 만들지 않음</span>'
        f'<span><strong>보고 상태</strong><br>{escape(_localize_display_value(target_status))}</span>'
        "</div>",
        unsafe_allow_html=True,
    )

    memo_items = (
        (
            "요약",
            f"{_localize_display_value(target_status)} / 월말 예상 {format_amount(selected_row.get('forecast_after_provision'))}",
        ),
        (
            "운영 판단",
            _operation_mode_label(target_status),
        ),
        (
            "리스크",
            f"{_localize_display_value(selected_row.get('risk_level', '계산 불가'))} / {_localize_display_value(selected_row.get('status', '계산 불가'))}",
        ),
        (
            "권장 액션",
            _home_next_action_text(target_status, validation_result, next_close_result),
        ),
        (
            "근거 수치",
            f"목표 대비 {_format_signed_amount(selected_row.get('target_variance'))}, 다음 마감 누적선 필요실적 {format_amount(next_close_result.get('required_to_recover_next_close_cum'))}",
        ),
        (
            "공유 전 확인",
            "복사용 보고문에서 문장 흐름을 확인하고 필요한 채널에 맞게 전체 선택 후 복사합니다.",
        ),
    )
    memo_cards = "".join(
        '<div class="report-memo-card">'
        f"<strong>{escape(title)}</strong>"
        f"<span>{escape(body)}</span>"
        "</div>"
        for title, body in memo_items
    )
    st.markdown(
        render_section_header(
            "판단 메모 · 내부 검토용",
            "요약, 운영 판단, 리스크, 권장 액션, 근거 수치를 분리해 검토합니다.",
        ),
        unsafe_allow_html=True,
    )
    st.markdown(f'<div class="report-memo-grid">{memo_cards}</div>', unsafe_allow_html=True)
    _render_report_glossary_panel()

    st.markdown(
        render_section_header(
            "복사용 보고문 · 공유용",
            "팀장/임원/공유 채널에 그대로 붙여넣기 위한 최종 문장입니다.",
        ),
        unsafe_allow_html=True,
    )
    st.caption("복사 버튼 대신 텍스트 영역을 전체 선택한 뒤 복사하세요. 보고문 생성 로직은 변경하지 않았습니다.")
    report_key = hashlib.sha1(report_text.encode("utf-8")).hexdigest()[:12]
    st.text_area("복사용 보고문", value=report_text, height=360, key=f"report_page_{report_key}")
    render_next_action_panel("report", context)


def _render_history_detail_page(context: Mapping[str, Any]) -> None:
    if _render_validation_guard(context):
        return
    _render_forecast_history_backtest_tab(
        _as_dataframe(context["scenario_df"]),
        str(context["metric"]),
        context["as_of_date"],
        audit_readonly=bool(context.get("audit_readonly", False)),
    )
    render_next_action_panel("history", context)


def _render_excel_freshness_badge(report_path: Path) -> None:
    stat = report_path.stat()
    generated_at = pd.Timestamp.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S")
    st.markdown(
        '<section class="excel-freshness-badge">'
        '<div>'
        '<div class="excel-freshness-badge__label">Excel freshness</div>'
        f"<strong>{escape(report_path.name)}</strong>"
        '</div>'
        f"<span>생성/수정시각 {escape(generated_at)} · {stat.st_size:,} bytes</span>"
        "</section>",
        unsafe_allow_html=True,
    )


def _render_excel_detail_page(context: Mapping[str, Any]) -> None:
    if _render_validation_guard(context):
        return

    audit_readonly = bool(context.get("audit_readonly", False))
    scenario_df = _as_dataframe(context["scenario_df"])
    revised_targets_df = _as_dataframe(context["revised_targets_df"])
    st.markdown(
        render_section_header(
            "Excel 공유",
            "읽기 전용 상태 조회가 기본이며, 파일 생성은 명시적 버튼 클릭 시에만 실행합니다.",
        ),
        unsafe_allow_html=True,
    )

    latest_files = list_latest_excel_outputs()
    latest_report_path = _latest_existing_report_path(latest_files)
    st.markdown(
        '<section class="excel-readonly-panel">'
        '<div class="excel-readonly-panel__label">읽기 전용</div>'
        '<h3>outputs/latest 상태 조회</h3>'
        '<p>이 페이지를 열거나 화면 캡처하는 동안 Excel 리포트는 생성하거나 수정하지 않습니다. '
        '파일 생성은 아래 재생성 버튼을 누를 때만 실행됩니다.</p>'
        "</section>",
        unsafe_allow_html=True,
    )
    if audit_readonly:
        st.info("읽기 전용 감리 모드: Excel 생성/재생성 버튼이 비활성화됩니다.")
    if latest_files.empty:
        st.info("outputs/latest에 공유 가능한 Excel 파일이 없습니다. 생성 필요 상태입니다.")
    else:
        st.dataframe(latest_files, hide_index=True, use_container_width=True)

    if latest_report_path is None:
        st.info("기존 daily_report Excel 리포트가 없어 생성 필요 상태입니다.")
    else:
        _render_excel_freshness_badge(latest_report_path)
        st.markdown(render_download_card(latest_report_path.name), unsafe_allow_html=True)
        st.download_button(
            "기존 Excel 리포트 다운로드",
            data=latest_report_path.read_bytes(),
            file_name=latest_report_path.name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    if st.button(
        "최신 리포트 재생성",
        key="regenerate_latest_excel_report",
        disabled=audit_readonly,
    ):
        report_bytes, report_name = build_excel_report_bytes(
            dict(context["summary_dict"]),
            scenario_df,
            revised_targets_df,
            _as_dataframe(context["close_cycle_df"]),
            build_display_validation_result(dict(context["validation_result"])),
            str(context["report_text"]),
            str(context["metric"]),
            context["as_of_date"],
        )
        st.success(f"Excel 리포트를 생성했습니다: {report_name}")
        st.download_button(
            "생성된 Excel 리포트 다운로드",
            data=report_bytes,
            file_name=report_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    st.caption("archive_invalid와 archive_old_format는 기본 공유 대상이 아닙니다.")
    render_next_action_panel("excel", context)
    st.markdown("**ScenarioGrid 최신 컬럼 점검**")
    st.dataframe(_scenario_grid_column_check_df(scenario_df), hide_index=True, use_container_width=True)


def _render_audit_detail_page(context: Mapping[str, Any]) -> None:
    st.markdown(
        render_section_header(
            "검증 · 운영관리",
            "테스트, Gate Runner, 금지 패턴, 보안 운영 정책을 요약합니다.",
        ),
        unsafe_allow_html=True,
    )
    validation_result = dict(context.get("validation_result") or {})
    if bool(context.get("audit_readonly", False)):
        st.info("읽기 전용 감리 모드: 이 페이지 조회 중 outputs 파일을 생성하거나 갱신하지 않습니다.")
    st.write(f"- 현재 입력 검증 오류: {len(validation_result.get('errors') or [])}건")
    st.write(f"- 현재 입력 검증 주의: {len(validation_result.get('warnings') or [])}건")
    st.write("- 현재 화면 보정 기준: U03-A1.1 Visual Correction + Excel Read-only Fix")
    st.write("- Auth Gate: 로컬 운영 기준이며, 외부 배포 전 보안 복원 예정입니다.")
    st.write("- `.streamlit/secrets.toml`은 로컬 전용이며 감리/배포 패키지에 포함하지 않습니다.")
    st.write("- 실데이터, 고객/계약/개인 식별정보의 외부 공개는 금지됩니다.")
    st.warning(SECURITY_WARNING_TEXT)

    st.markdown("**현재 검증 로그**")
    for label, path in (
        ("pytest", REPO_ROOT / "audit" / "logs" / "u03_a1_1_pytest.txt"),
        ("Gate Runner", REPO_ROOT / "audit" / "logs" / "u03_a1_1_gate_runner_all.json"),
        ("금지 패턴", REPO_ROOT / "audit" / "logs" / "u03_a1_1_forbidden_pattern_scan.txt"),
        ("outputs mtime", REPO_ROOT / "audit" / "logs" / "u03_a1_1_output_mtime_check.txt"),
    ):
        st.caption(f"{label}: {path}")
    render_next_action_panel("audit", context)


def _render_input_state_summary(
    df: pd.DataFrame,
    validation_result: dict[str, Any],
) -> None:
    close_flags = _coerce_is_close_day(df["is_close_day"]) if "is_close_day" in df.columns else pd.Series(dtype=bool)
    cols = st.columns(4)
    cols[0].metric("입력 행 수", len(df))
    cols[1].metric("마감일 행 수", int(close_flags.sum()) if not close_flags.empty else 0)
    cols[2].metric("월 목표", format_amount(validation_result.get("monthly_target")))
    cols[3].metric("현재 누적 실적", format_amount(validation_result.get("current_actual_cum")))
    with st.expander("입력표 확인", expanded=False):
        st.dataframe(_format_display_df(df), hide_index=True, use_container_width=True)


def _render_validation_guard(context: Mapping[str, Any]) -> bool:
    validation_result = dict(context.get("validation_result") or {})
    errors = list(validation_result.get("errors") or [])
    if not errors:
        return False
    st.warning("입력값에 고쳐야 할 항목이 있어 이 상세 페이지의 계산 결과를 표시하지 않습니다.")
    _render_validation(validation_result)
    return True


def prepare_scenario_grid_export_frame(data: pd.DataFrame | Any) -> pd.DataFrame:
    if callable(_excel_prepare_scenario_grid_export_frame):
        return _excel_prepare_scenario_grid_export_frame(data)
    return _prepare_scenario_grid_export_frame_fallback(data)


def _prepare_scenario_grid_export_frame_fallback(data: pd.DataFrame | Any) -> pd.DataFrame:
    df = _as_dataframe(data)
    if df.empty and len(df.columns) == 0:
        return pd.DataFrame(columns=SCENARIO_GRID_REQUIRED_COLUMNS)

    result = df.copy()
    result["scenario"] = _first_existing_series(result, ("scenario", "scenario_id"))
    result["forecast_model"] = _first_existing_series(result, ("forecast_model",))
    result["model_name"] = result["forecast_model"].map(get_forecast_model_label)
    result["strategy_code"] = [
        get_strategy_code(_scenario_strategy_source(row.to_dict()))
        for _, row in result.iterrows()
    ]
    if "strategy_type" not in result.columns:
        result["strategy_type"] = None
    result["strategy_label"] = result["strategy_code"].map(get_strategy_label)
    result["strategy_group"] = result["strategy_code"].map(get_strategy_group)
    result["expected_month_end_amount"] = [
        _first_non_missing_value(row, ("forecast_after_provision", "forecast_amount"))
        for _, row in result.iterrows()
    ]
    result["risk_note"] = [
        _first_non_missing_value(row, ("risk_note", "comment", "warnings", "recommended_action"))
        or ""
        for _, row in result.iterrows()
    ]

    for column in SCENARIO_GRID_REQUIRED_COLUMNS:
        if column not in result.columns:
            result[column] = None
    ordered_columns = [
        *[column for column in SCENARIO_GRID_REQUIRED_COLUMNS if column in result.columns],
        *[column for column in result.columns if column not in set(SCENARIO_GRID_REQUIRED_COLUMNS)],
    ]
    return result.loc[:, ordered_columns]


def _first_existing_series(df: pd.DataFrame, columns: tuple[str, ...]) -> pd.Series:
    for column in columns:
        if column in df.columns:
            return df[column]
    return pd.Series([None] * len(df), index=df.index)


def _first_non_missing_value(row: Mapping[str, Any], columns: tuple[str, ...]) -> object:
    for column in columns:
        value = row.get(column)
        if not _is_missing(value):
            return value
    return None


def _scenario_grid_column_check_df(scenario_df: pd.DataFrame) -> pd.DataFrame:
    export_frame = prepare_scenario_grid_export_frame(scenario_df)
    return pd.DataFrame(
        {
            "column": SCENARIO_GRID_REQUIRED_COLUMNS,
            "present": [column in export_frame.columns for column in SCENARIO_GRID_REQUIRED_COLUMNS],
        }
    )


def _require_access_password() -> bool:
    configured_password, configured_password_hash = _configured_access_credentials()
    is_configured = bool(configured_password or configured_password_hash)
    if st.session_state.get(ACCESS_SESSION_STATE_KEY) and is_configured:
        _render_access_logout()
        return True

    st.title("월마감 영업실적 예측툴")
    st.caption("링크를 받은 대상자에게만 제공되는 제한 배포 화면입니다.")

    if not is_configured:
        st.warning("접속 비밀번호가 설정되지 않아 앱을 열 수 없습니다.")
        st.code(
            "APP_ACCESS_PASSWORD=공유할_접속_비밀번호",
            language="text",
        )
        st.caption("배포 환경의 Secrets 또는 환경변수에 위 값을 설정한 뒤 앱을 다시 시작하세요.")
        return False

    with st.form("access_password_form", clear_on_submit=True):
        password = st.text_input("접속 비밀번호", type="password")
        submitted = st.form_submit_button("입장")

    if submitted:
        if verify_access_password(
            password,
            configured_password=configured_password,
            configured_password_hash=configured_password_hash,
        ):
            st.session_state[ACCESS_SESSION_STATE_KEY] = True
            _rerun_app()
            return False
        st.error("비밀번호가 일치하지 않습니다.")

    return False


def _render_access_logout() -> None:
    with st.sidebar:
        st.caption("제한 배포 모드")
        if st.button("로그아웃"):
            st.session_state.pop(ACCESS_SESSION_STATE_KEY, None)
            _rerun_app()


def _rerun_app() -> None:
    if hasattr(st, "rerun"):
        st.rerun()
    else:  # pragma: no cover - compatibility for older Streamlit runtimes.
        st.experimental_rerun()


def _configured_access_credentials() -> tuple[str | None, str | None]:
    return (
        _read_configured_setting(ACCESS_PASSWORD_SETTING_KEYS),
        _read_configured_setting(ACCESS_PASSWORD_HASH_SETTING_KEYS),
    )


def _read_configured_setting(setting_keys: tuple[str, ...]) -> str | None:
    for key in setting_keys:
        value = _clean_optional_secret(os.environ.get(key))
        if value:
            return value

    if st is None:
        return None

    for key in setting_keys:
        try:
            value = _clean_optional_secret(st.secrets.get(key))
        except Exception:  # noqa: BLE001 - Streamlit raises different errors when secrets are absent.
            continue
        if value:
            return value
    return None


def _clean_optional_secret(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def verify_access_password(
    candidate_password: str,
    *,
    configured_password: str | None = None,
    configured_password_hash: str | None = None,
) -> bool:
    """Verify an access password against plain or SHA-256 configured credentials."""
    candidate = str(candidate_password or "")
    expected_password = _clean_optional_secret(configured_password)
    expected_hash = _clean_optional_secret(configured_password_hash)

    if expected_password and hmac.compare_digest(candidate, expected_password):
        return True

    if expected_hash:
        candidate_hash = hashlib.sha256(candidate.encode("utf-8")).hexdigest()
        return hmac.compare_digest(candidate_hash, expected_hash.lower())

    return False


def _inject_app_styles() -> None:
    st.markdown(
        """
        <style>
        :root {
            --app-bg: #f6f8fb;
            --panel-bg: #ffffff;
            --line-soft: #d9e2ec;
            --line-strong: #b8c6d6;
            --text-main: #18212f;
            --text-muted: #667085;
            --accent: #1f6f78;
            --accent-soft: #e7f3f4;
            --font-title: 1.72rem;
            --font-section: 1.12rem;
            --font-subsection: 0.98rem;
            --font-body: 0.88rem;
            --font-control: 0.86rem;
            --font-caption: 0.78rem;
            --font-small: 0.72rem;
            --font-kpi-value: 1.02rem;
            --metric-card-height: 100px;
            --line-body: 1.48;
            --line-tight: 1.22;
        }

        .stApp {
            background: var(--app-bg);
            color: var(--text-main);
            font-size: var(--font-body);
            line-height: var(--line-body);
        }

        [data-testid="stHeader"] {
            background: rgba(246, 248, 251, 0.92);
            border-bottom: 1px solid rgba(217, 226, 236, 0.72);
        }

        footer {
            display: none !important;
            visibility: hidden !important;
        }

        .block-container {
            max-width: 1480px;
            padding-top: 0.78rem;
            padding-bottom: 3rem;
        }

        div[data-testid="stVerticalBlock"] {
            gap: 0.55rem;
        }

        h1 {
            color: var(--text-main);
            font-size: var(--font-title) !important;
            font-weight: 720 !important;
            line-height: var(--line-tight) !important;
            letter-spacing: 0 !important;
            margin-bottom: 1.1rem !important;
        }

        h2 {
            color: var(--text-main);
            font-size: var(--font-section) !important;
            font-weight: 680 !important;
            line-height: var(--line-tight) !important;
            letter-spacing: 0 !important;
            margin-top: 1.6rem !important;
            padding-top: 1.05rem !important;
            border-top: 1px solid var(--line-soft);
        }

        h3 {
            color: var(--text-main);
            font-size: var(--font-subsection) !important;
            font-weight: 650 !important;
            line-height: var(--line-tight) !important;
            letter-spacing: 0 !important;
            margin-top: 1rem !important;
            margin-bottom: 0.45rem !important;
        }

        p, li, label, div[data-testid="stMarkdownContainer"] {
            color: var(--text-main);
            font-size: var(--font-body) !important;
            line-height: var(--line-body) !important;
            letter-spacing: 0 !important;
        }

        div[data-testid="stCaptionContainer"],
        div[data-testid="stCaptionContainer"] p,
        small {
            color: var(--text-muted) !important;
            font-size: var(--font-caption) !important;
            line-height: 1.38 !important;
            letter-spacing: 0 !important;
        }

        [data-testid="stMetric"] {
            height: var(--metric-card-height);
            min-height: var(--metric-card-height);
            box-sizing: border-box;
            padding: 0.58rem 0.72rem;
            display: flex;
            flex-direction: column;
            justify-content: flex-start;
            background: var(--panel-bg);
            border: 1px solid var(--line-soft);
            border-left: 3px solid var(--accent);
            border-radius: 8px;
            box-shadow: 0 1px 2px rgba(16, 24, 40, 0.05);
        }

        [data-testid="stMetricLabel"],
        [data-testid="stMetricLabel"] p {
            color: var(--text-muted) !important;
            font-size: var(--font-small) !important;
            font-weight: 620 !important;
            line-height: var(--line-tight) !important;
            margin-bottom: 0.2rem !important;
        }

        [data-testid="stMetricValue"] {
            color: var(--text-main) !important;
            font-size: var(--font-kpi-value) !important;
            font-weight: 700 !important;
            line-height: var(--line-tight) !important;
            letter-spacing: 0 !important;
        }

        [data-testid="stMetricDelta"] {
            font-size: var(--font-small) !important;
            line-height: 1.15 !important;
        }

        div[data-testid="stAlert"] {
            border-radius: 8px;
            border: 1px solid var(--line-soft);
            box-shadow: 0 1px 2px rgba(16, 24, 40, 0.04);
        }

        div[data-testid="stDataFrame"],
        div[data-testid="stTable"] {
            border: 1px solid var(--line-soft);
            border-radius: 8px;
            overflow: hidden;
            background: var(--panel-bg);
        }

        div[data-testid="stDataFrame"] *,
        div[data-testid="stTable"] *,
        [data-testid="stDataEditor"] * {
            font-size: var(--font-control) !important;
            line-height: 1.36 !important;
        }

        div[data-testid="stTabs"] [role="tablist"] {
            gap: 0.2rem;
            border-bottom: 1px solid var(--line-soft);
        }

        button[data-baseweb="tab"] {
            border-radius: 6px 6px 0 0;
            padding: 0.42rem 0.76rem;
            color: var(--text-muted);
            font-size: var(--font-control) !important;
            font-weight: 620;
            line-height: var(--line-tight) !important;
        }

        button[data-baseweb="tab"][aria-selected="true"] {
            background: var(--accent-soft);
            color: var(--accent);
        }

        div[data-testid="stFileUploader"] section,
        div[data-testid="stExpander"] details {
            border: 1px solid var(--line-soft);
            border-radius: 8px;
            background: var(--panel-bg);
        }

        [data-testid="stFileUploaderDropzone"] button [data-testid="stIconMaterial"],
        [data-testid="stFileUploaderDropzone"] button [data-testid="stMarkdownContainer"],
        [data-testid="stFileUploaderDropzoneInstructions"] {
            display: none !important;
        }

        [data-testid="stFileUploaderDropzone"] button::after {
            content: "파일 선택";
            font-size: var(--font-control);
            font-weight: 650;
            color: var(--text-main);
        }

        [data-testid="stFileUploaderDropzone"]::after {
            content: "CSV 또는 XLSX 파일을 업로드할 수 있습니다.";
            color: var(--text-muted);
            font-size: var(--font-caption);
            line-height: 1.35;
        }

        div[data-testid="stExpander"] summary {
            font-size: var(--font-control) !important;
            font-weight: 650;
            color: var(--text-main);
            line-height: var(--line-tight) !important;
        }

        div[data-testid="stButton"] button,
        div[data-testid="stDownloadButton"] button {
            border-radius: 6px;
            border: 1px solid var(--line-strong);
            font-size: var(--font-control) !important;
            font-weight: 650;
            line-height: var(--line-tight) !important;
            box-shadow: 0 1px 2px rgba(16, 24, 40, 0.05);
        }

        div[data-testid="stSelectbox"] label,
        div[data-testid="stDateInput"] label,
        div[data-testid="stNumberInput"] label,
        div[data-testid="stTextInput"] label,
        div[data-testid="stFileUploader"] label {
            font-size: var(--font-control) !important;
            font-weight: 620 !important;
            line-height: var(--line-tight) !important;
        }

        div[data-baseweb="select"] *,
        div[data-baseweb="input"] *,
        div[data-baseweb="base-input"] *,
        div[data-testid="stDateInput"] * {
            font-size: var(--font-control) !important;
            line-height: 1.34 !important;
        }

        textarea,
        input {
            border-radius: 6px !important;
            font-size: var(--font-control) !important;
            line-height: 1.42 !important;
        }

        textarea[aria-label="보고 메모"] {
            font-size: var(--font-body) !important;
            line-height: 1.58 !important;
        }
        .page-header {
            margin: 0 0 8px !important;
            padding: 0 !important;
        }

        .page-header__eyebrow {
            font-size: 11px !important;
            line-height: 1.15 !important;
            margin-bottom: 2px !important;
        }

        .page-header h1 {
            margin: 0 0 5px !important;
            font-size: 25px !important;
            line-height: 1.14 !important;
        }

        .page-header__subtitle {
            font-size: 12px !important;
            line-height: 1.28 !important;
            margin-top: 0 !important;
        }

        .workbench-shell,
        .workbench-main {
            color: #25312f;
        }

        .workbench-shell.top-status-bar,
        .page-header-compact {
            border: 1px solid #d8d1c6;
            border-radius: 8px;
            background: #fffdf8;
            padding: 8px 12px;
            margin: 0 0 6px;
            box-shadow: 0 1px 2px rgba(54, 46, 36, 0.06);
        }

        .page-header-compact {
            display: flex;
            align-items: flex-end;
            justify-content: space-between;
            gap: 12px;
            padding: 0;
            margin: 0;
            border: 0;
            box-shadow: none;
        }

        .page-header-compact__eyebrow {
            color: #2f6f68;
            font-size: 11px;
            font-weight: 850;
            line-height: 1.2;
        }

        .page-header-compact h1 {
            margin: 2px 0 0 !important;
            font-size: 19px !important;
            line-height: 1.15 !important;
            font-weight: 850 !important;
            letter-spacing: 0 !important;
        }

        .page-header-compact p {
            margin: 2px 0 0;
            color: #66716e;
            font-size: 12px !important;
            line-height: 1.25 !important;
        }

        .workbench-fact-row {
            display: grid;
            grid-template-columns: repeat(5, minmax(130px, 1fr));
            gap: 8px;
            margin: 8px 0 10px;
        }

        .metric-card-compact {
            min-height: 52px;
            border: 1px solid #d8d1c6;
            border-radius: 8px;
            background: #faf8f3;
            padding: 8px 10px;
        }

        .metric-card-compact span {
            display: block;
            color: #66716e;
            font-size: 11px;
            font-weight: 800;
            line-height: 1.2;
        }

        .metric-card-compact strong {
            display: block;
            margin-top: 3px;
            color: #25312f;
            font-size: 14px;
            font-weight: 850;
            line-height: 1.24;
            overflow-wrap: anywhere;
        }

        .projection-chart-card,
        .decision-panel,
        .report-card,
        .excel-card,
        .empty-state {
            border: 1px solid #d8d1c6;
            border-radius: 8px;
            background: #fffdf8;
            box-shadow: 0 1px 2px rgba(54, 46, 36, 0.06);
        }

        .projection-chart-card {
            padding: 8px 10px 8px;
            min-height: 0;
        }

        .projection-chart-card__head {
            display: flex;
            align-items: flex-start;
            justify-content: space-between;
            gap: 12px;
            margin-bottom: 4px;
        }

        .projection-chart-card__label {
            color: #25312f;
            font-size: 16px;
            font-weight: 850;
            line-height: 1.25;
        }

        .projection-chart-card__copy {
            margin-top: 2px;
            color: #66716e;
            font-size: 12px;
            line-height: 1.3;
        }

        .chart-legend-row {
            display: flex;
            flex-wrap: wrap;
            gap: 8px 12px;
            padding: 8px 2px 0;
            color: #66716e;
            font-size: 12px;
            font-weight: 750;
        }

        .chart-legend-row span {
            display: inline-flex;
            align-items: center;
            gap: 6px;
        }

        .chart-legend-row i {
            display: inline-block;
            width: 18px;
            height: 0;
            border-top: 3px solid #2f6f68;
        }

        .chart-legend-row .legend-target {
            border-color: #8a8f98;
            border-top-style: dashed;
        }

        .chart-legend-row .legend-actual {
            border-color: #0f766e;
        }

        .chart-legend-row .legend-projection {
            border-color: #2f6f68;
            border-top-style: dashed;
        }

        .chart-legend-row .legend-band {
            height: 10px;
            border: 0;
            background: #dcebe8;
        }

        .chart-legend-row .legend-close {
            border-color: #b48632;
            border-top-style: dashed;
        }

        .chart-legend-row .legend-current {
            width: 10px;
            height: 10px;
            border: 0;
            border-radius: 999px;
            background: #25312f;
        }

        .decision-panel {
            padding: 10px 12px;
            min-height: 0;
        }

        .decision-panel__label {
            color: #2f6f68;
            font-size: 11px;
            font-weight: 900;
            line-height: 1.2;
            text-transform: uppercase;
        }

        .decision-panel h2 {
            margin: 4px 0 6px !important;
            padding: 0 !important;
            border: 0 !important;
            color: #25312f !important;
            font-size: 17px !important;
            line-height: 1.2 !important;
            font-weight: 850 !important;
        }

        .decision-panel__row {
            display: grid;
            grid-template-columns: 96px 1fr;
            gap: 7px;
            padding: 6px 0;
            border-top: 1px solid #e6ded2;
        }

        .decision-panel__row span {
            color: #66716e;
            font-size: 12px;
            font-weight: 800;
            line-height: 1.3;
        }

        .decision-panel__row strong {
            color: #25312f;
            font-size: 13px;
            font-weight: 800;
            line-height: 1.35;
            overflow-wrap: anywhere;
        }

        .strategy-card-row {
            display: grid;
            grid-template-columns: repeat(3, minmax(220px, 1fr));
            gap: 10px;
            margin: 8px 0 16px;
        }

        .strategy-section,
        .next-action-panel,
        .excel-readonly-panel,
        .compact-arrival-chart {
            border: 1px solid #d8d1c6;
            border-radius: 8px;
            background: #fffdf8;
            box-shadow: 0 1px 2px rgba(54, 46, 36, 0.05);
        }

        .strategy-section {
            padding: 8px;
            margin: 6px 0;
        }

        .strategy-section__head {
            display: flex;
            align-items: flex-start;
            justify-content: space-between;
            gap: 12px;
            margin-bottom: 8px;
        }

        .strategy-section__status,
        .next-action-panel__label,
        .excel-readonly-panel__label {
            color: #2f6f68;
            font-size: 11px;
            font-weight: 900;
            line-height: 1.2;
            text-transform: uppercase;
        }

        .strategy-section__head p {
            margin: 2px 0 0;
            color: #66716e;
            font-size: 12px !important;
            line-height: 1.32 !important;
        }

        .strategy-section__head span {
            flex: 0 0 auto;
            border: 1px solid #cfd8d5;
            border-radius: 999px;
            background: #eef5f3;
            color: #193b37;
            padding: 4px 8px;
            font-size: 11px;
            font-weight: 850;
        }

        .strategy-section__cards {
            display: grid;
            grid-template-columns: repeat(3, minmax(200px, 1fr));
            gap: 7px;
        }

        .strategy-card-shell {
            min-width: 0;
        }

        .strategy-section .scenario-card {
            min-height: 108px;
            border-radius: 8px;
            padding: 8px;
        }

        .strategy-section .scenario-card__topline {
            margin-bottom: 5px;
        }

        .strategy-section .scenario-card__name {
            font-size: 14px;
            line-height: 1.18;
            margin-bottom: 4px;
        }

        .strategy-section .scenario-card__description {
            min-height: 0;
            max-height: 30px;
            overflow: hidden;
            font-size: 11px;
            line-height: 1.32;
        }

        .strategy-section .scenario-card__metrics {
            gap: 6px;
            margin-top: 5px;
        }

        .strategy-section .scenario-card__metric {
            border-radius: 8px;
            padding: 6px;
        }

        .strategy-section .scenario-card__metric-label {
            font-size: 10px;
            margin-bottom: 2px;
        }

        .strategy-section .scenario-card__metric-value {
            font-size: 12px;
            line-height: 1.2;
        }

        .strategy-card-shell__code {
            margin-bottom: 4px;
            color: #53615e;
            font-size: 11px;
            font-weight: 850;
            line-height: 1.25;
            overflow-wrap: anywhere;
        }

        .strategy-card-inactive {
            opacity: 0.72;
        }

        .strategy-card-inactive .scenario-card {
            background: #fbfaf6;
        }

        .compact-arrival-chart {
            padding: 10px;
            margin: 6px 0 8px;
        }

        .scenario-inline-chart-title {
            display: flex;
            align-items: baseline;
            gap: 8px;
            margin: 4px 0 0;
            color: #25312f;
            font-size: 12px;
            line-height: 1.25;
        }

        .scenario-inline-chart-title strong {
            font-size: 13px;
            font-weight: 900;
        }

        .scenario-inline-chart-title span {
            color: #66716e;
            font-size: 11px;
            font-weight: 760;
        }

        .compact-arrival-target {
            margin-bottom: 7px;
            color: #636967;
            font-size: 12px;
            font-weight: 850;
        }

        .compact-arrival-row {
            display: grid;
            grid-template-columns: minmax(136px, 0.28fr) minmax(160px, 1fr) 86px;
            gap: 8px;
            align-items: center;
            padding: 4px 0;
        }

        .compact-arrival-row__label {
            color: #25312f;
            font-size: 12px;
            font-weight: 850;
            line-height: 1.22;
            overflow-wrap: anywhere;
        }

        .compact-arrival-row__label span {
            display: block;
            color: #66716e;
            font-size: 10px;
            font-weight: 760;
        }

        .compact-arrival-row__track {
            height: 14px;
            border-radius: 999px;
            background: #ece6dc;
            overflow: hidden;
        }

        .compact-arrival-row__bar {
            height: 100%;
            border-radius: 999px;
            background: #2f6f68;
        }

        .compact-arrival-row__bar--p {
            background: #b48632;
        }

        .compact-arrival-row__bar--o {
            background: #567c5d;
        }

        .compact-arrival-row__bar--n {
            background: #51758c;
        }

        .compact-arrival-row__bar.is-selected {
            box-shadow: inset 0 0 0 2px rgba(37, 49, 47, 0.34);
        }

        .compact-arrival-row__value {
            color: #25312f;
            font-size: 12px;
            font-weight: 850;
            text-align: right;
        }

        .next-action-panel {
            padding: 10px 12px;
            margin: 10px 0;
        }

        .next-action-panel h3,
        .excel-readonly-panel h3 {
            margin: 3px 0 6px !important;
            padding: 0 !important;
            border: 0 !important;
            color: #25312f !important;
            font-size: 15px !important;
            font-weight: 850 !important;
        }

        .next-action-panel ul {
            margin: 0;
            padding-left: 18px;
        }

        .next-action-panel li {
            margin: 2px 0;
            color: #25312f;
            font-size: 13px !important;
            line-height: 1.38 !important;
        }

        .excel-readonly-panel {
            padding: 10px 12px;
            margin: 8px 0;
        }

        .excel-readonly-panel p {
            margin: 0;
            color: #66716e;
            font-size: 12px !important;
            line-height: 1.38 !important;
        }

        .strategy-card,
        .strategy-card-under,
        .strategy-card-over,
        .strategy-card-neutral {
            border-radius: 8px;
        }

        .strategy-card-under {
            border-color: #d1b98a;
            background: #f6ead4;
        }

        .strategy-card-over {
            border-color: #9fc0a5;
            background: #dfe9de;
        }

        .strategy-card-neutral {
            border-color: #a6b9cb;
            background: #dce6ed;
        }

        section[data-testid="stSidebar"] .nav-rail {
            max-width: 244px;
            padding: 8px;
            border: 1px solid #d8d1c6;
            border-radius: 8px;
            background: #fffdf8;
        }

        section[data-testid="stSidebar"] .nav-rail__title {
            margin: 0 0 8px;
            color: #66716e;
            font-size: 11px;
            font-weight: 900;
            letter-spacing: 0;
        }

        section[data-testid="stSidebar"] .nav-rail a.nav-item,
        section[data-testid="stSidebar"] .nav-rail a.nav-item:visited,
        .mini-nav a.nav-item,
        .mini-nav a.nav-item:visited {
            display: block;
            position: relative;
            margin: 3px 0;
            padding: 8px 10px 8px 12px;
            border: 1px solid transparent;
            border-radius: 8px;
            background: transparent;
            color: #25312f !important;
            text-decoration: none !important;
            box-shadow: none;
        }

        section[data-testid="stSidebar"] .nav-rail a.nav-item:hover,
        .mini-nav a.nav-item:hover {
            border-color: #d2dfdc;
            background: #f3f6f4;
            color: #193b37 !important;
            text-decoration: none !important;
        }

        section[data-testid="stSidebar"] .nav-rail a.nav-item.active,
        .mini-nav a.nav-item.active {
            border-color: #a8c9c3;
            border-left: 4px solid #2f6f68;
            background: #dcebe8;
            color: #193b37 !important;
            font-weight: 850;
            text-decoration: none !important;
        }

        section[data-testid="stSidebar"] .nav-item__label,
        .mini-nav .nav-item__label {
            display: block;
            color: inherit !important;
            font-size: 13px;
            font-weight: 820;
            line-height: 1.25;
            text-decoration: none !important;
        }

        section[data-testid="stSidebar"] .nav-item__marker,
        .mini-nav .nav-item__marker {
            display: inline-block;
            margin-top: 3px;
            color: #2f6f68;
            font-size: 10px;
            font-weight: 900;
            line-height: 1;
        }

        .mini-nav {
            display: flex;
            flex-wrap: wrap;
            gap: 6px;
            margin: 6px 0 8px;
        }

        .mini-nav a.nav-item,
        .mini-nav a.nav-item:visited {
            display: inline-flex !important;
            align-items: center;
            justify-content: center;
            min-height: 30px;
            margin: 0;
            padding: 6px 10px;
            border-radius: 999px;
        }

        .mini-nav .nav-item__label {
            font-size: 12px;
            line-height: 1;
        }

        .mini-nav .nav-item__marker {
            display: none;
        }

        @media (max-width: 1180px) {
            .workbench-fact-row,
            .strategy-card-row,
            .strategy-section__cards {
                grid-template-columns: repeat(2, minmax(180px, 1fr));
            }

            .decision-panel,
            .projection-chart-card {
                min-height: auto;
            }
        }

        @media (max-width: 760px) {
            .workbench-fact-row,
            .strategy-card-row,
            .strategy-section__cards,
            .compact-arrival-row,
            .decision-panel__row {
                grid-template-columns: 1fr;
            }

            .page-header-compact {
                align-items: flex-start;
            }
        }
        </style>
        """,
        unsafe_allow_html=True,
    )
    st.markdown(get_pace_check_css(), unsafe_allow_html=True)
    inject_global_styles(st)
    st.markdown(
        """
        <style>
        :root {
            --font-sans: -apple-system, BlinkMacSystemFont, "Segoe UI", "Apple SD Gothic Neo", "Malgun Gothic", "Noto Sans KR", sans-serif;
            --font-app-title: 16px;
            --font-page-title: 20px;
            --font-section-title: 16px;
            --font-card-title: 14px;
            --font-body: 13px;
            --font-body-large: 14px;
            --font-caption: 12px;
            --font-overline: 11px;
            --font-metric-value: 17px;
            --font-nav-title: 13px;
            --font-nav-subtitle: 11px;
            --font-chart-axis: 11px;
            --font-chart-legend: 12px;
            --line-body: 1.62;
            --line-tight: 1.32;
            --line-card: 1.56;
            --app-bg: #f5f7fa;
            --surface: #ffffff;
            --surface-muted: #f7f9fb;
            --surface-soft: #eef6f5;
            --line-soft: #d8e0e7;
            --line-strong: #bcc8d4;
            --text-main: #202833;
            --text-muted: #65717f;
            --teal: #14756f;
            --teal-soft: #e6f3f1;
            --amber: #b7791f;
            --slate: #536170;
            --chart-bg: #f7f9fb;
            --bg: var(--app-bg);
            --ink: var(--text-main);
            --ink-2: var(--text-muted);
            --muted: #7d8793;
            --line: var(--line-soft);
            --surface-2: var(--surface-muted);
            --radius-lg: 8px;
            --radius-md: 8px;
        }

        *,
        *::before,
        *::after {
            box-sizing: border-box;
        }

        html,
        body,
        .stApp,
        .block-container,
        .page-shell,
        .workbench-shell,
        .same-window-top-status,
        .top-status-bar,
        .nav-rail,
        .metric-card,
        .metric-card-compact,
        .decision-panel,
        .report-card,
        .excel-card,
        .projection-chart-card,
        .section-header,
        textarea,
        input,
        button,
        table {
            font-family: var(--font-sans) !important;
        }

        .stApp {
            background: var(--app-bg) !important;
            color: var(--text-main) !important;
            font-size: var(--font-body) !important;
            line-height: var(--line-body) !important;
        }

        .block-container {
            max-width: 1440px;
            padding-top: 2.35rem !important;
            padding-left: clamp(1rem, 2vw, 1.75rem) !important;
            padding-right: clamp(1rem, 2vw, 1.75rem) !important;
            padding-bottom: 3rem !important;
        }

        .page-shell {
            display: block;
            padding-top: 20px;
        }

        [data-testid="stHeader"] {
            background: rgba(245, 247, 250, 0.96) !important;
            border-bottom: 1px solid var(--line-soft) !important;
        }

        h1,
        .page-header h1,
        .pace-hero-title {
            font-family: var(--font-sans) !important;
            font-size: var(--font-page-title) !important;
            font-weight: 700 !important;
            line-height: var(--line-tight) !important;
            letter-spacing: 0 !important;
            margin: 0 0 6px !important;
        }

        h2,
        .section-header__title {
            font-size: var(--font-section-title) !important;
            font-weight: 700 !important;
            line-height: var(--line-tight) !important;
            letter-spacing: 0 !important;
        }

        h3,
        .projection-chart-card__label,
        .decision-panel h2,
        .next-action-panel h3,
        .excel-readonly-panel h3 {
            font-size: var(--font-card-title) !important;
            font-weight: 700 !important;
            line-height: var(--line-tight) !important;
            letter-spacing: 0 !important;
        }

        p,
        li,
        label,
        div[data-testid="stMarkdownContainer"] {
            font-size: var(--font-body) !important;
            font-weight: 400 !important;
            line-height: var(--line-body) !important;
            letter-spacing: 0 !important;
            word-break: keep-all;
            overflow-wrap: anywhere;
        }

        div[data-testid="stCaptionContainer"],
        div[data-testid="stCaptionContainer"] p,
        small,
        .section-header__subtitle,
        .projection-chart-card__copy,
        .page-header__subtitle {
            color: var(--text-muted) !important;
            font-size: var(--font-caption) !important;
            font-weight: 400 !important;
            line-height: 1.55 !important;
        }

        .page-header {
            display: flow-root;
            min-width: 0;
            margin: 0 0 12px !important;
            padding: 14px 16px 15px !important;
            overflow: hidden;
        }

        .page-header-compact {
            min-width: 0;
            margin: 0 !important;
            padding: 0 !important;
            overflow: hidden;
        }

        .workbench-shell.top-status-bar {
            margin: 0 0 16px !important;
        }

        .page-header > *:first-child,
        .page-header-compact > *:first-child,
        .section-header > *:first-child,
        .metric-card-compact > *:first-child,
        .kpi-card > *:first-child,
        .scenario-card > *:first-child,
        .report-card > *:first-child,
        .excel-card > *:first-child,
        .next-action-panel > *:first-child,
        .excel-readonly-panel > *:first-child,
        .report-memo-card > *:first-child,
        .history-purpose-card > *:first-child {
            margin-top: 0 !important;
        }

        .page-header > *:last-child,
        .page-header-compact > *:last-child,
        .section-header > *:last-child,
        .metric-card-compact > *:last-child,
        .kpi-card > *:last-child,
        .scenario-card > *:last-child,
        .report-card > *:last-child,
        .excel-card > *:last-child,
        .next-action-panel > *:last-child,
        .excel-readonly-panel > *:last-child,
        .report-memo-card > *:last-child,
        .history-purpose-card > *:last-child {
            margin-bottom: 0 !important;
        }

        .page-header h1 {
            margin: 4px 0 0 !important;
            max-width: 100%;
            overflow-wrap: anywhere;
        }

        .page-header__subtitle {
            margin-top: 7px !important;
            max-width: 100%;
            overflow-wrap: anywhere;
        }

        .page-header__eyebrow,
        .page-header-compact__eyebrow,
        .projection-chart-card__label + .projection-chart-card__copy,
        .decision-panel__label,
        .next-action-panel__label,
        .excel-readonly-panel__label,
        .nav-rail__title {
            letter-spacing: 0.02em !important;
        }

        .page-header__eyebrow,
        .page-header-compact__eyebrow,
        .decision-panel__label,
        .next-action-panel__label,
        .excel-readonly-panel__label,
        .nav-rail__title {
            display: block;
            max-width: 100%;
            color: var(--teal) !important;
            font-size: var(--font-overline) !important;
            font-weight: 700 !important;
            line-height: 1.2 !important;
            text-transform: none;
            overflow-wrap: anywhere;
        }

        .same-window-top-status {
            display: grid;
            grid-template-columns: minmax(180px, 0.7fr) auto minmax(0, 1.3fr);
            align-items: center;
            gap: 12px;
            min-width: 0;
            margin: 22px 0 18px;
            padding: 13px 16px;
            border: 1px solid var(--line-soft);
            border-radius: 8px;
            background: var(--surface);
            box-shadow: 0 1px 2px rgba(28, 39, 49, 0.06);
        }

        .same-window-top-status__brand {
            display: inline-flex;
            align-items: center;
            gap: 8px;
            min-width: 0;
            color: var(--text-main);
            font-size: var(--font-app-title);
            font-weight: 700;
            line-height: 1.35;
        }

        .same-window-top-status__brand span:last-child {
            min-width: 0;
        }

        .same-window-top-status__brand small {
            display: block;
            margin-top: 1px;
            font-size: var(--font-nav-subtitle) !important;
            font-weight: 500;
        }

        .same-window-top-status__page {
            color: var(--text-main);
            font-size: var(--font-card-title);
            font-weight: 700;
            line-height: 1.35;
            min-width: 0;
            white-space: normal;
            overflow-wrap: anywhere;
        }

        .same-window-top-status__meta {
            display: flex;
            flex-wrap: wrap;
            justify-content: flex-end;
            gap: 6px;
            min-width: 0;
        }

        .pace-brand-mark {
            width: 24px !important;
            height: 24px !important;
            border-radius: 7px !important;
            flex: 0 0 auto;
        }

        .pace-pill {
            min-width: 0;
            min-height: 28px;
            padding: 5px 9px;
            border-radius: 999px;
            font-size: var(--font-nav-subtitle) !important;
            font-weight: 500 !important;
            line-height: 1.32 !important;
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
            max-width: 220px;
        }

        section[data-testid="stSidebar"] {
            background: var(--surface-muted) !important;
        }

        section[data-testid="stSidebar"] .nav-rail {
            max-width: 100%;
            min-width: 0;
            padding: 4px 0 8px;
            border: 0;
            background: transparent;
        }

        section[data-testid="stSidebar"] .nav-rail__title {
            margin: 0 0 8px;
        }

        section[data-testid="stSidebar"] div[data-testid="stButton"] button {
            min-width: 0;
            justify-content: flex-start;
            border-radius: 8px;
            padding: 8px 10px;
            font-size: var(--font-nav-title) !important;
            font-weight: 700 !important;
            line-height: 1.42 !important;
            white-space: normal;
            word-break: keep-all;
            overflow-wrap: anywhere;
            text-align: left;
        }

        div[data-testid="stHorizontalBlock"],
        div[data-testid="stHorizontalBlock"] > div,
        div[data-testid="column"],
        div[data-testid="stVerticalBlock"],
        div[data-testid="stVerticalBlock"] > div,
        div[data-testid="stDataFrame"],
        div[data-testid="stTable"],
        div[data-testid="stMetric"],
        .metric-card,
        .metric-card-compact,
        .kpi-grid,
        .kpi-card,
        .decision-panel,
        .report-card,
        .excel-card,
        .nav-item,
        .projection-chart-card,
        .section-header,
        .strategy-section,
        .strategy-section__head,
        .strategy-section__cards,
        .strategy-card-shell,
        .scenario-card,
        .scenario-card__metrics,
        .report-memo-card,
        .history-purpose-card {
            min-width: 0;
        }

        .text-wrap,
        .page-header,
        .page-header-compact,
        .section-header,
        .metric-card,
        .metric-card-compact,
        .decision-panel,
        .report-card,
        .excel-card,
        .projection-chart-card,
        .kpi-card,
        .scenario-card,
        .strategy-section,
        .compact-arrival-chart,
        .report-memo-card,
        .history-purpose-card {
            word-break: keep-all;
            overflow-wrap: anywhere;
        }

        .page-header,
        .section-header,
        .metric-card-compact,
        .kpi-card,
        .scenario-card,
        .report-card,
        .excel-card,
        .next-action-panel,
        .excel-readonly-panel,
        .report-memo-card,
        .history-purpose-card {
            overflow: hidden;
        }

        .text-truncate {
            overflow: hidden;
            text-overflow: ellipsis;
            white-space: nowrap;
        }

        .line-clamp-2 {
            display: -webkit-box;
            -webkit-line-clamp: 2;
            -webkit-box-orient: vertical;
            overflow: hidden;
        }

        .pace-mode-card {
            display: grid;
            grid-template-columns: minmax(0, 1fr) minmax(260px, 0.86fr);
            align-items: center;
            gap: 5px 14px;
            margin: 6px 0 12px !important;
            padding: 12px 14px !important;
            border: 1px solid #bdd8d4;
            border-left: 4px solid var(--teal);
            border-radius: 8px;
            background: var(--surface);
            box-shadow: 0 1px 2px rgba(28, 39, 49, 0.06);
        }

        .pace-mode-card.status-under-target {
            border-color: #e5c5bd;
            border-left-color: #b76351;
            background: #fff8f6;
        }

        .pace-mode-card.status-on-target {
            border-color: #e7d1a7;
            border-left-color: var(--amber);
            background: #fffaf0;
        }

        .pace-mode-card.status-over-target {
            border-color: #bfdcc9;
            border-left-color: #2d8b67;
            background: #f5fbf7;
        }

        .pace-mode-card__label {
            grid-column: 1;
            color: var(--text-muted);
            font-size: var(--font-overline) !important;
            font-weight: 700 !important;
            line-height: 1.2 !important;
        }

        .pace-mode-card__mode {
            grid-column: 1;
            margin-top: 1px;
            color: var(--text-main);
            font-size: 24px !important;
            font-weight: 800 !important;
            line-height: 1.15 !important;
            overflow-wrap: anywhere;
        }

        .pace-mode-card.status-under-target .pace-mode-card__mode {
            color: #854234;
        }

        .pace-mode-card.status-on-target .pace-mode-card__mode {
            color: #805113;
        }

        .pace-mode-card.status-over-target .pace-mode-card__mode {
            color: #1d6446;
        }

        .pace-mode-card__description {
            grid-column: 1;
            margin-top: 2px;
            color: var(--text-muted);
            font-size: var(--font-body) !important;
            line-height: 1.45 !important;
            overflow-wrap: anywhere;
        }

        .pace-mode-card__facts {
            grid-column: 2;
            grid-row: 1 / span 3;
            display: grid;
            grid-template-columns: repeat(2, minmax(0, 1fr));
            gap: 8px;
            margin-top: 0;
            min-width: 0;
        }

        .pace-mode-card__fact {
            min-width: 0;
            padding: 9px 10px;
            border: 1px solid var(--line-soft);
            border-radius: 8px;
            background: rgba(255, 255, 255, 0.72);
        }

        .pace-mode-card__fact small {
            display: block;
            color: var(--text-muted) !important;
            font-size: var(--font-overline) !important;
            font-weight: 500 !important;
            line-height: 1.25 !important;
        }

        .pace-mode-card__fact strong {
            display: block;
            margin-top: 4px;
            color: var(--text-main);
            font-size: var(--font-metric-value) !important;
            font-weight: 700 !important;
            line-height: 1.22 !important;
            overflow-wrap: anywhere;
        }

        [data-testid="stMetric"] {
            height: 100px !important;
            min-height: 100px !important;
            display: flex !important;
            flex-direction: column !important;
            justify-content: flex-start !important;
            border-radius: 8px !important;
            background: var(--surface) !important;
        }

        [data-testid="stMetric"] > div {
            min-height: 0 !important;
        }

        [data-testid="stMetricLabel"],
        [data-testid="stMetricLabel"] p {
            font-size: var(--font-overline) !important;
            line-height: 1.3 !important;
        }

        [data-testid="stMetricValue"] {
            font-size: var(--font-metric-value) !important;
            font-weight: 700 !important;
            line-height: 1.25 !important;
            margin-top: 8px !important;
            overflow-wrap: anywhere;
        }

        [data-testid="stMetricDelta"] {
            font-size: var(--font-overline) !important;
            line-height: 1.25 !important;
            margin-top: 6px !important;
        }

        .kpi-grid {
            grid-template-columns: repeat(7, minmax(0, 1fr)) !important;
            gap: 8px !important;
        }

        .kpi-card {
            min-height: 0 !important;
            padding: 10px 11px !important;
            border-radius: 8px !important;
            line-height: 1.5 !important;
        }

        .kpi-card__label {
            font-size: 12px !important;
            font-weight: 700 !important;
            line-height: 1.3 !important;
        }

        .kpi-card__value {
            font-size: 17px !important;
            font-weight: 700 !important;
            line-height: 1.25 !important;
            margin-top: 9px !important;
            overflow-wrap: anywhere;
        }

        .kpi-card__sub {
            font-size: 12px !important;
            font-weight: 400 !important;
            line-height: 1.38 !important;
            margin-top: 8px !important;
            overflow-wrap: anywhere;
        }

        .metric-card-compact {
            min-height: 0;
            height: auto;
            padding: 9px 10px;
            border: 1px solid var(--line-soft);
            border-radius: 8px;
            background: var(--surface);
        }

        .metric-card-compact span {
            color: var(--text-muted);
            font-size: var(--font-overline);
            font-weight: 700;
            line-height: 1.32;
        }

        .metric-card-compact strong {
            font-size: var(--font-metric-value);
            font-weight: 700;
            line-height: 1.3;
            overflow-wrap: anywhere;
        }

        .report-card__rail {
            margin-bottom: 16px !important;
        }

        .report-card__body {
            display: grid;
            gap: 0;
            margin: 0;
            color: var(--text-main);
            font-family: var(--font-sans) !important;
            white-space: normal;
        }

        .report-card__section {
            padding: 0 0 15px;
            margin: 0 0 15px;
            border-bottom: 1px solid var(--line-soft);
        }

        .report-card__section:last-child {
            padding-bottom: 0;
            margin-bottom: 0;
            border-bottom: 0;
        }

        .report-card__section-title {
            margin: 0 0 9px !important;
            color: var(--text-main);
            font-size: 13px !important;
            font-weight: 800 !important;
            line-height: 1.35 !important;
            letter-spacing: 0 !important;
        }

        .report-card__list {
            display: grid;
            gap: 8px;
            margin: 0;
            padding-left: 18px;
        }

        .report-card__list li,
        .report-card__paragraph {
            color: var(--text-main);
            font-size: 13px !important;
            line-height: 1.72 !important;
            word-break: keep-all;
            overflow-wrap: anywhere;
        }

        .report-card__paragraph {
            margin: 0;
        }

        .report-card__placeholder {
            color: var(--text-muted);
        }

        .chart-first-grid {
            display: grid;
            grid-template-columns: minmax(0, 1fr) minmax(300px, 340px);
            gap: 14px;
            align-items: stretch;
        }

        div[data-testid="stHorizontalBlock"]:has(.projection-chart-card__head) {
            display: grid !important;
            grid-template-columns: minmax(0, 1fr) minmax(300px, 340px) !important;
            gap: 16px !important;
            align-items: stretch !important;
            margin-top: 14px !important;
        }

        div[data-testid="stHorizontalBlock"]:has(.projection-chart-card__head) > div {
            display: flex !important;
            width: 100% !important;
            min-width: 0 !important;
        }

        div[data-testid="stHorizontalBlock"]:has(.projection-chart-card__head) > div > div,
        div[data-testid="stHorizontalBlock"]:has(.projection-chart-card__head) div[data-testid="stVerticalBlock"],
        div[data-testid="stHorizontalBlock"]:has(.projection-chart-card__head) div[data-testid="stMarkdownContainer"]:has(.decision-panel) {
            width: 100% !important;
            height: 100% !important;
        }

        div[data-testid="stVerticalBlockBorderWrapper"]:has(.projection-chart-card__head) {
            display: flex !important;
            flex-direction: column !important;
            height: 100% !important;
            border: 1px solid var(--line-soft) !important;
            border-radius: 8px !important;
            background: var(--chart-bg) !important;
            padding: 14px 16px 12px !important;
            box-shadow: 0 1px 2px rgba(28, 39, 49, 0.06) !important;
        }

        div[data-testid="stVerticalBlockBorderWrapper"]:has(.projection-chart-card__head) .projection-chart-card {
            margin: 0 !important;
            padding: 0 !important;
            border: 0 !important;
            border-radius: 0 !important;
            background: var(--chart-bg) !important;
            box-shadow: none !important;
        }

        div[data-testid="stVerticalBlockBorderWrapper"]:has(.projection-chart-card__head) .projection-chart-card__head {
            background: var(--chart-bg) !important;
        }

        div[data-testid="stVerticalBlockBorderWrapper"]:has(.projection-chart-card__head) > div,
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.projection-chart-card__head) div,
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.projection-chart-card__head) details,
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.projection-chart-card__head) [data-testid="stElementContainer"],
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.projection-chart-card__head) div[data-testid="stMarkdownContainer"],
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.projection-chart-card__head) div[data-testid="stVegaLiteChart"],
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.projection-chart-card__head) div[data-testid="stVegaLiteChart"] > div,
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.projection-chart-card__head) canvas,
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.projection-chart-card__head) svg {
            background: var(--chart-bg) !important;
        }

        .projection-chart-card,
        .decision-panel,
        .report-card,
        .excel-card,
        .empty-state,
        .strategy-section,
        .next-action-panel,
        .excel-readonly-panel,
        .compact-arrival-chart,
        .report-memo-card,
        .history-purpose-card {
            border: 1px solid var(--line-soft);
            border-radius: 8px;
            background: var(--surface);
            box-shadow: 0 1px 2px rgba(28, 39, 49, 0.06);
        }

        .projection-chart-card {
            padding: 0;
            min-height: 0;
            border: 0;
            box-shadow: none;
        }

        .projection-chart-card__head {
            gap: 12px;
            margin-bottom: 6px;
        }

        .projection-chart-card__copy {
            max-width: 760px;
            margin-top: 3px;
            line-height: 1.55 !important;
        }

        .chart-legend-row {
            gap: 8px 14px;
            padding: 9px 0 0;
            margin-bottom: 10px;
            color: var(--text-muted);
            font-size: var(--font-chart-legend) !important;
            font-weight: 500;
            line-height: 1.42;
        }

        .chart-legend-row span {
            min-width: 0;
            white-space: normal;
        }

        .chart-legend-row i {
            width: 18px;
            border-top-width: 3px;
        }

        .chart-legend-row .legend-target {
            border-color: #8a94a1;
            border-top-style: dashed;
        }

        .chart-legend-row .legend-actual {
            border-color: var(--teal);
        }

        .chart-legend-row .legend-projection {
            border-color: var(--amber);
            border-top-style: dashed;
        }

        .chart-legend-row .legend-band {
            background: rgba(183, 121, 31, 0.16);
        }

        .chart-legend-row .legend-close {
            border-color: var(--slate);
            border-top-style: dashed;
        }

        .chart-legend-row .legend-current {
            background: var(--text-main);
        }

        .projection-chart-caption {
            clear: both;
            display: block;
            margin: 8px 0 0;
            padding-top: 8px;
            border-top: 1px solid var(--line-soft);
            color: var(--text-muted);
            font-size: var(--font-caption) !important;
            line-height: 1.6 !important;
            word-break: keep-all;
            overflow-wrap: anywhere;
        }

        div[data-testid="stMarkdownContainer"]:has(.scenario-inline-chart-title) {
            display: block !important;
            margin: 10px 0 12px !important;
            position: relative;
            z-index: 1;
        }

        .scenario-inline-chart-title {
            display: grid !important;
            grid-template-columns: auto minmax(0, 1fr);
            align-items: end;
            gap: 4px 10px;
            width: 100%;
            min-height: 28px;
            margin: 0 !important;
            padding: 0 0 4px;
            color: var(--text-main);
            line-height: 1.35 !important;
            overflow: visible;
        }

        .scenario-inline-chart-title strong,
        .scenario-inline-chart-title span {
            display: block;
            min-width: 0;
            line-height: 1.35 !important;
            overflow-wrap: anywhere;
        }

        .scenario-inline-chart-title strong {
            font-size: var(--font-body) !important;
            font-weight: 800 !important;
        }

        .scenario-inline-chart-title span {
            color: var(--text-muted);
            font-size: var(--font-overline) !important;
            font-weight: 500 !important;
        }

        div[data-testid="stAlert"] {
            clear: both;
            position: relative;
            z-index: 0;
        }

        .decision-panel {
            display: flex;
            flex-direction: column;
            width: 100%;
            height: 100%;
            min-height: 100%;
            padding: 14px;
        }

        .decision-panel__row {
            grid-template-columns: 96px minmax(0, 1fr);
            align-items: center;
            gap: 12px;
            width: 100%;
            min-height: 34px;
            padding: 7px 0;
        }

        .decision-panel__row span {
            min-width: 0;
            font-size: var(--font-caption) !important;
            font-weight: 500 !important;
            line-height: 1.38 !important;
        }

        .decision-panel__row strong {
            font-size: var(--font-body) !important;
            font-weight: 600 !important;
            line-height: 1.38 !important;
            max-width: 100%;
            overflow-wrap: anywhere;
        }

        .decision-panel__row:last-child {
            align-items: start;
            padding-bottom: 0;
        }

        .workbench-fact-row {
            grid-template-columns: repeat(5, minmax(0, 1fr));
            gap: 10px;
        }

        .strategy-section .scenario-card__description {
            max-height: none;
        }

        .strategy-section {
            padding: 12px;
            margin: 10px 0 14px;
        }

        .strategy-section.is-active-management {
            border-top: 3px solid rgba(20, 117, 111, 0.55);
            background: #fbfefe;
        }

        .strategy-section__head {
            gap: 14px;
            margin-bottom: 10px;
        }

        .strategy-section__head p {
            line-height: 1.55 !important;
        }

        .strategy-section__head span {
            display: inline-flex;
            align-items: center;
            min-width: 0;
            border: 0;
            background: transparent;
            color: #0f665e;
            padding: 0;
            font-size: var(--font-overline) !important;
            font-weight: 700 !important;
            line-height: 1.3 !important;
            word-break: keep-all;
            overflow-wrap: anywhere;
        }

        .strategy-card-shell__badge {
            display: inline-flex;
            align-items: center;
            min-width: 0;
            border: 1px solid var(--line-soft);
            border-radius: 999px;
            background: var(--surface-muted);
            color: var(--text-muted);
            padding: 3px 7px;
            font-size: var(--font-overline) !important;
            font-weight: 700 !important;
            line-height: 1.3 !important;
            word-break: keep-all;
            overflow-wrap: anywhere;
        }

        .strategy-card-shell__head {
            display: flex;
            align-items: flex-start;
            justify-content: space-between;
            gap: 8px;
            min-width: 0;
            margin-bottom: 5px;
        }

        .strategy-card-shell__badge.is-reference {
            border-color: var(--line-soft);
            background: var(--surface-muted);
            color: var(--text-muted);
        }

        .strategy-card-shell__badge.is-recommended-badge {
            border-color: #bdd8d4;
            background: #e9f5f2;
            color: #0f665e;
        }

        .strategy-card-active {
            border-left: 0;
            border-radius: 0;
            padding-left: 0;
        }

        .strategy-card-active .scenario-card {
            border-color: var(--line-soft) !important;
            box-shadow: none !important;
        }

        .strategy-card-active.is-recommended .scenario-card {
            border-left: 1px solid rgba(20, 117, 111, 0.2) !important;
            background: #f7fcfa;
            box-shadow: inset 0 0 0 1px rgba(20, 117, 111, 0.16) !important;
        }

        .strategy-card-shell__code {
            line-height: 1.35 !important;
            overflow-wrap: anywhere;
        }

        .strategy-section .scenario-card {
            min-height: 0;
        }

        .strategy-section .scenario-card.is-emphasis {
            box-shadow: none !important;
        }

        .strategy-card-active.is-recommended .scenario-card.is-emphasis {
            box-shadow: inset 0 0 0 1px rgba(20, 117, 111, 0.16) !important;
        }

        .strategy-section .status-badge {
            border-color: transparent;
            background: transparent;
            color: #0f665e;
            padding: 0;
            gap: 5px;
            font-size: 11px !important;
            box-shadow: none;
        }

        .strategy-section .status-badge::before {
            width: 6px;
            height: 6px;
            opacity: 0.72;
        }

        .strategy-section .scenario-card__name {
            font-size: 14px !important;
            line-height: 1.35 !important;
        }

        .strategy-section .scenario-card__description {
            font-size: 12px !important;
            line-height: 1.5 !important;
        }

        .strategy-section .scenario-card__metric-label {
            font-size: 11px !important;
            line-height: 1.32 !important;
        }

        .strategy-section .scenario-card__metric-value {
            font-size: 12px !important;
            line-height: 1.35 !important;
        }

        .report-ia-note,
        .history-purpose-card {
            margin: 8px 0 12px;
            padding: 12px 14px;
            color: var(--text-main);
            font-size: var(--font-body) !important;
            line-height: var(--line-card) !important;
            word-break: keep-all;
            overflow-wrap: anywhere;
        }

        .report-meta-row,
        .history-question-grid,
        .history-next-actions {
            display: grid;
            grid-template-columns: repeat(3, minmax(0, 1fr));
            gap: 10px;
            margin: 8px 0 12px;
        }

        .report-meta-row span,
        .history-question-grid span,
        .history-next-actions span {
            display: block;
            min-width: 0;
            padding: 9px 10px;
            border: 1px solid var(--line-soft);
            border-radius: 8px;
            background: var(--surface-muted);
            color: var(--text-main);
            font-size: var(--font-caption) !important;
            line-height: 1.55 !important;
            word-break: keep-all;
            overflow-wrap: anywhere;
        }

        .report-memo-grid {
            display: grid;
            grid-template-columns: repeat(2, minmax(0, 1fr));
            gap: 10px;
            margin: 8px 0 12px;
        }

        .report-memo-card {
            padding: 12px;
            min-height: 0;
        }

        .report-memo-card strong {
            display: block;
            margin-bottom: 5px;
            color: var(--text-main);
            font-size: var(--font-card-title);
            font-weight: 700;
            line-height: var(--line-tight);
        }

        .report-memo-card span {
            display: block;
            color: var(--text-muted);
            font-size: var(--font-body);
            line-height: var(--line-card);
            word-break: keep-all;
            overflow-wrap: anywhere;
        }

        textarea[aria-label="복사용 보고문"],
        textarea[aria-label="보고 메모"] {
            min-height: 320px;
            font-family: var(--font-sans) !important;
            font-size: var(--font-body-large) !important;
            line-height: 1.6 !important;
            word-break: keep-all;
            overflow-wrap: anywhere;
        }

        .history-purpose-card h3 {
            margin: 0 0 6px !important;
            padding: 0 !important;
            border: 0 !important;
        }

        div[data-testid="stDataFrame"],
        div[data-testid="stTable"] {
            min-width: 0;
            overflow: auto;
        }

        div[data-testid="stDataFrame"] *,
        div[data-testid="stTable"] *,
        [data-testid="stDataEditor"] * {
            font-family: var(--font-sans) !important;
            font-size: var(--font-caption) !important;
            line-height: 1.5 !important;
        }

        @media (max-width: 1180px) {
            .chart-first-grid,
            div[data-testid="stHorizontalBlock"]:has(.projection-chart-card__head) {
                grid-template-columns: 1fr !important;
            }

            .decision-panel {
                width: 100%;
            }

            .workbench-fact-row,
            .report-meta-row,
            .history-question-grid,
            .history-next-actions {
                grid-template-columns: repeat(2, minmax(0, 1fr));
            }
        }

        @media (max-width: 900px) {
            section[data-testid="stSidebar"] {
                width: 100% !important;
            }

            .block-container {
                padding-top: 2rem !important;
            }

            .same-window-top-status {
                grid-template-columns: 1fr;
                align-items: start;
                margin-top: 18px;
            }

            .same-window-top-status__meta {
                justify-content: flex-start;
            }
        }

        @media (max-width: 760px) {
            .pace-mode-card {
                grid-template-columns: 1fr;
                padding: 12px 14px !important;
            }

            .pace-mode-card__mode {
                font-size: 22px !important;
            }

            .pace-mode-card__facts {
                grid-column: 1;
                grid-row: auto;
                grid-template-columns: 1fr;
                margin-top: 6px;
            }

            .workbench-fact-row,
            .report-meta-row,
            .report-memo-grid,
            .history-question-grid,
            .history-next-actions {
                grid-template-columns: 1fr;
            }

            .scenario-inline-chart-title {
                grid-template-columns: 1fr;
                align-items: start;
            }

            .pace-pill {
                max-width: 100%;
            }
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def calculate_validated_results(
    df: pd.DataFrame,
    as_of_date: object,
    metric: str,
    config: dict[str, Any],
) -> dict[str, Any]:
    """Validate input and run calculations only when no errors exist."""
    validation_result = validate_input(df, as_of_date, metric, config)
    if validation_result["errors"]:
        return {"validation": validation_result}

    return {
        "validation": validation_result,
        "scenario_df": run_scenario_grid(df, as_of_date, metric, config),
        "next_close_result": calculate_next_close_required(
            df,
            as_of_date,
            metric,
            config,
        ),
        "close_cycle_df": build_close_cycle_summary(df, metric, as_of_date),
    }


def default_as_of_date(
    df: pd.DataFrame,
    metric: str,
    today: object | None = None,
) -> pd.Timestamp:
    """Return the previous input business date before today."""
    dates = pd.to_datetime(df["date"], errors="raise").dt.normalize()
    current_date = _current_app_date(today)
    previous_mask = dates.dt.date < current_date

    previous_business_dates = dates.loc[previous_mask]
    if not previous_business_dates.empty:
        return previous_business_dates.iloc[-1]
    return dates.iloc[0]


def _current_app_date(today: object | None = None) -> object:
    if today is not None:
        return pd.Timestamp(today).date()
    return pd.Timestamp.now(tz=APP_TIMEZONE).date()


def normalize_direct_input_edits(df: pd.DataFrame) -> pd.DataFrame:
    """Return editor data with target and actual columns converted for calculation."""
    normalized = df.copy()
    for column in DIRECT_EDITABLE_COLUMNS:
        if column not in normalized.columns:
            continue
        values = normalized[column].replace(r"^\s*$", pd.NA, regex=True)
        normalized[column] = pd.to_numeric(values, errors="coerce").astype(float)
    return normalized


def load_saved_actuals(path: str | Path = SAVED_ACTUALS_PATH) -> pd.DataFrame:
    """Load locally saved cumulative actual values without creating or rewriting files."""
    saved_path = Path(path)
    if not saved_path.exists():
        return pd.DataFrame(columns=SAVED_ACTUAL_COLUMNS)
    return normalize_saved_actuals_schema(pd.read_csv(saved_path, encoding="utf-8-sig"))


def normalize_saved_actuals_schema(saved_actuals: pd.DataFrame) -> pd.DataFrame:
    """Normalize saved actuals in memory only."""
    return _normalize_saved_actuals(saved_actuals)


def save_saved_actuals(
    saved_actuals: pd.DataFrame,
    path: str | Path = SAVED_ACTUALS_PATH,
) -> Path:
    """Persist normalized saved actuals after an explicit user action."""
    saved_path = Path(path)
    saved_path.parent.mkdir(parents=True, exist_ok=True)
    normalize_saved_actuals_schema(saved_actuals).to_csv(
        saved_path,
        index=False,
        encoding="utf-8-sig",
    )
    return saved_path


def save_actual_values(
    df: pd.DataFrame,
    path: str | Path = SAVED_ACTUALS_PATH,
) -> Path:
    """Persist cumulative actual values for future app defaults."""
    saved_path = Path(path)
    saved_path.parent.mkdir(parents=True, exist_ok=True)
    saved_actuals = _build_saved_actuals(df)
    if saved_path.exists():
        saved_actuals = _merge_saved_actuals(
            load_saved_actuals(saved_path),
            saved_actuals,
        )
    return save_saved_actuals(saved_actuals, saved_path)


def save_current_input_defaults(
    df: pd.DataFrame,
    path: str | Path = SAVED_ACTUALS_PATH,
) -> dict[str, Any]:
    """Persist edited current input values as restart defaults."""
    normalized = normalize_direct_input_edits(df)
    saved_actuals_path = save_actual_values(normalized, path)
    operator_result = save_operator_sample("current_input", normalized)
    return {
        "ok": bool(operator_result.get("ok")),
        "df": normalized,
        "saved_actuals_path": saved_actuals_path,
        "operator_result": operator_result,
    }


def apply_saved_actuals(
    df: pd.DataFrame,
    saved_actuals: pd.DataFrame,
) -> pd.DataFrame:
    """Apply saved actual values to matching date and business-day rows."""
    saved = normalize_saved_actuals_schema(saved_actuals)
    if saved.empty:
        return df.copy()

    result = df.copy()
    result["_date_key"] = pd.to_datetime(result["date"], errors="coerce").dt.normalize()
    result["_business_day_key"] = pd.to_numeric(
        result["business_day_no"],
        errors="coerce",
    ).astype("Int64")

    saved = saved.rename(
        columns={
            "date": "_date_key",
            "business_day_no": "_business_day_key",
            **{column: f"{column}_saved" for column in ACTUAL_CUM_COLUMNS},
        }
    )
    saved["_saved_actual_row"] = True
    merged = result.merge(
        saved,
        on=["_date_key", "_business_day_key"],
        how="left",
        sort=False,
    )
    matched = merged["_saved_actual_row"].fillna(False).astype(bool)
    for column in ACTUAL_CUM_COLUMNS:
        saved_column = f"{column}_saved"
        if column in result.columns and saved_column in merged.columns:
            has_saved_value = matched & merged[saved_column].notna()
            if not has_saved_value.any():
                continue
            saved_values = pd.to_numeric(
                merged.loc[has_saved_value, saved_column],
                errors="coerce",
            )
            result.loc[has_saved_value, column] = saved_values.to_numpy(dtype="float64")

    return result.drop(columns=["_date_key", "_business_day_key"])


def clear_saved_actuals(path: str | Path = SAVED_ACTUALS_PATH) -> None:
    """Delete locally saved actual defaults when the user asks to reset them."""
    saved_path = Path(path)
    if saved_path.exists():
        saved_path.unlink()


def _build_saved_actuals(df: pd.DataFrame) -> pd.DataFrame:
    result = pd.DataFrame(columns=SAVED_ACTUAL_COLUMNS)
    if "date" not in df.columns or "business_day_no" not in df.columns:
        return result

    result["date"] = pd.to_datetime(df["date"], errors="coerce").dt.strftime("%Y-%m-%d")
    result["business_day_no"] = pd.to_numeric(
        df["business_day_no"],
        errors="coerce",
    ).astype("Int64")
    for column in ACTUAL_CUM_COLUMNS:
        if column not in df.columns:
            result[column] = pd.NA
            continue
        values = df[column].replace(r"^\s*$", pd.NA, regex=True)
        result[column] = pd.to_numeric(values, errors="coerce")

    valid_rows = result["date"].notna() & result["business_day_no"].notna()
    return _normalize_saved_actuals(result.loc[valid_rows, list(SAVED_ACTUAL_COLUMNS)])


def _merge_saved_actuals(
    previous_actuals: pd.DataFrame,
    latest_actuals: pd.DataFrame,
) -> pd.DataFrame:
    previous = _normalize_saved_actuals(previous_actuals)
    latest = _normalize_saved_actuals(latest_actuals)
    if previous.empty:
        return latest
    if latest.empty:
        return previous

    key_columns = ["date", "business_day_no"]
    merged = previous.merge(
        latest,
        on=key_columns,
        how="outer",
        suffixes=("_previous", "_latest"),
        sort=False,
    )
    result = merged.loc[:, key_columns].copy()
    for column in ACTUAL_CUM_COLUMNS:
        latest_column = f"{column}_latest"
        previous_column = f"{column}_previous"
        result[column] = merged[latest_column].combine_first(merged[previous_column])

    return _normalize_saved_actuals(result)


def _normalize_saved_actuals(saved_actuals: pd.DataFrame) -> pd.DataFrame:
    result = saved_actuals.copy()
    for column in SAVED_ACTUAL_COLUMNS:
        if column not in result.columns:
            result[column] = pd.NA

    result = result.loc[:, list(SAVED_ACTUAL_COLUMNS)]
    result["date"] = pd.to_datetime(result["date"], errors="coerce").dt.normalize()
    result["business_day_no"] = pd.to_numeric(
        result["business_day_no"],
        errors="coerce",
    ).astype("Int64")
    for column in ACTUAL_CUM_COLUMNS:
        values = result[column].replace(r"^\s*$", pd.NA, regex=True)
        result[column] = pd.to_numeric(values, errors="coerce")

    result = result.dropna(subset=["date", "business_day_no"])
    if result.empty:
        return result.reset_index(drop=True)

    result = (
        result.groupby(["date", "business_day_no"], as_index=False, sort=False)
        .agg({column: _last_non_empty_value for column in ACTUAL_CUM_COLUMNS})
        .loc[:, list(SAVED_ACTUAL_COLUMNS)]
    )
    has_actual_value = result.loc[:, list(ACTUAL_CUM_COLUMNS)].notna().any(axis=1)
    return result.loc[has_actual_value].reset_index(drop=True)


def _last_non_empty_value(values: pd.Series) -> object:
    filled_values = values.dropna()
    if filled_values.empty:
        return pd.NA
    return filled_values.iloc[-1]


def build_runtime_config(
    base_config: dict[str, Any],
    close_day_cap_rate: float,
    non_close_day_cap_rate: float,
) -> dict[str, Any]:
    """Return a calculation config without mutating the loaded YAML config."""
    runtime_config = dict(base_config)
    runtime_config["close_day_cap_rate"] = float(close_day_cap_rate)
    runtime_config["non_close_day_cap_rate"] = float(non_close_day_cap_rate)
    return runtime_config


def run_selected_scenario_detail(
    df: pd.DataFrame,
    as_of_date: object,
    metric: str,
    scenario_id: str,
    config: dict[str, Any],
) -> tuple[dict[str, object], dict[str, object]]:
    """Run the selected scenario pair and return detailed forecast and strategy data."""
    forecast_key, strategy_key = scenario_id.split("_", maxsplit=1)
    forecast_model = FORECAST_MODEL_OPTIONS[forecast_key]
    forecast_result = run_forecast_model(
        df,
        metric,
        as_of_date,
        forecast_model,
        config,
    )
    if strategy_key in PROVISION_STRATEGY_OPTIONS:
        strategy_result = run_provision_model(
            df,
            forecast_result,
            as_of_date,
            metric,
            PROVISION_STRATEGY_OPTIONS[strategy_key],
            config,
        )
    elif strategy_key in OVERACHIEVEMENT_STRATEGY_OPTIONS:
        strategy_result = run_overachievement_strategy(
            forecast_result,
            OVERACHIEVEMENT_STRATEGY_OPTIONS[strategy_key],
            config,
        )
    elif strategy_key in NEUTRAL_STRATEGY_OPTIONS:
        strategy_result = run_neutral_strategy(
            forecast_result,
            NEUTRAL_STRATEGY_OPTIONS[strategy_key],
        )
    else:
        raise ValueError(f"지원하지 않는 시나리오 전략입니다: {strategy_key}")
    return forecast_result, strategy_result


def build_scenario_matrix(scenario_df: pd.DataFrame) -> pd.DataFrame:
    """Return a 3x3 display matrix for forecast/strategy scenarios."""
    matrix = pd.DataFrame(
        index=["F1", "F2", "F3"],
        columns=_ordered_strategy_keys(scenario_df),
    )
    for _, row in scenario_df.iterrows():
        scenario_id = str(row.get("scenario_id", ""))
        if "_" not in scenario_id:
            continue
        forecast_key, strategy_key = scenario_id.split("_", maxsplit=1)
        matrix.loc[forecast_key, strategy_key] = (
            f"{format_amount(row.get('forecast_after_provision'))} / "
            f"{_localize_display_value(row.get('risk_level', ''))} / "
            f"{_scenario_matrix_mode_label(row, strategy_key)}"
        )
    return matrix


def _scenario_matrix_mode_label(row: pd.Series, strategy_key: str) -> object:
    strategy_type = str(row.get("strategy_type", ""))
    if strategy_type == OVERACHIEVEMENT or strategy_key in {"O1", "O2", "O3"}:
        strategy_value = row.get("overachievement_strategy")
        return OVERACHIEVEMENT_MATRIX_LABELS.get(
            str(strategy_value),
            OVERACHIEVEMENT_MATRIX_LABELS.get(
                strategy_key,
                _operation_mode_label(row.get("target_status")),
            ),
        )
    return _localize_display_value(row.get("status", ""))


def build_scenario_value_matrix(
    scenario_df: pd.DataFrame,
    value_column: str = "forecast_after_provision",
) -> pd.DataFrame:
    """Return a numeric 3x3 scenario matrix for quick visual comparison."""
    matrix = pd.DataFrame(
        index=["F1", "F2", "F3"],
        columns=_ordered_strategy_keys(scenario_df),
        dtype="float64",
    )
    if value_column not in scenario_df.columns:
        return matrix

    for _, row in scenario_df.iterrows():
        scenario_id = str(row.get("scenario_id", ""))
        if "_" not in scenario_id:
            continue
        forecast_key, strategy_key = scenario_id.split("_", maxsplit=1)
        if forecast_key in matrix.index and strategy_key in matrix.columns:
            matrix.loc[forecast_key, strategy_key] = _as_float(row.get(value_column))
    return matrix


def build_historical_forecast_comparison(
    scenario_df: pd.DataFrame,
    historical_context: Mapping[str, object],
    selected_scenario_id: str = "",
) -> pd.DataFrame:
    """Return current forecasts and historical-performance forecasts in one comparison table."""
    scenarios = _as_dataframe(scenario_df)
    benchmark = dict(historical_context.get("benchmark") or {})
    if scenarios.empty or not benchmark:
        return pd.DataFrame(
            columns=[
                "comparison_group",
                "basis",
                "forecast_amount",
                "monthly_target",
                "forecast_rate",
                "diff_vs_target",
                "diff_vs_historical_median",
            ]
        )

    monthly_target = _as_float(benchmark.get("current_monthly_target"))
    if not math.isfinite(monthly_target) and "monthly_target" in scenarios.columns:
        target_values = pd.to_numeric(scenarios["monthly_target"], errors="coerce").dropna()
        if not target_values.empty:
            monthly_target = _as_float(target_values.iloc[0])
    historical_median = _as_float(benchmark.get("historical_forecast_median"))

    rows: list[dict[str, object]] = []

    def append_row(group: str, basis: str, amount: object) -> None:
        value = _as_float(amount)
        if not math.isfinite(value):
            return
        rows.append(
            {
                "comparison_group": group,
                "basis": basis,
                "forecast_amount": value,
                "monthly_target": monthly_target,
                "forecast_rate": safe_divide(value, monthly_target),
                "diff_vs_target": (
                    value - monthly_target if math.isfinite(monthly_target) else float("nan")
                ),
                "diff_vs_historical_median": (
                    value - historical_median
                    if math.isfinite(historical_median)
                    else float("nan")
                ),
            }
        )

    if selected_scenario_id and "scenario_id" in scenarios.columns:
        selected_row = _selected_scenario_row(scenarios, selected_scenario_id)
        append_row(
            "현재 예측",
            f"선택 시나리오 {selected_scenario_id}",
            selected_row.get("forecast_after_provision", selected_row.get("forecast_amount")),
        )

    forecast_summary = _forecast_summary(scenarios)
    for forecast_key in ("F1", "F2", "F3"):
        definition = FORECAST_MODEL_DEFINITIONS.get(forecast_key, {})
        append_row(
            "F모델 기본 예측",
            f"{forecast_key} {definition.get('name', forecast_key)}",
            forecast_summary.get(forecast_key),
        )

    for key, label in (
        ("historical_forecast_lower", "과거 하위 25%"),
        ("historical_forecast_median", "과거 중앙값"),
        ("historical_forecast_upper", "과거 상위 25%"),
    ):
        append_row("과거 실적 기반", label, benchmark.get(key))

    return pd.DataFrame(rows)


def build_historical_forecast_decision_summary(comparison_df: pd.DataFrame) -> dict[str, object]:
    """Condense historical comparison rows into one report-ready decision summary."""
    source = _as_dataframe(comparison_df)
    if source.empty:
        return {"has_data": False}

    def first_amount(mask: pd.Series, column: str = "forecast_amount") -> float:
        rows = source.loc[mask]
        if rows.empty or column not in rows.columns:
            return float("nan")
        return _as_float(rows.iloc[0].get(column))

    if "basis" not in source.columns:
        return {"has_data": False}

    selected_amount = first_amount(source["basis"].astype(str).str.startswith("선택 시나리오"))
    historical_lower = first_amount(source["basis"] == "과거 하위 25%")
    historical_median = first_amount(source["basis"] == "과거 중앙값")
    historical_upper = first_amount(source["basis"] == "과거 상위 25%")
    monthly_target_values = (
        pd.to_numeric(source["monthly_target"], errors="coerce").dropna()
        if "monthly_target" in source.columns
        else pd.Series(dtype="float64")
    )
    monthly_target = (
        _as_float(monthly_target_values.iloc[0])
        if not monthly_target_values.empty
        else float("nan")
    )
    reference_amount = selected_amount if math.isfinite(selected_amount) else historical_median
    target_delta = (
        reference_amount - monthly_target
        if math.isfinite(reference_amount) and math.isfinite(monthly_target)
        else float("nan")
    )
    history_delta = (
        selected_amount - historical_median
        if math.isfinite(selected_amount) and math.isfinite(historical_median)
        else float("nan")
    )
    threshold = 0.5
    if math.isfinite(monthly_target):
        threshold = max(threshold, abs(monthly_target) * 0.005)

    if math.isfinite(selected_amount) and math.isfinite(historical_median):
        report_low = min(selected_amount, historical_median)
        report_high = max(selected_amount, historical_median)
        report_basis = "선택 예측과 과거 중앙값 사이"
    elif math.isfinite(historical_lower) and math.isfinite(historical_upper):
        report_low = historical_lower
        report_high = historical_upper
        report_basis = "과거 하위 25%~상위 25% 범위"
    elif math.isfinite(reference_amount):
        report_low = reference_amount
        report_high = reference_amount
        report_basis = "사용 가능한 단일 예측값"
    else:
        return {"has_data": False}

    report_range = (
        format_amount(report_low)
        if math.isclose(report_low, report_high, rel_tol=1e-9, abs_tol=1e-9)
        else f"{format_amount(report_low)} ~ {format_amount(report_high)}"
    )

    if math.isfinite(history_delta) and history_delta > threshold:
        forecast_position = (
            f"선택 예측이 과거 중앙값보다 {format_amount(history_delta)} 높아 공격적인 전망입니다."
        )
    elif math.isfinite(history_delta) and history_delta < -threshold:
        forecast_position = (
            f"선택 예측이 과거 중앙값보다 {format_amount(abs(history_delta))} 낮아 보수적인 전망입니다."
        )
    elif math.isfinite(history_delta):
        forecast_position = "선택 예측이 과거 중앙값과 유사해 기준 전망으로 쓰기 좋은 구간입니다."
    else:
        forecast_position = "과거 기준 범위만으로 전망을 읽는 상태입니다."

    if math.isfinite(target_delta) and target_delta > threshold:
        target_position = f"목표보다 {format_amount(target_delta)} 높은 초과 예상입니다."
        action = "초과분은 안전버퍼와 Stretch 전환분으로 나누고, 실적 인정 가능성을 함께 점검합니다."
    elif math.isfinite(target_delta) and target_delta < -threshold:
        target_position = f"목표보다 {format_amount(abs(target_delta))} 낮은 미달 리스크입니다."
        action = "잔여 목표 보정과 다음 마감 누적선 회복을 우선 검토합니다."
    elif math.isfinite(target_delta):
        target_position = "목표에 근접한 구간입니다."
        action = "현재 페이스를 유지하면서 취소, 철회, 실적인정 리스크를 모니터링합니다."
    else:
        target_position = "목표 대비 차이를 계산할 수 없습니다."
        action = "월 목표와 과거 비교 데이터 입력 상태를 먼저 확인합니다."

    return {
        "has_data": True,
        "headline": f"기준 보고 범위는 {report_range}입니다.",
        "report_range": report_range,
        "report_basis": report_basis,
        "forecast_position": forecast_position,
        "target_position": target_position,
        "action": action,
        "reference_amount": reference_amount,
        "monthly_target": monthly_target,
        "history_delta": history_delta,
        "target_delta": target_delta,
    }


def build_historical_forecast_axis_domain(comparison_df: pd.DataFrame) -> list[float] | None:
    """Return a zoomed x-axis domain for the historical forecast comparison chart."""
    source = _as_dataframe(comparison_df)
    if source.empty:
        return None

    values: list[pd.Series] = []
    for column in ("forecast_amount", "monthly_target"):
        if column in source.columns:
            values.append(pd.to_numeric(source[column], errors="coerce"))
    if not values:
        return None

    return build_auto_axis_domain(pd.concat(values, ignore_index=True))


def _ordered_strategy_keys(scenario_df: pd.DataFrame) -> list[str]:
    default_order = ["P1", "P2", "P3", "O1", "O2", "O3", "N1", "N2", "N3"]
    suffixes: set[str] = set()
    if "scenario_id" in scenario_df:
        for scenario_id in scenario_df["scenario_id"].astype(str):
            if "_" in scenario_id:
                suffixes.add(scenario_id.split("_", maxsplit=1)[1])
    ordered = [suffix for suffix in default_order if suffix in suffixes]
    return ordered or ["P1", "P2", "P3"]


def build_scenario_chart_data(scenario_df: pd.DataFrame) -> pd.DataFrame:
    """Return scenario-level values ready for Streamlit charts."""
    return _build_indexed_numeric_chart_data(
        scenario_df,
        "scenario_id",
        (
            "monthly_target",
            "forecast_amount",
            "forecast_after_provision",
            "revised_monthly_target",
            "target_variance",
            "gap_to_target",
            "surplus_to_target",
            "required_uplift",
            "gap_after_provision",
        ),
    )


def build_scenario_daily_forecast_source(
    df: pd.DataFrame,
    scenario_df: pd.DataFrame,
    as_of_date: object,
    metric: str,
    config: dict[str, Any],
    selected_scenario_id: str | None = None,
) -> pd.DataFrame:
    """Return cumulative daily actuals and scenario forecasts for progress charts."""
    columns = get_metric_columns(metric)
    required_columns = {"date", "is_close_day", columns["target_daily"], columns["actual_cum"]}
    if df.empty or scenario_df.empty or not required_columns.issubset(df.columns):
        return pd.DataFrame(columns=SCENARIO_DAILY_FORECAST_COLUMNS)
    if "scenario_id" not in scenario_df.columns:
        return pd.DataFrame(columns=SCENARIO_DAILY_FORECAST_COLUMNS)

    working = df.copy()
    try:
        working["date"] = pd.to_datetime(working["date"], errors="raise").dt.normalize()
        working = working.sort_values(["date"], kind="mergesort").reset_index(drop=True)
        working["target_daily_numeric"] = pd.to_numeric(
            working[columns["target_daily"]],
            errors="raise",
        ).astype("float64")
        actual_values = working[columns["actual_cum"]].replace(r"^\s*$", pd.NA, regex=True)
        working["actual_cum_numeric"] = pd.to_numeric(actual_values, errors="coerce")
        working["is_close_day_bool"] = _coerce_is_close_day(working["is_close_day"])
    except Exception:  # noqa: BLE001 - visual source should fail closed.
        return pd.DataFrame(columns=SCENARIO_DAILY_FORECAST_COLUMNS)

    working["target_cum"] = working["target_daily_numeric"].cumsum()
    monthly_target = float(working["target_daily_numeric"].sum())
    as_of_timestamp = pd.Timestamp(as_of_date).normalize()
    selected_id = str(selected_scenario_id or "")

    rows: list[dict[str, object]] = []
    actual_rows = working.loc[
        (working["date"] <= as_of_timestamp) & working["actual_cum_numeric"].notna()
    ]
    for _, day_row in actual_rows.iterrows():
        rows.append(
            _build_daily_forecast_row(
                day_row=day_row,
                scenario_id="확정 실적",
                series_type="확정 실적",
                line_group="actual",
                daily_expected=float("nan"),
                forecast_cum=day_row["actual_cum_numeric"],
                monthly_target=monthly_target,
                selected_id=selected_id,
                as_of_timestamp=as_of_timestamp,
                risk_level_label="",
            )
        )

    scenario_lookup = scenario_df.set_index(scenario_df["scenario_id"].astype(str), drop=False)
    for scenario_id in scenario_df["scenario_id"].astype(str).tolist():
        forecast_result, strategy_result = run_selected_scenario_detail(
            df,
            as_of_date,
            metric,
            scenario_id,
            config,
        )
        current_actual_cum = _as_float(forecast_result.get("current_actual_cum"))
        if not math.isfinite(current_actual_cum):
            continue

        as_of_rows = working.loc[working["date"] == as_of_timestamp]
        if not as_of_rows.empty:
            rows.append(
                _build_daily_forecast_row(
                    day_row=as_of_rows.iloc[-1],
                    scenario_id=scenario_id,
                    series_type="시나리오 예상",
                    line_group=scenario_id,
                    daily_expected=0.0,
                    forecast_cum=current_actual_cum,
                    monthly_target=monthly_target,
                    selected_id=selected_id,
                    as_of_timestamp=as_of_timestamp,
                    risk_level_label=_scenario_daily_risk_label(
                        scenario_lookup,
                        scenario_id,
                    ),
                )
            )

        expected_by_day = _build_scenario_expected_daily_map(
            working,
            forecast_result,
            strategy_result,
            as_of_timestamp,
        )
        cumulative = current_actual_cum
        for _, day_row in working.loc[working["date"] > as_of_timestamp].iterrows():
            day = pd.Timestamp(day_row["date"]).normalize()
            daily_expected = expected_by_day.get(day, float("nan"))
            if not math.isfinite(_as_float(daily_expected)):
                continue
            cumulative += float(daily_expected)
            rows.append(
                _build_daily_forecast_row(
                    day_row=day_row,
                    scenario_id=scenario_id,
                    series_type="시나리오 예상",
                    line_group=scenario_id,
                    daily_expected=daily_expected,
                    forecast_cum=cumulative,
                    monthly_target=monthly_target,
                    selected_id=selected_id,
                    as_of_timestamp=as_of_timestamp,
                    risk_level_label=_scenario_daily_risk_label(
                        scenario_lookup,
                        scenario_id,
                    ),
                )
            )

    if not rows:
        return pd.DataFrame(columns=SCENARIO_DAILY_FORECAST_COLUMNS)
    return pd.DataFrame(rows, columns=SCENARIO_DAILY_FORECAST_COLUMNS)


def build_scenario_weekly_forecast_source(daily_source: pd.DataFrame) -> pd.DataFrame:
    """Return weekly endpoint rows from the daily cumulative scenario source."""
    if daily_source.empty:
        return pd.DataFrame(columns=SCENARIO_DAILY_FORECAST_COLUMNS)

    working = daily_source.copy()
    working["date"] = pd.to_datetime(working["date"], errors="coerce")
    working = working.dropna(subset=["date", "forecast_cum"])
    if working.empty:
        return pd.DataFrame(columns=SCENARIO_DAILY_FORECAST_COLUMNS)

    working = working.sort_values(["line_group", "date"], kind="mergesort")
    group_columns = [
        "line_group",
        "scenario_id",
        "series_type",
        "week_start",
        "week_end",
        "week_label",
        "week_no",
    ]
    weekly = (
        working.groupby(group_columns, dropna=False, sort=False)
        .tail(1)
        .sort_values(["date", "line_group"], kind="mergesort")
        .reset_index(drop=True)
    )
    return weekly.loc[:, list(SCENARIO_DAILY_FORECAST_COLUMNS)]


def build_selected_scenario_daily_detail_source(
    daily_source: pd.DataFrame,
    selected_scenario_id: str | None,
) -> pd.DataFrame:
    """Return actual and selected scenario daily rows for the integrated detail table."""
    if daily_source.empty:
        return pd.DataFrame(columns=SCENARIO_DAILY_DETAIL_COLUMNS)

    selected_id = str(selected_scenario_id or "")
    source = daily_source.loc[
        (daily_source["series_type"] == "확정 실적")
        | (daily_source["scenario_id"].astype(str) == selected_id)
    ].copy()
    if source.empty:
        return pd.DataFrame(columns=SCENARIO_DAILY_DETAIL_COLUMNS)

    source["date"] = pd.to_datetime(source["date"], errors="coerce").dt.date
    for column in (
        "daily_expected",
        "forecast_cum",
        "target_cum",
        "achievement_rate",
        "target_achievement_rate",
    ):
        source[column] = pd.to_numeric(source[column], errors="coerce")
    return source.loc[:, list(SCENARIO_DAILY_DETAIL_COLUMNS)].reset_index(drop=True)


def build_remaining_operation_direction_source(
    df: pd.DataFrame,
    as_of_date: object,
    metric: str,
    scenario_id: str,
    forecast_result: dict[str, object],
    strategy_result: dict[str, object],
) -> pd.DataFrame:
    """Return remaining-day operating direction rows for one selected scenario."""
    allocation_by_day = _as_dataframe(strategy_result.get("allocation_by_day"))
    if not allocation_by_day.empty:
        return _build_allocation_operation_direction_source(
            allocation_by_day,
            scenario_id,
            strategy_result,
        )

    return _build_maintenance_operation_direction_source(
        df,
        as_of_date,
        metric,
        scenario_id,
        forecast_result,
        strategy_result,
    )


def _build_allocation_operation_direction_source(
    allocation_by_day: pd.DataFrame,
    scenario_id: str,
    strategy_result: dict[str, object],
) -> pd.DataFrame:
    available_columns = [
        column
        for column in (
            "date",
            "is_close_day",
            "close_type",
            "original_target",
            "uplift",
            "revised_target",
            "expected_after_revision",
            "expected_rate",
            "cap_exceeded",
        )
        if column in allocation_by_day.columns
    ]
    source = allocation_by_day.loc[:, available_columns].copy()
    source["date"] = pd.to_datetime(source["date"], errors="coerce").dt.normalize()
    for column in (
        "original_target",
        "uplift",
        "revised_target",
        "expected_after_revision",
        "expected_rate",
    ):
        if column not in source.columns:
            source[column] = 0.0
        source[column] = pd.to_numeric(source[column], errors="coerce").fillna(0.0)
    if "cap_exceeded" not in source.columns:
        source["cap_exceeded"] = False
    if "is_close_day" not in source.columns:
        source["is_close_day"] = False
    if "close_type" not in source.columns:
        source["close_type"] = ""

    source["scenario_id"] = scenario_id
    source["strategy_type"] = str(strategy_result.get("strategy_type", PROVISION))
    source["operation_mode"] = "업리프트 배분"
    source["expected_daily"] = source["expected_after_revision"]
    source["day_type"] = source["is_close_day"].map(
        lambda value: "마감일" if not _is_missing(value) and bool(value) else "일반일"
    )
    source["direction"] = source.apply(_allocation_direction_label, axis=1)
    source["direction_detail"] = source.apply(_allocation_direction_detail, axis=1)
    source["date_label"] = source["date"].map(_format_chart_index)
    return source.loc[:, list(REMAINING_OPERATION_DIRECTION_COLUMNS)].dropna(
        subset=["date"]
    ).reset_index(drop=True)


def _build_maintenance_operation_direction_source(
    df: pd.DataFrame,
    as_of_date: object,
    metric: str,
    scenario_id: str,
    forecast_result: dict[str, object],
    strategy_result: dict[str, object],
) -> pd.DataFrame:
    columns = get_metric_columns(metric)
    required_columns = {"date", "is_close_day", columns["target_daily"]}
    if df.empty or not required_columns.issubset(df.columns):
        return pd.DataFrame(columns=REMAINING_OPERATION_DIRECTION_COLUMNS)

    working = df.copy()
    try:
        working["date"] = pd.to_datetime(working["date"], errors="raise").dt.normalize()
        working["is_close_day_bool"] = _coerce_is_close_day(working["is_close_day"])
        working["original_target"] = pd.to_numeric(
            working[columns["target_daily"]],
            errors="raise",
        ).astype("float64")
    except Exception:  # noqa: BLE001 - visual source should fail closed.
        return pd.DataFrame(columns=REMAINING_OPERATION_DIRECTION_COLUMNS)

    remaining = working.loc[working["date"] > pd.Timestamp(as_of_date).normalize()].copy()
    if remaining.empty:
        return pd.DataFrame(columns=REMAINING_OPERATION_DIRECTION_COLUMNS)

    expected_rates = _expected_rate_map(forecast_result)
    remaining["expected_rate"] = remaining["date"].map(
        lambda value: expected_rates.get(pd.Timestamp(value).normalize(), float("nan"))
    )
    remaining["expected_daily"] = remaining["original_target"] * remaining["expected_rate"]
    remaining["uplift"] = 0.0
    remaining["revised_target"] = remaining["original_target"]
    remaining["scenario_id"] = scenario_id
    remaining["strategy_type"] = str(strategy_result.get("strategy_type", NEUTRAL))
    remaining["operation_mode"] = _maintenance_operation_mode(strategy_result)
    remaining["day_type"] = remaining["is_close_day_bool"].map(
        lambda value: "마감일" if bool(value) else "일반일"
    )
    remaining["close_type"] = remaining["close_type"] if "close_type" in remaining else ""
    remaining["direction"] = _maintenance_direction_label(strategy_result)
    remaining["direction_detail"] = remaining.apply(
        lambda row: _maintenance_direction_detail(row, strategy_result),
        axis=1,
    )
    remaining["date_label"] = remaining["date"].map(_format_chart_index)
    return remaining.loc[:, list(REMAINING_OPERATION_DIRECTION_COLUMNS)].dropna(
        subset=["date"]
    ).reset_index(drop=True)


def _expected_rate_map(forecast_result: dict[str, object]) -> dict[pd.Timestamp, float]:
    expected_rate_by_day = forecast_result.get("expected_rate_by_day", {})
    if not isinstance(expected_rate_by_day, dict):
        return {}
    return {
        pd.Timestamp(day).normalize(): _as_float(rate)
        for day, rate in expected_rate_by_day.items()
    }


def _allocation_direction_label(row: pd.Series) -> str:
    if bool(row.get("cap_exceeded", False)):
        return "상한 점검"
    if _as_float(row.get("uplift")) > 0:
        return "추가 배분"
    return "기존 목표 유지"


def _allocation_direction_detail(row: pd.Series) -> str:
    if bool(row.get("cap_exceeded", False)):
        return "배분 요청이 일별 상한에 근접하거나 초과합니다. 목표 조정 가능성을 먼저 확인합니다."
    uplift = _as_float(row.get("uplift"))
    if math.isfinite(uplift) and uplift > 0:
        return f"기존 일 목표에 {format_amount(uplift)}를 추가 배분해 잔여 부족분을 회복합니다."
    return "추가 상향 없이 기존 일 목표를 유지합니다."


def _maintenance_operation_mode(strategy_result: dict[str, object]) -> str:
    strategy_type = str(strategy_result.get("strategy_type", NEUTRAL))
    if strategy_type == OVERACHIEVEMENT:
        return "목표 유지/버퍼 관리"
    return "목표 유지/모니터링"


def _maintenance_direction_label(strategy_result: dict[str, object]) -> str:
    strategy_type = str(strategy_result.get("strategy_type", NEUTRAL))
    strategy_id = str(strategy_result.get("strategy_id", ""))
    if strategy_type == OVERACHIEVEMENT and "STRETCH" in strategy_id:
        return "Stretch 후보"
    if strategy_type == OVERACHIEVEMENT:
        return "버퍼 방어"
    return "유지 모니터링"


def _maintenance_direction_detail(
    row: pd.Series,
    strategy_result: dict[str, object],
) -> str:
    direction = _maintenance_direction_label(strategy_result)
    expected_daily = format_amount(row.get("expected_daily"))
    if direction == "Stretch 후보":
        return f"기존 일 목표는 유지하되 예상 일실적 {expected_daily}를 기준으로 초과분 전환 여지를 봅니다."
    if direction == "버퍼 방어":
        return f"기존 일 목표를 유지하고 예상 일실적 {expected_daily} 대비 취소/철회/미결제 리스크를 점검합니다."
    return f"기존 일 목표를 유지하며 예상 일실적 {expected_daily} 흐름이 계획선에서 이탈하는지 봅니다."


def _build_scenario_expected_daily_map(
    working: pd.DataFrame,
    forecast_result: dict[str, object],
    strategy_result: dict[str, object],
    as_of_timestamp: pd.Timestamp,
) -> dict[pd.Timestamp, float]:
    allocation_by_day = _as_dataframe(strategy_result.get("allocation_by_day"))
    if not allocation_by_day.empty and {"date", "expected_after_revision"}.issubset(
        allocation_by_day.columns
    ):
        allocation = allocation_by_day.loc[:, ["date", "expected_after_revision"]].copy()
        allocation["date"] = pd.to_datetime(allocation["date"], errors="coerce").dt.normalize()
        allocation["expected_after_revision"] = pd.to_numeric(
            allocation["expected_after_revision"],
            errors="coerce",
        )
        allocation = allocation.dropna(subset=["date", "expected_after_revision"])
        return {
            pd.Timestamp(day).normalize(): float(value)
            for day, value in allocation.groupby("date")["expected_after_revision"].sum().items()
        }

    expected_rate_by_day = forecast_result.get("expected_rate_by_day", {})
    if not isinstance(expected_rate_by_day, dict):
        return {}
    normalized_rates = {
        pd.Timestamp(day).normalize(): _as_float(rate)
        for day, rate in expected_rate_by_day.items()
    }
    result: dict[pd.Timestamp, float] = {}
    remaining_rows = working.loc[working["date"] > as_of_timestamp]
    for _, day_row in remaining_rows.iterrows():
        day = pd.Timestamp(day_row["date"]).normalize()
        rate = normalized_rates.get(day, float("nan"))
        target = _as_float(day_row.get("target_daily_numeric"))
        if math.isfinite(rate) and math.isfinite(target):
            result[day] = target * rate
    return result


def _build_daily_forecast_row(
    *,
    day_row: pd.Series,
    scenario_id: str,
    series_type: str,
    line_group: str,
    daily_expected: object,
    forecast_cum: object,
    monthly_target: float,
    selected_id: str,
    as_of_timestamp: pd.Timestamp,
    risk_level_label: str,
) -> dict[str, object]:
    forecast_value = _as_float(forecast_cum)
    target_cum = _as_float(day_row.get("target_cum"))
    achievement_rate = safe_divide(forecast_value, monthly_target)
    target_achievement_rate = safe_divide(target_cum, monthly_target)
    date_value = pd.Timestamp(day_row.get("date")).normalize()
    week_start, week_end, week_no, week_label = _week_bucket_values(date_value)
    is_close_day = bool(day_row.get("is_close_day_bool", False))

    return {
        "date": date_value,
        "date_label": _format_chart_index(date_value),
        "week_start": week_start,
        "week_end": week_end,
        "week_label": week_label,
        "week_no": week_no,
        "business_day_no": day_row.get("business_day_no", pd.NA),
        "is_close_day": is_close_day,
        "day_type": "마감일" if bool(is_close_day) else "일반일",
        "close_type": day_row.get("close_type", ""),
        "scenario_id": scenario_id,
        "series_type": series_type,
        "line_group": line_group,
        "daily_expected": _as_float(daily_expected),
        "forecast_cum": forecast_value,
        "target_cum": target_cum,
        "monthly_target": monthly_target,
        "achievement_rate": achievement_rate,
        "target_achievement_rate": target_achievement_rate,
        "achievement_label": format_rate(achievement_rate),
        "target_achievement_label": format_rate(target_achievement_rate),
        "forecast_label": format_amount(forecast_value),
        "target_cum_label": format_amount(target_cum),
        "risk_level_label": risk_level_label,
        "is_selected": series_type == "시나리오 예상" and scenario_id == selected_id,
        "is_as_of_date": date_value == as_of_timestamp,
    }


def _week_bucket_values(date_value: pd.Timestamp) -> tuple[pd.Timestamp, pd.Timestamp, int, str]:
    week_start = (date_value - pd.Timedelta(days=int(date_value.dayofweek))).normalize()
    week_end = week_start + pd.Timedelta(days=6)
    iso = date_value.isocalendar()
    week_no = int(iso.week)
    week_label = f"{week_start.strftime('%m/%d')}~{week_end.strftime('%m/%d')}"
    return week_start, week_end, week_no, week_label


def _scenario_daily_risk_label(
    scenario_lookup: pd.DataFrame,
    scenario_id: str,
) -> str:
    if scenario_id not in scenario_lookup.index:
        return ""
    return str(_localize_display_value(scenario_lookup.loc[scenario_id].get("risk_level", "")))


def build_scenario_target_position_source(
    scenario_df: pd.DataFrame,
    selected_scenario_id: str | None = None,
) -> pd.DataFrame:
    """Return scenario rows focused on target-line comparison."""
    required = ("scenario_id", "monthly_target")
    if scenario_df.empty or not set(required).issubset(scenario_df.columns):
        return pd.DataFrame(
            columns=[
                "scenario_id",
                "monthly_target",
                "forecast_after_provision",
                "target_variance",
                "target_status_label",
                "risk_level_label",
                "is_selected",
                "forecast_label",
                "variance_label",
            ]
        )

    value_column = (
        "forecast_after_provision"
        if "forecast_after_provision" in scenario_df.columns
        else "forecast_amount"
    )
    available_columns = [
        column
        for column in (
            "scenario_id",
            "monthly_target",
            value_column,
            "target_variance",
            "target_status",
            "risk_level",
        )
        if column in scenario_df.columns
    ]
    source = scenario_df.loc[:, available_columns].copy()
    if value_column != "forecast_after_provision":
        source["forecast_after_provision"] = source[value_column]

    for column in ("monthly_target", "forecast_after_provision", "target_variance"):
        if column in source.columns:
            source[column] = pd.to_numeric(source[column], errors="coerce")
    if "target_variance" not in source.columns:
        source["target_variance"] = (
            source["forecast_after_provision"] - source["monthly_target"]
        )

    source["scenario_id"] = source["scenario_id"].astype(str)
    source["target_status_label"] = source.get(
        "target_status",
        pd.Series(["UNKNOWN_TARGET_STATUS"] * len(source), index=source.index),
    ).map(_localize_display_value)
    source["risk_level_label"] = source.get(
        "risk_level",
        pd.Series(["N/A"] * len(source), index=source.index),
    ).map(_localize_display_value)
    source["is_selected"] = source["scenario_id"] == str(selected_scenario_id or "")
    source["forecast_label"] = source["forecast_after_provision"].map(format_amount)
    source["variance_label"] = source["target_variance"].map(_format_signed_amount)
    source = source.dropna(subset=["forecast_after_provision", "monthly_target"])
    return source.reset_index(drop=True)


def build_scenario_heatmap_source(
    scenario_df: pd.DataFrame,
    selected_scenario_id: str | None = None,
) -> pd.DataFrame:
    """Return scenario matrix rows with signed target variance."""
    source = build_scenario_target_position_source(scenario_df, selected_scenario_id)
    if source.empty:
        return pd.DataFrame(
            columns=[
                "scenario_id",
                "forecast_key",
                "strategy_key",
                "target_variance",
                "variance_label",
                "target_status_label",
                "is_selected",
            ]
        )

    rows: list[dict[str, object]] = []
    for row in source.to_dict("records"):
        scenario_id = str(row.get("scenario_id", ""))
        forecast_key, strategy_key = _split_scenario_id(scenario_id)
        rows.append(
            {
                "scenario_id": scenario_id,
                "forecast_key": forecast_key,
                "strategy_key": strategy_key,
                "target_variance": row.get("target_variance"),
                "variance_label": row.get("variance_label"),
                "target_status_label": row.get("target_status_label"),
                "is_selected": row.get("is_selected"),
            }
        )
    return pd.DataFrame(rows)


def build_remaining_target_chart_data(revised_targets_df: pd.DataFrame) -> pd.DataFrame:
    """Return remaining daily target values ready for Streamlit charts."""
    return _build_indexed_numeric_chart_data(
        revised_targets_df,
        "date",
        (
            "original_target",
            "uplift",
            "revised_target",
            "cap_target",
            "expected_after_revision",
        ),
    )


def build_remaining_target_daily_source(revised_targets_df: pd.DataFrame) -> pd.DataFrame:
    """Return remaining-day rows with labels for intuitive target allocation charts."""
    if revised_targets_df.empty or "date" not in revised_targets_df.columns:
        return pd.DataFrame(
            columns=[
                "date",
                "date_label",
                "day_type",
                "close_type",
                "original_target",
                "uplift",
                "revised_target",
                "cap_target",
                "expected_after_revision",
                "cap_exceeded",
            ]
        )

    available_columns = [
        column
        for column in (
            "date",
            "day_name",
            "is_close_day",
            "close_type",
            "original_target",
            "uplift",
            "revised_target",
            "cap_target",
            "expected_after_revision",
            "cap_exceeded",
        )
        if column in revised_targets_df.columns
    ]
    source = revised_targets_df.loc[:, available_columns].copy()
    for column in (
        "original_target",
        "uplift",
        "revised_target",
        "cap_target",
        "expected_after_revision",
    ):
        if column in source.columns:
            source[column] = pd.to_numeric(source[column], errors="coerce")
        else:
            source[column] = pd.NA
    if "cap_exceeded" not in source.columns:
        source["cap_exceeded"] = False
    if "close_type" not in source.columns:
        source["close_type"] = ""

    source["date_label"] = source["date"].map(_format_chart_index)
    if "is_close_day" in source.columns:
        source["day_type"] = source["is_close_day"].map(
            lambda value: "마감일" if not _is_missing(value) and bool(value) else "일반일"
        )
    else:
        source["day_type"] = "잔여일"
    return source.reset_index(drop=True)


def build_remaining_target_stack_source(revised_targets_df: pd.DataFrame) -> pd.DataFrame:
    """Return long-form existing target and uplift rows for stacked bars."""
    daily = build_remaining_target_daily_source(revised_targets_df)
    if daily.empty:
        return pd.DataFrame(columns=["date_label", "target_part", "value"])

    stack_source = daily.melt(
        id_vars=[
            "date",
            "date_label",
            "day_type",
            "close_type",
            "revised_target",
            "cap_target",
            "expected_after_revision",
            "cap_exceeded",
        ],
        value_vars=["original_target", "uplift"],
        var_name="target_part",
        value_name="value",
    )
    stack_source["target_part"] = stack_source["target_part"].map(
        {
            "original_target": "기존 일 목표",
            "uplift": "추가 배분 목표",
        }
    )
    return stack_source.dropna(subset=["value"]).reset_index(drop=True)


def build_strategy_level_table(
    scenario_df: pd.DataFrame,
    selected_scenario_id: str | None = None,
) -> pd.DataFrame:
    """Return strategy-level target values when no daily reallocation is produced."""
    if scenario_df.empty:
        return pd.DataFrame(columns=STRATEGY_LEVEL_COLUMNS)

    result = scenario_df.copy(deep=False)
    forecast_key = _selected_forecast_key(selected_scenario_id)
    if forecast_key and "scenario_id" in result.columns:
        result = result.loc[
            result["scenario_id"].astype(str).str.startswith(f"{forecast_key}_")
        ]
    result = result.copy()
    if "scenario_id" in result.columns:
        result["forecast_model"] = result["scenario_id"].astype(str).map(
            lambda value: _split_scenario_id(value)[0]
        )
    result["strategy_difference_summary"] = result.apply(
        _strategy_difference_summary,
        axis=1,
    )

    available_columns = [
        column for column in STRATEGY_LEVEL_COLUMNS if column in result.columns
    ]
    if not available_columns:
        return pd.DataFrame(columns=STRATEGY_LEVEL_COLUMNS)
    return result.loc[:, available_columns].reset_index(drop=True)


def _strategy_difference_summary(row: pd.Series) -> str:
    strategy_id = str(row.get("provision_strategy") or row.get("overachievement_strategy") or "")
    scenario_id = str(row.get("scenario_id") or "")
    strategy_key = _split_scenario_id(scenario_id)[1]
    stretch = format_amount(row.get("stretch_uplift"))
    buffer = format_amount(row.get("remaining_surplus_buffer"))
    revised_target = format_amount(row.get("revised_monthly_target"))
    minimum = format_amount(row.get("minimum_remaining_to_hit_target"))
    relief = format_amount(row.get("relief_amount"))

    if strategy_id == O1_TARGET_HOLD_BUFFER or strategy_key == "O1":
        return f"공식 목표 유지, 초과 예상분 {buffer}을 안전버퍼로 관리"
    if strategy_id == O2_STRETCH_TARGET_CAPTURE or strategy_key == "O2":
        return f"{stretch}을 Stretch 목표로 전환, 운영 월 목표 {revised_target}"
    if strategy_id == O3_QUALITY_GUARD_RELIEF or strategy_key == "O3":
        return f"목표 달성 최소 잔여 {minimum}, 품질관리 여유 {relief} 분리"
    if strategy_key.startswith("P"):
        return "목표 미달분을 잔여 일자에 배분"
    if strategy_key.startswith("N"):
        return "목표 유지와 변동 리스크 모니터링"
    return "운영 기준 비교"


def build_strategy_effect_table(
    scenario_df: pd.DataFrame,
    selected_scenario_id: str | None = None,
) -> pd.DataFrame:
    """Return O/P/N strategy differences with the forecast model held fixed."""
    strategy_table = build_strategy_level_table(scenario_df, selected_scenario_id)
    if strategy_table.empty or "scenario_id" not in strategy_table.columns:
        return pd.DataFrame()

    result = strategy_table.copy()
    result["forecast_basis"] = result["scenario_id"].astype(str).map(
        lambda value: _split_scenario_id(value)[0]
    )
    result["strategy_key"] = result["scenario_id"].astype(str).map(
        lambda value: _split_scenario_id(value)[1]
    )
    result["strategy_effect_type"] = result["strategy_key"].map(_strategy_effect_type_label)

    preferred_columns = (
        "forecast_basis",
        "strategy_key",
        "provision_strategy",
        "strategy_effect_type",
        "strategy_difference_summary",
        "stretch_uplift",
        "revised_monthly_target",
        "remaining_surplus_buffer",
        "minimum_remaining_to_hit_target",
        "relief_amount",
        "revised_remaining_target",
    )
    available_columns = [column for column in preferred_columns if column in result.columns]
    return result.loc[:, available_columns].reset_index(drop=True)


def _strategy_effect_type_label(strategy_key: object) -> str:
    text = str(strategy_key or "")
    if text == "O1":
        return "버퍼 유지"
    if text == "O2":
        return "Stretch 전환"
    if text == "O3":
        return "품질 방어"
    if text.startswith("P"):
        return "미달 보정"
    if text.startswith("N"):
        return "유지/모니터링"
    return "운영 기준"


def build_strategy_level_chart_data(
    scenario_df: pd.DataFrame,
    selected_scenario_id: str | None = None,
) -> pd.DataFrame:
    """Return strategy-level values ready for Streamlit charts."""
    strategy_table = build_strategy_level_table(scenario_df, selected_scenario_id)
    chart_data = _build_indexed_numeric_chart_data(
        strategy_table,
        "scenario_id",
        STRATEGY_LEVEL_CHART_COLUMNS,
    )
    if chart_data.empty:
        return chart_data
    return chart_data.dropna(axis=1, how="all")


def build_grouped_bar_chart_source(
    chart_data: pd.DataFrame,
    metric_columns: tuple[str, ...],
) -> pd.DataFrame:
    """Return long-form chart data so bars are compared side-by-side, not stacked."""
    available_columns = _available_chart_columns(chart_data, metric_columns)
    if chart_data.empty or not available_columns:
        return pd.DataFrame(columns=["category", "metric", "범례", "value"])

    source = chart_data.loc[:, list(available_columns)].reset_index()
    index_column = str(source.columns[0])
    source = source.rename(columns={index_column: "category"})
    long_source = source.melt(
        id_vars="category",
        value_vars=list(available_columns),
        var_name="metric",
        value_name="value",
    )
    labels = _chart_labels(available_columns)
    long_source["범례"] = long_source["metric"].map(labels)
    long_source["value"] = pd.to_numeric(long_source["value"], errors="coerce")
    return long_source.dropna(subset=["value"]).reset_index(drop=True)


def build_auto_axis_domain(values: pd.Series) -> list[float] | None:
    """Return a padded value-axis domain that emphasizes chart differences."""
    numeric_values = pd.to_numeric(values, errors="coerce").dropna()
    numeric_values = numeric_values.loc[numeric_values.map(math.isfinite)]
    if numeric_values.empty:
        return None

    value_min = float(numeric_values.min())
    value_max = float(numeric_values.max())
    magnitude = max(abs(value_min), abs(value_max), 1.0)
    if math.isclose(value_min, value_max, rel_tol=0.0, abs_tol=1e-9):
        padding = max(magnitude * 0.05, 0.1)
        return [value_min - padding, value_max + padding]

    span = value_max - value_min
    padding = max(span * 0.12, magnitude * 0.02, 0.1)
    lower = value_min - padding
    upper = value_max + padding

    if value_min < 0 < value_max:
        return [min(lower, 0.0), max(upper, 0.0)]
    if value_min >= 0:
        return [max(0.0, lower), upper]
    return [lower, min(0.0, upper)]


def chart_value_format(unit: str) -> str:
    """Return an Altair number format for values already stored in display units."""
    _ = unit
    return ",.1f"


def build_close_cycle_chart_data(close_cycle_df: pd.DataFrame) -> pd.DataFrame:
    """Return close-cycle values ready for Streamlit charts."""
    cumulative_source = build_close_cycle_cumulative_source(close_cycle_df)
    return _build_indexed_numeric_chart_data(
        cumulative_source,
        "cycle_end_date",
        (
            "target_sum",
            "actual_sum",
            "achievement_rate",
            "target_cum",
            "actual_cum",
            "cumulative_achievement_rate",
        ),
    )


def build_visual_metric_definition_df(metric_columns: tuple[str, ...]) -> pd.DataFrame:
    """Return user-facing definitions for chart series."""
    rows = []
    for column in metric_columns:
        definition = VISUAL_METRIC_DEFINITIONS.get(column, {})
        rows.append(
            {
                "범례": definition.get("label", column),
                "단위": definition.get("unit", ""),
                "수치 의미": definition.get("definition", "정의가 등록되지 않은 수치입니다."),
            }
        )
    return pd.DataFrame(rows)


def build_visual_reading_guide(guide_key: str) -> dict[str, object]:
    """Return the reading logic for a visual block."""
    guide = VISUAL_READING_GUIDES.get(guide_key, {})
    return {
        "title": guide.get("title", ""),
        "steps": tuple(guide.get("steps", ())),
        "decision": guide.get("decision", ""),
    }


def build_visual_headline(
    selected_row: pd.Series,
    validation_result: dict[str, Any],
    next_close_result: dict[str, Any],
) -> str:
    """Return the first sentence users should read before the charts."""
    _ = validation_result
    target_status = str(selected_row.get("target_status", "UNKNOWN_TARGET_STATUS"))
    target_variance = _as_float(selected_row.get("target_variance"))
    surplus = _as_float(selected_row.get("surplus_to_target"))
    next_close_sentence = _visual_next_close_sentence(next_close_result)

    if target_status == "UNDER_TARGET" or (
        math.isfinite(target_variance) and target_variance < 0
    ):
        shortage = abs(target_variance) if math.isfinite(target_variance) else selected_row.get("gap_to_target")
        return (
            f"결론: 목표선보다 {format_amount(shortage)} 부족할 가능성이 큽니다. "
            "먼저 전략 반영 후 예상이 공식 월 목표선까지 회복되는지 보고, "
            f"그다음 잔여 일자별 추가 배분이 감당 가능한지 확인하세요. {next_close_sentence}"
        )

    if target_status == "OVER_TARGET" or (
        math.isfinite(target_variance) and target_variance > 0
    ):
        surplus_amount = surplus if math.isfinite(surplus) and surplus > 0 else target_variance
        return (
            f"결론: 목표선보다 {format_amount(surplus_amount)} 여유가 예상됩니다. "
            "초과분을 안전버퍼로 남길지, Stretch 목표로 전환할지 차트에서 확인하세요. "
            f"{next_close_sentence}"
        )

    return (
        "결론: 목표선 근처의 유지/모니터링 구간입니다. "
        "시각화에서는 예측모델별 흔들림과 다음 마감 누적선을 함께 확인하세요. "
        f"{next_close_sentence}"
    )


def build_visual_decision_summary(
    selected_row: pd.Series,
    validation_result: dict[str, Any],
    next_close_result: dict[str, Any],
) -> pd.DataFrame:
    """Return chart interpretation rows in a fixed reading order."""
    target_status = selected_row.get("target_status", "UNKNOWN_TARGET_STATUS")
    risk_level = selected_row.get("risk_level", "N/A")
    operation_mode = _operation_mode_label(target_status)
    monthly_target = validation_result.get("monthly_target")
    forecast_after = selected_row.get("forecast_after_provision")
    target_variance = selected_row.get("target_variance")
    next_close_date = next_close_result.get("next_close_date")
    next_close_required = next_close_result.get("required_to_recover_next_close_cum")

    return pd.DataFrame(
        [
            {
                "확인 순서": "1",
                "볼 것": "목표 판정",
                "현재 값": (
                    f"{_localize_display_value(target_status)} / "
                    f"위험 {_localize_display_value(risk_level)}"
                ),
                "해석": _visual_status_sentence(target_status, operation_mode),
            },
            {
                "확인 순서": "2",
                "볼 것": "목표선 대비 예상 실적",
                "현재 값": (
                    f"{format_amount(forecast_after)} / "
                    f"목표 {format_amount(monthly_target)}"
                ),
                "해석": "막대가 목표선보다 낮으면 잔여 목표 보정이 필요하고, 높으면 초과분 관리가 핵심입니다.",
            },
            {
                "확인 순서": "3",
                "볼 것": "목표 대비 차이",
                "현재 값": _format_signed_amount(target_variance),
                "해석": _visual_variance_sentence(target_variance),
            },
            {
                "확인 순서": "4",
                "볼 것": "다음 마감선",
                "현재 값": (
                    f"{_format_date(next_close_date)} / "
                    f"{format_amount(next_close_required)}"
                ),
                "해석": _visual_next_close_sentence(next_close_result),
            },
        ]
    )


def _visual_status_sentence(target_status: object, operation_mode: object) -> str:
    status = str(target_status)
    mode = str(operation_mode)
    if status == "UNDER_TARGET":
        return f"{mode} 상태입니다. 시나리오별 예상 탭에서 어떤 F/P 조합이 부족분을 줄이는지 보세요."
    if status == "OVER_TARGET":
        return f"{mode} 상태입니다. 초과분을 버퍼로 둘지, 상향 목표로 전환할지 전략 수준 탭에서 보세요."
    if status == "ON_TARGET":
        return f"{mode} 상태입니다. 목표선은 맞지만 마감차수 흐름이 흔들리는지 함께 확인하세요."
    return "목표 판정에 필요한 값이 부족합니다. 입력값 점검 결과를 먼저 확인하세요."


def _visual_variance_sentence(value: object) -> str:
    number = _as_float(value)
    if not math.isfinite(number):
        return "목표 대비 차이를 계산할 수 없습니다. 선택 시나리오 상세 값을 확인하세요."
    if number < 0:
        return f"월말 기준 {format_amount(abs(number))}를 더 채워야 목표선에 도달합니다."
    if number > 0:
        return f"월말 기준 {format_amount(number)}가 목표선 위에 있어 버퍼 또는 Stretch 후보입니다."
    return "월말 예상이 공식 월 목표와 거의 같습니다. 남은 기간의 변동 리스크를 봅니다."


def _visual_next_close_sentence(next_close_result: dict[str, Any]) -> str:
    next_close_date = next_close_result.get("next_close_date")
    required = _as_float(next_close_result.get("required_to_recover_next_close_cum"))
    if _is_missing(next_close_date) or not math.isfinite(required):
        return "다음 마감 기준선은 계산 가능한 데이터가 있을 때 표시됩니다."
    if required <= 0:
        return f"{_format_date(next_close_date)}까지 다음 마감 기준선에 대한 추가 회복 부담은 없습니다."
    return f"{_format_date(next_close_date)}까지 누적 기준으로 최소 {format_amount(required)}을 더 확보해야 합니다."


def _format_signed_amount(value: object) -> str:
    number = _as_float(value)
    if not math.isfinite(number):
        return "계산 불가"
    sign = "+" if number > 0 else ""
    return f"{sign}{number:.1f}억 원"


def build_forecast_definition_df() -> pd.DataFrame:
    """Return display definitions for F1/F2/F3 forecast models."""
    return pd.DataFrame(
        [
            {
                "model": model_id,
                "name": definition["name"],
                "description": definition["description"],
                "formula": definition["formula"],
            }
            for model_id, definition in FORECAST_MODEL_DEFINITIONS.items()
        ]
    )


def build_provision_definition_df() -> pd.DataFrame:
    """Return display definitions for P1/P2/P3 provision strategies."""
    return pd.DataFrame(
        [
            {
                "strategy": strategy_id,
                "name": definition["name"],
                "description": definition["description"],
            }
            for strategy_id, definition in PROVISION_STRATEGY_DEFINITIONS.items()
        ]
    )


def build_overachievement_definition_df() -> pd.DataFrame:
    """Return display definitions for O1/O2/O3 overachievement strategies."""
    return pd.DataFrame(
        [
            {
                "strategy": strategy_id,
                "name": definition["name"],
                "description": definition["description"],
            }
            for strategy_id, definition in OVERACHIEVEMENT_STRATEGY_DEFINITIONS.items()
        ]
    )


def build_neutral_definition_df() -> pd.DataFrame:
    """Return display definitions for N1/N2/N3 neutral strategies."""
    return pd.DataFrame(
        [
            {
                "strategy": strategy_id,
                "name": definition["name"],
                "description": definition["description"],
            }
            for strategy_id, definition in NEUTRAL_STRATEGY_DEFINITIONS.items()
        ]
    )


def build_report_glossary_df() -> pd.DataFrame:
    """Return the fixed glossary used by the generated report."""
    rows: list[dict[str, str]] = []
    for group, definitions in REPORT_GLOSSARY_GROUPS:
        for code, definition in definitions.items():
            rows.append(
                {
                    "구분": group,
                    "코드": str(code),
                    "정의": definition,
                }
            )
    return pd.DataFrame(rows)


def build_risk_definition_df() -> pd.DataFrame:
    """Return display definitions for scenario risk levels."""
    return pd.DataFrame(
        [
            {
                "risk_level": risk_level,
                "definition": definition,
            }
            for risk_level, definition in RISK_LEVEL_DEFINITIONS.items()
        ]
    )


def build_selected_scenario_explanation(
    scenario_id: str,
    selected_row: pd.Series,
) -> pd.DataFrame:
    """Return a concise, non-duplicative summary for the selected scenario."""
    forecast_key, strategy_key = _split_scenario_id(scenario_id)
    forecast_name = FORECAST_MODEL_DEFINITIONS.get(forecast_key, {}).get("name", forecast_key)
    strategy_name = get_strategy_label(strategy_key)
    risk_level = str(selected_row.get("risk_level", "계산 불가"))
    status = str(selected_row.get("status", "계산 불가"))
    strategy_type = str(selected_row.get("strategy_type", PROVISION))
    target_status = str(selected_row.get("target_status", "계산 불가"))
    display_risk_level = _localize_display_value(risk_level)
    display_status = _localize_display_value(status)
    display_strategy_type = _localize_display_value(strategy_type)
    display_target_status = _localize_display_value(target_status)
    display_operation_mode = _operation_mode_label(target_status)

    return pd.DataFrame(
        [
            {
                "item": "선택 조합",
                "value": scenario_id,
            },
            {
                "item": "예측 모델",
                "value": f"{forecast_key} {forecast_name}",
            },
            {
                "item": "운영 전략",
                "value": f"{strategy_key} {strategy_name}",
            },
            {
                "item": "전략 구분",
                "value": display_strategy_type,
            },
            {
                "item": "조합 의미",
                "value": _selected_strategy_meaning(
                    forecast_key,
                    strategy_key,
                    strategy_type,
                    display_target_status,
                ),
            },
            {
                "item": "목표 상태",
                "value": display_target_status,
            },
            {
                "item": "위험등급 / 운영모드",
                "value": f"{display_risk_level} / {display_operation_mode}",
            },
            {
                "item": "계산 상태",
                "value": display_status,
            },
            {
                "item": "월말 예상 실적",
                "value": format_amount(selected_row.get("forecast_amount")),
            },
            {
                "item": "전략 반영 후 예상",
                "value": format_amount(selected_row.get("forecast_after_provision")),
            },
            {
                "item": "목표 대비 차이",
                "value": format_amount(selected_row.get("target_variance")),
            },
            {
                "item": "필요 상향",
                "value": format_amount(selected_row.get("required_uplift")),
            },
            {
                "item": "초과 예상분",
                "value": format_amount(selected_row.get("surplus_to_target")),
            },
            {
                "item": "권장 조치",
                "value": selected_row.get("recommended_action", "계산 결과를 확인합니다."),
            },
        ]
    )


def _selected_strategy_meaning(
    forecast_key: str,
    strategy_key: str,
    strategy_type: str,
    target_status: str,
) -> str:
    if strategy_type == OVERACHIEVEMENT:
        return (
            f"{forecast_key}로 월말 예상 실적을 산출한 결과 {target_status}로 판단되어, "
            f"{strategy_key} 초과달성 운영전략을 적용합니다."
        )
    if strategy_type == NEUTRAL:
        return (
            f"{forecast_key}로 월말 예상 실적을 산출한 결과 {target_status}로 판단되어, "
            f"{strategy_key} 유지/모니터링 전략을 적용합니다."
        )
    return (
        f"{forecast_key}로 월말 예상 실적을 산출한 뒤, "
        f"{strategy_key} 방식으로 목표 상향분을 잔여 일자에 배분합니다."
    )


def format_scenario_option_label(scenario_id: str) -> str:
    """Return a scenario option label with F/P definitions for the selectbox."""
    forecast_key, strategy_key = _split_scenario_id(scenario_id)
    forecast_name = FORECAST_MODEL_DEFINITIONS.get(forecast_key, {}).get(
        "name",
        "정의 없음",
    )
    strategy_name = get_strategy_label(strategy_key) or "정의 없음"
    return f"{scenario_id} - {forecast_key} {forecast_name} / {strategy_key} {strategy_name}"


def format_validation_messages(messages: list[object]) -> list[str]:
    """Return user-facing Korean validation messages."""
    return [format_validation_message(message) for message in messages]


def format_validation_message(message: object) -> str:
    """Translate validation and warning messages for non-technical users."""
    text = str(message)
    if text in VALIDATION_MESSAGE_TRANSLATIONS:
        return VALIDATION_MESSAGE_TRANSLATIONS[text]

    if text.startswith("Missing required input column: "):
        column = _strip_sentence_end(text.removeprefix("Missing required input column: "))
        return f"{_format_column_label(column)} 열이 없습니다. 입력표에 해당 열을 추가해 주세요."

    if text.startswith("Missing required input columns: "):
        columns = _split_column_list(
            _strip_sentence_end(text.removeprefix("Missing required input columns: "))
        )
        column_text = ", ".join(_format_column_label(column) for column in columns)
        return f"{column_text} 열이 없습니다. 입력표에 해당 열을 추가해 주세요."

    if text.startswith("metric must be one of: "):
        return "지표 선택값을 확인해 주세요. 판매실적(sales) 또는 인정실적(recognized) 중 하나를 선택해야 합니다."

    if text.endswith(" contains invalid numeric values."):
        column = text.removesuffix(" contains invalid numeric values.")
        return f"{_format_column_label(column)} 칸에 숫자로 읽을 수 없는 값이 있습니다. 숫자만 입력해 주세요."

    if text.endswith(" contains duplicate values."):
        column = text.removesuffix(" contains duplicate values.")
        return f"{_format_column_label(column)} 값이 중복되어 있습니다. 같은 값이 두 번 들어간 행을 확인해 주세요."

    if text.startswith("Actual daily calculation failed: "):
        detail = _replace_validation_terms(text.removeprefix("Actual daily calculation failed: "))
        return f"일 실적 계산 중 문제가 발생했습니다. {detail}"

    if text.startswith("Close-cycle summary failed: "):
        detail = _replace_validation_terms(text.removeprefix("Close-cycle summary failed: "))
        return f"마감차수 요약 계산 중 문제가 발생했습니다. {detail}"

    if text.startswith("Next close calculation failed: "):
        detail = _replace_validation_terms(text.removeprefix("Next close calculation failed: "))
        return f"다음 마감 누적선 필요실적 계산 중 문제가 발생했습니다. {detail}"

    if text.startswith("Calculation error: "):
        detail = _replace_validation_terms(text.removeprefix("Calculation error: "))
        return f"계산 중 문제가 발생했습니다. {detail}"

    return f"확인 필요: {_replace_validation_terms(text)}"


def build_display_validation_result(validation_result: dict[str, Any]) -> dict[str, Any]:
    """Return a validation result with user-facing messages and original metrics."""
    display_result = dict(validation_result)
    display_result["errors"] = format_validation_messages(
        list(validation_result.get("errors", []))
    )
    display_result["warnings"] = format_validation_messages(
        list(validation_result.get("warnings", []))
    )
    return display_result


def _format_column_label(column: str) -> str:
    column = column.strip()
    label = VALIDATION_COLUMN_LABELS.get(column)
    if label is None:
        return column
    return f"{label}({column})"


def _replace_validation_terms(text: str) -> str:
    replaced = text
    for term in sorted(VALIDATION_TERM_REPLACEMENTS, key=len, reverse=True):
        replaced = replaced.replace(term, VALIDATION_TERM_REPLACEMENTS[term])
    return replaced


def _strip_sentence_end(text: str) -> str:
    return text.strip().removesuffix(".").strip()


def _split_column_list(text: str) -> list[str]:
    return [column.strip() for column in text.split(",") if column.strip()]


def build_summary_dict(
    validation_result: dict[str, Any],
    selected_row: pd.Series,
    next_close_result: dict[str, Any],
    metric: str,
    as_of_date: object,
) -> dict[str, Any]:
    """Build summary values for the KPI area and Excel export."""
    achievement_rate = safe_divide(
        validation_result.get("current_actual_cum"),
        validation_result.get("current_target_cum"),
    )
    return {
        "metric": metric,
        "as_of_date": pd.Timestamp(as_of_date).date(),
        "monthly_target": validation_result.get("monthly_target"),
        "current_target_cum": validation_result.get("current_target_cum"),
        "current_actual_cum": validation_result.get("current_actual_cum"),
        "cumulative_achievement_rate": achievement_rate,
        "selected_scenario_id": selected_row.get("scenario_id"),
        "forecast_after_provision": selected_row.get("forecast_after_provision"),
        "target_status": selected_row.get("target_status"),
        "target_variance": selected_row.get("target_variance"),
        "surplus_to_target": selected_row.get("surplus_to_target"),
        "strategy_type": selected_row.get("strategy_type"),
        "risk_level": selected_row.get("risk_level"),
        "status": selected_row.get("status"),
        "next_close_date": next_close_result.get("next_close_date"),
        "next_close_required": next_close_result.get(
            "required_to_recover_next_close_cum"
        ),
        "validation_errors": format_validation_messages(
            list(validation_result.get("errors", []))
        ),
        "validation_warnings": format_validation_messages(
            list(validation_result.get("warnings", []))
        ),
    }


def list_latest_excel_outputs(output_dir: Path | None = None) -> pd.DataFrame:
    """Return read-only metadata for existing Excel files under outputs/latest."""
    latest_output_dir = Path(output_dir) if output_dir is not None else OUTPUT_DIR / "latest"
    columns = ["파일명", "경로", "수정시각", "크기(bytes)", "공유구분"]
    if not latest_output_dir.exists():
        return pd.DataFrame(columns=columns)

    rows: list[dict[str, object]] = []
    for path in sorted(latest_output_dir.iterdir(), key=lambda item: item.name.lower()):
        if not path.is_file() or path.suffix.lower() != ".xlsx":
            continue
        stat = path.stat()
        category = "daily_report" if path.name.startswith("daily_report_") else "reference"
        rows.append(
            {
                "파일명": path.name,
                "경로": f"outputs/latest/{path.name}",
                "수정시각": pd.Timestamp.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
                "크기(bytes)": stat.st_size,
                "공유구분": category,
            }
        )
    return pd.DataFrame(rows, columns=columns)


def _latest_existing_report_path(latest_files: pd.DataFrame) -> Path | None:
    if latest_files.empty or "파일명" not in latest_files.columns:
        return None
    candidates = latest_files.loc[latest_files["파일명"].astype(str).str.startswith("daily_report_")]
    if candidates.empty:
        return None
    paths = [OUTPUT_DIR / "latest" / str(name) for name in candidates["파일명"].tolist()]
    existing = [path for path in paths if path.exists() and path.is_file()]
    if not existing:
        return None
    return max(existing, key=lambda path: path.stat().st_mtime)


def build_excel_report_bytes(
    summary_dict: dict[str, Any],
    scenario_df: pd.DataFrame,
    revised_targets_df: pd.DataFrame,
    close_cycle_df: pd.DataFrame,
    validation_result: dict[str, Any],
    report_text: str,
    metric: str,
    as_of_date: object,
) -> tuple[bytes, str]:
    """Export the calculated report to outputs and return bytes for download."""
    latest_output_dir = OUTPUT_DIR / "latest"
    latest_output_dir.mkdir(parents=True, exist_ok=True)
    date_token = pd.Timestamp(as_of_date).strftime("%Y%m%d")
    report_name = f"daily_report_{metric}_{date_token}.xlsx"
    output_path = latest_output_dir / report_name
    saved_path = export_daily_report(
        output_path,
        summary_dict,
        scenario_df,
        revised_targets_df,
        close_cycle_df,
        validation_result,
        report_text,
        overwrite=True,
    )
    return saved_path.read_bytes(), saved_path.name


def load_forecast_history_for_app(path: str | Path | None = None) -> pd.DataFrame:
    """Load forecast history for the UI, preserving optional future columns."""
    history_path = (
        Path(path)
        if path is not None
        else _history_storage_path(history_schema.FORECAST_HISTORY)
    )
    columns = list(history_schema.FORECAST_HISTORY_COLUMNS)
    if not history_path.exists() or history_path.stat().st_size == 0:
        return pd.DataFrame(columns=columns)

    loaded = pd.read_csv(history_path, encoding="utf-8-sig")
    history_schema.validate_required_columns(loaded.columns, history_schema.FORECAST_HISTORY)
    optional_columns = [column for column in loaded.columns if column not in columns]
    return loaded.loc[:, [*columns, *optional_columns]].copy()


def load_history_tables_for_app(
    forecast_history_path: str | Path | None = None,
    final_actuals_path: str | Path | None = None,
) -> dict[str, pd.DataFrame]:
    """Load history tables without failing when either storage file is absent."""
    final_path = (
        Path(final_actuals_path)
        if final_actuals_path is not None
        else _history_storage_path(history_schema.FINAL_ACTUALS)
    )
    return {
        "forecast_history": load_forecast_history_for_app(forecast_history_path),
        "final_actuals": load_final_actuals(final_path),
    }


def save_forecast_history_snapshot(
    scenario_df: pd.DataFrame,
    as_of_date: object,
    metric: str,
    path: str | Path | None = None,
) -> pd.DataFrame:
    """Save the current scenario grid as one forecast history snapshot."""
    target_path = (
        Path(path)
        if path is not None
        else _history_storage_path(history_schema.FORECAST_HISTORY)
    )
    run_context = {
        "run_id": None,
        "run_datetime": pd.Timestamp.now(tz=APP_TIMEZONE),
        "target_month": pd.Timestamp(as_of_date).strftime("%Y-%m"),
        "as_of_date": pd.Timestamp(as_of_date).date().isoformat(),
        "metric": metric,
    }
    rows = build_forecast_history_rows(scenario_df, run_context)
    return append_forecast_history(rows, target_path)


def build_backtest_insights(
    forecast_history: pd.DataFrame,
    final_actuals: pd.DataFrame,
    model_summary: pd.DataFrame,
) -> list[str]:
    """Return concise interpretation messages for the Backtest tab."""
    if forecast_history.empty:
        return ["저장된 forecast_history가 아직 없습니다. 계산 후 예측 이력 저장을 누르면 Backtest 표본이 쌓입니다."]
    if final_actuals.empty:
        return ["final_actuals가 아직 없어 예측값과 확정 실적의 오차율은 계산하지 않았습니다."]
    if model_summary.empty:
        return ["forecast_history와 final_actuals의 target_month/metric이 아직 매칭되지 않아 Backtest 요약을 만들 수 없습니다."]

    ranked = model_summary.copy()
    ranked["mean_error_rate"] = pd.to_numeric(ranked["mean_error_rate"], errors="coerce")
    ranked = ranked.loc[ranked["mean_error_rate"].notna()].sort_values(
        ["mean_error_rate", "forecast_model"],
        kind="mergesort",
    )
    if ranked.empty:
        return ["Backtest 표본은 있지만 유효한 평균 오차율이 없어 모델 우열을 판단하지 않았습니다."]

    best = ranked.iloc[0]
    bias = _as_float(best.get("bias"))
    bias_message = "bias는 거의 중립입니다."
    if math.isfinite(bias) and bias > 0:
        bias_message = "bias가 양수라 평균적으로 과대 예측 경향이 있습니다."
    elif math.isfinite(bias) and bias < 0:
        bias_message = "bias가 음수라 평균적으로 과소 예측 경향이 있습니다."

    return [
        f"현재 표본 기준 최저 평균 오차율 모델은 {best.get('forecast_model')}이며 평균 오차율은 {format_rate(best.get('mean_error_rate'))}입니다.",
        bias_message,
        "동적 가중 예측은 이 모델별 오차율과 bias가 충분히 누적된 뒤 적용 후보로 판단할 수 있습니다.",
    ]


def _history_storage_path(schema_name: str) -> Path:
    return history_schema.get_storage_paths(repo_root=REPO_ROOT)[schema_name]


def format_amount(value: object) -> str:
    """Format an amount in hundred-million KRW."""
    number = _as_float(value)
    if not math.isfinite(number):
        return "계산 불가"
    return f"{number:.1f}억 원"


def format_rate(value: object) -> str:
    """Format a decimal ratio as a percentage."""
    number = _as_float(value)
    if not math.isfinite(number):
        return "계산 불가"
    return f"{number * 100:.1f}%"


def safe_divide(numerator: object, denominator: object) -> float:
    """Return numerator / denominator or NaN for unavailable ratios."""
    numerator_value = _as_float(numerator)
    denominator_value = _as_float(denominator)
    if not math.isfinite(numerator_value) or not math.isfinite(denominator_value):
        return float("nan")
    if denominator_value == 0:
        return float("nan")
    return numerator_value / denominator_value


def build_historical_monthly_summary(
    historical_df: pd.DataFrame,
    metric: str,
) -> pd.DataFrame:
    """Return month-level final target and achievement summaries."""
    prepared = _prepare_historical_metric_frame(historical_df, metric)
    if prepared.empty:
        return pd.DataFrame(
            columns=[
                "month",
                "row_count",
                "completed_actual_days",
                "final_business_day_no",
                "monthly_target",
                "final_actual_cum",
                "final_achievement_rate",
                "close_day_count",
            ]
        )

    rows: list[dict[str, object]] = []
    for month, month_df in prepared.groupby("_month", sort=True):
        completed = month_df.dropna(subset=["_actual_cum"])
        final_business_day_no = (
            completed["_business_day_no"].max()
            if not completed.empty
            else month_df["_business_day_no"].max()
        )
        final_actual_cum = (
            completed.sort_values(["_business_day_no", "_date"])["_actual_cum"].iloc[-1]
            if not completed.empty
            else float("nan")
        )
        monthly_target = month_df["_target_daily"].sum()
        rows.append(
            {
                "month": str(month),
                "row_count": int(len(month_df)),
                "completed_actual_days": int(len(completed)),
                "final_business_day_no": int(final_business_day_no),
                "monthly_target": float(monthly_target),
                "final_actual_cum": float(final_actual_cum),
                "final_achievement_rate": safe_divide(final_actual_cum, monthly_target),
                "close_day_count": int(month_df["_is_close_day"].sum()),
            }
        )
    return pd.DataFrame(rows)


def build_historical_stage_benchmark(
    historical_df: pd.DataFrame,
    current_df: pd.DataFrame,
    as_of_date: object,
    metric: str,
    validation_result: dict[str, Any] | None = None,
) -> dict[str, object]:
    """Compare the current month with historical months at the same business-day stage."""
    stage_df = build_historical_stage_comparison(
        historical_df,
        current_df,
        as_of_date,
        metric,
    )
    current_business_day_no = _current_business_day_no(current_df, as_of_date)
    current_monthly_target = _historical_context_value(
        validation_result,
        "monthly_target",
        _current_monthly_target(current_df, metric),
    )
    current_target_cum = _historical_context_value(
        validation_result,
        "current_target_cum",
        _current_target_cum(current_df, as_of_date, metric),
    )
    current_actual_cum = _historical_context_value(
        validation_result,
        "current_actual_cum",
        _current_actual_cum(current_df, as_of_date, metric),
    )
    current_rate = safe_divide(current_actual_cum, current_target_cum)

    result: dict[str, object] = {
        "month_count": int(len(stage_df)),
        "current_business_day_no": current_business_day_no,
        "current_monthly_target": current_monthly_target,
        "current_target_cum": current_target_cum,
        "current_actual_cum": current_actual_cum,
        "current_achievement_rate": current_rate,
        "historical_stage_median_rate": float("nan"),
        "historical_stage_p25_rate": float("nan"),
        "historical_stage_p75_rate": float("nan"),
        "historical_final_median_rate": float("nan"),
        "historical_forecast_lower": float("nan"),
        "historical_forecast_median": float("nan"),
        "historical_forecast_upper": float("nan"),
        "current_stage_percentile": float("nan"),
        "stage_df": stage_df,
    }
    if stage_df.empty:
        return result

    stage_rates = pd.to_numeric(
        stage_df["as_of_achievement_rate"],
        errors="coerce",
    ).dropna()
    final_rates = pd.to_numeric(
        stage_df["final_achievement_rate"],
        errors="coerce",
    ).dropna()
    if not stage_rates.empty:
        result["historical_stage_median_rate"] = float(stage_rates.median())
        result["historical_stage_p25_rate"] = float(stage_rates.quantile(0.25))
        result["historical_stage_p75_rate"] = float(stage_rates.quantile(0.75))
        if math.isfinite(current_rate):
            result["current_stage_percentile"] = float((stage_rates <= current_rate).mean())
    if not final_rates.empty:
        result["historical_final_median_rate"] = float(final_rates.median())

    multiplier_source = stage_df.copy()
    multiplier_source["stage_to_final_multiplier"] = [
        safe_divide(final_rate, stage_rate)
        for final_rate, stage_rate in zip(
            multiplier_source["final_achievement_rate"],
            multiplier_source["as_of_achievement_rate"],
        )
    ]
    multipliers = pd.to_numeric(
        multiplier_source["stage_to_final_multiplier"],
        errors="coerce",
    ).dropna()
    multipliers = multipliers.loc[multipliers.map(math.isfinite)]
    multipliers = multipliers.loc[multipliers > 0]
    if not multipliers.empty and math.isfinite(current_rate):
        result["historical_forecast_lower"] = float(
            current_monthly_target * current_rate * multipliers.quantile(0.25)
        )
        result["historical_forecast_median"] = float(
            current_monthly_target * current_rate * multipliers.median()
        )
        result["historical_forecast_upper"] = float(
            current_monthly_target * current_rate * multipliers.quantile(0.75)
        )
    return result


def build_historical_stage_comparison(
    historical_df: pd.DataFrame,
    current_df: pd.DataFrame,
    as_of_date: object,
    metric: str,
) -> pd.DataFrame:
    """Return historical month rows matched to the current business-day stage."""
    prepared = _prepare_historical_metric_frame(historical_df, metric)
    if prepared.empty:
        return pd.DataFrame()

    current_business_day_no = _current_business_day_no(current_df, as_of_date)
    if not math.isfinite(current_business_day_no):
        return pd.DataFrame()

    monthly_summary = build_historical_monthly_summary(historical_df, metric)
    monthly_summary = monthly_summary.set_index("month", drop=False)
    rows: list[dict[str, object]] = []
    for month, month_df in prepared.groupby("_month", sort=True):
        candidates = month_df.loc[
            (month_df["_business_day_no"] <= current_business_day_no)
            & month_df["_actual_cum"].notna()
            & (month_df["_target_cum"] > 0)
        ]
        if candidates.empty or str(month) not in monthly_summary.index:
            continue

        stage_row = candidates.sort_values(["_business_day_no", "_date"]).iloc[-1]
        final_row = monthly_summary.loc[str(month)]
        rows.append(
            {
                "month": str(month),
                "matched_business_day_no": int(stage_row["_business_day_no"]),
                "as_of_target_cum": float(stage_row["_target_cum"]),
                "as_of_actual_cum": float(stage_row["_actual_cum"]),
                "as_of_achievement_rate": safe_divide(
                    stage_row["_actual_cum"],
                    stage_row["_target_cum"],
                ),
                "monthly_target": float(final_row["monthly_target"]),
                "final_actual_cum": float(final_row["final_actual_cum"]),
                "final_achievement_rate": float(final_row["final_achievement_rate"]),
                "remaining_actual_growth": float(
                    final_row["final_actual_cum"] - stage_row["_actual_cum"]
                ),
            }
        )
    return pd.DataFrame(rows)


def build_historical_progress_profile(
    historical_df: pd.DataFrame,
    metric: str,
) -> pd.DataFrame:
    """Return historical cumulative achievement bands by business day."""
    prepared = _prepare_historical_metric_frame(historical_df, metric)
    if prepared.empty:
        return pd.DataFrame()

    valid = prepared.dropna(subset=["_achievement_rate"])
    if valid.empty:
        return pd.DataFrame()

    grouped = valid.groupby("_business_day_no")["_achievement_rate"]
    profile = pd.DataFrame(
        {
            "business_day_no": grouped.median().index.astype(int),
            "historical_p25_rate": grouped.quantile(0.25).to_numpy(),
            "historical_median_rate": grouped.median().to_numpy(),
            "historical_p75_rate": grouped.quantile(0.75).to_numpy(),
            "month_count": grouped.count().to_numpy(),
        }
    )
    return profile.reset_index(drop=True)


def build_historical_progress_chart_data(
    historical_df: pd.DataFrame,
    current_df: pd.DataFrame,
    as_of_date: object,
    metric: str,
) -> pd.DataFrame:
    """Return long-form current and historical progress data for charting."""
    profile = build_historical_progress_profile(historical_df, metric)
    current_progress = _build_current_progress_series(current_df, as_of_date, metric)
    rows: list[dict[str, object]] = []

    for _, row in profile.iterrows():
        business_day_no = int(row["business_day_no"])
        for column, label in (
            ("historical_p25_rate", "과거 하위 25%"),
            ("historical_median_rate", "과거 중앙값"),
            ("historical_p75_rate", "과거 상위 25%"),
        ):
            rows.append(
                {
                    "business_day_no": business_day_no,
                    "series": label,
                    "achievement_rate": float(row[column]),
                }
            )

    for _, row in current_progress.iterrows():
        rows.append(
            {
                "business_day_no": int(row["business_day_no"]),
                "series": "현재 월",
                "achievement_rate": float(row["achievement_rate"]),
            }
        )
    return pd.DataFrame(rows)


def build_historical_context(
    historical_df: pd.DataFrame,
    current_df: pd.DataFrame,
    as_of_date: object,
    metric: str,
    validation_result: dict[str, Any],
    source_label: str = "",
) -> dict[str, object]:
    """Return all historical analysis artifacts used by the UI."""
    history = _as_dataframe(historical_df)
    if history.empty:
        return {"has_data": False, "source_label": source_label}

    monthly_summary = build_historical_monthly_summary(history, metric)
    benchmark = build_historical_stage_benchmark(
        history,
        current_df,
        as_of_date,
        metric,
        validation_result,
    )
    chart_data = build_historical_progress_chart_data(
        history,
        current_df,
        as_of_date,
        metric,
    )
    return {
        "has_data": True,
        "source_label": source_label,
        "row_count": int(len(history)),
        "monthly_summary": monthly_summary,
        "benchmark": benchmark,
        "progress_chart_data": chart_data,
        "interpretation": build_historical_interpretation(benchmark),
    }


def build_historical_interpretation(benchmark: dict[str, object]) -> list[str]:
    """Return business-readable interpretation sentences for historical context."""
    month_count = int(benchmark.get("month_count") or 0)
    if month_count == 0:
        return [
            "현재 기준 영업일차와 비교할 수 있는 과거 월 데이터가 아직 부족합니다.",
            "과거 파일에는 같은 영업일차까지 누적 실적이 입력된 완료 월을 포함해 주세요.",
        ]

    current_rate = _as_float(benchmark.get("current_achievement_rate"))
    median_rate = _as_float(benchmark.get("historical_stage_median_rate"))
    p25_rate = _as_float(benchmark.get("historical_stage_p25_rate"))
    p75_rate = _as_float(benchmark.get("historical_stage_p75_rate"))
    current_business_day_no = benchmark.get("current_business_day_no")

    messages = [
        f"현재 월은 {current_business_day_no}영업일차 기준으로 과거 {month_count}개 월과 비교합니다.",
    ]
    if math.isfinite(current_rate) and math.isfinite(median_rate):
        diff = current_rate - median_rate
        messages.append(
            f"현재 누적 달성률은 {format_rate(current_rate)}이고, 같은 영업일차 과거 중앙값은 {format_rate(median_rate)}입니다."
        )
        if math.isfinite(p25_rate) and current_rate < p25_rate:
            messages.append("현재 흐름은 과거 하위 25%보다 낮아 월말 리스크를 보수적으로 보는 편이 좋습니다.")
        elif math.isfinite(p75_rate) and current_rate > p75_rate:
            messages.append("현재 흐름은 과거 상위 25%보다 높아 초과분 관리와 품질 리스크를 함께 볼 구간입니다.")
        else:
            messages.append("현재 흐름은 과거 일반 범위 안에 있어 기존 예측모델과 함께 균형 있게 해석할 수 있습니다.")
        if abs(diff) >= 0.01:
            messages.append(f"과거 중앙값 대비 차이는 {diff * 100:+.1f}%p입니다.")

    lower = _as_float(benchmark.get("historical_forecast_lower"))
    median = _as_float(benchmark.get("historical_forecast_median"))
    upper = _as_float(benchmark.get("historical_forecast_upper"))
    if all(math.isfinite(value) for value in (lower, median, upper)):
        messages.append(
            "과거의 같은 시점 이후 월말 전환 흐름을 적용하면 "
            f"{format_amount(lower)} ~ {format_amount(upper)} 범위, 중앙값 {format_amount(median)} 수준으로 읽을 수 있습니다."
        )
    return messages


def _prepare_historical_metric_frame(
    df: pd.DataFrame,
    metric: str,
) -> pd.DataFrame:
    metric_columns = get_metric_columns(metric)
    target_column = metric_columns["target_daily"]
    actual_column = metric_columns["actual_cum"]
    required_columns = {"date", "business_day_no", "is_close_day", target_column, actual_column}
    if df.empty or not required_columns.issubset(df.columns):
        return pd.DataFrame()

    result = df.copy()
    result["_date"] = pd.to_datetime(result["date"], errors="coerce")
    result["_month"] = result["_date"].dt.to_period("M").astype(str)
    result["_business_day_no"] = pd.to_numeric(
        result["business_day_no"],
        errors="coerce",
    )
    result["_target_daily"] = pd.to_numeric(result[target_column], errors="coerce")
    result["_actual_cum"] = pd.to_numeric(result[actual_column], errors="coerce")
    result["_is_close_day"] = result["is_close_day"].astype(bool)
    result = result.dropna(subset=["_date", "_business_day_no", "_target_daily"])
    if result.empty:
        return pd.DataFrame()

    result = result.sort_values(["_month", "_business_day_no", "_date"]).reset_index(drop=True)
    result["_target_cum"] = result.groupby("_month")["_target_daily"].cumsum()
    result["_achievement_rate"] = [
        safe_divide(actual, target)
        for actual, target in zip(result["_actual_cum"], result["_target_cum"])
    ]
    return result


def _build_current_progress_series(
    current_df: pd.DataFrame,
    as_of_date: object,
    metric: str,
) -> pd.DataFrame:
    prepared = _prepare_historical_metric_frame(current_df, metric)
    if prepared.empty:
        return pd.DataFrame(columns=["business_day_no", "achievement_rate"])

    as_of_timestamp = pd.Timestamp(as_of_date).normalize()
    current = prepared.loc[
        (prepared["_date"].dt.normalize() <= as_of_timestamp)
        & prepared["_actual_cum"].notna()
        & prepared["_achievement_rate"].notna()
    ].copy()
    if current.empty:
        return pd.DataFrame(columns=["business_day_no", "achievement_rate"])
    return pd.DataFrame(
        {
            "business_day_no": current["_business_day_no"].astype(int),
            "achievement_rate": current["_achievement_rate"].astype(float),
        }
    )


def _current_business_day_no(df: pd.DataFrame, as_of_date: object) -> float:
    if df.empty or "date" not in df.columns or "business_day_no" not in df.columns:
        return float("nan")

    dates = pd.to_datetime(df["date"], errors="coerce").dt.normalize()
    as_of_timestamp = pd.Timestamp(as_of_date).normalize()
    rows = df.loc[dates == as_of_timestamp]
    if rows.empty:
        return float("nan")
    return float(pd.to_numeric(rows.iloc[0]["business_day_no"], errors="coerce"))


def _current_monthly_target(df: pd.DataFrame, metric: str) -> float:
    metric_columns = get_metric_columns(metric)
    target_column = metric_columns["target_daily"]
    if df.empty or target_column not in df.columns:
        return float("nan")
    return float(pd.to_numeric(df[target_column], errors="coerce").sum())


def _current_target_cum(
    df: pd.DataFrame,
    as_of_date: object,
    metric: str,
) -> float:
    metric_columns = get_metric_columns(metric)
    target_column = metric_columns["target_daily"]
    if df.empty or "date" not in df.columns or target_column not in df.columns:
        return float("nan")

    dates = pd.to_datetime(df["date"], errors="coerce").dt.normalize()
    as_of_timestamp = pd.Timestamp(as_of_date).normalize()
    values = pd.to_numeric(df.loc[dates <= as_of_timestamp, target_column], errors="coerce")
    if values.empty:
        return float("nan")
    return float(values.sum())


def _current_actual_cum(
    df: pd.DataFrame,
    as_of_date: object,
    metric: str,
) -> float:
    metric_columns = get_metric_columns(metric)
    actual_column = metric_columns["actual_cum"]
    if df.empty or "date" not in df.columns or actual_column not in df.columns:
        return float("nan")

    dates = pd.to_datetime(df["date"], errors="coerce").dt.normalize()
    as_of_timestamp = pd.Timestamp(as_of_date).normalize()
    rows = df.loc[dates == as_of_timestamp]
    if rows.empty:
        return float("nan")
    return float(pd.to_numeric(pd.Series([rows.iloc[0][actual_column]]), errors="coerce").iloc[0])


def _historical_context_value(
    validation_result: dict[str, Any] | None,
    key: str,
    fallback: float,
) -> float:
    if validation_result is None:
        return fallback
    value = _as_float(validation_result.get(key))
    if math.isfinite(value):
        return value
    return fallback


def _build_template_workbook_bytes(
    sheet_title: str,
    sample_rows: tuple[tuple[object, ...], ...],
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_title
    worksheet.append(list(INPUT_TEMPLATE_HEADERS))
    for row in sample_rows:
        worksheet.append(list(row))

    header_fill = PatternFill(fill_type="solid", fgColor="44546A")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:J{worksheet.max_row}"
    for cell in worksheet[1]:
        cell_width = max(12, min(28, len(str(cell.value)) + 4))
        worksheet.column_dimensions[cell.column_letter].width = cell_width

    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def build_input_template_bytes() -> bytes:
    return _build_template_workbook_bytes("InputTemplate", INPUT_TEMPLATE_SAMPLE_ROWS)


def build_historical_input_template_bytes() -> bytes:
    return _build_template_workbook_bytes(
        "HistoricalInputTemplate",
        _load_historical_input_template_sample_rows(),
    )


def _load_historical_input_template_sample_rows() -> tuple[tuple[object, ...], ...]:
    try:
        sample_df = pd.read_csv(HISTORICAL_SAMPLE_INPUT_PATH, encoding="utf-8-sig")
    except (OSError, UnicodeDecodeError, pd.errors.ParserError):
        return HISTORICAL_INPUT_TEMPLATE_SAMPLE_ROWS

    missing_columns = [
        column for column in INPUT_TEMPLATE_HEADERS if column not in sample_df.columns
    ]
    if missing_columns:
        return HISTORICAL_INPUT_TEMPLATE_SAMPLE_ROWS

    rows: list[tuple[object, ...]] = []
    for row in sample_df.loc[:, list(INPUT_TEMPLATE_HEADERS)].itertuples(
        index=False,
        name=None,
    ):
        rows.append(tuple(None if pd.isna(value) else value for value in row))

    return tuple(rows) or HISTORICAL_INPUT_TEMPLATE_SAMPLE_ROWS


def _render_input_template_download() -> None:
    st.download_button(
        "엑셀 업로드 양식 다운로드",
        data=build_input_template_bytes(),
        file_name=INPUT_TEMPLATE_FILENAME,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _render_historical_input_template_download() -> None:
    st.download_button(
        "과거 월 누적 업로드 양식 다운로드",
        data=build_historical_input_template_bytes(),
        file_name=HISTORICAL_INPUT_TEMPLATE_FILENAME,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _render_file_upload() -> tuple[pd.DataFrame | None, str]:
    st.header("1. 파일 업로드")
    _render_input_template_download()
    st.caption(
        "업로드 파일은 현재 화면에 먼저 적용됩니다. 앱 리부트 후 기본 입력값으로 쓰려면 "
        "입력 수정 영역의 완료월 실제 실적 저장 버튼으로 현재 입력값을 저장하세요."
    )
    uploaded_file = st.file_uploader("입력 파일 업로드", type=["csv", "xlsx"])
    uploaded_name = getattr(uploaded_file, "name", None)
    if uploaded_name != st.session_state.get("uploaded_file_name"):
        st.session_state["uploaded_file_name"] = uploaded_name
        st.session_state["force_sample_input"] = False
    if st.button("샘플 데이터 로딩"):
        st.session_state["force_sample_input"] = True

    try:
        if uploaded_file is not None and not st.session_state.get("force_sample_input", False):
            return _load_uploaded_input(uploaded_file), uploaded_file.name
        if st.session_state.get("force_sample_input", False):
            return _load_packaged_sample_for_app("current_input"), SAMPLE_INPUT_SOURCE_LABEL
        return _get_current_input_state()
    except Exception as exc:  # noqa: BLE001 - surface load errors in the UI.
        st.error(f"입력 파일을 로딩할 수 없습니다: {exc}")
        return None, ""


def _render_historical_upload() -> tuple[pd.DataFrame, str]:
    st.header("1-1. 과거 월 누적 데이터")
    with st.expander("과거 월 데이터 업로드", expanded=False):
        st.caption(
            "과거 샘플 데이터는 기본으로 비교 계산에 반영합니다. "
            "별도 CSV/XLSX를 업로드하면 업로드 파일을 우선 사용합니다. "
            "과거 월 파일은 비교 계산에만 사용하고 최신 기본값 저장 대상에서는 제외합니다."
        )
        _render_historical_input_template_download()
        uploaded_file = st.file_uploader(
            "과거 월 누적 파일 업로드",
            type=["csv", "xlsx"],
            key="historical_month_upload",
        )
        sample_col, clear_col = st.columns(2)
        if sample_col.button("기본 샘플 다시 적용", key="load_historical_sample"):
            st.session_state["use_historical_sample_input"] = True
            st.session_state[HISTORICAL_SAMPLE_DISABLED_SESSION_KEY] = False
        if clear_col.button("과거 데이터 비우기", key="clear_historical_sample"):
            st.session_state["use_historical_sample_input"] = False
            st.session_state[HISTORICAL_SAMPLE_DISABLED_SESSION_KEY] = True
            return pd.DataFrame(), ""

        try:
            if uploaded_file is not None:
                st.session_state["use_historical_sample_input"] = False
                st.session_state[HISTORICAL_SAMPLE_DISABLED_SESSION_KEY] = False
                historical_df = _load_uploaded_input(
                    uploaded_file,
                    sort_by="date",
                    strict_business_day_no=False,
                )
                st.success(f"과거 월 데이터 {len(historical_df)}행을 불러왔습니다.")
                return historical_df, uploaded_file.name

            if not st.session_state.get(HISTORICAL_SAMPLE_DISABLED_SESSION_KEY, False):
                if st.session_state.get("use_historical_sample_input", False):
                    historical_df = _load_packaged_sample_for_app("historical_input")
                    st.info(f"과거 샘플 데이터 {len(historical_df)}행을 기본 반영 중입니다.")
                    return historical_df, HISTORICAL_SAMPLE_INPUT_SOURCE_LABEL
                return _get_historical_input_state()
        except Exception as exc:  # noqa: BLE001 - surface load errors in the UI.
            st.error(f"과거 월 데이터를 로딩할 수 없습니다: {exc}")
            return pd.DataFrame(), ""

        st.info("과거 데이터를 비운 상태입니다. 기본 샘플 다시 적용을 누르면 비교값이 다시 표시됩니다.")
    return pd.DataFrame(), ""


def _render_operator_sample_management(
    current_df: pd.DataFrame,
    current_source_label: str,
    historical_df: pd.DataFrame,
    historical_source_label: str,
    audit_readonly: bool = False,
) -> tuple[pd.DataFrame, str, pd.DataFrame, str]:
    with st.expander("운영 샘플 관리", expanded=False):
        st.caption(
            "운영 기본값으로 저장하면 앱 리부트 후에도 해당 데이터가 기본 입력값으로 로드됩니다. "
            "단, 여러 사용자가 같은 앱을 쓰는 경우 이 저장본은 공용 기본값으로 적용됩니다."
        )
        current_tab, historical_tab = st.tabs(["현재 입력 샘플", "과거 샘플"])
        with current_tab:
            current_df, current_source_label = _render_operator_sample_panel(
                "current_input",
                current_df,
                current_source_label,
                save_button_label="현재 입력값을 운영 기본값으로 저장",
                download_file_name="current_input_sample.csv",
                audit_readonly=audit_readonly,
            )
        with historical_tab:
            historical_df, historical_source_label = _render_operator_sample_panel(
                "historical_input",
                historical_df,
                historical_source_label,
                save_button_label="과거 샘플을 운영 기본값으로 저장",
                download_file_name="historical_input_sample.csv",
                audit_readonly=audit_readonly,
            )
    return current_df, current_source_label, historical_df, historical_source_label


def _render_operator_sample_panel(
    kind: str,
    df: pd.DataFrame,
    source_label: str,
    *,
    save_button_label: str,
    download_file_name: str,
    audit_readonly: bool,
) -> tuple[pd.DataFrame, str]:
    metadata = dict(read_operator_metadata().get(kind) or {})
    source_display = _sample_source_display_label(source_label)
    saved_at = str(metadata.get("saved_at") or metadata.get("reset_at") or "-")
    saved_rows = metadata.get("rows")
    saved_rows_display = "-" if saved_rows is None else str(saved_rows)
    operator_location = get_operator_sample_location(kind)

    status_cols = st.columns(4)
    status_cols[0].metric("현재 데이터 소스", source_display)
    status_cols[1].metric("마지막 저장 시각", saved_at)
    status_cols[2].metric("화면 row 수", str(len(df)))
    status_cols[3].metric("저장 row 수", saved_rows_display)
    st.caption(f"운영 저장 위치: {_short_display_path(operator_location)}")

    editor_key = f"operator_sample_editor_{kind}"
    edited_df = st.data_editor(
        df,
        column_config=_input_editor_column_config(),
        disabled=audit_readonly,
        hide_index=True,
        key=editor_key,
        num_rows="dynamic",
        use_container_width=True,
    )
    working_df = _as_dataframe(edited_df)

    save_col, reload_col, packaged_col, download_col = st.columns(4)
    if save_col.button(
        save_button_label,
        key=f"save_operator_sample_{kind}",
        disabled=audit_readonly,
    ):
        result = save_operator_sample(kind, working_df)
        if result.get("ok"):
            loaded_df, _source_info = load_sample_with_source(kind)
            source_label = OPERATOR_SAMPLE_SOURCE_LABEL
            working_df = loaded_df
            _store_operator_sample_state(kind, working_df, source_label)
            saved_metadata = dict(result.get("metadata") or {})
            st.success(
                "운영 기본값 저장 완료: "
                f"{_short_display_path(str(result.get('path')))} / "
                f"{saved_metadata.get('saved_at', '-')} / "
                f"{result.get('rows', len(working_df))}행"
            )
            _render_operator_sample_warnings(result.get("warnings") or [])
        else:
            st.error("운영 기본값으로 저장하지 못했습니다. 아래 검증 오류를 확인해 주세요.")
            _render_operator_sample_errors(result.get("errors") or [])
            _render_operator_sample_warnings(result.get("warnings") or [])

    if reload_col.button(
        "저장된 운영 기본값 다시 불러오기",
        key=f"reload_operator_sample_{kind}",
    ):
        loaded_df, source_info = load_sample_with_source(kind)
        if source_info.get("source") in {"operator", "github"}:
            working_df = loaded_df
            source_label = OPERATOR_SAMPLE_SOURCE_LABEL
            _store_operator_sample_state(kind, working_df, source_label)
            st.success(f"저장된 운영 기본값 {len(working_df)}행을 다시 불러왔습니다.")
        else:
            st.warning("저장된 운영 기본값을 불러오지 못해 내장 샘플을 유지합니다.")
            _render_operator_sample_warnings(source_info.get("warnings") or [])

    if packaged_col.button(
        "내장 샘플로 화면 초기화",
        key=f"reset_operator_sample_screen_{kind}",
    ):
        working_df = _load_packaged_sample_for_app(kind)
        source_label = (
            SAMPLE_INPUT_SOURCE_LABEL
            if kind == "current_input"
            else HISTORICAL_SAMPLE_INPUT_SOURCE_LABEL
        )
        _store_operator_sample_state(kind, working_df, source_label)
        st.info("내장 샘플을 화면에 다시 불러왔습니다. 저장 버튼을 누르기 전에는 운영 기본값이 바뀌지 않습니다.")

    download_col.download_button(
        "CSV 다운로드",
        data=_operator_sample_csv_bytes(working_df),
        file_name=download_file_name,
        mime="text/csv",
        key=f"download_operator_sample_{kind}",
    )
    return normalize_direct_input_edits(working_df), source_label


def _store_operator_sample_state(kind: str, df: pd.DataFrame, source_label: str) -> None:
    if kind == "current_input":
        st.session_state["force_sample_input"] = source_label == SAMPLE_INPUT_SOURCE_LABEL
        _store_current_input_state(df, source_label)
        return
    st.session_state["use_historical_sample_input"] = source_label == HISTORICAL_SAMPLE_INPUT_SOURCE_LABEL
    st.session_state[HISTORICAL_SAMPLE_DISABLED_SESSION_KEY] = False
    _store_historical_input_state(df, source_label)


def _load_packaged_sample_for_app(kind: str) -> pd.DataFrame:
    path = get_packaged_sample_path(kind)
    if kind == "historical_input":
        return load_input(path, sort_by="date", strict_business_day_no=False)
    return load_input(path)


def _sample_source_display_label(source_label: str) -> str:
    if source_label == OPERATOR_SAMPLE_SOURCE_LABEL:
        return OPERATOR_SAMPLE_SOURCE_LABEL
    if source_label == SAVED_ACTUALS_SOURCE_LABEL:
        return SAVED_ACTUALS_SOURCE_LABEL
    if source_label in {SAMPLE_INPUT_SOURCE_LABEL, HISTORICAL_SAMPLE_INPUT_SOURCE_LABEL}:
        return PACKAGED_SAMPLE_DISPLAY_LABEL
    if source_label:
        return UPLOAD_SAMPLE_DISPLAY_LABEL
    return "-"


def _warn_operator_sample_fallback(title: str, source_info: Mapping[str, Any]) -> None:
    warnings = list(source_info.get("warnings") or [])
    if source_info.get("source") != "packaged" or not warnings:
        return
    warning_fn = getattr(st, "warning", None)
    if callable(warning_fn):
        warning_fn(f"{title} 운영 저장본을 불러오지 못해 내장 샘플을 사용합니다.")


def _operator_sample_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8-sig")


def _render_operator_sample_errors(messages: object) -> None:
    for message in list(messages or []):
        st.markdown(f"- {escape(str(message))}")


def _render_operator_sample_warnings(messages: object) -> None:
    warning_messages = [str(message) for message in list(messages or []) if str(message)]
    if not warning_messages:
        return
    st.warning("확인 필요: " + " / ".join(warning_messages))


def _short_display_path(path: str | Path) -> str:
    if isinstance(path, str):
        if path.startswith("github://") or len(path) <= 80:
            return path
        return f".../{path[-77:]}"
    try:
        return path.resolve().relative_to(REPO_ROOT.resolve()).as_posix()
    except ValueError:
        text = str(path)
    if len(text) <= 80:
        return text
    return f".../{path.parent.name}/{path.name}"


def _render_input_editor(
    df: pd.DataFrame,
    source_label: str,
    audit_readonly: bool = False,
) -> pd.DataFrame:
    st.header("2. 입력 수정")
    saved_actuals = _load_saved_actuals_for_ui()
    df, default_source = apply_latest_upload_policy(
        df,
        source_label,
        saved_actuals,
        persist_uploaded_defaults=False,
    )
    if default_source == "uploaded":
        st.caption(
            "업로드 입력값을 현재 화면에 적용했습니다. 완료월 실제 실적 저장 버튼을 누르면 "
            "현재 입력표가 다음 시작 기본값으로 저장됩니다."
        )
    elif default_source == "saved":
        st.caption(
            f"저장된 실적 기본값 {len(saved_actuals)}건을 내장 샘플 위에 적용했습니다. "
            "완료월 실제 실적 저장 버튼을 누르면 현재 입력표가 다음 시작 기본값으로 저장됩니다."
        )

    editor_key = "direct_input_editor"
    source_key = "direct_input_editor_source"
    source_token = _input_source_token(df, source_label)
    if st.session_state.get(source_key) != source_token:
        st.session_state[source_key] = source_token
        st.session_state.pop(editor_key, None)

    if audit_readonly:
        st.info("읽기 전용 감리 모드: 저장된 실적값 저장/삭제 버튼이 비활성화됩니다.")

    reset_col, clear_col, save_col = st.columns(3)
    if reset_col.button("입력값 초기화", key="reset_direct_input_editor"):
        st.session_state.pop(editor_key, None)
    if clear_col.button(
        "저장된 실적값 삭제",
        key="clear_saved_actuals",
        disabled=audit_readonly,
    ):
        clear_saved_actuals()
        st.session_state.pop(editor_key, None)
        st.session_state.pop(source_key, None)
        st.rerun()

    edited_df = st.data_editor(
        df,
        column_config=_input_editor_column_config(),
        disabled=_non_editable_input_columns(df),
        hide_index=True,
        key=editor_key,
        num_rows="fixed",
        use_container_width=True,
    )
    normalized = normalize_direct_input_edits(_as_dataframe(edited_df))
    if save_col.button(
        "완료월 실제 실적 저장",
        key="save_saved_actuals_explicit",
        disabled=audit_readonly,
    ):
        result = save_current_input_defaults(normalized)
        saved_path = Path(str(result.get("saved_actuals_path")))
        operator_result = dict(result.get("operator_result") or {})
        operator_path = str(operator_result.get("path") or get_operator_sample_path("current_input"))
        if result.get("ok"):
            st.session_state[CURRENT_INPUT_SOURCE_OVERRIDE_SESSION_KEY] = OPERATOR_SAMPLE_SOURCE_LABEL
            _store_current_input_state(normalized, OPERATOR_SAMPLE_SOURCE_LABEL)
            st.success(
                "현재 입력값을 다음 시작 기본값으로 저장했습니다: "
                f"{_short_display_path(operator_path)}. "
                f"실적 보조 저장 파일도 갱신했습니다: {_short_display_path(saved_path)}."
            )
            _render_operator_sample_warnings(operator_result.get("warnings") or [])
        else:
            st.warning(
                "실적 보조 저장 파일은 갱신했지만, 리부트 기본 입력값 저장에는 실패했습니다. "
                f"저장 파일: {_short_display_path(saved_path)}"
            )
            _render_operator_sample_errors(operator_result.get("errors") or [])
            _render_operator_sample_warnings(operator_result.get("warnings") or [])
    return normalized


def apply_latest_upload_policy(
    df: pd.DataFrame,
    source_label: str,
    saved_actuals: pd.DataFrame,
    path: str | Path = SAVED_ACTUALS_PATH,
    persist_uploaded_defaults: bool = True,
) -> tuple[pd.DataFrame, str]:
    """Apply the latest-value policy for uploaded current-month inputs."""
    if _is_current_upload_source(source_label):
        normalized = normalize_direct_input_edits(df)
        if persist_uploaded_defaults:
            save_actual_values(normalized, path)
        return normalized, "uploaded"

    if not saved_actuals.empty:
        return apply_saved_actuals(df, saved_actuals), "saved"

    return df.copy(), "none"


def _is_current_upload_source(source_label: str) -> bool:
    non_upload_sources = {
        SAMPLE_INPUT_SOURCE_LABEL,
        HISTORICAL_SAMPLE_INPUT_SOURCE_LABEL,
        OPERATOR_SAMPLE_SOURCE_LABEL,
        SAVED_ACTUALS_SOURCE_LABEL,
        "",
    }
    return bool(source_label) and source_label not in non_upload_sources


def _load_saved_actuals_for_ui() -> pd.DataFrame:
    try:
        return load_saved_actuals()
    except Exception as exc:  # noqa: BLE001 - keep the app usable if the store is corrupt.
        st.warning(f"저장된 실적값을 불러올 수 없습니다: {exc}")
        return pd.DataFrame(columns=SAVED_ACTUAL_COLUMNS)


def _input_source_token(df: pd.DataFrame, source_label: str) -> str:
    parts = [source_label, str(len(df))]
    for column in ("date", "business_day_no", "is_close_day", "close_type"):
        if column in df.columns:
            parts.append(f"{column}:{','.join(df[column].astype(str).tolist())}")
    return "|".join(parts)


def _input_editor_column_config() -> dict[str, object]:
    if st is None:
        return {}

    return {
        "date": st.column_config.DateColumn("날짜", format="YYYY-MM-DD"),
        "day_name": st.column_config.TextColumn("요일(표시용)"),
        "business_day_no": st.column_config.NumberColumn("영업일 번호", format="%d"),
        "is_close_day": st.column_config.CheckboxColumn("마감일 여부"),
        "close_type": st.column_config.TextColumn("마감 유형"),
        "sales_target_daily": st.column_config.NumberColumn(
            "판매실적 일 목표",
            min_value=0.0,
            step=0.1,
            format="%.1f",
        ),
        "recognized_target_daily": st.column_config.NumberColumn(
            "인정실적 일 목표",
            min_value=0.0,
            step=0.1,
            format="%.1f",
        ),
        "sales_actual_cum": st.column_config.NumberColumn(
            "판매실적 누적 실적",
            min_value=0.0,
            step=0.1,
            format="%.1f",
        ),
        "recognized_actual_cum": st.column_config.NumberColumn(
            "인정실적 누적 실적",
            min_value=0.0,
            step=0.1,
            format="%.1f",
        ),
        "memo": st.column_config.TextColumn("메모"),
    }


def _non_editable_input_columns(df: pd.DataFrame) -> list[str]:
    return [column for column in df.columns if column not in DIRECT_EDITABLE_COLUMNS]


def _render_settings(
    df: pd.DataFrame,
    base_config: dict[str, Any],
) -> tuple[str, pd.Timestamp, str, str, dict[str, Any]]:
    st.header("3. 기준 설정")
    metric_default, as_of_default, forecast_default, provision_default, _ = _normalize_app_settings(
        df,
        base_config,
    )
    metric = st.selectbox(
        "지표 선택",
        ["sales", "recognized"],
        index=["sales", "recognized"].index(metric_default),
        format_func=lambda value: METRIC_DISPLAY_LABELS.get(value, value),
        key=PACE_METRIC_SESSION_KEY,
    )
    dates = pd.to_datetime(df["date"], errors="raise")
    date_values = [timestamp.date() for timestamp in dates]
    default_date = pd.Timestamp(as_of_default).date()
    default_index = date_values.index(default_date) if default_date in date_values else 0
    as_of_date = st.selectbox(
        "기준일 선택",
        date_values,
        index=default_index,
        key=PACE_AS_OF_DATE_SESSION_KEY,
    )
    st.caption("기본 기준일은 입력표 기준 직전 영업일로 매일 갱신됩니다. 필요하면 다른 입력일을 선택할 수 있습니다.")
    forecast_choice = st.selectbox(
        "예측모델 선택",
        ["F1", "F2", "F3", COMPARE_LABEL],
        index=["F1", "F2", "F3", COMPARE_LABEL].index(forecast_default),
        key=PACE_FORECAST_CHOICE_SESSION_KEY,
    )
    provision_choice = st.selectbox(
        "운영 전략 선택",
        ["P1", "P2", "P3", "O1", "O2", "O3", "N1", "N2", "N3", COMPARE_LABEL],
        index=["P1", "P2", "P3", "O1", "O2", "O3", "N1", "N2", "N3", COMPARE_LABEL].index(provision_default),
        key=PACE_STRATEGY_CHOICE_SESSION_KEY,
    )

    col1, col2 = st.columns(2)
    close_day_cap_rate = col1.number_input(
        "마감일 목표 상한 배율",
        min_value=0.0,
        value=_session_float(PACE_CLOSE_CAP_SESSION_KEY, base_config.get("close_day_cap_rate", 1.30)),
        step=0.05,
        format="%.2f",
        key=PACE_CLOSE_CAP_SESSION_KEY,
    )
    non_close_day_cap_rate = col2.number_input(
        "비마감일 목표 상한 배율",
        min_value=0.0,
        value=_session_float(PACE_NON_CLOSE_CAP_SESSION_KEY, base_config.get("non_close_day_cap_rate", 1.50)),
        step=0.05,
        format="%.2f",
        key=PACE_NON_CLOSE_CAP_SESSION_KEY,
    )
    config = build_runtime_config(
        base_config,
        close_day_cap_rate,
        non_close_day_cap_rate,
    )
    return metric, pd.Timestamp(as_of_date), forecast_choice, provision_choice, config


def _render_validation(validation_result: dict[str, Any]) -> None:
    errors = format_validation_messages(list(validation_result.get("errors", [])))
    warnings = format_validation_messages(list(validation_result.get("warnings", [])))
    if errors:
        st.error("수정이 필요한 항목")
        for message in errors:
            st.write(f"- {message}")
    else:
        st.success("계산을 막는 입력 문제 없음")

    if warnings:
        st.warning("확인하면 좋은 항목")
        for message in warnings:
            st.write(f"- {message}")
    else:
        st.info("추가로 확인할 주의 사항 없음")


def _render_selected_scenario_picker(
    scenario_df: pd.DataFrame,
    forecast_choice: str,
    provision_choice: str,
) -> str:
    candidates = _filter_scenarios(scenario_df, forecast_choice, provision_choice)
    scenario_ids = candidates["scenario_id"].astype(str).tolist()
    if len(scenario_ids) == 1:
        st.caption(f"선택 시나리오: {format_scenario_option_label(scenario_ids[0])}")
        return scenario_ids[0]
    return st.selectbox(
        "선택 시나리오 상세",
        scenario_ids,
        format_func=format_scenario_option_label,
        index=0,
    )


def _render_pace_check_header(df: pd.DataFrame, as_of_date: object) -> None:
    context = _pace_header_context(df, as_of_date)
    st.markdown(
        render_pace_header(
            as_of_date=as_of_date,
            current_business_day_no=context["current_business_day_no"],
            total_business_days=context["total_business_days"],
            close_day_label=context["close_day_label"],
        ),
        unsafe_allow_html=True,
    )


def _pace_header_context(df: pd.DataFrame, as_of_date: object) -> dict[str, object]:
    if df.empty:
        return {
            "current_business_day_no": "계산 불가",
            "total_business_days": "계산 불가",
            "close_day_label": "입력 없음",
        }

    total_business_days = len(df)
    if "business_day_no" in df.columns:
        numeric_business_days = pd.to_numeric(df["business_day_no"], errors="coerce")
        if numeric_business_days.notna().any():
            total_business_days = int(numeric_business_days.max())

    as_of_timestamp = pd.Timestamp(as_of_date).normalize()
    dates = pd.to_datetime(df["date"], errors="coerce").dt.normalize()
    rows = df.loc[dates == as_of_timestamp]
    if rows.empty:
        return {
            "current_business_day_no": "입력 행 없음",
            "total_business_days": total_business_days,
            "close_day_label": "입력 행 없음",
        }

    selected = rows.iloc[0]
    current_business_day_no = selected.get("business_day_no", "계산 불가")
    is_close_day = bool(_coerce_is_close_day(pd.Series([selected.get("is_close_day")])).iloc[0])
    close_day_label = "마감일" if is_close_day else "일반 영업일"
    return {
        "current_business_day_no": current_business_day_no,
        "total_business_days": total_business_days,
        "close_day_label": close_day_label,
    }


def _render_kpis(
    validation_result: dict[str, Any],
    scenario_df: pd.DataFrame,
    next_close_result: dict[str, Any],
    selected_row: pd.Series,
) -> None:
    st.markdown(
        render_operation_mode_card(
            selected_row.get("target_status"),
            target_variance=selected_row.get("target_variance"),
            surplus_to_target=selected_row.get("surplus_to_target"),
        ),
        unsafe_allow_html=True,
    )
    _render_kpi_grid(validation_result, next_close_result, selected_row)

    with st.expander("KPI 상세", expanded=False):
        for row in build_kpi_rows(
            validation_result,
            scenario_df,
            next_close_result,
            selected_row,
        ):
            cols = st.columns(len(row))
            for col, (label, value) in zip(cols, row):
                help_text = KPI_HELP_TEXTS.get(label)
                if help_text:
                    col.metric(label, value, help=help_text)
                else:
                    col.metric(label, value)


def _render_kpi_grid(
    validation_result: dict[str, Any],
    next_close_result: dict[str, Any],
    selected_row: pd.Series,
) -> None:
    target_status = selected_row.get("target_status")
    kpi_cards = (
        render_kpi_card(
            "현재 누적 실적",
            format_krw(validation_result.get("current_actual_cum")),
            sub="입력 기준 누적 실적",
            target_status=target_status,
        ),
        render_kpi_card(
            "월마감 예상 실적",
            format_krw(selected_row.get("forecast_after_provision")),
            sub=str(selected_row.get("scenario_id", "")),
            focus=True,
            target_status=target_status,
        ),
        render_kpi_card(
            "월 목표",
            format_krw(validation_result.get("monthly_target")),
            sub="입력 일별 목표 합계",
            target_status=target_status,
        ),
        render_kpi_card(
            "목표 대비 차이",
            format_krw(selected_row.get("target_variance")),
            sub="양수는 초과, 음수는 미달",
            focus=True,
            target_status=target_status,
        ),
        render_kpi_card(
            "목표 상태",
            status_label(target_status),
            sub="목표 보정/유지/초과달성 운영 구분",
            target_status=target_status,
        ),
        render_kpi_card(
            "초과 예상분",
            format_krw(selected_row.get("surplus_to_target")),
            sub="초과달성 운영 버퍼",
            target_status=target_status,
        ),
        render_kpi_card(
            "다음 마감 누적선 필요실적",
            format_krw(
                next_close_result.get("required_to_recover_next_close_cum")
            ),
            sub="월 부족분이 아니라 다음 마감일까지의 누적 계획선 기준입니다.",
            focus=True,
            target_status=target_status,
        ),
    )
    st.markdown(
        f'<div class="kpi-grid">{"".join(kpi_cards)}</div>',
        unsafe_allow_html=True,
    )


def _render_scenario_check(scenario_df: pd.DataFrame) -> None:
    if scenario_df.empty:
        st.info("시나리오 카드로 표시할 데이터가 없습니다.")
        return

    cards = "".join(render_scenario_card(row) for row in scenario_df.to_dict("records"))
    st.markdown(
        render_section_header(
            "시나리오 체크",
            "P1/P2/P3는 목표 보정, O1/O2/O3는 초과달성 운영 전략으로 구분합니다.",
        ),
        unsafe_allow_html=True,
    )
    st.markdown(
        f'<div class="scenario-grid">{cards}</div>',
        unsafe_allow_html=True,
    )


def _render_report_memo_card(report_text: str) -> None:
    st.markdown(
        render_section_header(
            "보고 메모",
            "복사해서 공유하기 쉬운 운영 요약입니다.",
        ),
        unsafe_allow_html=True,
    )
    st.markdown(render_report_card(report_text), unsafe_allow_html=True)


def _render_history_insight_header() -> None:
    st.markdown(
        render_section_header(
            "예측 이력",
            "예측 이력은 과거 완료월의 예측값과 실제 마감 실적을 비교해, 현재 월 예측의 신뢰도와 모델별 편향을 점검하는 화면입니다.",
        ),
        unsafe_allow_html=True,
    )
    st.markdown(
        '<section class="history-purpose-card">'
        '<h3>이 탭에서 확인할 질문</h3>'
        '<div class="history-question-grid">'
        "<span>과거 같은 영업일차에서 현재 페이스는 빠른가, 느린가?</span>"
        "<span>F1/F2/F3 중 어느 모델이 최근 완료월에서 더 안정적이었는가?</span>"
        "<span>현재 보고에서 보수/기준/상향 중 어느 관점으로 설명해야 하는가?</span>"
        "</div>"
        "</section>",
        unsafe_allow_html=True,
    )


def build_kpi_rows(
    validation_result: dict[str, Any],
    scenario_df: pd.DataFrame,
    next_close_result: dict[str, Any],
    selected_row: pd.Series,
) -> tuple[tuple[tuple[str, object], ...], ...]:
    achievement_rate = safe_divide(
        validation_result.get("current_actual_cum"),
        validation_result.get("current_target_cum"),
    )
    forecast_summary = _forecast_summary(scenario_df)
    next_close_date = next_close_result.get("next_close_date")
    next_close_required = next_close_result.get("required_to_recover_next_close_cum")
    target_status = selected_row.get("target_status", "계산 불가")

    return (
        (
            ("월 목표", format_amount(validation_result.get("monthly_target"))),
            ("기준일 누적 목표", format_amount(validation_result.get("current_target_cum"))),
            ("기준일 누적 실적", format_amount(validation_result.get("current_actual_cum"))),
            ("누적 달성률", format_rate(achievement_rate)),
        ),
        (
            ("F1예상", format_amount(forecast_summary.get("F1"))),
            ("F2예상", format_amount(forecast_summary.get("F2"))),
            ("F3예상", format_amount(forecast_summary.get("F3"))),
        ),
        (
            ("다음 마감일", _format_date(next_close_date)),
            (NEXT_CLOSE_REQUIRED_LABEL, format_amount(next_close_required)),
        ),
        (
            ("목표상태", _localize_display_value(target_status)),
            ("목표대비 차이", format_amount(selected_row.get("target_variance"))),
            ("초과 예상분", format_amount(selected_row.get("surplus_to_target"))),
            ("위험등급", _localize_display_value(selected_row.get("risk_level", "계산 불가"))),
            ("운영모드", _operation_mode_label(target_status)),
        ),
    )


def _render_body(
    scenario_df: pd.DataFrame,
    selected_scenario_id: str,
    selected_row: pd.Series,
    forecast_result: dict[str, object],
    provision_result: dict[str, object],
    revised_targets_df: Any,
    close_cycle_df: pd.DataFrame,
    next_close_result: dict[str, Any],
    validation_result: dict[str, Any],
    historical_context: dict[str, object],
    metric: str,
    as_of_date: object,
    df: pd.DataFrame,
    config: dict[str, Any],
) -> None:
    _render_scenario_check(scenario_df)

    _render_visuals(
        scenario_df,
        selected_scenario_id,
        selected_row,
        _as_dataframe(revised_targets_df),
        close_cycle_df,
        next_close_result,
        validation_result,
        metric,
        as_of_date,
        df,
        config,
    )
    _render_historical_context_panel(
        historical_context,
        scenario_df,
        selected_scenario_id,
    )

    st.subheader("시나리오 매트릭스")
    st.dataframe(build_scenario_matrix(scenario_df), use_container_width=True)

    _render_selected_scenario_summary(selected_scenario_id, selected_row)

    with st.expander("선택 시나리오 상세", expanded=False):
        detail_df = selected_row.to_frame(name="value").reset_index()
        detail_df.columns = ["item", "value"]
        st.dataframe(_format_display_df(detail_df), use_container_width=True)

    warnings = [
        *list(forecast_result.get("warnings", [])),
        *list(provision_result.get("warnings", [])),
    ]
    if warnings:
        st.warning("선택 시나리오 확인 사항")
        for message in dict.fromkeys(str(warning) for warning in warnings):
            st.write(f"- {format_validation_message(message)}")

    _render_target_or_strategy_table(
        scenario_df,
        selected_scenario_id,
        revised_targets_df,
    )

    st.subheader("보고 메모")
    report_text = build_daily_report_text(
        scenario_df,
        next_close_result,
        selected_scenario_id=selected_scenario_id,
    )
    _render_report_memo_card(report_text)
    _render_report_glossary_panel()
    report_key = hashlib.sha1(report_text.encode("utf-8")).hexdigest()[:12]
    st.text_area("보고 메모", value=report_text, height=320, key=f"auto_report_{report_key}")

    st.subheader("입력값 점검 결과")
    _render_validation(validation_result)


def _render_report_glossary_panel() -> None:
    glossary_df = build_report_glossary_df()
    with st.expander("보고 메모 고정 용어 정의", expanded=False):
        st.caption("보고문 본문과 분리해 별도로 운영하는 고정 기준값입니다.")
        tabs = st.tabs([group for group, _ in REPORT_GLOSSARY_GROUPS])
        for tab, (group, _) in zip(tabs, REPORT_GLOSSARY_GROUPS):
            with tab:
                group_df = glossary_df.loc[glossary_df["구분"] == group, ["코드", "정의"]]
                st.dataframe(group_df, hide_index=True, use_container_width=True)


def _render_selected_scenario_summary(
    selected_scenario_id: str,
    selected_row: pd.Series,
) -> None:
    st.subheader("선택 시나리오 요약")
    st.dataframe(
        build_selected_scenario_explanation(selected_scenario_id, selected_row),
        hide_index=True,
        use_container_width=True,
    )


def _render_target_or_strategy_table(
    scenario_df: pd.DataFrame,
    selected_scenario_id: str,
    revised_targets_df: Any,
    *,
    show_all_forecast_models: bool = False,
) -> None:
    revised_targets = _as_dataframe(revised_targets_df)
    if not revised_targets.empty:
        st.subheader("잔여 일자별 수정 목표표")
        st.caption("목표 미달 보정전략(P1~P3)에서 입력표에 있는 잔여 일자별 목표를 어떻게 수정하는지 표시합니다.")
        st.dataframe(
            _format_display_df(revised_targets),
            use_container_width=True,
        )
        return

    strategy_table = build_strategy_level_table(
        scenario_df,
        None if show_all_forecast_models else selected_scenario_id,
    )
    if show_all_forecast_models:
        strategy_effect_table = build_strategy_effect_table(scenario_df, selected_scenario_id)
        if not strategy_effect_table.empty:
            forecast_key = _selected_forecast_key(selected_scenario_id) or str(
                strategy_effect_table["forecast_basis"].iloc[0]
            )
            st.subheader(f"{forecast_key} 고정 O전략 차이 요약")
            st.caption(
                f"{forecast_key} 예측값은 그대로 두고, O1/O2/O3가 운영 목표·버퍼·품질관리 여유를 어떻게 나누는지 비교합니다. "
                "O전략은 월말 예상 실적을 새로 예측하지 않습니다."
            )
            st.dataframe(
                _format_display_df(strategy_effect_table),
                hide_index=True,
                use_container_width=True,
            )

        st.subheader("F예측 × O전략 전체 매트릭스")
        st.caption(
            "세로로 반복되는 차이는 F1/F2/F3 예측 차이이고, 같은 F 안에서 O1/O2/O3를 비교할 때가 O전략 차이입니다. "
            "각 O전략은 월말 예상 실적을 바꾸지 않고 Stretch 전환분, 운영전략 월 목표, 안전버퍼, 품질관리 여유분을 바꿉니다."
        )
    else:
        forecast_key = _selected_forecast_key(selected_scenario_id) or "선택 모델"
        st.subheader(f"{forecast_key} 기준 전략 목표·버퍼·리스크 수준표")
        st.caption(
            "같은 예측모델의 월말 예상 실적은 전략별로 바꾸지 않습니다. "
            "이 표는 Stretch 전환분, 운영전략 월 목표, 안전버퍼, 품질관리 여유분처럼 실제로 달라지는 운영 기준을 먼저 비교합니다."
        )
    if strategy_table.empty:
        st.info("운영전략별 목표 수준 데이터 없음")
        return
    st.dataframe(
        _format_display_df(strategy_table),
        hide_index=True,
        use_container_width=True,
    )


def _render_forecast_strategy_chart_tabs(
    df: pd.DataFrame,
    scenario_df: pd.DataFrame,
    selected_scenario_id: str,
    selected_row: pd.Series,
    revised_targets_df: pd.DataFrame,
    close_cycle_df: pd.DataFrame,
    next_close_result: dict[str, Any],
    validation_result: dict[str, Any],
    metric: str,
    as_of_date: object,
    config: dict[str, Any],
) -> None:
    _ = selected_row, next_close_result, validation_result
    st.subheader("시각화 상세")
    scenario_tab, target_tab, close_cycle_tab = st.tabs(
        ["시나리오별 예상", "잔여 목표/전략 수준", "마감차수 흐름"]
    )

    with scenario_tab:
        scenario_chart_data = build_scenario_chart_data(scenario_df)
        if scenario_chart_data.empty:
            st.info("시나리오 차트 데이터 없음")
        else:
            _render_visual_metric_definitions(
                ("daily_forecast_cum", "daily_target_cum", "daily_achievement_rate")
            )
            _render_chart_reading_guide("scenario_daily_progress")
            _render_forecast_model_scenario_tabs(
                df,
                scenario_df,
                as_of_date,
                metric,
                config,
                selected_scenario_id,
            )

            with st.expander("시나리오 숫자표", expanded=False):
                value_matrix = build_scenario_value_matrix(scenario_df)
                st.dataframe(value_matrix.map(format_amount), use_container_width=True)

    with target_tab:
        st.caption(f"선택 시나리오: {selected_scenario_id}")
        target_chart_data = build_remaining_target_chart_data(revised_targets_df)
        if target_chart_data.empty:
            _render_strategy_level_visuals(scenario_df, selected_scenario_id)
        else:
            _render_visual_metric_definitions(
                ("original_target", "uplift", "revised_target", "cap_target")
            )
            _render_chart_reading_guide("target_stack")
            _render_remaining_target_stack_chart(revised_targets_df)

    with close_cycle_tab:
        close_cycle_bar_columns = ("target_sum", "actual_sum")
        close_cycle_cumulative_amount_columns = ("target_cum", "actual_cum")
        close_cycle_rate_columns = ("achievement_rate",)
        close_cycle_cumulative_rate_columns = ("cumulative_achievement_rate",)
        _render_visual_metric_definitions(
            (
                *close_cycle_bar_columns,
                *close_cycle_cumulative_amount_columns,
                *close_cycle_rate_columns,
                *close_cycle_cumulative_rate_columns,
            )
        )
        close_cycle_chart_data = build_close_cycle_chart_data(close_cycle_df)
        if close_cycle_chart_data.empty:
            st.info("마감 사이클 차트 데이터 없음")
        else:
            _render_chart_reading_guide("close_cycle_amount")
            _render_grouped_bar_chart(close_cycle_chart_data, close_cycle_bar_columns)
            st.markdown("**CloseCycle 누적 목표선/누적 실적**")
            st.caption("입력표에 있는 마감차수 row만 사용해 누적 목표선과 누적 실적을 함께 표시합니다.")
            _render_line_chart(close_cycle_chart_data, close_cycle_cumulative_amount_columns)
            _render_chart_reading_guide("close_cycle_rate")
            _render_line_chart(close_cycle_chart_data, close_cycle_rate_columns)
            st.markdown("**누적 달성률**")
            _render_rate_ratio_line_chart(close_cycle_chart_data, close_cycle_cumulative_rate_columns)


def _render_visuals(
    scenario_df: pd.DataFrame,
    selected_scenario_id: str,
    selected_row: pd.Series,
    revised_targets_df: pd.DataFrame,
    close_cycle_df: pd.DataFrame,
    next_close_result: dict[str, Any],
    validation_result: dict[str, Any],
    metric: str,
    as_of_date: object,
    df: pd.DataFrame,
    config: dict[str, Any],
    audit_readonly: bool = False,
) -> None:
    st.subheader("시각화")
    st.caption("각 탭은 결론 확인 → 기준선 확인 → 차이 확인 → 실행 판단 순서로 읽습니다.")
    _render_visual_decision_panel(
        selected_scenario_id,
        selected_row,
        validation_result,
        next_close_result,
    )
    scenario_tab, target_tab, close_cycle_tab, history_tab = st.tabs(
        ["시나리오별 예상", "잔여 목표/전략 수준", "마감차수 흐름", HISTORY_TAB_LABEL]
    )

    with scenario_tab:
        scenario_chart_data = build_scenario_chart_data(scenario_df)
        if scenario_chart_data.empty:
            st.info("시나리오 차트 데이터 없음")
        else:
            _render_visual_metric_definitions(
                ("daily_forecast_cum", "daily_target_cum", "daily_achievement_rate")
            )
            _render_chart_reading_guide("scenario_daily_progress")
            _render_forecast_model_scenario_tabs(
                df,
                scenario_df,
                as_of_date,
                metric,
                config,
                selected_scenario_id,
            )

            with st.expander("시나리오 숫자표", expanded=False):
                value_matrix = build_scenario_value_matrix(scenario_df)
                st.dataframe(value_matrix.map(format_amount), use_container_width=True)

    with target_tab:
        st.caption(f"선택 시나리오: {selected_scenario_id}")
        target_chart_data = build_remaining_target_chart_data(revised_targets_df)
        if target_chart_data.empty:
            _render_strategy_level_visuals(scenario_df, selected_scenario_id)
        else:
            _render_visual_metric_definitions(
                ("original_target", "uplift", "revised_target", "cap_target")
            )
            _render_chart_reading_guide("target_stack")
            _render_remaining_target_stack_chart(revised_targets_df)

    with close_cycle_tab:
        close_cycle_bar_columns = ("target_sum", "actual_sum")
        close_cycle_cumulative_amount_columns = ("target_cum", "actual_cum")
        close_cycle_rate_columns = ("achievement_rate",)
        close_cycle_cumulative_rate_columns = ("cumulative_achievement_rate",)
        _render_visual_metric_definitions(
            (
                *close_cycle_bar_columns,
                *close_cycle_cumulative_amount_columns,
                *close_cycle_rate_columns,
                *close_cycle_cumulative_rate_columns,
            )
        )
        close_cycle_chart_data = build_close_cycle_chart_data(close_cycle_df)
        if close_cycle_chart_data.empty:
            st.info("마감 사이클 차트 데이터 없음")
        else:
            _render_chart_reading_guide("close_cycle_amount")
            _render_grouped_bar_chart(close_cycle_chart_data, close_cycle_bar_columns)
            st.markdown("**CloseCycle 누적 목표선/누적 실적**")
            st.caption("입력표에 있는 마감차수 row만 사용해 누적 목표선과 누적 실적을 함께 표시합니다.")
            _render_line_chart(close_cycle_chart_data, close_cycle_cumulative_amount_columns)
            _render_chart_reading_guide("close_cycle_rate")
            _render_line_chart(close_cycle_chart_data, close_cycle_rate_columns)
            st.markdown("**누적 달성률**")
            _render_rate_ratio_line_chart(close_cycle_chart_data, close_cycle_cumulative_rate_columns)

    with history_tab:
        _render_forecast_history_backtest_tab(
            scenario_df,
            metric,
            as_of_date,
            audit_readonly=audit_readonly,
        )


def _render_visual_decision_panel(
    selected_scenario_id: str,
    selected_row: pd.Series,
    validation_result: dict[str, Any],
    next_close_result: dict[str, Any],
) -> None:
    st.markdown("**결론 먼저 보기**")
    st.info(build_visual_headline(selected_row, validation_result, next_close_result))

    cols = st.columns(4)
    cols[0].metric("선택 시나리오", selected_scenario_id)
    cols[1].metric(
        "전략 반영 후 예상",
        format_amount(selected_row.get("forecast_after_provision")),
        delta=_format_signed_amount(selected_row.get("target_variance")),
        help="delta는 공식 월 목표 대비 차이입니다.",
    )
    cols[2].metric("목표 상태", _localize_display_value(selected_row.get("target_status")))
    cols[3].metric("위험등급", _localize_display_value(selected_row.get("risk_level", "N/A")))

    with st.expander("차트 해석 순서", expanded=True):
        st.dataframe(
            build_visual_decision_summary(
                selected_row,
                validation_result,
                next_close_result,
            ),
            hide_index=True,
            use_container_width=True,
        )


def _render_forecast_history_backtest_tab(
    scenario_df: pd.DataFrame,
    metric: str,
    as_of_date: object,
    audit_readonly: bool = False,
) -> None:
    st.subheader(HISTORY_TAB_LABEL)
    _render_history_insight_header()
    st.caption("현재 계산 결과를 저장하고, 월마감 확정 실적과 매칭되는 이력은 Backtest로 비교합니다.")
    if audit_readonly:
        st.info("읽기 전용 감리 모드: 예측 이력 저장 버튼이 비활성화됩니다.")

    save_col, path_col = st.columns([1, 3])
    if save_col.button(
        "예측 이력 저장",
        key=f"save_forecast_history_{metric}_{pd.Timestamp(as_of_date).date()}",
        disabled=audit_readonly,
    ):
        try:
            saved_history = save_forecast_history_snapshot(scenario_df, as_of_date, metric)
            st.success(f"예측 이력을 저장했습니다. forecast_history {len(saved_history)}건")
        except Exception as exc:  # noqa: BLE001 - Streamlit should stay usable.
            st.warning(f"예측 이력을 저장하지 못했습니다: {exc}")

    path_col.caption(
        f"forecast_history: {_history_storage_path(history_schema.FORECAST_HISTORY)} | "
        f"final_actuals: {_history_storage_path(history_schema.FINAL_ACTUALS)}"
    )

    tables = _load_history_tables_for_ui()
    forecast_history = tables["forecast_history"]
    final_actuals = tables["final_actuals"]
    focused_history = _focus_history_for_current_context(
        forecast_history,
        metric,
        as_of_date,
    )

    if forecast_history.empty and final_actuals.empty:
        st.info("완료월 데이터가 쌓이면 예측 이력과 모델 신뢰도 비교가 표시됩니다.")
        st.caption("월마감 후 실제 실적을 저장하면 다음 달부터 비교 기준으로 사용할 수 있습니다.")

    st.markdown("**완료월 비교**")
    st.caption("저장된 예측값과 월마감 후 확정 실적이 같은 대상 월/지표로 연결되는지 확인합니다.")
    if forecast_history.empty:
        st.info("forecast_history 파일이 없거나 저장된 예측 이력이 없습니다. 예측 이력 저장 후 이 표에 누적됩니다.")
    else:
        st.markdown("forecast_history 테이블")
        st.dataframe(
            _format_display_df(forecast_history.tail(200)),
            hide_index=True,
            use_container_width=True,
        )

    if final_actuals.empty:
        st.info("final_actuals가 아직 없습니다. 월마감 확정 실적이 저장되면 Backtest 오차율을 계산합니다.")
    else:
        st.markdown("final_actuals 테이블")
        st.dataframe(
            _format_display_df(final_actuals.tail(200)),
            hide_index=True,
            use_container_width=True,
        )

    st.markdown("**같은 영업일차 Benchmark**")
    if focused_history.empty:
        st.info("현재 기준월/지표와 같은 조건의 예측 이력이 아직 없어 같은 영업일차 Benchmark를 표시할 수 없습니다.")
    else:
        benchmark_columns = _available_columns(
            focused_history,
            ("target_month", "as_of_date", "metric", "forecast_model", "forecast_amount", "target_status"),
        )
        st.dataframe(
            _format_display_df(focused_history.loc[:, benchmark_columns].tail(80)),
            hide_index=True,
            use_container_width=True,
        )

    backtest_df = build_backtest_dataset(forecast_history, final_actuals)
    model_summary = summarize_by_forecast_model(backtest_df)

    st.markdown("**Backtest Summary**")
    if model_summary.empty:
        st.info("예측 이력과 확정 실적의 대상 월/지표가 아직 매칭되지 않아 모델별 평균 오차율과 bias를 표시할 수 없습니다.")
    else:
        st.dataframe(
            _format_display_df(model_summary),
            hide_index=True,
            use_container_width=True,
        )

    _render_forecast_history_trend_chart(focused_history)
    _render_target_status_distribution(focused_history)
    _render_gap_surplus_trend_chart(focused_history)
    _render_optional_weighted_forecast(focused_history)
    _render_optional_confidence_band(focused_history)

    st.markdown("**Insights**")
    for message in build_backtest_insights(forecast_history, final_actuals, model_summary):
        st.write(f"- {message}")
    st.markdown(
        '<div class="history-next-actions">'
        "<span>완료월 실제 실적 저장</span>"
        "<span>현재 예측과 과거 같은 영업일차 비교</span>"
        "<span>보고 메모에 신뢰도 코멘트 반영</span>"
        "</div>",
        unsafe_allow_html=True,
    )


def _load_history_tables_for_ui() -> dict[str, pd.DataFrame]:
    try:
        return load_history_tables_for_app()
    except Exception as exc:  # noqa: BLE001 - keep the app usable if storage is corrupt.
        st.warning(f"예측 이력/Backtest 저장소를 불러오지 못했습니다: {exc}")
        return {
            "forecast_history": pd.DataFrame(columns=history_schema.FORECAST_HISTORY_COLUMNS),
            "final_actuals": pd.DataFrame(columns=history_schema.FINAL_ACTUALS_COLUMNS),
        }


def _focus_history_for_current_context(
    forecast_history: pd.DataFrame,
    metric: str,
    as_of_date: object,
) -> pd.DataFrame:
    if forecast_history.empty:
        return forecast_history

    focused = forecast_history.copy()
    if "metric" in focused.columns:
        metric_rows = focused.loc[focused["metric"].astype(str) == str(metric)]
        if not metric_rows.empty:
            focused = metric_rows

    if "target_month" in focused.columns:
        target_month = pd.Timestamp(as_of_date).strftime("%Y-%m")
        month_rows = focused.loc[focused["target_month"].astype(str) == target_month]
        if not month_rows.empty:
            focused = month_rows
    return focused.reset_index(drop=True)


def _render_forecast_history_trend_chart(forecast_history: pd.DataFrame) -> None:
    st.markdown("**예측 추이 그래프**")
    source = _history_long_metric_source(forecast_history, ("forecast_amount",))
    if source.empty:
        st.info("예측 추이 그래프를 만들 forecast_amount 이력이 없습니다.")
        return

    chart = (
        alt.Chart(source)
        .mark_line(point=True, strokeWidth=2.4)
        .encode(
            x=alt.X("as_of_date:T", title="기준일"),
            y=alt.Y(
                "value:Q",
                title="예측값",
                scale=_auto_value_scale(source),
                axis=alt.Axis(format=chart_value_format("억원")),
            ),
            color=alt.Color("forecast_model:N", title="예측 모델"),
            tooltip=[
                alt.Tooltip("as_of_date:T", title="기준일"),
                alt.Tooltip("forecast_model:N", title="예측 모델"),
                alt.Tooltip("value:Q", title="예측값", format=chart_value_format("억원")),
            ],
        )
        .properties(height=300)
    )
    st.altair_chart(chart, use_container_width=True)


def _render_target_status_distribution(forecast_history: pd.DataFrame) -> None:
    st.markdown("**target_status 분포**")
    if forecast_history.empty or "target_status" not in forecast_history.columns:
        st.info("target_status 분포를 만들 예측 이력이 없습니다.")
        return

    source = (
        forecast_history["target_status"]
        .fillna("UNKNOWN")
        .astype(str)
        .value_counts()
        .rename_axis("target_status")
        .reset_index(name="count")
    )
    source["target_status_label"] = source["target_status"].map(_localize_display_value)
    chart = (
        alt.Chart(source)
        .mark_bar(cornerRadiusTopLeft=2, cornerRadiusTopRight=2)
        .encode(
            x=alt.X("target_status_label:N", title="목표 상태", sort=None),
            y=alt.Y("count:Q", title="건수", axis=alt.Axis(format="d")),
            color=alt.Color("target_status_label:N", title="목표 상태"),
            tooltip=[
                alt.Tooltip("target_status_label:N", title="목표 상태"),
                alt.Tooltip("count:Q", title="건수", format="d"),
            ],
        )
        .properties(height=260)
    )
    st.altair_chart(chart, use_container_width=True)


def _render_gap_surplus_trend_chart(forecast_history: pd.DataFrame) -> None:
    st.markdown("**gap_to_target / surplus_to_target 추이**")
    source = _history_long_metric_source(
        forecast_history,
        ("gap_to_target", "surplus_to_target"),
    )
    if source.empty:
        st.info("gap_to_target / surplus_to_target 추이를 만들 이력이 없습니다.")
        return

    chart = (
        alt.Chart(source)
        .mark_line(point=True, strokeWidth=2.3)
        .encode(
            x=alt.X("as_of_date:T", title="기준일"),
            y=alt.Y(
                "value:Q",
                title="금액",
                scale=_auto_value_scale(source),
                axis=alt.Axis(format=chart_value_format("억원")),
            ),
            color=alt.Color("metric_label:N", title="항목"),
            strokeDash=alt.StrokeDash("forecast_model:N", title="예측 모델"),
            tooltip=[
                alt.Tooltip("as_of_date:T", title="기준일"),
                alt.Tooltip("forecast_model:N", title="예측 모델"),
                alt.Tooltip("metric_label:N", title="항목"),
                alt.Tooltip("value:Q", title="금액", format=chart_value_format("억원")),
            ],
        )
        .properties(height=300)
    )
    st.altair_chart(chart, use_container_width=True)


def _render_optional_weighted_forecast(forecast_history: pd.DataFrame) -> None:
    st.markdown("**ModelWeights**")
    weighted_columns = _optional_columns_by_token(
        forecast_history,
        WEIGHTED_FORECAST_COLUMN_TOKENS,
    )
    if not weighted_columns:
        st.caption("Weighted Forecast 데이터가 아직 없습니다.")
        return

    display_columns = _available_columns(
        forecast_history,
        ("target_month", "as_of_date", "metric", *weighted_columns),
    )
    weighted_rows = forecast_history.loc[:, display_columns].dropna(
        how="all",
        subset=weighted_columns,
    )
    if weighted_rows.empty:
        st.caption("Weighted Forecast 컬럼은 있지만 표시할 값이 없습니다.")
        return
    st.dataframe(_format_display_df(weighted_rows.tail(100)), hide_index=True, use_container_width=True)


def _render_optional_confidence_band(forecast_history: pd.DataFrame) -> None:
    st.markdown("**ConfidenceBand**")
    band_pair = _confidence_band_pair(forecast_history)
    if band_pair is None:
        st.caption("Confidence Band 데이터가 아직 없습니다.")
        return

    lower_column, upper_column = band_pair
    display_columns = _available_columns(
        forecast_history,
        ("target_month", "as_of_date", "metric", "forecast_model", "forecast_amount", lower_column, upper_column),
    )
    band_rows = forecast_history.loc[:, display_columns].dropna(
        how="all",
        subset=[lower_column, upper_column],
    )
    if band_rows.empty:
        st.caption("Confidence Band 컬럼은 있지만 표시할 값이 없습니다.")
        return
    st.dataframe(_format_display_df(band_rows.tail(100)), hide_index=True, use_container_width=True)


def _history_long_metric_source(
    forecast_history: pd.DataFrame,
    value_columns: tuple[str, ...],
) -> pd.DataFrame:
    if forecast_history.empty or "as_of_date" not in forecast_history.columns:
        return pd.DataFrame()

    available_value_columns = [
        column
        for column in value_columns
        if column in forecast_history.columns
    ]
    if not available_value_columns:
        return pd.DataFrame()

    working = forecast_history.copy()
    working["as_of_date"] = pd.to_datetime(working["as_of_date"], errors="coerce")
    if "forecast_model" not in working.columns:
        working["forecast_model"] = "ALL"
    for column in available_value_columns:
        working[column] = pd.to_numeric(working[column], errors="coerce")

    grouped = (
        working.dropna(subset=["as_of_date"])
        .groupby(["as_of_date", "forecast_model"], dropna=False)[available_value_columns]
        .mean()
        .reset_index()
    )
    source = grouped.melt(
        id_vars=["as_of_date", "forecast_model"],
        value_vars=available_value_columns,
        var_name="metric",
        value_name="value",
    ).dropna(subset=["value"])
    source["metric_label"] = source["metric"].map(_display_column_label)
    return source.reset_index(drop=True)


def _optional_columns_by_token(
    df: pd.DataFrame,
    tokens: tuple[str, ...],
) -> list[str]:
    if df.empty:
        return []
    lower_tokens = tuple(token.lower() for token in tokens)
    return [
        column
        for column in df.columns
        if any(token in str(column).lower() for token in lower_tokens)
    ]


def _confidence_band_pair(df: pd.DataFrame) -> tuple[str, str] | None:
    if df.empty:
        return None
    columns = set(df.columns)
    for lower_column, upper_column in CONFIDENCE_BAND_COLUMN_PAIRS:
        if lower_column in columns and upper_column in columns:
            return lower_column, upper_column
    return None


def _available_columns(
    df: pd.DataFrame,
    preferred_columns: tuple[str, ...],
) -> list[str]:
    return [column for column in preferred_columns if column in df.columns]


def _render_historical_context_panel(
    historical_context: dict[str, object],
    scenario_df: pd.DataFrame | None = None,
    selected_scenario_id: str = "",
) -> None:
    st.subheader("과거 실적 기반 예측 비교")
    if not historical_context.get("has_data"):
        st.info("과거 월 데이터가 없어 과거 실적 기반 예측값을 계산할 수 없습니다.")
        return

    comparison_df = build_historical_forecast_comparison(
        _as_dataframe(scenario_df),
        historical_context,
        selected_scenario_id,
    )
    if comparison_df.empty:
        st.info("같은 영업일차의 과거 월말 전환 데이터가 부족해 비교값을 만들 수 없습니다.")
    else:
        decision_summary = build_historical_forecast_decision_summary(comparison_df)
        _render_historical_forecast_decision_card(decision_summary)
        selected_rows = comparison_df.loc[
            comparison_df["basis"].astype(str).str.startswith("선택 시나리오")
        ]
        historical_rows = comparison_df.loc[comparison_df["basis"] == "과거 중앙값"]
        selected_amount = (
            selected_rows.iloc[0]["forecast_amount"] if not selected_rows.empty else float("nan")
        )
        historical_median = (
            historical_rows.iloc[0]["forecast_amount"] if not historical_rows.empty else float("nan")
        )
        diff_vs_historical = (
            selected_amount - historical_median
            if math.isfinite(_as_float(selected_amount)) and math.isfinite(_as_float(historical_median))
            else float("nan")
        )
        diff_vs_target = (
            selected_rows.iloc[0]["diff_vs_target"] if not selected_rows.empty else float("nan")
        )

        cols = st.columns(4)
        cols[0].metric("선택 예측", format_amount(selected_amount))
        cols[1].metric("과거 중앙값 예측", format_amount(historical_median))
        cols[2].metric("과거 중앙값 대비", format_amount(diff_vs_historical))
        cols[3].metric("목표 대비", format_amount(diff_vs_target))

        _render_historical_forecast_comparison_chart(comparison_df)
        st.dataframe(
            _format_historical_forecast_comparison_df(comparison_df),
            hide_index=True,
            use_container_width=True,
        )

    benchmark = dict(historical_context.get("benchmark") or {})
    source_label = str(historical_context.get("source_label") or "과거 월 데이터")
    row_count = int(historical_context.get("row_count") or 0)
    month_count = int(benchmark.get("month_count") or 0)

    st.markdown("**과거 월 누적 기준 해석**")
    st.caption(
        f"과거 데이터: {source_label} | {row_count}행 | 같은 영업일차 비교 가능 월 {month_count}개"
    )

    lower = benchmark.get("historical_forecast_lower")
    median = benchmark.get("historical_forecast_median")
    upper = benchmark.get("historical_forecast_upper")
    forecast_range = (
        f"{format_amount(lower)} ~ {format_amount(upper)}"
        if math.isfinite(_as_float(lower)) and math.isfinite(_as_float(upper))
        else "계산 불가"
    )

    cols = st.columns(4)
    cols[0].metric("현재 누적 달성률", format_rate(benchmark.get("current_achievement_rate")))
    cols[1].metric("과거 중앙값", format_rate(benchmark.get("historical_stage_median_rate")))
    cols[2].metric("과거 보정 예상 범위", forecast_range)
    cols[3].metric("과거 보정 중앙값", format_amount(median))

    for message in historical_context.get("interpretation", []):
        st.write(f"- {message}")

    progress_chart_data = _as_dataframe(historical_context.get("progress_chart_data"))
    _render_historical_progress_chart(progress_chart_data)

    stage_df = _as_dataframe(benchmark.get("stage_df"))
    if not stage_df.empty:
        with st.expander("같은 영업일차 과거 월 비교표", expanded=False):
            st.dataframe(
                _format_historical_stage_df(stage_df),
                hide_index=True,
                use_container_width=True,
            )

    monthly_summary = _as_dataframe(historical_context.get("monthly_summary"))
    if not monthly_summary.empty:
        with st.expander("과거 월별 최종 요약", expanded=False):
            st.dataframe(
                _format_historical_monthly_summary_df(monthly_summary),
                hide_index=True,
                use_container_width=True,
        )


def _render_historical_forecast_decision_card(summary: Mapping[str, object]) -> None:
    if not summary.get("has_data"):
        return

    rows = (
        ("기준 보고 범위", str(summary.get("headline") or "")),
        ("예측 성격", str(summary.get("forecast_position") or "")),
        ("목표 판단", str(summary.get("target_position") or "")),
        ("운영 액션", str(summary.get("action") or "")),
    )
    row_html = "".join(
        '<span>'
        f"<strong>{escape(label)}</strong><br>"
        f"{escape(value)}"
        "</span>"
        for label, value in rows
        if value
    )
    st.markdown(
        '<section class="history-purpose-card">'
        "<h3>최종 판단 요약</h3>"
        f'<div class="history-question-grid">{row_html}</div>'
        "</section>",
        unsafe_allow_html=True,
    )


def _render_historical_forecast_comparison_chart(comparison_df: pd.DataFrame) -> None:
    source = _as_dataframe(comparison_df)
    if source.empty:
        return

    source = source.copy()
    source["sort_order"] = range(len(source))
    source["forecast_amount"] = pd.to_numeric(source["forecast_amount"], errors="coerce")
    source = source.dropna(subset=["forecast_amount"])
    if source.empty:
        return
    axis_domain = build_historical_forecast_axis_domain(source)
    x_scale = (
        alt.Scale(domain=axis_domain, zero=False, nice=True)
        if axis_domain is not None
        else alt.Scale(zero=False, nice=True)
    )
    axis_floor = (
        axis_domain[0]
        if axis_domain is not None
        else float(source["forecast_amount"].min())
    )
    source["axis_floor"] = axis_floor
    st.caption("차이가 보이는 구간만 확대해서 표시합니다. 0 기준 전체 막대가 아니라 월말 예상값 주변 비교용 축입니다.")

    bar = (
        alt.Chart(source)
        .mark_bar(cornerRadiusEnd=3)
        .encode(
            y=alt.Y("basis:N", title=None, sort=list(source["basis"])),
            x=alt.X(
                "axis_floor:Q",
                title="월말 예상 실적",
                axis=alt.Axis(format=chart_value_format("억원")),
                scale=x_scale,
            ),
            x2=alt.X2("forecast_amount:Q"),
            color=alt.Color(
                "comparison_group:N",
                title="구분",
                scale=alt.Scale(
                    domain=["현재 예측", "F모델 기본 예측", "과거 실적 기반"],
                    range=["#DC2626", "#2563EB", "#059669"],
                ),
            ),
            tooltip=[
                alt.Tooltip("comparison_group:N", title="구분"),
                alt.Tooltip("basis:N", title="기준"),
                alt.Tooltip("forecast_amount:Q", title="예측값", format=chart_value_format("억원")),
                alt.Tooltip("diff_vs_target:Q", title="목표 대비", format=chart_value_format("억원")),
                alt.Tooltip(
                    "diff_vs_historical_median:Q",
                    title="과거 중앙값 대비",
                    format=chart_value_format("억원"),
                ),
            ],
        )
        .properties(height=max(260, min(420, len(source) * 42)))
    )

    rules: list[alt.Chart] = []
    monthly_target_values = pd.to_numeric(source["monthly_target"], errors="coerce").dropna()
    monthly_target = (
        _as_float(monthly_target_values.iloc[0])
        if not monthly_target_values.empty
        else float("nan")
    )
    if math.isfinite(monthly_target):
        rules.append(
            alt.Chart(pd.DataFrame({"value": [monthly_target], "label": ["월 목표"]}))
            .mark_rule(strokeDash=[6, 4], color="#111827")
            .encode(x="value:Q", tooltip=[alt.Tooltip("value:Q", title="월 목표", format=chart_value_format("억원"))])
        )

    historical_median_rows = source.loc[source["basis"] == "과거 중앙값"]
    if not historical_median_rows.empty:
        historical_median = _as_float(historical_median_rows.iloc[0]["forecast_amount"])
        if math.isfinite(historical_median):
            rules.append(
                alt.Chart(pd.DataFrame({"value": [historical_median], "label": ["과거 중앙값"]}))
                .mark_rule(strokeDash=[3, 3], color="#059669")
                .encode(
                    x="value:Q",
                    tooltip=[
                        alt.Tooltip(
                            "value:Q",
                            title="과거 중앙값",
                            format=chart_value_format("억원"),
                        )
                    ],
                )
            )

    chart = bar
    for rule in rules:
        chart = chart + rule
    st.altair_chart(chart, use_container_width=True)


def _format_historical_forecast_comparison_df(comparison_df: pd.DataFrame) -> pd.DataFrame:
    result = comparison_df.copy()
    columns = [
        "comparison_group",
        "basis",
        "forecast_amount",
        "forecast_rate",
        "diff_vs_target",
        "diff_vs_historical_median",
    ]
    result = result.loc[:, [column for column in columns if column in result.columns]]
    for column in ("forecast_amount", "diff_vs_target", "diff_vs_historical_median"):
        if column in result.columns:
            result[column] = result[column].map(format_amount)
    if "forecast_rate" in result.columns:
        result["forecast_rate"] = result["forecast_rate"].map(format_rate)
    return result.rename(
        columns={
            "comparison_group": "구분",
            "basis": "비교 기준",
            "forecast_amount": "월말 예상 실적",
            "forecast_rate": "월 목표 대비",
            "diff_vs_target": "목표 대비",
            "diff_vs_historical_median": "과거 중앙값 대비",
        }
    )


def _render_historical_progress_chart(chart_data: pd.DataFrame) -> None:
    if chart_data.empty:
        st.info("과거 누적 흐름 차트 데이터 없음")
        return

    chart = (
        alt.Chart(chart_data)
        .mark_line(point=True, strokeWidth=2.3)
        .encode(
            x=alt.X(
                "business_day_no:O",
                title="영업일차",
                sort=None,
                axis=alt.Axis(labelAngle=0),
            ),
            y=alt.Y(
                "achievement_rate:Q",
                title="누적 달성률",
                axis=alt.Axis(format=".0%"),
                scale=alt.Scale(zero=False, nice=True),
            ),
            color=alt.Color(
                "series:N",
                title="범례",
                scale=alt.Scale(
                    domain=["현재 월", "과거 중앙값", "과거 하위 25%", "과거 상위 25%"],
                    range=["#DC2626", "#2563EB", "#94A3B8", "#64748B"],
                ),
            ),
            tooltip=[
                alt.Tooltip("business_day_no:O", title="영업일차"),
                alt.Tooltip("series:N", title="범례"),
                alt.Tooltip("achievement_rate:Q", title="누적 달성률", format=".1%"),
            ],
        )
        .properties(height=320)
        .configure_axis(labelFontSize=11, titleFontSize=12)
        .configure_legend(labelFontSize=11, titleFontSize=12)
    )
    st.altair_chart(chart, use_container_width=True)


def _format_historical_stage_df(stage_df: pd.DataFrame) -> pd.DataFrame:
    result = stage_df.copy()
    columns = [
        "month",
        "matched_business_day_no",
        "as_of_target_cum",
        "as_of_actual_cum",
        "as_of_achievement_rate",
        "monthly_target",
        "final_actual_cum",
        "final_achievement_rate",
        "remaining_actual_growth",
    ]
    result = result.loc[:, [column for column in columns if column in result.columns]]
    for column in (
        "as_of_target_cum",
        "as_of_actual_cum",
        "monthly_target",
        "final_actual_cum",
        "remaining_actual_growth",
    ):
        if column in result.columns:
            result[column] = result[column].map(format_amount)
    for column in ("as_of_achievement_rate", "final_achievement_rate"):
        if column in result.columns:
            result[column] = result[column].map(format_rate)
    return result.rename(
        columns={
            "month": "월",
            "matched_business_day_no": "비교 영업일차",
            "as_of_target_cum": "당시 누적 목표",
            "as_of_actual_cum": "당시 누적 실적",
            "as_of_achievement_rate": "당시 누적 달성률",
            "monthly_target": "월 목표",
            "final_actual_cum": "최종 누적 실적",
            "final_achievement_rate": "최종 달성률",
            "remaining_actual_growth": "비교일 이후 증가 실적",
        }
    )


def _format_historical_monthly_summary_df(monthly_summary: pd.DataFrame) -> pd.DataFrame:
    result = monthly_summary.copy()
    columns = [
        "month",
        "row_count",
        "completed_actual_days",
        "final_business_day_no",
        "monthly_target",
        "final_actual_cum",
        "final_achievement_rate",
        "close_day_count",
    ]
    result = result.loc[:, [column for column in columns if column in result.columns]]
    for column in ("monthly_target", "final_actual_cum"):
        if column in result.columns:
            result[column] = result[column].map(format_amount)
    if "final_achievement_rate" in result.columns:
        result["final_achievement_rate"] = result["final_achievement_rate"].map(format_rate)
    return result.rename(
        columns={
            "month": "월",
            "row_count": "행 수",
            "completed_actual_days": "실적 입력일 수",
            "final_business_day_no": "최종 영업일차",
            "monthly_target": "월 목표",
            "final_actual_cum": "최종 누적 실적",
            "final_achievement_rate": "최종 달성률",
            "close_day_count": "마감일 수",
        }
    )


def _render_strategy_level_visuals(
    scenario_df: pd.DataFrame,
    selected_scenario_id: str,
) -> None:
    strategy_table = build_strategy_level_table(scenario_df, selected_scenario_id)
    if strategy_table.empty:
        st.info("운영전략별 목표 수준 데이터 없음")
        return

    _render_visual_metric_definitions(
        (
            "stretch_uplift",
            "revised_monthly_target",
            "remaining_surplus_buffer",
            "relief_amount",
        )
    )
    _render_chart_reading_guide("scenario_target_position")
    strategy_compare_source = build_strategy_arrival_compare_source(
        strategy_table,
        selected_scenario_id,
    )
    if bool(strategy_compare_source.attrs.get("fallback_used")):
        _render_strategy_compare_fallback(strategy_compare_source)
    else:
        _render_scenario_target_position_chart(strategy_table, selected_scenario_id)

    st.dataframe(_format_display_df(strategy_table), hide_index=True, use_container_width=True)


def _render_forecast_model_scenario_tabs(
    df: pd.DataFrame,
    scenario_df: pd.DataFrame,
    as_of_date: object,
    metric: str,
    config: dict[str, Any],
    selected_scenario_id: str | None,
) -> None:
    forecast_keys = [
        forecast_key
        for forecast_key in ("F1", "F2", "F3")
        if not _scenario_df_for_forecast_key(scenario_df, forecast_key).empty
    ]
    if not forecast_keys:
        st.info("예측모델별 시나리오 데이터 없음")
        return

    tabs = st.tabs([f"{forecast_key} 예측" for forecast_key in forecast_keys])
    for tab, forecast_key in zip(tabs, forecast_keys):
        with tab:
            model_scenario_df = _scenario_df_for_forecast_key(scenario_df, forecast_key)
            model_selected_id = _default_scenario_for_forecast_key(
                model_scenario_df,
                forecast_key,
                selected_scenario_id,
            )
            st.caption(
                f"{forecast_key} 기준으로 운영전략만 비교합니다. "
                "그래프는 해당 예측모델의 전략선만 표시합니다."
            )
            strategy_scenario_id = _render_model_strategy_selector(
                model_scenario_df,
                forecast_key,
                model_selected_id,
            )
            daily_forecast_source = build_scenario_daily_forecast_source(
                df,
                model_scenario_df,
                as_of_date,
                metric,
                config,
                strategy_scenario_id,
            )
            _render_scenario_daily_forecast_chart(
                daily_forecast_source,
                model_scenario_df,
                strategy_scenario_id,
            )
            _render_remaining_operation_direction_panel(
                df,
                as_of_date,
                metric,
                config,
                strategy_scenario_id,
            )
            _render_selected_scenario_daily_detail_table(
                daily_forecast_source,
                strategy_scenario_id,
            )


def _scenario_df_for_forecast_key(
    scenario_df: pd.DataFrame,
    forecast_key: str,
) -> pd.DataFrame:
    if scenario_df.empty or "scenario_id" not in scenario_df.columns:
        return pd.DataFrame(columns=scenario_df.columns)
    return scenario_df.loc[
        scenario_df["scenario_id"].astype(str).str.startswith(f"{forecast_key}_")
    ].reset_index(drop=True)


def _default_scenario_for_forecast_key(
    model_scenario_df: pd.DataFrame,
    forecast_key: str,
    selected_scenario_id: str | None,
) -> str:
    scenario_ids = model_scenario_df["scenario_id"].astype(str).tolist()
    selected_id = str(selected_scenario_id or "")
    if selected_id.startswith(f"{forecast_key}_") and selected_id in scenario_ids:
        return selected_id
    if scenario_ids:
        return scenario_ids[0]
    return ""


def _render_model_strategy_selector(
    model_scenario_df: pd.DataFrame,
    forecast_key: str,
    selected_scenario_id: str,
) -> str:
    scenario_ids = model_scenario_df["scenario_id"].astype(str).tolist()
    if not scenario_ids:
        return ""

    default_index = (
        scenario_ids.index(selected_scenario_id)
        if selected_scenario_id in scenario_ids
        else 0
    )
    return st.selectbox(
        "잔여기간 운영전략",
        scenario_ids,
        format_func=format_scenario_option_label,
        index=default_index,
        key=f"remaining_strategy_{forecast_key}_{'_'.join(scenario_ids)}",
    )


def _render_remaining_operation_direction_panel(
    df: pd.DataFrame,
    as_of_date: object,
    metric: str,
    config: dict[str, Any],
    scenario_id: str,
) -> None:
    if not scenario_id:
        st.info("잔여기간 운영전략을 선택할 수 없습니다.")
        return

    forecast_result, strategy_result = run_selected_scenario_detail(
        df,
        as_of_date,
        metric,
        scenario_id,
        config,
    )
    direction_source = build_remaining_operation_direction_source(
        df,
        as_of_date,
        metric,
        scenario_id,
        forecast_result,
        strategy_result,
    )
    if direction_source.empty:
        st.info("잔여기간 일자별 운영 방향 데이터 없음")
        return

    st.markdown("**잔여기간 일자별 운영 방향**")
    _render_remaining_operation_direction_chart(direction_source)
    with st.expander("잔여기간 운영 방향표", expanded=False):
        st.dataframe(
            _format_remaining_operation_direction_df(direction_source),
            hide_index=True,
            use_container_width=True,
        )


def _render_remaining_operation_direction_chart(source: pd.DataFrame) -> None:
    chart_source = source.copy()
    chart_source["date"] = pd.to_datetime(chart_source["date"], errors="coerce")
    chart_source = chart_source.dropna(subset=["date"])
    if chart_source.empty:
        st.info("잔여기간 운영 방향 차트 데이터 없음")
        return

    for column in ("original_target", "uplift", "revised_target", "expected_daily"):
        chart_source[column] = pd.to_numeric(chart_source[column], errors="coerce")

    date_order = chart_source["date_label"].tolist()
    bar = (
        alt.Chart(chart_source)
        .mark_bar(cornerRadiusTopLeft=2, cornerRadiusTopRight=2)
        .encode(
            x=alt.X(
                "date_label:N",
                title=None,
                sort=date_order,
                axis=alt.Axis(labelAngle=-30, labelLimit=110),
            ),
            y=alt.Y(
                "revised_target:Q",
                title="일자별 관리 목표(억원)",
                axis=alt.Axis(format=chart_value_format("억원")),
            ),
            color=alt.Color(
                "direction:N",
                title="운영 방향",
                scale=alt.Scale(
                    domain=[
                        "추가 배분",
                        "상한 점검",
                        "기존 목표 유지",
                        "버퍼 방어",
                        "Stretch 후보",
                        "유지 모니터링",
                    ],
                    range=[
                        "#2563EB",
                        "#DC2626",
                        "#94A3B8",
                        "#059669",
                        "#7C3AED",
                        "#0F766E",
                    ],
                ),
            ),
            tooltip=[
                alt.Tooltip("date_label:N", title="날짜"),
                alt.Tooltip("day_type:N", title="일자 구분"),
                alt.Tooltip("operation_mode:N", title="운영 모드"),
                alt.Tooltip("direction:N", title="운영 방향"),
                alt.Tooltip("original_target:Q", title="기존 일 목표", format=chart_value_format("억원")),
                alt.Tooltip("uplift:Q", title="업리프트", format=chart_value_format("억원")),
                alt.Tooltip("revised_target:Q", title="관리 목표", format=chart_value_format("억원")),
                alt.Tooltip("expected_daily:Q", title="예상 일실적", format=chart_value_format("억원")),
                alt.Tooltip("direction_detail:N", title="해석"),
            ],
        )
    )
    expected_line = (
        alt.Chart(chart_source)
        .mark_line(point=True, color="#111827", strokeWidth=2)
        .encode(
            x=alt.X("date_label:N", sort=date_order),
            y=alt.Y("expected_daily:Q"),
            tooltip=[
                alt.Tooltip("date_label:N", title="날짜"),
                alt.Tooltip("expected_daily:Q", title="예상 일실적", format=chart_value_format("억원")),
            ],
        )
    )
    chart = (
        (bar + expected_line)
        .properties(height=280)
        .configure_axis(labelFontSize=11, titleFontSize=12)
        .configure_legend(labelFontSize=11, titleFontSize=12)
    )
    st.altair_chart(chart, use_container_width=True)


def _format_remaining_operation_direction_df(source: pd.DataFrame) -> pd.DataFrame:
    result = source.copy()
    result = result.loc[:, list(REMAINING_OPERATION_DIRECTION_COLUMNS)]
    result["date"] = result["date"].map(_format_date)
    for column in ("original_target", "uplift", "revised_target", "expected_daily"):
        result[column] = result[column].map(format_amount)
    result["expected_rate"] = result["expected_rate"].map(format_rate)
    return result.rename(
        columns={
            "date": "날짜",
            "date_label": "날짜 라벨",
            "scenario_id": "시나리오",
            "strategy_type": "전략 구분",
            "operation_mode": "운영 모드",
            "day_type": "일자 구분",
            "close_type": "마감 유형",
            "original_target": "기존 일 목표",
            "uplift": "업리프트",
            "revised_target": "관리 목표",
            "expected_daily": "예상 일실적",
            "expected_rate": "예상 달성률",
            "direction": "운영 방향",
            "direction_detail": "방향 해석",
        }
    )


def _render_scenario_daily_forecast_chart(
    source: pd.DataFrame,
    scenario_df: pd.DataFrame,
    selected_scenario_id: str | None = None,
) -> None:
    if source.empty:
        st.info("주간 시나리오 누적 전망 데이터 없음")
        return

    chart_source = source.copy()
    chart_source["date"] = pd.to_datetime(chart_source["date"], errors="coerce")
    chart_source["week_start"] = pd.to_datetime(chart_source["week_start"], errors="coerce")
    chart_source["week_end"] = pd.to_datetime(chart_source["week_end"], errors="coerce")
    chart_source = chart_source.dropna(subset=["date", "forecast_cum"])
    if chart_source.empty:
        st.info("주간 시나리오 누적 전망 데이터 없음")
        return

    target_source = (
        chart_source.loc[
            :,
            [
                "date",
                "week_end",
                "week_label",
                "target_cum",
                "monthly_target",
                "target_achievement_rate",
                "target_achievement_label",
            ],
        ]
        .drop_duplicates("date")
        .sort_values("date")
    )
    week_marker_source = (
        chart_source.loc[:, ["week_start", "week_label"]]
        .dropna(subset=["week_start"])
        .drop_duplicates("week_start")
        .sort_values("week_start")
    )
    actual_source = chart_source.loc[chart_source["series_type"] == "확정 실적"]
    forecast_source = chart_source.loc[chart_source["series_type"] == "시나리오 예상"]
    position_source = build_scenario_target_position_source(scenario_df, selected_scenario_id)
    forecast_source = _attach_scenario_position_fields(forecast_source, position_source)
    close_day_source = source.loc[
        source["is_close_day"],
        ["date", "date_label", "close_type"],
    ].drop_duplicates("date")
    close_day_source["date"] = pd.to_datetime(close_day_source["date"], errors="coerce")
    close_day_source = close_day_source.dropna(subset=["date"])
    close_day_source["band_start"] = close_day_source["date"] - pd.Timedelta(hours=12)
    close_day_source["band_end"] = close_day_source["date"] + pd.Timedelta(hours=12)

    y_max = max(
        _finite_max(chart_source["achievement_rate"]),
        _finite_max(target_source["target_achievement_rate"]),
        1.0,
    )
    scenario_order = forecast_source["scenario_id"].drop_duplicates().tolist()
    selected_id = str(selected_scenario_id or "")
    final_label_source = _selected_daily_final_label_source(forecast_source, selected_id)
    as_of_dates = source.loc[source["is_as_of_date"], "date"].drop_duplicates()
    as_of_source = pd.DataFrame({"as_of_date": as_of_dates.tolist()[:1]})
    zoom = alt.selection_interval(bind="scales", encodings=["x"], name="weekly_zoom")
    x_axis = alt.Axis(
        format="%m/%d",
        labelAngle=0,
        tickCount=alt.TimeIntervalStep("week", 1),
    )

    close_day_band = (
        alt.Chart(close_day_source)
        .mark_rect(color="#FEF3C7", opacity=0.55)
        .encode(
            x=alt.X("band_start:T", title=None),
            x2="band_end:T",
            tooltip=[
                alt.Tooltip("date_label:N", title="마감일"),
                alt.Tooltip("close_type:N", title="마감 유형"),
            ],
        )
    )
    week_markers = (
        alt.Chart(week_marker_source)
        .mark_rule(color="#E2E8F0", strokeWidth=1)
        .encode(
            x=alt.X("week_start:T", title=None),
            tooltip=[alt.Tooltip("week_label:N", title="주간")],
        )
    )
    target_line = (
        alt.Chart(target_source)
        .mark_line(color="#94A3B8", strokeDash=[6, 5], strokeWidth=1.8, interpolate="linear")
        .encode(
            x=alt.X(
                "date:T",
                title=None,
                axis=x_axis,
            ),
            y=alt.Y(
                "target_achievement_rate:Q",
                title="월 목표 달성률",
                scale=alt.Scale(domain=[0, y_max * 1.08], nice=True),
                axis=alt.Axis(format=".0%"),
            ),
            tooltip=[
                alt.Tooltip("week_label:N", title="주간"),
                alt.Tooltip("target_cum:Q", title="누적 목표선", format=chart_value_format("억원")),
                alt.Tooltip("target_achievement_rate:Q", title="누적 목표 달성률", format=".1%"),
            ],
        )
    )
    actual_line = (
        alt.Chart(actual_source)
        .mark_line(color="#1D4ED8", strokeWidth=3, interpolate="linear")
        .encode(
            x=alt.X(
                "date:T",
                title=None,
                axis=x_axis,
            ),
            y=alt.Y("achievement_rate:Q"),
            tooltip=_daily_forecast_tooltips("확정 누적 실적", include_variance=False),
        )
    )
    forecast_lines = (
        alt.Chart(forecast_source)
        .mark_line(interpolate="linear")
        .encode(
            x=alt.X(
                "date:T",
                title=None,
                axis=x_axis,
            ),
            y=alt.Y("achievement_rate:Q"),
            color=alt.Color(
                "scenario_id:N",
                title="시나리오",
                sort=scenario_order,
                scale=alt.Scale(range=list(CHART_COLOR_RANGE)),
            ),
            detail="line_group:N",
            opacity=alt.condition("datum.is_selected", alt.value(1.0), alt.value(0.38)),
            strokeWidth=alt.condition("datum.is_selected", alt.value(3.2), alt.value(1.5)),
            tooltip=_daily_forecast_tooltips("누적 예상"),
        )
    )
    monthly_rule = (
        alt.Chart(pd.DataFrame({"goal_rate": [1.0]}))
        .mark_rule(color="#DC2626", strokeDash=[7, 5], strokeWidth=2)
        .encode(
            y="goal_rate:Q",
            tooltip=[
                alt.Tooltip(
                    "goal_rate:Q",
                    title="공식 월 목표선",
                    format=".0%",
                )
            ],
        )
    )
    as_of_rule = (
        alt.Chart(as_of_source)
        .mark_rule(color="#64748B", strokeDash=[2, 3], strokeWidth=1.4)
        .encode(
            x="as_of_date:T",
            tooltip=[alt.Tooltip("as_of_date:T", title="기준일", format="%Y-%m-%d")],
        )
    )
    final_label = (
        alt.Chart(final_label_source)
        .mark_text(align="right", baseline="middle", dx=-7, dy=-8, fontSize=11, fontWeight="bold")
        .encode(
            x="date:T",
            y="achievement_rate:Q",
            text="final_label:N",
            color=alt.value("#111827"),
        )
    )
    progress_chart = (
        (
            close_day_band
            + week_markers
            + target_line
            + actual_line
            + forecast_lines
            + monthly_rule
            + as_of_rule
            + final_label
        )
        .properties(height=330)
        .add_params(zoom)
    )

    chart = (
        progress_chart
        .configure_axis(labelFontSize=11, titleFontSize=12)
        .configure_legend(labelFontSize=11, titleFontSize=12)
    )
    st.caption("차트 위에서 마우스 휠 또는 드래그로 주간 구간을 확대/축소할 수 있습니다.")
    st.altair_chart(chart, use_container_width=True)


def _render_selected_scenario_daily_detail_table(
    daily_source: pd.DataFrame,
    selected_scenario_id: str | None,
) -> None:
    detail = build_selected_scenario_daily_detail_source(daily_source, selected_scenario_id)
    if detail.empty:
        return

    with st.expander("선택 시나리오 일자별 추정 내역", expanded=False):
        st.dataframe(
            _format_daily_forecast_detail_df(detail),
            hide_index=True,
            use_container_width=True,
        )


def _format_daily_forecast_detail_df(detail: pd.DataFrame) -> pd.DataFrame:
    result = detail.copy()
    result["date"] = result["date"].map(_format_date)
    result["daily_expected"] = result["daily_expected"].map(_format_optional_amount)
    result["forecast_cum"] = result["forecast_cum"].map(format_amount)
    result["target_cum"] = result["target_cum"].map(format_amount)
    result["achievement_rate"] = result["achievement_rate"].map(format_rate)
    result["target_achievement_rate"] = result["target_achievement_rate"].map(format_rate)
    return result.rename(
        columns={
            "date": "날짜",
            "scenario_id": "시나리오",
            "series_type": "구분",
            "day_type": "일자 구분",
            "close_type": "마감 유형",
            "daily_expected": "당일 추정",
            "forecast_cum": "누적 실적/예상",
            "target_cum": "누적 목표선",
            "achievement_rate": "월 목표 달성률",
            "target_achievement_rate": "계획선 달성률",
        }
    )


def _format_optional_amount(value: object) -> str:
    number = _as_float(value)
    if not math.isfinite(number):
        return "-"
    return format_amount(number)


def _daily_forecast_tooltips(
    value_title: str,
    *,
    include_variance: bool = True,
) -> list[alt.Tooltip]:
    tooltips = [
        alt.Tooltip("week_label:N", title="주간"),
        alt.Tooltip("date_label:N", title="대표일"),
        alt.Tooltip("scenario_id:N", title="시나리오"),
        alt.Tooltip("day_type:N", title="일자 구분"),
        alt.Tooltip("close_type:N", title="마감 유형"),
        alt.Tooltip("daily_expected:Q", title="당일 예상", format=chart_value_format("억원")),
        alt.Tooltip("forecast_cum:Q", title=value_title, format=chart_value_format("억원")),
        alt.Tooltip("target_cum:Q", title="누적 목표선", format=chart_value_format("억원")),
        alt.Tooltip("achievement_rate:Q", title="월 목표 달성률", format=".1%"),
        alt.Tooltip("target_achievement_rate:Q", title="계획선 달성률", format=".1%"),
        alt.Tooltip("risk_level_label:N", title="위험등급"),
    ]
    if include_variance:
        tooltips.insert(-1, alt.Tooltip("variance_label:N", title="월말 목표 대비 차이"))
    return tooltips


def _finite_max(values: pd.Series) -> float:
    numeric_values = pd.to_numeric(values, errors="coerce")
    numeric_values = numeric_values.loc[numeric_values.map(math.isfinite)]
    if numeric_values.empty:
        return 0.0
    return float(numeric_values.max())


def _selected_daily_final_label_source(
    forecast_source: pd.DataFrame,
    selected_scenario_id: str,
) -> pd.DataFrame:
    if forecast_source.empty:
        return pd.DataFrame(columns=[*SCENARIO_DAILY_FORECAST_COLUMNS, "final_label"])
    selected = forecast_source.loc[forecast_source["scenario_id"] == selected_scenario_id]
    if selected.empty:
        selected = forecast_source
    last_rows = (
        selected.sort_values(["scenario_id", "date"])
        .groupby("scenario_id", as_index=False)
        .tail(1)
        .copy()
    )
    if selected_scenario_id and selected_scenario_id in set(last_rows["scenario_id"].astype(str)):
        last_rows = last_rows.loc[last_rows["scenario_id"].astype(str) == selected_scenario_id]
    variance_label = last_rows.get(
        "variance_label",
        pd.Series([""] * len(last_rows), index=last_rows.index),
    ).fillna("")
    last_rows["final_label"] = (
        last_rows["scenario_id"].astype(str)
        + " "
        + last_rows["achievement_label"].astype(str)
        + " / "
        + variance_label.astype(str)
    )
    return last_rows


def _attach_scenario_position_fields(
    forecast_source: pd.DataFrame,
    position_source: pd.DataFrame,
) -> pd.DataFrame:
    if forecast_source.empty or position_source.empty:
        return forecast_source

    fields = [
        column
        for column in (
            "scenario_id",
            "target_variance",
            "variance_label",
            "target_status_label",
            "forecast_after_provision",
        )
        if column in position_source.columns
    ]
    if len(fields) <= 1:
        return forecast_source
    return forecast_source.merge(
        position_source.loc[:, fields].drop_duplicates("scenario_id"),
        on="scenario_id",
        how="left",
    )


def _render_scenario_target_position_chart(
    scenario_df: pd.DataFrame,
    selected_scenario_id: str | None = None,
) -> None:
    source = build_scenario_target_position_source(scenario_df, selected_scenario_id)
    if source.empty:
        st.info("목표선 대비 예상 실적 차트 데이터 없음")
        return

    scenario_order = source["scenario_id"].tolist()
    target_value = source["monthly_target"].dropna().iloc[0]
    x_max = max(
        float(source["forecast_after_provision"].max()),
        float(source["monthly_target"].max()),
        1.0,
    )
    bar = (
        alt.Chart(source)
        .mark_bar(cornerRadiusTopRight=3, cornerRadiusBottomRight=3)
        .encode(
            y=alt.Y(
                "scenario_id:N",
                title=None,
                sort=scenario_order,
                axis=alt.Axis(labelLimit=120),
            ),
            x=alt.X(
                "forecast_after_provision:Q",
                title="전략 반영 후 예상(억원)",
                scale=alt.Scale(domain=[0, x_max * 1.08], nice=True),
                axis=alt.Axis(format=chart_value_format("억원")),
            ),
            color=alt.Color(
                "target_status_label:N",
                title="목표 상태",
                scale=alt.Scale(
                    domain=["목표 미달", "목표 달성", "목표 초과", "계산 불가"],
                    range=["#DC2626", "#2563EB", "#059669", "#6B7280"],
                ),
            ),
            opacity=alt.condition("datum.is_selected", alt.value(1.0), alt.value(0.68)),
            stroke=alt.condition("datum.is_selected", alt.value("#111827"), alt.value("#ffffff")),
            strokeWidth=alt.condition("datum.is_selected", alt.value(2.2), alt.value(0.4)),
            tooltip=[
                alt.Tooltip("scenario_id:N", title="시나리오"),
                alt.Tooltip("target_status_label:N", title="목표 상태"),
                alt.Tooltip("risk_level_label:N", title="위험등급"),
                alt.Tooltip(
                    "forecast_after_provision:Q",
                    title="전략 반영 후 예상",
                    format=chart_value_format("억원"),
                ),
                alt.Tooltip(
                    "monthly_target:Q",
                    title="공식 월 목표",
                    format=chart_value_format("억원"),
                ),
                alt.Tooltip("variance_label:N", title="목표 대비 차이"),
            ],
        )
    )
    target_rule = (
        alt.Chart(pd.DataFrame({"monthly_target": [target_value]}))
        .mark_rule(color="#DC2626", strokeDash=[6, 4], strokeWidth=2)
        .encode(
            x="monthly_target:Q",
            tooltip=[
                alt.Tooltip(
                    "monthly_target:Q",
                    title="공식 월 목표",
                    format=chart_value_format("억원"),
                )
            ],
        )
    )
    labels = (
        alt.Chart(source)
        .mark_text(align="left", baseline="middle", dx=5, fontSize=11)
        .encode(
            y=alt.Y("scenario_id:N", sort=scenario_order),
            x="forecast_after_provision:Q",
            text="forecast_label:N",
        )
    )
    chart = (
        (bar + target_rule + labels)
        .properties(height=max(260, len(source) * 30))
        .configure_axis(labelFontSize=11, titleFontSize=12)
        .configure_legend(labelFontSize=11, titleFontSize=12)
    )
    st.altair_chart(chart, use_container_width=True)


def _render_scenario_gap_chart(
    scenario_df: pd.DataFrame,
    selected_scenario_id: str | None = None,
) -> None:
    source = build_scenario_target_position_source(scenario_df, selected_scenario_id)
    if source.empty:
        st.info("부족/초과 금액 차트 데이터 없음")
        return

    scenario_order = source["scenario_id"].tolist()
    magnitude = max(
        float(source["target_variance"].abs().max(skipna=True) or 0.0),
        1.0,
    )
    x_domain = [-magnitude * 1.2, magnitude * 1.2]
    bar = (
        alt.Chart(source)
        .mark_bar(cornerRadiusTopRight=3, cornerRadiusBottomRight=3)
        .encode(
            y=alt.Y(
                "scenario_id:N",
                title=None,
                sort=scenario_order,
                axis=alt.Axis(labelLimit=120),
            ),
            x=alt.X(
                "target_variance:Q",
                title="목표 대비 차이(억원)",
                scale=alt.Scale(domain=x_domain, nice=True),
                axis=alt.Axis(format=chart_value_format("억원")),
            ),
            color=alt.condition(
                "datum.target_variance >= 0",
                alt.value("#059669"),
                alt.value("#DC2626"),
            ),
            opacity=alt.condition("datum.is_selected", alt.value(1.0), alt.value(0.7)),
            tooltip=[
                alt.Tooltip("scenario_id:N", title="시나리오"),
                alt.Tooltip("target_status_label:N", title="목표 상태"),
                alt.Tooltip("variance_label:N", title="목표 대비 차이"),
            ],
        )
    )
    zero_rule = alt.Chart(pd.DataFrame({"zero": [0]})).mark_rule(
        color="#111827",
        strokeWidth=1.4,
    ).encode(x="zero:Q")
    positive_text = (
        alt.Chart(source.loc[source["target_variance"] >= 0])
        .mark_text(align="left", baseline="middle", dx=4, fontSize=11)
        .encode(
            y=alt.Y("scenario_id:N", sort=scenario_order),
            x="target_variance:Q",
            text="variance_label:N",
        )
    )
    negative_text = (
        alt.Chart(source.loc[source["target_variance"] < 0])
        .mark_text(align="right", baseline="middle", dx=-4, fontSize=11)
        .encode(
            y=alt.Y("scenario_id:N", sort=scenario_order),
            x="target_variance:Q",
            text="variance_label:N",
        )
    )
    chart = (
        (bar + zero_rule + positive_text + negative_text)
        .properties(height=max(260, len(source) * 30))
        .configure_axis(labelFontSize=11, titleFontSize=12)
    )
    st.altair_chart(chart, use_container_width=True)


def _render_scenario_heatmap(
    scenario_df: pd.DataFrame,
    selected_scenario_id: str | None = None,
) -> None:
    source = build_scenario_heatmap_source(scenario_df, selected_scenario_id)
    if source.empty:
        st.info("시나리오 조합 지도 데이터 없음")
        return

    strategy_order = _ordered_strategy_keys(scenario_df)
    magnitude = max(
        float(pd.to_numeric(source["target_variance"], errors="coerce").abs().max() or 0.0),
        1.0,
    )
    heatmap = (
        alt.Chart(source)
        .mark_rect(cornerRadius=2)
        .encode(
            x=alt.X("strategy_key:N", title="운영전략", sort=strategy_order),
            y=alt.Y("forecast_key:N", title="예측모델", sort=["F1", "F2", "F3"]),
            color=alt.Color(
                "target_variance:Q",
                title="목표 대비 차이",
                scale=alt.Scale(
                    domain=[-magnitude, 0, magnitude],
                    range=["#DC2626", "#F8FAFC", "#059669"],
                ),
            ),
            stroke=alt.condition("datum.is_selected", alt.value("#111827"), alt.value("#ffffff")),
            strokeWidth=alt.condition("datum.is_selected", alt.value(2.4), alt.value(0.8)),
            tooltip=[
                alt.Tooltip("scenario_id:N", title="시나리오"),
                alt.Tooltip("target_status_label:N", title="목표 상태"),
                alt.Tooltip("variance_label:N", title="목표 대비 차이"),
            ],
        )
    )
    text = (
        alt.Chart(source)
        .mark_text(fontSize=12, fontWeight="bold")
        .encode(
            x=alt.X("strategy_key:N", sort=strategy_order),
            y=alt.Y("forecast_key:N", sort=["F1", "F2", "F3"]),
            text="variance_label:N",
            color=alt.value("#111827"),
        )
    )
    chart = (
        (heatmap + text)
        .properties(height=210)
        .configure_axis(labelFontSize=12, titleFontSize=12)
        .configure_legend(labelFontSize=11, titleFontSize=12)
    )
    st.altair_chart(chart, use_container_width=True)


def _render_remaining_uplift_chart(revised_targets_df: pd.DataFrame) -> None:
    source = build_remaining_target_daily_source(revised_targets_df)
    if source.empty:
        st.info("일자별 추가 부담 차트 데이터 없음")
        return

    date_order = source["date_label"].tolist()
    chart = (
        alt.Chart(source)
        .mark_bar(cornerRadiusTopLeft=3, cornerRadiusTopRight=3)
        .encode(
            x=alt.X(
                "date_label:N",
                title=None,
                sort=date_order,
                axis=alt.Axis(labelAngle=-30, labelLimit=120),
            ),
            y=alt.Y(
                "uplift:Q",
                title="추가 배분 목표(억원)",
                axis=alt.Axis(format=chart_value_format("억원")),
            ),
            color=alt.Color(
                "day_type:N",
                title="일자 구분",
                scale=alt.Scale(
                    domain=["마감일", "일반일", "잔여일"],
                    range=["#D97706", "#2563EB", "#6B7280"],
                ),
            ),
            tooltip=[
                alt.Tooltip("date_label:N", title="날짜"),
                alt.Tooltip("day_type:N", title="일자 구분"),
                alt.Tooltip("close_type:N", title="마감 유형"),
                alt.Tooltip("original_target:Q", title="기존 일 목표", format=chart_value_format("억원")),
                alt.Tooltip("uplift:Q", title="추가 배분 목표", format=chart_value_format("억원")),
                alt.Tooltip("revised_target:Q", title="수정 후 일 목표", format=chart_value_format("억원")),
            ],
        )
        .properties(height=300)
        .configure_axis(labelFontSize=11, titleFontSize=12)
        .configure_legend(labelFontSize=11, titleFontSize=12)
    )
    st.altair_chart(chart, use_container_width=True)


def _render_remaining_target_stack_chart(revised_targets_df: pd.DataFrame) -> None:
    daily = build_remaining_target_daily_source(revised_targets_df)
    stack_source = build_remaining_target_stack_source(revised_targets_df)
    if daily.empty or stack_source.empty:
        st.info("기존 목표와 추가 배분 차트 데이터 없음")
        return

    date_order = daily["date_label"].tolist()
    bars = (
        alt.Chart(stack_source)
        .mark_bar(cornerRadiusTopLeft=2, cornerRadiusTopRight=2)
        .encode(
            x=alt.X(
                "date_label:N",
                title=None,
                sort=date_order,
                axis=alt.Axis(labelAngle=-30, labelLimit=120),
            ),
            y=alt.Y(
                "value:Q",
                title="수정 후 일 목표(억원)",
                stack="zero",
                axis=alt.Axis(format=chart_value_format("억원")),
            ),
            color=alt.Color(
                "target_part:N",
                title="목표 구성",
                scale=alt.Scale(
                    domain=["기존 일 목표", "추가 배분 목표"],
                    range=["#CBD5E1", "#D97706"],
                ),
            ),
            tooltip=[
                alt.Tooltip("date_label:N", title="날짜"),
                alt.Tooltip("target_part:N", title="구성"),
                alt.Tooltip("value:Q", title="금액", format=chart_value_format("억원")),
                alt.Tooltip("revised_target:Q", title="수정 후 일 목표", format=chart_value_format("억원")),
                alt.Tooltip("cap_target:Q", title="일별 허용 상한", format=chart_value_format("억원")),
            ],
        )
    )
    cap_line = (
        alt.Chart(daily)
        .mark_line(point=True, color="#DC2626", strokeDash=[5, 4], strokeWidth=2)
        .encode(
            x=alt.X("date_label:N", sort=date_order),
            y=alt.Y("cap_target:Q"),
            tooltip=[
                alt.Tooltip("date_label:N", title="날짜"),
                alt.Tooltip("cap_target:Q", title="일별 허용 상한", format=chart_value_format("억원")),
            ],
        )
    )
    expected_line = (
        alt.Chart(daily)
        .mark_line(point=True, color="#0F766E", strokeWidth=2)
        .encode(
            x=alt.X("date_label:N", sort=date_order),
            y=alt.Y("expected_after_revision:Q"),
            tooltip=[
                alt.Tooltip("date_label:N", title="날짜"),
                alt.Tooltip(
                    "expected_after_revision:Q",
                    title="수정 후 예상 일 실적",
                    format=chart_value_format("억원"),
                ),
            ],
        )
    )
    chart = (
        (bars + cap_line + expected_line)
        .properties(height=320)
        .configure_axis(labelFontSize=11, titleFontSize=12)
        .configure_legend(labelFontSize=11, titleFontSize=12)
    )
    st.altair_chart(chart, use_container_width=True)


def _render_grouped_bar_chart(
    chart_data: pd.DataFrame,
    metric_columns: tuple[str, ...],
) -> None:
    source = build_grouped_bar_chart_source(chart_data, metric_columns)
    if source.empty:
        st.info("막대 차트 데이터 없음")
        return

    label_order = [
        _chart_labels(metric_columns)[column]
        for column in metric_columns
        if column in set(source["metric"])
    ]
    value_format = chart_value_format(_chart_unit(metric_columns))
    chart = (
        alt.Chart(source)
        .mark_bar(cornerRadiusTopLeft=2, cornerRadiusTopRight=2)
        .encode(
            x=alt.X(
                "category:N",
                title=None,
                sort=None,
                axis=alt.Axis(labelAngle=-30, labelLimit=120),
            ),
            xOffset=alt.XOffset("범례:N", sort=label_order),
            y=alt.Y(
                "value:Q",
                title=_chart_unit(metric_columns) or "값",
                stack=None,
                scale=_auto_value_scale(source),
                axis=alt.Axis(format=value_format),
            ),
            color=alt.Color(
                "범례:N",
                title="범례",
                sort=label_order,
                scale=alt.Scale(range=list(CHART_COLOR_RANGE)),
            ),
            tooltip=[
                alt.Tooltip("category:N", title="항목"),
                alt.Tooltip("범례:N", title="수치"),
                alt.Tooltip("value:Q", title=_chart_tooltip_value_title(metric_columns), format=value_format),
            ],
        )
        .properties(height=320)
        .configure_axis(labelFontSize=11, titleFontSize=12)
        .configure_legend(labelFontSize=11, titleFontSize=12)
    )
    st.altair_chart(chart, use_container_width=True)


def _render_line_chart(
    chart_data: pd.DataFrame,
    metric_columns: tuple[str, ...],
) -> None:
    source = build_grouped_bar_chart_source(chart_data, metric_columns)
    if source.empty:
        st.info("선 차트 데이터 없음")
        return

    label_order = [
        _chart_labels(metric_columns)[column]
        for column in metric_columns
        if column in set(source["metric"])
    ]
    unit = _chart_unit(metric_columns)
    value_format = chart_value_format(unit)
    tooltip_title = _chart_tooltip_value_title(metric_columns)
    chart = (
        alt.Chart(source)
        .mark_line(point=True, strokeWidth=2.5)
        .encode(
            x=alt.X(
                "category:N",
                title=None,
                sort=None,
                axis=alt.Axis(labelAngle=-30, labelLimit=120),
            ),
            y=alt.Y(
                "value:Q",
                title=unit or "값",
                scale=_auto_value_scale(source),
                axis=alt.Axis(format=value_format),
            ),
            color=alt.Color(
                "범례:N",
                title="범례",
                sort=label_order,
                scale=alt.Scale(range=list(CHART_COLOR_RANGE)),
            ),
            tooltip=[
                alt.Tooltip("category:N", title="항목"),
                alt.Tooltip("범례:N", title="수치"),
                alt.Tooltip("value:Q", title=tooltip_title, format=value_format),
            ],
        )
        .properties(height=300)
        .configure_axis(labelFontSize=11, titleFontSize=12)
        .configure_legend(labelFontSize=11, titleFontSize=12)
    )
    st.altair_chart(chart, use_container_width=True)


def _render_rate_ratio_line_chart(
    chart_data: pd.DataFrame,
    metric_columns: tuple[str, ...],
) -> None:
    source = build_grouped_bar_chart_source(chart_data, metric_columns)
    if source.empty:
        st.info("누적 달성률 차트 데이터 없음")
        return

    label_order = [
        _chart_labels(metric_columns)[column]
        for column in metric_columns
        if column in set(source["metric"])
    ]
    chart = (
        alt.Chart(source)
        .mark_line(point=True, strokeWidth=2.6, color="#14756f")
        .encode(
            x=alt.X(
                "category:N",
                title=None,
                sort=None,
                axis=alt.Axis(labelAngle=-30, labelLimit=120),
            ),
            y=alt.Y(
                "value:Q",
                title="누적 달성률",
                scale=_auto_value_scale(source),
                axis=alt.Axis(format=".0%"),
            ),
            color=alt.Color(
                "범례:N",
                title="범례",
                sort=label_order,
                scale=alt.Scale(range=["#14756f"]),
            ),
            tooltip=[
                alt.Tooltip("category:N", title="항목"),
                alt.Tooltip("범례:N", title="수치"),
                alt.Tooltip("value:Q", title="누적 달성률", format=".1%"),
            ],
        )
        .properties(height=280)
        .configure_axis(labelFontSize=11, titleFontSize=12)
        .configure_legend(labelFontSize=11, titleFontSize=12)
    )
    st.altair_chart(chart, use_container_width=True)


def _auto_value_scale(source: pd.DataFrame) -> alt.Scale:
    axis_domain = build_auto_axis_domain(source["value"])
    if axis_domain is None:
        return alt.Scale(zero=False)
    return alt.Scale(domain=axis_domain, zero=False, nice=True)


def _chart_unit(metric_columns: tuple[str, ...]) -> str:
    units = {
        VISUAL_METRIC_DEFINITIONS.get(column, {}).get("unit", "")
        for column in metric_columns
    }
    units.discard("")
    if len(units) == 1:
        return next(iter(units))
    return ""


def _chart_tooltip_value_title(metric_columns: tuple[str, ...]) -> str:
    unit = _chart_unit(metric_columns)
    if unit:
        return f"값({unit})"
    return "값"


def _render_chart_reading_guide(guide_key: str) -> None:
    guide = build_visual_reading_guide(guide_key)
    title = str(guide.get("title") or "").strip()
    steps = tuple(guide.get("steps") or ())
    decision = str(guide.get("decision") or "").strip()
    if not title or not steps:
        return

    st.markdown(f"**{title} 읽는 법**")
    for index, step in enumerate(steps, start=1):
        st.markdown(f"{index}. {step}")
    if decision:
        st.caption(f"판단 기준: {decision}")


def _render_visual_metric_definitions(metric_columns: tuple[str, ...]) -> None:
    definition_df = build_visual_metric_definition_df(metric_columns)
    if definition_df.empty:
        return

    with st.expander("범례와 수치 정의", expanded=False):
        for row in definition_df.to_dict("records"):
            st.markdown(f"- **{row['범례']} ({row['단위']})**: {row['수치 의미']}")


def _load_uploaded_input(
    uploaded_file: Any,
    sort_by: str = "business_day_no",
    strict_business_day_no: bool = True,
) -> pd.DataFrame:
    suffix = Path(uploaded_file.name).suffix.lower()
    if suffix not in {".csv", ".xlsx"}:
        raise ValueError("Unsupported input file type. Use CSV or XLSX.")

    with tempfile.TemporaryDirectory() as tmp_dir:
        temp_path = Path(tmp_dir) / f"uploaded{suffix}"
        temp_path.write_bytes(uploaded_file.getvalue())
        if sort_by not in {"business_day_no", "date"}:
            raise ValueError("sort_by must be either 'business_day_no' or 'date'.")
        return load_input(
            temp_path,
            sort_by=sort_by,
            strict_business_day_no=strict_business_day_no,
        )


def _filter_scenarios(
    scenario_df: pd.DataFrame,
    forecast_choice: str,
    provision_choice: str,
) -> pd.DataFrame:
    result = scenario_df.copy(deep=False)
    if forecast_choice != COMPARE_LABEL:
        result = result.loc[result["scenario_id"].astype(str).str.startswith(forecast_choice)]
    if provision_choice != COMPARE_LABEL:
        strategy_filtered = result.loc[
            result["scenario_id"].astype(str).str.endswith(provision_choice)
        ]
        if not strategy_filtered.empty:
            result = strategy_filtered
    return result.reset_index(drop=True)


def _split_scenario_id(scenario_id: str) -> tuple[str, str]:
    if "_" not in scenario_id:
        return scenario_id, ""
    return scenario_id.split("_", maxsplit=1)


def _selected_forecast_key(selected_scenario_id: str | None) -> str:
    if not selected_scenario_id:
        return ""
    forecast_key, _ = _split_scenario_id(str(selected_scenario_id))
    if forecast_key in FORECAST_MODEL_OPTIONS:
        return forecast_key
    return ""


def _selected_scenario_row(scenario_df: pd.DataFrame, scenario_id: str) -> pd.Series:
    rows = scenario_df.loc[scenario_df["scenario_id"].astype(str) == scenario_id]
    if rows.empty:
        return scenario_df.iloc[0]
    return rows.iloc[0]


def _forecast_summary(scenario_df: pd.DataFrame) -> dict[str, object]:
    summary: dict[str, object] = {}
    for _, row in scenario_df.iterrows():
        scenario_id = str(row.get("scenario_id", ""))
        forecast_key = scenario_id.split("_", maxsplit=1)[0]
        if forecast_key in {"F1", "F2", "F3"} and forecast_key not in summary:
            summary[forecast_key] = row.get("forecast_amount")
    return summary


def _build_indexed_numeric_chart_data(
    df: pd.DataFrame,
    index_column: str,
    value_columns: tuple[str, ...],
) -> pd.DataFrame:
    if df.empty or index_column not in df.columns:
        return pd.DataFrame()

    available_columns = [column for column in value_columns if column in df.columns]
    if not available_columns:
        return pd.DataFrame()

    result = df.loc[:, [index_column, *available_columns]].copy()
    result[index_column] = result[index_column].map(_format_chart_index)
    for column in available_columns:
        result[column] = pd.to_numeric(result[column], errors="coerce")
    return result.set_index(index_column)


def _available_chart_columns(
    chart_data: pd.DataFrame,
    preferred_columns: tuple[str, ...],
) -> tuple[str, ...]:
    if chart_data.empty:
        return ()
    return tuple(
        column
        for column in preferred_columns
        if column in chart_data.columns and not chart_data[column].isna().all()
    )


def _rename_chart_columns(
    df: pd.DataFrame,
    labels: dict[str, str],
) -> pd.DataFrame:
    return df.rename(columns=labels)


def _chart_labels(metric_columns: tuple[str, ...]) -> dict[str, str]:
    return {
        column: VISUAL_METRIC_DEFINITIONS.get(column, {}).get("label", column)
        for column in metric_columns
    }


def _format_chart_index(value: object) -> str:
    if _is_missing(value):
        return ""
    if isinstance(value, pd.Timestamp):
        return str(value.date())
    return str(value)


def _format_display_df(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    result = df.copy()
    if {"item", "value"}.issubset(result.columns):
        original_items = result["item"].astype(str)
        result["value"] = [
            _format_named_value(item, value)
            for item, value in zip(result["item"], result["value"])
        ]
        result["value"] = result["value"].map(_localize_display_value)
        result["item"] = original_items.map(_display_column_label)
        return result

    for column in result.columns:
        if column in TECHNICAL_CODE_COLUMNS:
            result[column] = result[column].map(lambda value: "" if _is_missing(value) else str(value))
        elif column in AMOUNT_COLUMNS:
            result[column] = result[column].map(format_amount)
        elif column in RATE_COLUMNS:
            result[column] = result[column].map(format_rate)
        elif "date" in str(column).lower():
            result[column] = result[column].map(_format_date)
        else:
            result[column] = result[column].map(_localize_display_value)
    return result.rename(columns={column: _display_column_label(column) for column in result.columns})


def _format_named_value(name: object, value: object) -> object:
    column_name = str(name)
    if column_name in AMOUNT_COLUMNS:
        return format_amount(value)
    if column_name in RATE_COLUMNS:
        return format_rate(value)
    if "date" in column_name.lower():
        return _format_date(value)
    return value


def _display_column_label(column: object) -> str:
    text = str(column)
    return DISPLAY_COLUMN_LABELS.get(text, get_metric_label(text))


def _localize_display_value(value: object) -> object:
    if _is_missing(value):
        return value
    if isinstance(value, bool):
        return "예" if value else "아니오"
    if isinstance(value, (list, tuple, set)):
        localized = [format_validation_message(item) for item in value]
        return ", ".join(str(item) for item in localized if str(item))
    text = str(value)
    if text in {"UNDER_TARGET", "ON_TARGET", "OVER_TARGET", "UNKNOWN_TARGET_STATUS"}:
        return get_status_label(text)
    strategy_label = get_strategy_label(text)
    if strategy_label != text:
        return strategy_label
    return DISPLAY_VALUE_LABELS.get(text, value)


def _operation_mode_label(target_status: object) -> object:
    if _is_missing(target_status):
        return "계산 불가"
    return get_operation_mode(target_status)


def _format_date(value: object) -> str:
    if _is_missing(value):
        return "계산 불가"
    try:
        return str(pd.Timestamp(value).date())
    except Exception:  # noqa: BLE001 - display only.
        return str(value)


def _as_dataframe(value: Any) -> pd.DataFrame:
    if isinstance(value, pd.DataFrame):
        return value
    if value is None:
        return pd.DataFrame()
    return pd.DataFrame(value)


def _as_series(value: Any) -> pd.Series:
    if isinstance(value, pd.Series):
        return value
    if isinstance(value, Mapping):
        return pd.Series(dict(value))
    return pd.Series(dtype=object)


def _as_float(value: object) -> float:
    if _is_missing(value):
        return float("nan")
    try:
        return float(value)
    except (TypeError, ValueError):
        return float("nan")


def _is_missing(value: object) -> bool:
    if value is None:
        return True
    try:
        return bool(pd.isna(value))
    except (TypeError, ValueError):
        return False


if __name__ == "__main__":
    main()

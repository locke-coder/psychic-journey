from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
import pytest
import yaml
from openpyxl import load_workbook

import app
from app import (
    HISTORY_TAB_LABEL,
    SAMPLE_INPUT_PATH,
    build_runtime_config,
    build_scenario_chart_data,
    build_scenario_value_matrix,
    calculate_validated_results,
    default_as_of_date,
    load_history_tables_for_app,
    run_selected_scenario_detail,
)
from src import history_schema
from src.backtest_engine import build_backtest_dataset, summarize_by_forecast_model
from src.excel_exporter import SHEET_NAMES, export_daily_report
from src.final_actual_store import (
    build_final_actual_record,
    load_final_actuals,
    upsert_final_actual,
)
from src.history_store import (
    append_forecast_history,
    build_forecast_history_rows,
    load_forecast_history,
)
from src.loader import load_input
from src.schema import load_model_config
from tools import gate_runner


def test_history_backtest_export_app_e2e_flow(tmp_path: Path) -> None:
    input_df = load_input(SAMPLE_INPUT_PATH)
    config = build_runtime_config(load_model_config(), 1.30, 1.50)
    as_of_date = default_as_of_date(input_df, "sales", today="2026-06-11")

    results = calculate_validated_results(input_df, as_of_date, "sales", config)
    assert results["validation"]["errors"] == []
    scenario_df = results["scenario_df"]
    assert scenario_df.shape[0] == 9
    assert {"F1_CUMULATIVE_RATE", "F2_LAST_TWO_CLOSES", "F3_DAY_CLOSE_WEIGHTED"} <= set(
        scenario_df["forecast_model"]
    )

    forecast_history_path = tmp_path / "history" / "forecast_history.csv"
    final_actuals_path = tmp_path / "history" / "final_actuals.csv"
    run_context = {
        "run_id": "E2E-HISTORY-BACKTEST-RUN",
        "run_datetime": datetime(2026, 6, 11, 9, 0),
        "target_month": "2026-06",
        "as_of_date": as_of_date,
        "metric": "sales",
    }
    history_rows = build_forecast_history_rows(scenario_df, run_context)
    saved_history = append_forecast_history(history_rows, forecast_history_path)

    assert forecast_history_path.is_file()
    assert saved_history.shape[0] == scenario_df.shape[0]
    assert tuple(saved_history.columns) == history_schema.FORECAST_HISTORY_COLUMNS

    monthly_target = float(scenario_df["monthly_target"].iloc[0])
    final_actual = round(monthly_target * 1.01, 1)
    final_record = build_final_actual_record(
        target_month="2026-06",
        metric="sales",
        final_actual=final_actual,
        monthly_target=monthly_target,
        updated_at="2026-07-01T09:00:00",
    )
    upsert_final_actual(final_record, final_actuals_path)

    assert final_actuals_path.is_file()
    loaded_history = load_forecast_history(forecast_history_path)
    loaded_actuals = load_final_actuals(final_actuals_path)
    assert loaded_actuals.loc[0, "final_actual"] == pytest.approx(final_actual)

    backtest_df = build_backtest_dataset(loaded_history, loaded_actuals)
    model_summary = summarize_by_forecast_model(backtest_df)

    assert backtest_df.shape[0] == scenario_df.shape[0]
    assert {"forecast_error", "abs_error", "error_rate", "signed_error_rate"} <= set(
        backtest_df.columns
    )
    first_row = backtest_df.iloc[0]
    expected_error_rate = abs(first_row["forecast_amount"] - final_actual) / final_actual
    assert first_row["error_rate"] == pytest.approx(expected_error_rate)
    assert set(model_summary["forecast_model"]) == {
        "F1_CUMULATIVE_RATE",
        "F2_LAST_TWO_CLOSES",
        "F3_DAY_CLOSE_WEIGHTED",
    }
    assert model_summary["sample_count"].tolist() == [3, 3, 3]

    scenario_chart_df = build_scenario_chart_data(scenario_df)
    scenario_value_matrix = build_scenario_value_matrix(scenario_df)
    assert scenario_chart_df.shape[0] == scenario_df.shape[0]
    assert {"forecast_amount", "gap_to_target", "surplus_to_target"} <= set(
        scenario_chart_df.columns
    )
    assert scenario_value_matrix.shape == (3, 3)

    _, selected_detail = run_selected_scenario_detail(
        input_df,
        as_of_date,
        "sales",
        "F1_P1",
        config,
    )
    report_path = export_daily_report(
        tmp_path / "e2e_history_backtest.xlsx",
        {"metric": "sales", "monthly_target": monthly_target},
        scenario_df,
        selected_detail["allocation_by_day"],
        results["close_cycle_df"],
        results["validation"],
        "E2E history backtest report",
        forecast_history_df=loaded_history,
        final_actuals_df=loaded_actuals,
        backtest_summary_df=model_summary,
    )

    workbook = load_workbook(report_path)
    assert workbook.sheetnames == list(SHEET_NAMES)
    assert workbook["ForecastHistory"].max_row == len(loaded_history) + 1
    assert workbook["FinalActuals"].max_row == len(loaded_actuals) + 1
    assert workbook["BacktestSummary"].max_row == len(model_summary) + 1
    assert workbook["ModelWeights"].max_row >= 2
    assert workbook["Insights"].max_row >= 2

    imported_app = app
    assert imported_app.HISTORY_TAB_LABEL == HISTORY_TAB_LABEL
    assert HISTORY_TAB_LABEL == "예측 이력 / Backtest"


def test_history_ui_handles_missing_storage_files_safely(tmp_path: Path) -> None:
    tables = load_history_tables_for_app(
        tmp_path / "missing_forecast_history.csv",
        tmp_path / "missing_final_actuals.csv",
    )

    assert tables["forecast_history"].empty
    assert tables["final_actuals"].empty
    assert tuple(tables["forecast_history"].columns) == history_schema.FORECAST_HISTORY_COLUMNS
    assert tuple(tables["final_actuals"].columns) == history_schema.FINAL_ACTUALS_COLUMNS


def test_new_history_backtest_gates_are_declared() -> None:
    catalog_path = Path(__file__).resolve().parents[1] / "config" / "gate_audit_catalog.yaml"
    catalog = yaml.safe_load(catalog_path.read_text(encoding="utf-8"))
    gates = {gate["gate_id"]: gate for gate in catalog["gates"]}

    assert {"G19", "G20", "G21", "G22"} <= set(gates)
    assert {
        "src/history_store.py",
        "src/final_actual_store.py",
        "tests/test_e2e_history_backtest.py",
    } <= set(gates["G19"]["required_files"])
    assert {
        "src/backtest_engine.py",
        "tests/test_backtest_engine.py",
        "tests/test_e2e_history_backtest.py",
    } <= set(gates["G20"]["required_files"])
    assert {
        "src/visualization_builder.py",
        "src/model_weight_engine.py",
        "src/confidence_band.py",
        "src/insight_engine.py",
    } <= set(gates["G21"]["required_files"])
    assert {
        "app.py",
        "tests/test_app_smoke.py",
        "tests/test_e2e_history_backtest.py",
    } <= set(gates["G22"]["required_files"])
    assert {"G18", "G19", "G20", "G21", "G22"} <= set(
        gates["ALL"].get("required_gates", [])
    )


@pytest.mark.parametrize("gate_id", ["G19", "G20", "G22"])
def test_new_implemented_gates_pass_static_checks_without_pytest(gate_id: str) -> None:
    result = gate_runner.run_gate(gate_id, execute_pytest=False)

    assert result["status"] == "PASS"
    assert result["required_files_missing"] == []
    assert result["required_keywords_missing"] == []
    assert result["errors"] == []

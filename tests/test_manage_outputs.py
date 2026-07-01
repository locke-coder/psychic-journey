from pathlib import Path

from openpyxl import Workbook, load_workbook

from tools import manage_outputs


def _write_latest_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.active.title = "Summary"
    for sheet_name in sorted(manage_outputs.REQUIRED_LATEST_SHEETS - {"Summary"}):
        workbook.create_sheet(sheet_name)
    workbook.save(path)
    workbook.close()


def _write_old_format_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.active.title = "Summary"
    workbook.save(path)
    workbook.close()


def test_valid_latest_xlsx_is_classified_as_latest(tmp_path: Path) -> None:
    outputs_root = tmp_path / "outputs"
    workbook_path = outputs_root / "daily_report_sales.xlsx"
    _write_latest_workbook(workbook_path)

    result = manage_outputs.classify_workbook(workbook_path, outputs_root)

    assert result["status"] == manage_outputs.STATUS_LATEST
    assert result["advanced_sheet_count"] == 0


def test_valid_old_format_xlsx_is_classified_as_old_format(tmp_path: Path) -> None:
    outputs_root = tmp_path / "outputs"
    workbook_path = outputs_root / "daily_report_old.xlsx"
    _write_old_format_workbook(workbook_path)

    result = manage_outputs.classify_workbook(workbook_path, outputs_root)

    assert result["status"] == manage_outputs.STATUS_OLD_FORMAT
    assert "missing latest report sheets" in result["reason"]


def test_legacy_named_report_is_classified_as_old_format(tmp_path: Path) -> None:
    outputs_root = tmp_path / "outputs"
    workbook_path = outputs_root / "daily_report_sales_20260604.xlsx"
    _write_latest_workbook(workbook_path)

    result = manage_outputs.classify_workbook(workbook_path, outputs_root)

    assert result["status"] == manage_outputs.STATUS_OLD_FORMAT
    assert result["reason"] == "legacy output filename pattern"


def test_invalid_xlsx_is_classified_as_invalid(tmp_path: Path) -> None:
    outputs_root = tmp_path / "outputs"
    workbook_path = outputs_root / "broken.xlsx"
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    workbook_path.write_text("not a workbook", encoding="utf-8")

    result = manage_outputs.classify_workbook(workbook_path, outputs_root)

    assert result["status"] == manage_outputs.STATUS_INVALID
    assert "BadZipFile" in result["reason"]


def test_dry_run_does_not_move_files(tmp_path: Path) -> None:
    outputs_root = tmp_path / "outputs"
    workbook_path = outputs_root / "daily_report_sales.xlsx"
    _write_latest_workbook(workbook_path)

    result = manage_outputs.organize_outputs(outputs_root, dry_run=True)

    assert result["status"] == "DRY_RUN"
    assert result["moves"] == [
        {
            "source": "daily_report_sales.xlsx",
            "destination": "latest/daily_report_sales.xlsx",
            "status": manage_outputs.STATUS_LATEST,
        }
    ]
    assert workbook_path.is_file()
    assert not (outputs_root / "latest" / "daily_report_sales.xlsx").exists()


def test_organize_creates_managed_output_dirs(tmp_path: Path) -> None:
    outputs_root = tmp_path / "outputs"
    _write_latest_workbook(outputs_root / "daily_report_sales.xlsx")

    manage_outputs.organize_outputs(outputs_root)

    assert (outputs_root / "latest" / ".gitkeep").is_file()
    assert (outputs_root / "archive_old_format" / ".gitkeep").is_file()
    assert (outputs_root / "archive_invalid" / ".gitkeep").is_file()
    assert (outputs_root / "latest" / "daily_report_sales.xlsx").is_file()


def test_organize_adds_suffix_when_destination_exists(tmp_path: Path) -> None:
    outputs_root = tmp_path / "outputs"
    existing_path = outputs_root / "latest" / "daily_report_sales.xlsx"
    source_path = outputs_root / "daily_report_sales.xlsx"
    _write_latest_workbook(existing_path)
    _write_latest_workbook(source_path)

    result = manage_outputs.organize_outputs(outputs_root)

    assert existing_path.is_file()
    assert (outputs_root / "latest" / "daily_report_sales_1.xlsx").is_file()
    assert {
        "source": "daily_report_sales.xlsx",
        "destination": "latest/daily_report_sales_1.xlsx",
        "status": manage_outputs.STATUS_LATEST,
    } in result["moves"]


def test_invalid_input_template_is_archived_and_recreated(tmp_path: Path) -> None:
    outputs_root = tmp_path / "outputs"
    invalid_template = (
        outputs_root
        / "input_forms_20260602"
        / manage_outputs.INPUT_TEMPLATE_NAME
    )
    invalid_template.parent.mkdir(parents=True, exist_ok=True)
    invalid_template.write_text("not a workbook", encoding="utf-8")

    result = manage_outputs.organize_outputs(outputs_root)
    new_template = outputs_root / "latest" / manage_outputs.INPUT_TEMPLATE_NAME

    assert result["created_template"] == f"latest/{manage_outputs.INPUT_TEMPLATE_NAME}"
    assert not invalid_template.exists()
    assert (
        outputs_root
        / "archive_invalid"
        / manage_outputs.INPUT_TEMPLATE_NAME
    ).is_file()
    assert manage_outputs.is_valid_input_template_file(new_template)

    workbook = load_workbook(new_template, read_only=True, data_only=True)
    try:
        headers = [
            cell
            for cell in next(
                workbook[workbook.sheetnames[0]].iter_rows(
                    min_row=1,
                    max_row=1,
                    values_only=True,
                )
            )
            if cell is not None
        ]
    finally:
        workbook.close()
    assert set(manage_outputs.INPUT_TEMPLATE_HEADERS).issubset(headers)

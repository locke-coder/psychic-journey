import io
import inspect
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

import app as app_module
from app import (
    ACTUAL_CUM_COLUMNS,
    DIRECT_EDITABLE_COLUMNS,
    HISTORICAL_INPUT_TEMPLATE_FILENAME,
    HISTORICAL_SAMPLE_INPUT_PATH,
    HISTORY_TAB_LABEL,
    INPUT_TEMPLATE_FILENAME,
    INPUT_TEMPLATE_HEADERS,
    SAMPLE_INPUT_SOURCE_LABEL,
    apply_latest_upload_policy,
    apply_saved_actuals,
    build_auto_axis_domain,
    build_close_cycle_chart_data,
    build_forecast_definition_df,
    build_grouped_bar_chart_source,
    build_historical_context,
    build_historical_forecast_axis_domain,
    build_historical_forecast_decision_summary,
    build_historical_forecast_comparison,
    build_historical_monthly_summary,
    build_historical_progress_chart_data,
    build_historical_stage_benchmark,
    build_neutral_definition_df,
    build_kpi_rows,
    build_overachievement_definition_df,
    build_remaining_operation_direction_source,
    build_remaining_target_chart_data,
    build_remaining_target_daily_source,
    build_remaining_target_stack_source,
    build_report_glossary_df,
    build_provision_definition_df,
    build_risk_definition_df,
    build_runtime_config,
    build_display_validation_result,
    build_historical_input_template_bytes,
    build_input_template_bytes,
    build_summary_dict,
    build_selected_scenario_explanation,
    build_selected_scenario_daily_detail_source,
    build_scenario_chart_data,
    build_scenario_daily_forecast_source,
    build_scenario_heatmap_source,
    build_scenario_matrix,
    build_scenario_operation_matrix,
    build_scenario_target_position_source,
    build_scenario_value_matrix,
    build_scenario_weekly_forecast_source,
    build_strategy_level_chart_data,
    build_strategy_effect_table,
    build_strategy_level_table,
    build_visual_decision_summary,
    build_visual_headline,
    build_visual_metric_definition_df,
    build_visual_reading_guide,
    calculate_validated_results,
    chart_value_format,
    clear_saved_actuals,
    default_as_of_date,
    format_validation_message,
    format_scenario_option_label,
    load_saved_actuals,
    load_history_tables_for_app,
    normalize_direct_input_edits,
    run_selected_scenario_detail,
    save_actual_values,
)
from src import history_schema
from src.loader import load_input
from src.overachievement_models import OVERACHIEVEMENT
from src.schema import load_model_config
from src.visualization_builder import build_strategy_arrival_compare_source


SAMPLE_INPUT_PATH = Path(__file__).resolve().parent / "fixtures" / "input_sample_2026_06.csv"


def _over_target_input_df() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "date": pd.to_datetime(
                [
                    "2026-06-01",
                    "2026-06-02",
                    "2026-06-03",
                    "2026-06-04",
                    "2026-06-05",
                    "2026-06-06",
                    "2026-06-07",
                    "2026-06-08",
                    "2026-06-09",
                ]
            ),
            "day_name": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun", "Mon", "Tue"],
            "business_day_no": list(range(1, 10)),
            "is_close_day": ["N", "Y", "N", "Y", "N", "Y", "N", "Y", "N"],
            "close_type": ["", "first", "", "mid", "", "final", "", "next", ""],
            "sales_target_daily": [
                10.0,
                20.0,
                10.0,
                20.0,
                10.0,
                20.0,
                10.0,
                20.0,
                10.0,
            ],
            "recognized_target_daily": [1.0] * 9,
            "sales_actual_cum": [
                20.0,
                45.0,
                70.0,
                95.0,
                120.0,
                145.0,
                170.0,
                pd.NA,
                pd.NA,
            ],
            "recognized_actual_cum": [
                1.0,
                2.0,
                3.0,
                4.0,
                5.0,
                6.0,
                7.0,
                pd.NA,
                pd.NA,
            ],
            "memo": [""] * 9,
        }
    )


def _under_target_input_df() -> pd.DataFrame:
    df = _over_target_input_df()
    df["sales_actual_cum"] = [2.0, 4.0, 6.0, 8.0, 10.0, 12.0, 14.0, pd.NA, pd.NA]
    return df


def test_app_sample_smoke_runs_validation_and_scenarios() -> None:
    df = load_input(SAMPLE_INPUT_PATH)
    config = build_runtime_config(load_model_config(), 1.30, 1.50)
    as_of_date = default_as_of_date(df, "sales", today="2026-06-11")

    results = calculate_validated_results(df, as_of_date, "sales", config)

    assert as_of_date == pd.Timestamp("2026-06-10")
    assert results["validation"]["errors"] == []
    assert results["scenario_df"].shape[0] == 9
    assert build_scenario_matrix(results["scenario_df"]).shape == (3, 3)

    _, provision_result = run_selected_scenario_detail(
        df,
        as_of_date,
        "sales",
        "F1_P1",
        config,
    )

    assert "allocation_by_day" in provision_result
    assert not provision_result["allocation_by_day"].empty


def test_app_main_requires_password_before_rendering_data_pages() -> None:
    app_source = Path("app.py").read_text(encoding="utf-8")
    navigation_source = Path("src/ui_navigation.py").read_text(encoding="utf-8")
    main_source = inspect.getsource(app_module.main)

    assert "마감 페이스 체크" in app_source
    assert "입력 · 데이터" in navigation_source
    assert "forecast_strategy" in navigation_source
    assert "예측 · 전략 통합" in navigation_source
    assert '"forecast": "forecast_strategy"' in navigation_source
    assert '"scenarios": "forecast_strategy"' in navigation_source
    assert "KPI · 예측" not in navigation_source
    assert '"title": "시나리오"' not in navigation_source
    assert "_render_forecast_strategy_detail_page" in app_source
    assert '"render_forecast_strategy_page"' in app_source
    assert "_render_forecast_detail_page" in app_source
    assert "_render_scenarios_detail_page" in app_source
    assert "if not _require_access_password():" in main_source
    assert main_source.index("_require_access_password()") < main_source.index(
        "base_config = load_model_config()"
    )
    assert "st.text_input(\"접속 비밀번호\"" not in main_source


def test_input_template_download_workbook_has_required_headers() -> None:
    workbook = load_workbook(io.BytesIO(build_input_template_bytes()), read_only=True)
    try:
        worksheet = workbook.active
        headers = [cell.value for cell in next(worksheet.iter_rows(max_row=1))]
        example_row = [cell.value for cell in next(worksheet.iter_rows(min_row=2, max_row=2))]

        assert INPUT_TEMPLATE_FILENAME.endswith(".xlsx")
        assert tuple(headers) == INPUT_TEMPLATE_HEADERS
        assert example_row[0] == "YYYY-MM-DD"
        assert example_row[2] == 1
    finally:
        workbook.close()


def test_historical_input_template_download_workbook_has_monthly_example_rows() -> None:
    sample_df = pd.read_csv(HISTORICAL_SAMPLE_INPUT_PATH)
    workbook = load_workbook(
        io.BytesIO(build_historical_input_template_bytes()),
        read_only=True,
        data_only=True,
    )
    try:
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))

        assert HISTORICAL_INPUT_TEMPLATE_FILENAME.endswith(".xlsx")
        assert rows[0] == INPUT_TEMPLATE_HEADERS
        assert len(rows) == len(sample_df) + 1
        assert rows[1][2] == 1
        assert rows[1][0] == sample_df.loc[0, "date"]
        assert rows[-1][0] == sample_df.iloc[-1]["date"]
        assert "전산침해 이슈" in {row[9] for row in rows[1:]}
    finally:
        workbook.close()


def test_app_validation_errors_stop_before_calculation() -> None:
    df = load_input(SAMPLE_INPUT_PATH)
    df["is_close_day"] = False
    config = build_runtime_config(load_model_config(), 1.30, 1.50)

    results = calculate_validated_results(df, "2026-06-10", "sales", config)

    assert results["validation"]["errors"]
    assert "scenario_df" not in results


def test_default_as_of_date_prefers_previous_input_business_day() -> None:
    df = load_input(SAMPLE_INPUT_PATH)

    as_of_date = default_as_of_date(df, "sales", today="2026-06-04")

    assert as_of_date == pd.Timestamp("2026-06-02")


def test_default_as_of_date_uses_previous_input_date_when_today_is_absent() -> None:
    df = load_input(SAMPLE_INPUT_PATH)

    as_of_date = default_as_of_date(df, "sales", today="2026-06-03")

    assert as_of_date == pd.Timestamp("2026-06-02")


def test_default_as_of_date_uses_previous_input_business_day_even_without_actual() -> None:
    df = load_input(SAMPLE_INPUT_PATH)

    as_of_date = default_as_of_date(df, "sales", today="2026-06-18")

    assert as_of_date == pd.Timestamp("2026-06-17")


def test_as_of_date_default_refreshes_daily_but_preserves_same_day_choice(
    monkeypatch,
) -> None:
    class FakeStreamlit:
        def __init__(self) -> None:
            self.session_state: dict[str, object] = {}

    fake_st = FakeStreamlit()
    df = load_input(SAMPLE_INPUT_PATH)
    config = load_model_config()
    app_date = pd.Timestamp("2026-06-18").date()

    monkeypatch.setattr(app_module, "st", fake_st)
    monkeypatch.setattr(app_module, "_current_app_date", lambda today=None: app_date)

    _, as_of_date, *_ = app_module._normalize_app_settings(df, config)
    assert as_of_date == pd.Timestamp("2026-06-17")

    fake_st.session_state[app_module.PACE_AS_OF_DATE_SESSION_KEY] = pd.Timestamp("2026-06-10").date()
    _, as_of_date, *_ = app_module._normalize_app_settings(df, config)
    assert as_of_date == pd.Timestamp("2026-06-10")

    app_date = pd.Timestamp("2026-06-19").date()
    _, as_of_date, *_ = app_module._normalize_app_settings(df, config)
    assert as_of_date == pd.Timestamp("2026-06-18")


def test_default_as_of_date_uses_first_input_date_when_no_prior_business_day() -> None:
    df = load_input(SAMPLE_INPUT_PATH)

    as_of_date = default_as_of_date(df, "sales", today="2026-06-01")

    assert as_of_date == pd.Timestamp("2026-06-01")


def test_app_builds_visual_chart_data_from_sample_results() -> None:
    df = load_input(SAMPLE_INPUT_PATH)
    config = build_runtime_config(load_model_config(), 1.30, 1.50)
    as_of_date = default_as_of_date(df, "sales", today="2026-06-11")

    results = calculate_validated_results(df, as_of_date, "sales", config)
    _, provision_result = run_selected_scenario_detail(
        df,
        as_of_date,
        "sales",
        "F1_P1",
        config,
    )

    scenario_chart_data = build_scenario_chart_data(results["scenario_df"])
    scenario_value_matrix = build_scenario_value_matrix(results["scenario_df"])
    remaining_target_chart_data = build_remaining_target_chart_data(
        provision_result["allocation_by_day"]
    )
    close_cycle_chart_data = build_close_cycle_chart_data(results["close_cycle_df"])

    assert scenario_chart_data.shape[0] == 9
    assert {
        "target_variance",
        "gap_to_target",
        "surplus_to_target",
        "revised_monthly_target",
    }.issubset(scenario_chart_data.columns)
    assert scenario_value_matrix.shape == (3, 3)
    assert scenario_value_matrix.notna().all().all()
    assert remaining_target_chart_data.index[0] == "2026-06-11"
    assert "revised_target" in remaining_target_chart_data
    assert "achievement_rate" in close_cycle_chart_data


def test_app_builds_intuitive_visual_sources_from_sample_results() -> None:
    df = _under_target_input_df()
    config = build_runtime_config(load_model_config(), 1.30, 1.50)
    as_of_date = pd.Timestamp("2026-06-07")

    results = calculate_validated_results(df, as_of_date, "sales", config)
    forecast_result, provision_result = run_selected_scenario_detail(
        df,
        as_of_date,
        "sales",
        "F3_P1",
        config,
    )

    target_source = build_scenario_target_position_source(results["scenario_df"], "F3_P1")
    heatmap_source = build_scenario_heatmap_source(results["scenario_df"], "F3_P1")
    daily_forecast_source = build_scenario_daily_forecast_source(
        df,
        results["scenario_df"],
        as_of_date,
        "sales",
        config,
        "F3_P1",
    )
    daily_source = build_remaining_target_daily_source(provision_result["allocation_by_day"])
    stack_source = build_remaining_target_stack_source(provision_result["allocation_by_day"])
    weekly_forecast_source = build_scenario_weekly_forecast_source(daily_forecast_source)
    selected_detail_source = build_selected_scenario_daily_detail_source(
        daily_forecast_source,
        "F3_P1",
    )
    selected_final = daily_forecast_source.loc[
        (daily_forecast_source["scenario_id"] == "F3_P1")
        & (daily_forecast_source["series_type"] == "시나리오 예상")
    ].sort_values("date").iloc[-1]
    selected_row = results["scenario_df"].loc[
        results["scenario_df"]["scenario_id"] == "F3_P1"
    ].iloc[0]

    assert not forecast_result["warnings"]
    assert target_source["is_selected"].sum() == 1
    assert {"target_status_label", "variance_label", "forecast_label"}.issubset(
        target_source.columns
    )
    assert {"forecast_key", "strategy_key", "target_variance"}.issubset(
        heatmap_source.columns
    )
    assert {"확정 실적", "시나리오 예상"}.issubset(set(daily_forecast_source["series_type"]))
    assert set(daily_forecast_source.loc[daily_forecast_source["is_close_day"], "day_type"]) == {
        "마감일"
    }
    assert {"week_start", "week_end", "week_label"}.issubset(weekly_forecast_source.columns)
    assert weekly_forecast_source["week_start"].nunique() < daily_forecast_source["date"].nunique()
    assert selected_detail_source["scenario_id"].isin(["확정 실적", "F3_P1"]).all()
    assert abs(selected_final["forecast_cum"] - selected_row["forecast_after_provision"]) < 1e-9
    assert selected_final["achievement_rate"] == (
        selected_final["forecast_cum"] / selected_final["monthly_target"]
    )
    assert set(daily_source["day_type"]) == {"마감일", "일반일"}
    first_day = daily_source.iloc[0]
    first_stack_total = stack_source.loc[
        stack_source["date_label"] == first_day["date_label"],
        "value",
    ].sum()
    assert abs(first_stack_total - first_day["revised_target"]) < 1e-9


def test_scenario_progress_chart_keeps_daily_line_points() -> None:
    df = load_input(SAMPLE_INPUT_PATH)
    config = build_runtime_config(load_model_config(), 1.30, 1.50)
    as_of_date = default_as_of_date(df, "sales", today="2026-06-11")
    results = calculate_validated_results(df, as_of_date, "sales", config)
    daily_forecast_source = build_scenario_daily_forecast_source(
        df,
        results["scenario_df"],
        as_of_date,
        "sales",
        config,
        "F1_P1",
    )

    class FakeStreamlit:
        def __init__(self) -> None:
            self.chart_specs: list[dict[str, object]] = []

        def info(self, *_args, **_kwargs) -> None:
            return None

        def caption(self, *_args, **_kwargs) -> None:
            return None

        def altair_chart(self, chart, *_args, **_kwargs) -> None:
            self.chart_specs.append(chart.to_dict())

    fake_st = FakeStreamlit()
    original_st = app_module.st
    app_module.st = fake_st
    try:
        app_module._render_scenario_daily_forecast_chart(
            daily_forecast_source,
            results["scenario_df"],
            "F1_P1",
        )
    finally:
        app_module.st = original_st

    assert fake_st.chart_specs
    line_x_fields = {
        layer["encoding"]["x"]["field"]
        for layer in fake_st.chart_specs[0]["layer"]
        if layer.get("mark", {}).get("type") == "line"
    }
    assert line_x_fields == {"date"}


def test_forecast_model_view_filters_scenarios_and_builds_operation_direction() -> None:
    df = load_input(SAMPLE_INPUT_PATH)
    config = build_runtime_config(load_model_config(), 1.30, 1.50)
    as_of_date = default_as_of_date(df, "sales", today="2026-06-11")
    results = calculate_validated_results(df, as_of_date, "sales", config)
    f1_scenarios = results["scenario_df"].loc[
        results["scenario_df"]["scenario_id"].astype(str).str.startswith("F1_")
    ]

    daily_forecast_source = build_scenario_daily_forecast_source(
        df,
        f1_scenarios,
        as_of_date,
        "sales",
        config,
        "F1_P1",
    )
    forecast_result, strategy_result = run_selected_scenario_detail(
        df,
        as_of_date,
        "sales",
        "F1_P1",
        config,
    )
    direction_source = build_remaining_operation_direction_source(
        df,
        as_of_date,
        "sales",
        "F1_P1",
        forecast_result,
        strategy_result,
    )

    assert set(daily_forecast_source["scenario_id"]) <= {
        "확정 실적",
        *set(f1_scenarios["scenario_id"]),
    }
    assert all(
        scenario_id == "확정 실적" or str(scenario_id).startswith("F1_")
        for scenario_id in daily_forecast_source["scenario_id"].unique()
    )
    assert not direction_source.empty
    assert {"추가 배분", "기존 목표 유지"}.intersection(set(direction_source["direction"]))
    assert (direction_source["scenario_id"] == "F1_P1").all()


def test_kpi_rows_follow_requested_operating_layout() -> None:
    scenario_df = pd.DataFrame(
        {
            "scenario_id": ["F1_P1", "F2_P1", "F3_P1"],
            "forecast_amount": [95.0, 98.0, 102.0],
        }
    )

    rows = build_kpi_rows(
        {
            "monthly_target": 100.0,
            "current_target_cum": 70.0,
            "current_actual_cum": 72.0,
        },
        scenario_df,
        {
            "next_close_date": pd.Timestamp("2026-06-15"),
            "required_to_recover_next_close_cum": 8.0,
        },
        pd.Series(
            {
                "target_status": "OVER_TARGET",
                "target_variance": 2.0,
                "surplus_to_target": 2.0,
                "risk_level": "Green",
            }
        ),
    )

    assert [[label for label, _ in row] for row in rows] == [
        ["월 목표", "기준일 누적 목표", "기준일 누적 실적", "누적 달성률"],
        ["F1예상", "F2예상", "F3예상"],
        ["다음 마감일", "다음 마감 누적선 필요실적"],
        ["목표상태", "목표대비 차이", "초과 예상분", "위험등급", "운영모드"],
    ]
    assert rows[1][0][1] == "95.0억 원"
    assert rows[3][0][1] == "초과달성 관리"
    assert rows[3][4][1] == "초과달성 관리"


def test_over_target_strategy_level_table_replaces_blank_daily_allocation() -> None:
    df = _over_target_input_df()
    config = build_runtime_config(load_model_config(), 10.0, 10.0)
    results = calculate_validated_results(df, "2026-06-07", "sales", config)
    _, strategy_result = run_selected_scenario_detail(
        df,
        "2026-06-07",
        "sales",
        "F1_O2",
        config,
    )

    strategy_table = build_strategy_level_table(results["scenario_df"], "F1_O2")
    strategy_chart_data = build_strategy_level_chart_data(
        results["scenario_df"],
        "F1_O2",
    )

    assert strategy_result["allocation_by_day"].empty
    assert strategy_table["scenario_id"].tolist() == ["F1_O1", "F1_O2", "F1_O3"]
    assert list(strategy_table.columns[:9]) == [
        "scenario_id",
        "forecast_model",
        "target_status",
        "provision_strategy",
        "strategy_difference_summary",
        "stretch_uplift",
        "revised_monthly_target",
        "remaining_surplus_buffer",
        "minimum_remaining_to_hit_target",
    ]
    assert strategy_table["forecast_model"].tolist() == ["F1", "F1", "F1"]
    assert strategy_table["strategy_type"].eq(OVERACHIEVEMENT).all()
    assert (strategy_table["surplus_to_target"] > 0).all()
    assert strategy_table.loc[
        strategy_table["scenario_id"].eq("F1_O2"),
        "strategy_difference_summary",
    ].iloc[0].endswith("Stretch 목표로 전환, 운영 월 목표 175.5억 원")
    assert strategy_table.loc[
        strategy_table["scenario_id"].eq("F1_O3"),
        "strategy_difference_summary",
    ].iloc[0].startswith("목표 달성 최소 잔여")
    assert "recommended_action" in strategy_table.columns
    assert strategy_chart_data.index.tolist() == ["F1_O1", "F1_O2", "F1_O3"]
    assert "forecast_after_provision" not in strategy_chart_data.columns
    assert "revised_monthly_target" in strategy_chart_data.columns
    assert "remaining_surplus_buffer" in strategy_chart_data.columns


def test_scenario_operation_matrix_keeps_nine_rows_and_display_labels() -> None:
    df = _over_target_input_df()
    config = build_runtime_config(load_model_config(), 10.0, 10.0)
    results = calculate_validated_results(df, "2026-06-07", "sales", config)

    matrix = build_scenario_operation_matrix(results["scenario_df"], "F1_O2")

    assert matrix.shape[0] == 9
    assert {
        "scenario",
        "forecast_model",
        "target_status",
        "strategy_code",
        "strategy_label",
        "strategy_group",
        "expected_month_end_amount",
        "target_variance",
        "surplus_to_target",
        "recommended_action",
        "risk_note",
    }.issubset(matrix.columns)
    assert {"O1", "O2", "O3"}.issubset(set(matrix["strategy_code"]))
    assert {"버퍼 유지", "Stretch 전환", "품질 방어"}.issubset(
        set(matrix["strategy_label"])
    )
    assert matrix.loc[matrix["scenario"] == "F1_O2", "recommended"].iloc[0] == "추천"

    under_results = calculate_validated_results(
        _under_target_input_df(),
        "2026-06-07",
        "sales",
        config,
    )
    under_matrix = build_scenario_operation_matrix(under_results["scenario_df"], "F1_P1")
    assert under_matrix.shape[0] == 9
    assert {"P1", "P2", "P3"}.issubset(set(under_matrix["strategy_code"]))
    assert {"잔여목표 균등 배분", "마감일 집중 보정", "비마감일 분산 보정"}.issubset(
        set(under_matrix["strategy_label"])
    )


def test_strategy_level_table_can_match_all_forecast_model_rows() -> None:
    df = _over_target_input_df()
    config = build_runtime_config(load_model_config(), 10.0, 10.0)
    results = calculate_validated_results(df, "2026-06-07", "sales", config)

    strategy_table = build_strategy_level_table(results["scenario_df"], None)

    assert strategy_table["forecast_model"].tolist() == [
        "F1",
        "F1",
        "F1",
        "F2",
        "F2",
        "F2",
        "F3",
        "F3",
        "F3",
    ]
    assert strategy_table["scenario_id"].tolist() == [
        "F1_O1",
        "F1_O2",
        "F1_O3",
        "F2_O1",
        "F2_O2",
        "F2_O3",
        "F3_O1",
        "F3_O2",
        "F3_O3",
    ]


def test_strategy_effect_table_holds_forecast_fixed_for_o_strategy_difference() -> None:
    df = _over_target_input_df()
    config = build_runtime_config(load_model_config(), 10.0, 10.0)
    results = calculate_validated_results(df, "2026-06-07", "sales", config)

    effect_table = build_strategy_effect_table(results["scenario_df"], "F1_O2")

    assert effect_table["forecast_basis"].tolist() == ["F1", "F1", "F1"]
    assert effect_table["strategy_key"].tolist() == ["O1", "O2", "O3"]
    assert effect_table["strategy_effect_type"].tolist() == [
        "버퍼 유지",
        "Stretch 전환",
        "품질 방어",
    ]
    assert effect_table.loc[
        effect_table["strategy_key"].eq("O2"),
        "stretch_uplift",
    ].iloc[0] > 0
    assert effect_table.loc[
        effect_table["strategy_key"].eq("O3"),
        "relief_amount",
    ].iloc[0] > 0


def test_strategy_compare_fallback_separates_forecast_from_operating_target(monkeypatch) -> None:
    df = _over_target_input_df()
    config = build_runtime_config(load_model_config(), 10.0, 10.0)
    results = calculate_validated_results(df, "2026-06-07", "sales", config)
    focused = results["scenario_df"].loc[
        results["scenario_df"]["scenario_id"].astype(str).str.startswith("F1_")
    ]
    source = build_strategy_arrival_compare_source(focused, "F1_O2")
    captured: dict[str, object] = {}

    class FakeStreamlit:
        def info(self, message: str) -> None:
            captured["info"] = message

        def caption(self, message: str) -> None:
            captured["caption"] = message

        def dataframe(self, df: pd.DataFrame, **kwargs) -> None:
            captured["df"] = df

    monkeypatch.setattr(app_module, "st", FakeStreamlit())

    app_module._render_strategy_compare_fallback(source)

    display = captured["df"]
    assert "월말 예상 실적을 다시 예측하지 않습니다" in str(captured["info"])
    assert "F예측 월말 예상" in display.columns
    assert "운영 기준 목표" in display.columns
    assert "O전략 차이" in display.columns
    assert "운영 기준 목표 비교 기준값" in display.columns
    assert display.loc[display["전략 코드"].eq("O2"), "O전략 차이"].iloc[0] == "Stretch 전환"
    assert str(display.loc[display["전략 코드"].eq("O2"), "운영 기준 목표"].iloc[0]).endswith("억 원")


def test_scenario_matrix_uses_distinct_overachievement_labels() -> None:
    scenario_df = pd.DataFrame(
        [
            {
                "scenario_id": scenario_id,
                "forecast_after_provision": 170.5,
                "risk_level": "Green",
                "status": "OVER_TARGET_MANAGED",
                "target_status": "OVER_TARGET",
                "strategy_type": OVERACHIEVEMENT,
                "overachievement_strategy": strategy,
            }
            for scenario_id, strategy in [
                ("F1_O1", "O1_TARGET_HOLD_BUFFER"),
                ("F1_O2", "O2_STRETCH_TARGET_CAPTURE"),
                ("F1_O3", "O3_QUALITY_GUARD_RELIEF"),
            ]
        ]
    )

    matrix = build_scenario_matrix(scenario_df)

    assert matrix.loc["F1", "O1"] == "170.5억 원 / 낮음 / 버퍼 유지"
    assert matrix.loc["F1", "O2"] == "170.5억 원 / 낮음 / Stretch 전환"
    assert matrix.loc["F1", "O3"] == "170.5억 원 / 낮음 / 품질 방어"


def test_grouped_bar_chart_source_keeps_metrics_separate_not_stacked() -> None:
    df = load_input(SAMPLE_INPUT_PATH)
    config = build_runtime_config(load_model_config(), 1.30, 1.50)
    _, provision_result = run_selected_scenario_detail(
        df,
        "2026-06-10",
        "sales",
        "F1_P1",
        config,
    )
    chart_data = build_remaining_target_chart_data(provision_result["allocation_by_day"])

    source = build_grouped_bar_chart_source(
        chart_data,
        ("original_target", "uplift", "revised_target"),
    )
    first_category = chart_data.index[0]
    first_values = source.loc[source["category"] == first_category].set_index("metric")[
        "value"
    ]

    assert set(source["metric"]) == {
        "original_target",
        "uplift",
        "revised_target",
    }
    assert first_values["original_target"] == chart_data.loc[first_category, "original_target"]
    assert first_values["uplift"] == chart_data.loc[first_category, "uplift"]
    assert first_values["revised_target"] == chart_data.loc[first_category, "revised_target"]
    assert first_values["revised_target"] < first_values.sum()


def test_auto_axis_domain_zooms_positive_values_and_preserves_zero_crossing() -> None:
    positive_domain = build_auto_axis_domain(pd.Series([100.0, 102.0, 101.0]))
    crossing_domain = build_auto_axis_domain(pd.Series([-2.0, 0.5, 4.0]))
    flat_domain = build_auto_axis_domain(pd.Series([12.0, 12.0]))

    assert positive_domain is not None
    assert positive_domain[0] > 0
    assert positive_domain[0] < 100.0
    assert positive_domain[1] > 102.0
    assert crossing_domain is not None
    assert crossing_domain[0] < 0 < crossing_domain[1]
    assert flat_domain is not None
    assert flat_domain[0] < 12.0 < flat_domain[1]


def test_forecast_model_comparison_chart_uses_target_delta_bars() -> None:
    source = pd.DataFrame(
        [
            {
                "forecast_model": "F1_CUMULATIVE_RATE",
                "label": "F1",
                "value": 150.0,
                "target_status_label": "목표 초과",
                "is_selected_model": False,
            },
            {
                "forecast_model": "F2_LAST_TWO_CLOSES",
                "label": "F2",
                "value": 157.0,
                "target_status_label": "목표 초과",
                "is_selected_model": True,
            },
            {
                "forecast_model": "F3_DAY_CLOSE_WEIGHTED",
                "label": "F3",
                "value": 154.0,
                "target_status_label": "목표 초과",
                "is_selected_model": False,
            },
        ]
    )
    target_source = pd.DataFrame({"value": [149.6], "label": ["목표선"]})
    representative_source = pd.DataFrame({"label": ["F2"], "value": [157.0]})
    scale_source = pd.concat(
        [source[["value"]], target_source[["value"]], representative_source[["value"]]],
        ignore_index=True,
    )

    spec = app_module._build_forecast_model_comparison_chart(
        source,
        target_source,
        representative_source,
        scale_source,
    ).to_dict()
    marks = [layer["mark"]["type"] for layer in spec["layer"]]
    bar_layer = spec["layer"][2]
    y_domain = bar_layer["encoding"]["y"]["scale"]["domain"]

    assert "line" not in marks
    assert {"bar", "point", "rule", "text"}.issubset(set(marks))
    assert bar_layer["encoding"]["y"]["field"] == "target_delta"
    assert bar_layer["encoding"]["y2"]["datum"] == 0
    assert bar_layer["encoding"]["y"]["title"] == "목표 대비 차이"
    assert bar_layer["encoding"]["y"]["axis"]["grid"] is False
    assert y_domain[0] < 0 < y_domain[1]
    assert y_domain[1] > 7.4


def test_home_projection_chart_supports_zoom_and_hover_guides(monkeypatch) -> None:
    source = pd.DataFrame(
        {
            "date": pd.to_datetime(
                ["2026-06-01", "2026-06-02", "2026-06-03", "2026-06-04"]
            ),
            "business_day_no": [1, 2, 3, 4],
            "is_close_day": [False, False, True, False],
            "sales_target_daily": [10.0, 10.0, 10.0, 10.0],
            "sales_actual_cum": [10.0, 21.0, pd.NA, pd.NA],
        }
    )
    projection = app_module.build_pace_projection_chart_data(
        source,
        {"F1": 39.0, "F2": 42.0, "F3": 45.0, "forecast_mid": 42.0},
        "OVER_TARGET",
        current_day_no=2,
    )
    captured: dict[str, object] = {}

    class FakeStreamlit:
        def info(self, message: str) -> None:
            captured["info"] = message

        def altair_chart(self, chart, use_container_width: bool = True) -> None:
            captured["spec"] = chart.to_dict(validate=True)
            captured["use_container_width"] = use_container_width

        def markdown(self, *args, **kwargs) -> None:
            return None

    monkeypatch.setattr(app_module, "st", FakeStreamlit())

    app_module._render_pace_projection_chart(projection)

    spec = captured["spec"]
    params = {param["name"]: param for param in spec["params"]}

    assert {"projection_zoom", "projection_hover", "projection_click"}.issubset(params)
    assert params["projection_zoom"]["select"]["type"] == "interval"
    assert params["projection_zoom"]["bind"] == "scales"
    assert spec["layer"][0]["encoding"]["x"]["axis"]["grid"] is False
    assert spec["layer"][0]["encoding"]["x"]["axis"]["title"] is None
    assert spec["layer"][0]["encoding"]["x"]["axis"]["labelExpr"] == "format(datum.value, 'd') + 'WD'"
    assert spec["layer"][0]["encoding"]["x"]["scale"]["domain"][0] < 1
    assert spec["background"] == "#f7f9fb"
    assert spec["config"]["view"]["fill"] == "#f7f9fb"
    assert spec["layer"][0]["encoding"]["y"]["axis"]["labels"] is False
    assert spec["layer"][0]["encoding"]["y"]["axis"]["grid"] is False
    assert any(
        layer.get("transform", [{}])[0].get("filter", {}).get("param") == "projection_hover"
        for layer in spec["layer"]
    )
    assert any(
        layer.get("transform", [{}])[0].get("filter", {}).get("param") == "projection_click"
        for layer in spec["layer"]
    )


def test_close_cycle_achievement_rate_chart_uses_percent_points_not_ratio() -> None:
    df = load_input(SAMPLE_INPUT_PATH)
    config = build_runtime_config(load_model_config(), 1.30, 1.50)
    results = calculate_validated_results(df, "2026-06-10", "sales", config)
    close_cycle_chart_data = build_close_cycle_chart_data(results["close_cycle_df"])

    source = build_grouped_bar_chart_source(
        close_cycle_chart_data,
        ("achievement_rate",),
    )

    assert close_cycle_chart_data["achievement_rate"].max() > 1.0
    assert source["value"].max() > 1.0
    assert chart_value_format("%") == ",.1f"


def test_visual_metric_definitions_use_business_friendly_terms() -> None:
    definitions = build_visual_metric_definition_df(
        ("forecast_amount", "gap_after_provision", "achievement_rate")
    )

    assert definitions["범례"].tolist() == [
        "월말 예상 실적(보정 전)",
        "보정 후 목표 차이(+부족/-초과)",
        "마감차수 달성률",
    ]
    assert definitions["단위"].tolist() == ["억원", "억원", "%"]
    assert "초과달성 전략에서는 부족분 대신 초과 예상분과 버퍼" in definitions.loc[
        definitions["범례"] == "보정 후 목표 차이(+부족/-초과)",
        "수치 의미",
    ].iloc[0]


def test_visual_reading_guides_cover_rendered_graphs() -> None:
    guide_keys = (
        "scenario_amount",
        "scenario_daily_progress",
        "scenario_target_position",
        "scenario_gap_position",
        "scenario_heatmap",
        "scenario_status",
        "scenario_matrix",
        "target_allocation",
        "target_uplift",
        "target_stack",
        "target_cap",
        "close_cycle_amount",
        "close_cycle_rate",
        "strategy_amount",
        "strategy_buffer",
    )

    for guide_key in guide_keys:
        guide = build_visual_reading_guide(guide_key)
        assert guide["title"]
        assert len(guide["steps"]) >= 3
        assert guide["decision"]

    close_rate_guide = build_visual_reading_guide("close_cycle_rate")
    assert "100%" in close_rate_guide["steps"][0]


def test_visual_decision_summary_leads_with_plain_language() -> None:
    selected_row = pd.Series(
        {
            "target_status": "UNDER_TARGET",
            "risk_level": "Yellow",
            "forecast_after_provision": 92.0,
            "target_variance": -8.0,
            "gap_to_target": 8.0,
        }
    )
    validation_result = {"monthly_target": 100.0}
    next_close_result = {
        "next_close_date": "2026-06-15",
        "required_to_recover_next_close_cum": 6.0,
    }

    headline = build_visual_headline(selected_row, validation_result, next_close_result)
    summary = build_visual_decision_summary(
        selected_row,
        validation_result,
        next_close_result,
    )

    assert "부족" in headline
    assert "잔여 일자별 추가 배분" in headline
    assert list(summary.columns) == ["확인 순서", "볼 것", "현재 값", "해석"]
    summary_text = "\n".join(summary.astype(str).agg(" ".join, axis=1))
    assert "목표 보정 필요" in summary_text
    assert "2026-06-15" in summary_text
    assert "-8.0억 원" in summary_text


def test_historical_monthly_summary_uses_completed_months() -> None:
    historical_df = load_input(HISTORICAL_SAMPLE_INPUT_PATH, sort_by="date")

    summary = build_historical_monthly_summary(historical_df, "sales")

    expected_months = historical_df["date"].dt.to_period("M").astype(str)
    expected_row_counts = expected_months.value_counts(sort=False).tolist()

    assert summary["month"].tolist() == expected_months.drop_duplicates().tolist()
    assert summary["row_count"].tolist() == expected_row_counts
    assert summary["final_achievement_rate"].notna().all()
    assert summary["final_achievement_rate"].max() > 1.0


def test_historical_stage_benchmark_compares_same_business_day() -> None:
    current_df = load_input(SAMPLE_INPUT_PATH)
    historical_df = load_input(HISTORICAL_SAMPLE_INPUT_PATH, sort_by="date")
    config = build_runtime_config(load_model_config(), 1.30, 1.50)
    results = calculate_validated_results(current_df, "2026-06-10", "sales", config)

    benchmark = build_historical_stage_benchmark(
        historical_df,
        current_df,
        "2026-06-10",
        "sales",
        results["validation"],
    )

    assert benchmark["month_count"] == historical_df["date"].dt.to_period("M").nunique()
    assert benchmark["current_business_day_no"] == 7
    assert benchmark["historical_stage_median_rate"] > 0
    assert benchmark["historical_forecast_lower"] <= benchmark["historical_forecast_median"]
    assert benchmark["historical_forecast_median"] <= benchmark["historical_forecast_upper"]
    assert not benchmark["stage_df"].empty


def test_historical_progress_chart_data_includes_current_and_bands() -> None:
    current_df = load_input(SAMPLE_INPUT_PATH)
    historical_df = load_input(HISTORICAL_SAMPLE_INPUT_PATH, sort_by="date")

    chart_data = build_historical_progress_chart_data(
        historical_df,
        current_df,
        "2026-06-10",
        "sales",
    )

    assert {"현재 월", "과거 중앙값", "과거 하위 25%", "과거 상위 25%"}.issubset(
        set(chart_data["series"])
    )
    assert chart_data["business_day_no"].max() == historical_df["business_day_no"].max()


def test_historical_context_returns_interpretation_messages() -> None:
    current_df = load_input(SAMPLE_INPUT_PATH)
    historical_df = load_input(HISTORICAL_SAMPLE_INPUT_PATH, sort_by="date")
    config = build_runtime_config(load_model_config(), 1.30, 1.50)
    results = calculate_validated_results(current_df, "2026-06-10", "sales", config)

    context = build_historical_context(
        historical_df,
        current_df,
        "2026-06-10",
        "sales",
        results["validation"],
        "과거 샘플 데이터",
    )

    assert context["has_data"] is True
    assert context["source_label"] == "과거 샘플 데이터"
    assert context["interpretation"]
    assert "progress_chart_data" in context


def test_historical_input_state_defaults_to_sample(monkeypatch) -> None:
    class FakeStreamlit:
        def __init__(self) -> None:
            self.session_state: dict[str, object] = {}

    fake_st = FakeStreamlit()
    monkeypatch.setattr(app_module, "st", fake_st)

    historical_df, source_label = app_module._get_historical_input_state()

    assert source_label == "과거 샘플 데이터"
    assert len(historical_df) > 0
    assert app_module.HISTORICAL_INPUT_DF_SESSION_KEY in fake_st.session_state


def test_historical_input_state_recovers_empty_unintentional_state(monkeypatch) -> None:
    class FakeStreamlit:
        def __init__(self) -> None:
            self.session_state: dict[str, object] = {
                app_module.HISTORICAL_INPUT_DF_SESSION_KEY: pd.DataFrame(),
                app_module.HISTORICAL_INPUT_SOURCE_SESSION_KEY: "",
            }

    fake_st = FakeStreamlit()
    monkeypatch.setattr(app_module, "st", fake_st)

    historical_df, source_label = app_module._get_historical_input_state()

    assert source_label == "과거 샘플 데이터"
    assert len(historical_df) > 0


def test_historical_input_state_can_be_explicitly_disabled(monkeypatch) -> None:
    class FakeStreamlit:
        def __init__(self) -> None:
            self.session_state: dict[str, object] = {
                app_module.HISTORICAL_INPUT_DF_SESSION_KEY: pd.DataFrame(),
                app_module.HISTORICAL_INPUT_SOURCE_SESSION_KEY: "",
                app_module.HISTORICAL_SAMPLE_DISABLED_SESSION_KEY: True,
            }

    fake_st = FakeStreamlit()
    monkeypatch.setattr(app_module, "st", fake_st)

    historical_df, source_label = app_module._get_historical_input_state()

    assert historical_df.empty
    assert source_label == ""


def test_historical_forecast_comparison_includes_current_models_and_history() -> None:
    current_df = load_input(SAMPLE_INPUT_PATH)
    historical_df = load_input(HISTORICAL_SAMPLE_INPUT_PATH, sort_by="date")
    config = build_runtime_config(load_model_config(), 1.30, 1.50)
    results = calculate_validated_results(current_df, "2026-06-10", "sales", config)
    context = build_historical_context(
        historical_df,
        current_df,
        "2026-06-10",
        "sales",
        results["validation"],
        "과거 샘플 데이터",
    )
    selected_scenario_id = str(results["scenario_df"].iloc[0]["scenario_id"])

    comparison = build_historical_forecast_comparison(
        results["scenario_df"],
        context,
        selected_scenario_id,
    )

    assert not comparison.empty
    assert {"현재 예측", "F모델 기본 예측", "과거 실적 기반"}.issubset(
        set(comparison["comparison_group"])
    )
    assert "과거 중앙값" in set(comparison["basis"])
    assert comparison["forecast_amount"].notna().all()
    median_row = comparison.loc[comparison["basis"] == "과거 중앙값"].iloc[0]
    assert abs(median_row["diff_vs_historical_median"]) < 1e-9


def test_historical_forecast_decision_summary_wraps_final_judgment() -> None:
    comparison = pd.DataFrame(
        [
            {
                "comparison_group": "현재 예측",
                "basis": "선택 시나리오 F1_P1",
                "forecast_amount": 155.0,
                "monthly_target": 150.0,
                "forecast_rate": 155.0 / 150.0,
                "diff_vs_target": 5.0,
                "diff_vs_historical_median": 4.0,
            },
            {
                "comparison_group": "과거 실적 기반",
                "basis": "과거 하위 25%",
                "forecast_amount": 148.0,
                "monthly_target": 150.0,
                "forecast_rate": 148.0 / 150.0,
                "diff_vs_target": -2.0,
                "diff_vs_historical_median": -3.0,
            },
            {
                "comparison_group": "과거 실적 기반",
                "basis": "과거 중앙값",
                "forecast_amount": 151.0,
                "monthly_target": 150.0,
                "forecast_rate": 151.0 / 150.0,
                "diff_vs_target": 1.0,
                "diff_vs_historical_median": 0.0,
            },
            {
                "comparison_group": "과거 실적 기반",
                "basis": "과거 상위 25%",
                "forecast_amount": 158.0,
                "monthly_target": 150.0,
                "forecast_rate": 158.0 / 150.0,
                "diff_vs_target": 8.0,
                "diff_vs_historical_median": 7.0,
            },
        ]
    )

    summary = build_historical_forecast_decision_summary(comparison)

    assert summary["has_data"] is True
    assert "151.0억 원 ~ 155.0억 원" in summary["headline"]
    assert "공격적인 전망" in summary["forecast_position"]
    assert "초과 예상" in summary["target_position"]
    assert "안전버퍼" in summary["action"]


def test_historical_forecast_axis_domain_zooms_to_difference_area() -> None:
    comparison = pd.DataFrame(
        {
            "forecast_amount": [155.0, 154.9, 152.8, 155.0, 152.7, 151.7, 156.0],
            "monthly_target": [150.0] * 7,
        }
    )

    domain = build_historical_forecast_axis_domain(comparison)

    assert domain is not None
    assert domain[0] > 0.0
    assert domain[0] < 150.0
    assert domain[1] > 156.0
    assert domain[1] - domain[0] < 30.0


def test_historical_forecast_chart_uses_visible_ranged_bars(monkeypatch) -> None:
    comparison = pd.DataFrame(
        {
            "comparison_group": ["현재 예측", "F모델 기본 예측", "과거 실적 기반"],
            "basis": ["선택 시나리오 F1_P1", "F1 누적 달성률 모델", "과거 중앙값"],
            "forecast_amount": [152.9, 152.9, 154.2],
            "monthly_target": [149.6, 149.6, 149.6],
            "diff_vs_target": [3.3, 3.3, 4.6],
            "diff_vs_historical_median": [-1.3, -1.3, 0.0],
        }
    )
    captured: dict[str, object] = {}

    class FakeStreamlit:
        def caption(self, message: str) -> None:
            captured["caption"] = message

        def altair_chart(self, chart, use_container_width: bool = True) -> None:
            captured["spec"] = chart.to_dict(validate=True)
            captured["use_container_width"] = use_container_width

    monkeypatch.setattr(app_module, "st", FakeStreamlit())

    app_module._render_historical_forecast_comparison_chart(comparison)

    spec = captured["spec"]
    bar_encoding = spec["layer"][0]["encoding"]
    assert "확대" in captured["caption"]
    assert bar_encoding["x"]["field"] == "axis_floor"
    assert bar_encoding["x2"]["field"] == "forecast_amount"
    assert bar_encoding["x"]["scale"]["domain"][0] > 0
    assert bar_encoding["x"]["scale"]["domain"][1] > 154.2


def test_historical_forecast_comparison_empty_without_benchmark() -> None:
    comparison = build_historical_forecast_comparison(
        pd.DataFrame(),
        {"has_data": False},
        "",
    )

    assert comparison.empty
    assert {
        "comparison_group",
        "basis",
        "forecast_amount",
        "diff_vs_historical_median",
    }.issubset(set(comparison.columns))


def test_direct_input_edits_are_normalized_and_used_for_calculation() -> None:
    df = load_input(SAMPLE_INPUT_PATH)
    edited_df = df.copy().astype(
        {
            "sales_target_daily": "object",
            "sales_actual_cum": "object",
        }
    )
    edited_df.loc[edited_df["business_day_no"] == 7, "sales_actual_cum"] = "80.0"
    edited_df.loc[edited_df["business_day_no"] == 8, "sales_actual_cum"] = ""
    edited_df.loc[edited_df["business_day_no"] == 8, "sales_target_daily"] = "12.0"

    normalized = normalize_direct_input_edits(edited_df)
    config = build_runtime_config(load_model_config(), 1.30, 1.50)
    results = calculate_validated_results(normalized, "2026-06-10", "sales", config)

    assert DIRECT_EDITABLE_COLUMNS == (
        "sales_target_daily",
        "recognized_target_daily",
        "sales_actual_cum",
        "recognized_actual_cum",
    )
    assert results["validation"]["errors"] == []
    assert results["validation"]["current_actual_cum"] == 80.0
    assert pd.isna(
        normalized.loc[normalized["business_day_no"] == 8, "sales_actual_cum"].iloc[0]
    )
    assert (
        normalized.loc[normalized["business_day_no"] == 8, "sales_target_daily"].iloc[0]
        == 12.0
    )


def test_saved_actuals_are_applied_as_future_defaults(tmp_path) -> None:
    df = load_input(SAMPLE_INPUT_PATH)
    edited_df = df.copy()
    edited_df.loc[edited_df["business_day_no"] == 8, "sales_actual_cum"] = 88.8
    edited_df.loc[edited_df["business_day_no"] == 8, "recognized_actual_cum"] = 77.7
    edited_df.loc[edited_df["business_day_no"] == 8, "sales_target_daily"] = 12.0
    saved_path = tmp_path / "saved_actuals.csv"

    save_actual_values(edited_df, saved_path)
    saved_actuals = load_saved_actuals(saved_path)
    applied_df = apply_saved_actuals(df, saved_actuals)

    default_row = df.loc[df["business_day_no"] == 8].iloc[0]
    applied_row = applied_df.loc[applied_df["business_day_no"] == 8].iloc[0]
    assert ACTUAL_CUM_COLUMNS == ("sales_actual_cum", "recognized_actual_cum")
    assert applied_row["sales_actual_cum"] == 88.8
    assert applied_row["recognized_actual_cum"] == 77.7
    assert applied_row["sales_target_daily"] == default_row["sales_target_daily"]

    clear_saved_actuals(saved_path)
    assert not saved_path.exists()


def test_saved_actuals_do_not_overwrite_input_with_blank_cells() -> None:
    df = load_input(SAMPLE_INPUT_PATH)
    original_sales_actual = df.loc[df["business_day_no"] == 7, "sales_actual_cum"].iloc[0]
    saved_actuals = pd.DataFrame(
        {
            "date": [pd.Timestamp("2026-06-10")],
            "business_day_no": [7],
            "sales_actual_cum": [pd.NA],
            "recognized_actual_cum": [99.9],
        }
    )

    applied_df = apply_saved_actuals(df, saved_actuals)
    applied_row = applied_df.loc[applied_df["business_day_no"] == 7].iloc[0]

    assert applied_row["sales_actual_cum"] == original_sales_actual
    assert applied_row["recognized_actual_cum"] == 99.9


def test_save_actual_values_preserves_existing_values_when_latest_input_is_blank(
    tmp_path,
) -> None:
    df = load_input(SAMPLE_INPUT_PATH)
    saved_path = tmp_path / "saved_actuals.csv"
    first_edit_df = df.copy()
    first_edit_df.loc[first_edit_df["business_day_no"] == 8, "sales_actual_cum"] = 88.8
    first_edit_df.loc[
        first_edit_df["business_day_no"] == 8,
        "recognized_actual_cum",
    ] = 77.7
    blank_edit_df = df.copy()
    blank_edit_df.loc[blank_edit_df["business_day_no"] == 8, "sales_actual_cum"] = pd.NA
    blank_edit_df.loc[
        blank_edit_df["business_day_no"] == 8,
        "recognized_actual_cum",
    ] = pd.NA

    save_actual_values(first_edit_df, saved_path)
    save_actual_values(blank_edit_df, saved_path)
    latest_actuals = load_saved_actuals(saved_path)
    latest_row = latest_actuals.loc[latest_actuals["business_day_no"] == 8].iloc[0]

    assert latest_row["sales_actual_cum"] == 88.8
    assert latest_row["recognized_actual_cum"] == 77.7


def test_uploaded_input_values_become_latest_defaults(tmp_path) -> None:
    df = load_input(SAMPLE_INPUT_PATH)
    saved_df = df.copy()
    saved_df.loc[saved_df["business_day_no"] == 8, "sales_actual_cum"] = 55.5
    upload_df = df.copy()
    upload_df.loc[upload_df["business_day_no"] == 8, "sales_actual_cum"] = 99.9
    saved_path = tmp_path / "saved_actuals.csv"

    save_actual_values(saved_df, saved_path)
    saved_actuals = load_saved_actuals(saved_path)
    prepared_df, default_source = apply_latest_upload_policy(
        upload_df,
        "daily_upload.xlsx",
        saved_actuals,
        saved_path,
    )
    latest_actuals = load_saved_actuals(saved_path)

    prepared_row = prepared_df.loc[prepared_df["business_day_no"] == 8].iloc[0]
    latest_row = latest_actuals.loc[latest_actuals["business_day_no"] == 8].iloc[0]
    assert default_source == "uploaded"
    assert prepared_row["sales_actual_cum"] == 99.9
    assert latest_row["sales_actual_cum"] == 99.9


def test_sample_input_still_uses_saved_actual_defaults(tmp_path) -> None:
    df = load_input(SAMPLE_INPUT_PATH)
    saved_df = df.copy()
    saved_df.loc[saved_df["business_day_no"] == 8, "sales_actual_cum"] = 66.6
    saved_path = tmp_path / "saved_actuals.csv"

    save_actual_values(saved_df, saved_path)
    saved_actuals = load_saved_actuals(saved_path)
    prepared_df, default_source = apply_latest_upload_policy(
        df,
        SAMPLE_INPUT_SOURCE_LABEL,
        saved_actuals,
        saved_path,
    )

    prepared_row = prepared_df.loc[prepared_df["business_day_no"] == 8].iloc[0]
    assert default_source == "saved"
    assert prepared_row["sales_actual_cum"] == 66.6


def test_model_and_risk_explanations_are_available() -> None:
    selected_row = pd.Series(
        {
            "scenario_id": "F2_P3",
            "risk_level": "Yellow",
            "status": "OK",
            "target_status": "UNDER_TARGET",
        }
    )

    forecast_definitions = build_forecast_definition_df()
    provision_definitions = build_provision_definition_df()
    overachievement_definitions = build_overachievement_definition_df()
    neutral_definitions = build_neutral_definition_df()
    risk_definitions = build_risk_definition_df()
    selected_explanation = build_selected_scenario_explanation("F2_P3", selected_row)

    assert forecast_definitions["model"].tolist() == ["F1", "F2", "F3"]
    assert provision_definitions["strategy"].tolist() == ["P1", "P2", "P3"]
    assert overachievement_definitions["strategy"].tolist() == ["O1", "O2", "O3"]
    assert neutral_definitions["strategy"].tolist() == ["N1", "N2", "N3"]
    assert set(risk_definitions["risk_level"]) == {"Green", "Yellow", "Red", "Black", "N/A"}
    assert selected_explanation.loc[
        selected_explanation["item"] == "조합 의미",
        "value",
    ].iloc[0].startswith("F2로 월말 예상 실적")
    assert selected_explanation.loc[
        selected_explanation["item"] == "예측 모델",
        "value",
    ].iloc[0] == "F2 직전 2개 완료 마감차수 모델"
    assert selected_explanation.loc[
        selected_explanation["item"] == "운영 전략",
        "value",
    ].iloc[0] == "P3 비마감일 분산 보정"
    assert "주의 / 목표 보정 필요" in selected_explanation.loc[
        selected_explanation["item"] == "위험등급 / 운영모드",
        "value",
    ].iloc[0]
    assert selected_explanation.loc[
        selected_explanation["item"] == "계산 상태",
        "value",
    ].iloc[0] == "정상"
    assert selected_explanation.loc[
        selected_explanation["item"] == "전략 구분",
        "value",
    ].iloc[0] == "목표 보정"
    selected_text = "\n".join(selected_explanation["value"].astype(str))
    assert "Yellow" not in selected_text
    assert "OK" not in selected_text
    assert "PROVISION" not in selected_text
    assert "기준일까지 완료된 최근 2개" not in selected_text
    assert "예상 부족분을 기준일 이후" not in selected_text
    assert "예상 달성률이 95% 이상" not in selected_text


def test_selected_scenario_explanation_maps_operation_mode_by_target_status() -> None:
    expected_modes = {
        "UNDER_TARGET": "목표 보정 필요",
        "ON_TARGET": "유지/모니터링",
        "OVER_TARGET": "초과달성 관리",
    }

    for target_status, expected_mode in expected_modes.items():
        selected_row = pd.Series(
            {
                "scenario_id": "F1_P1",
                "risk_level": "Green",
                "status": "OK",
                "target_status": target_status,
            }
        )

        selected_explanation = build_selected_scenario_explanation("F1_P1", selected_row)

        assert selected_explanation.loc[
            selected_explanation["item"] == "위험등급 / 운영모드",
            "value",
        ].iloc[0] == f"낮음 / {expected_mode}"


def test_next_close_cumulative_required_label_exists_in_app() -> None:
    app_source = Path("app.py").read_text(encoding="utf-8")

    assert "다음 마감 누적선 필요실적" in app_source


def test_security_warning_text_exists_without_secret_contents() -> None:
    app_source = Path("app.py").read_text(encoding="utf-8")

    assert "Public Streamlit 또는 외부 공개 URL에는 실제 영업실적을 업로드하지 마세요" in app_source
    assert ".streamlit/secrets.toml`은 로컬 전용" in app_source
    assert "APP_ACCESS_PASSWORD=공유할_접속_비밀번호" in app_source


def test_report_glossary_panel_uses_fixed_definition_groups() -> None:
    glossary = build_report_glossary_df()

    assert set(glossary.columns) == {"구분", "코드", "정의"}
    assert set(glossary["구분"]) == {
        "예측모델(F)",
        "목표 보정 전략(P)",
        "초과달성 운영전략(O)",
        "유지/모니터링 전략(N)",
        "위험등급",
    }
    assert {"F1", "F2", "F3", "P1", "P2", "P3", "O1", "O2", "O3"}.issubset(
        set(glossary["코드"])
    )
    assert glossary.loc[glossary["코드"] == "F1", "정의"].iloc[0].startswith(
        "누적 달성률 모델"
    )


def test_scenario_option_label_includes_definitions() -> None:
    label = format_scenario_option_label("F1_P1")
    over_label = format_scenario_option_label("F1_O1")

    assert label.startswith("F1_P1 -")
    assert "누적 달성률 모델" in label
    assert "잔여목표 균등 배분" in label
    assert "버퍼 유지" in over_label


def test_validation_messages_are_displayed_in_plain_korean() -> None:
    assert (
        format_validation_message("actual_cum must be populated through as_of_date.")
        == "기준일까지의 누적 실적에 빈값이 있습니다. 기준일 이전과 기준일의 누적 실적을 모두 입력해 주세요."
    )

    missing_column_message = format_validation_message(
        "Missing required input column: sales_actual_cum."
    )
    assert "판매실적 누적 실적(sales_actual_cum) 열이 없습니다" in missing_column_message

    fallback_message = format_validation_message(
        "F2_LAST_TWO_CLOSES fallback to F1_CUMULATIVE_RATE: "
        "fewer than two completed close cycles are available."
    )
    assert "F2" in fallback_message
    assert "F1 방식으로 대신 계산" in fallback_message


def test_forecast_history_backtest_tab_label_exists() -> None:
    app_source = Path("app.py").read_text(encoding="utf-8")

    assert HISTORY_TAB_LABEL == "예측 이력 / Backtest"
    assert "예측 이력 / Backtest" in app_source
    assert "forecast_history 테이블" in app_source
    assert "final_actuals 테이블" in app_source


def test_missing_history_files_are_loaded_as_empty_tables(tmp_path: Path) -> None:
    tables = load_history_tables_for_app(
        tmp_path / "missing_forecast_history.csv",
        tmp_path / "missing_final_actuals.csv",
    )

    assert tables["forecast_history"].empty
    assert tables["final_actuals"].empty
    assert tuple(tables["forecast_history"].columns) == history_schema.FORECAST_HISTORY_COLUMNS
    assert tuple(tables["final_actuals"].columns) == history_schema.FINAL_ACTUALS_COLUMNS


def test_target_status_kpi_value_is_preserved_in_summary_dict() -> None:
    summary = build_summary_dict(
        {
            "monthly_target": 100.0,
            "current_target_cum": 70.0,
            "current_actual_cum": 72.0,
            "errors": [],
            "warnings": [],
        },
        pd.Series(
            {
                "scenario_id": "F1_O1",
                "forecast_after_provision": 115.0,
                "target_status": "OVER_TARGET",
                "target_variance": 15.0,
                "surplus_to_target": 15.0,
                "strategy_type": "OVERACHIEVEMENT",
                "risk_level": "Green",
                "status": "OVER_TARGET_MANAGED",
            }
        ),
        {},
        "sales",
        "2026-06-10",
    )

    assert summary["target_status"] == "OVER_TARGET"


def test_display_validation_result_keeps_numbers_and_translates_messages() -> None:
    display_result = build_display_validation_result(
        {
            "errors": ["sales_actual_cum contains invalid numeric values."],
            "warnings": ["actual_cum after as_of_date is populated."],
            "monthly_target": 100.0,
        }
    )

    assert display_result["monthly_target"] == 100.0
    assert "숫자로 읽을 수 없는 값" in display_result["errors"][0]
    assert "기준일 이후 날짜" in display_result["warnings"][0]

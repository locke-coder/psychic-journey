from pathlib import Path

from openpyxl import Workbook

from tools import check_outputs_latest


def _write_latest_report(path: Path, *, missing_column: str | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.active.title = "Summary"
    scenario_grid = workbook.create_sheet("ScenarioGrid")
    headers = [
        "scenario_id",
        *check_outputs_latest.REQUIRED_REPORT_COLUMNS,
    ]
    if missing_column is not None:
        headers = [header for header in headers if header != missing_column]
    scenario_grid.append(headers)
    for sheet_name in check_outputs_latest.ADVANCED_SHEETS:
        workbook.create_sheet(sheet_name)
    workbook.save(path)
    workbook.close()


def _write_input_template(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.active.title = "InputTemplate"
    workbook.active.append(check_outputs_latest.INPUT_TEMPLATE_HEADERS)
    workbook.save(path)
    workbook.close()


def _write_old_format_report(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.active.title = "Summary"
    scenario_grid = workbook.create_sheet("ScenarioGrid")
    scenario_grid.append(["scenario_id", "forecast_amount"])
    workbook.save(path)
    workbook.close()


def test_latest_report_is_classified(tmp_path: Path) -> None:
    outputs_root = tmp_path / "outputs"
    report_path = outputs_root / "latest" / "daily_report_sales.xlsx"
    _write_latest_report(report_path)

    result = check_outputs_latest.classify_latest_file(report_path, outputs_root)

    assert result["category"] == check_outputs_latest.CATEGORY_LATEST_REPORT
    assert result["missing_required_columns"] == []
    assert result["missing_advanced_sheets"] == []


def test_input_template_is_classified_without_scenario_grid(tmp_path: Path) -> None:
    outputs_root = tmp_path / "outputs"
    template_path = outputs_root / "latest" / "month_close_forecast_input_template.xlsx"
    _write_input_template(template_path)

    result = check_outputs_latest.classify_latest_file(template_path, outputs_root)

    assert result["category"] == check_outputs_latest.CATEGORY_INPUT_TEMPLATE


def test_pre_metadata_file_is_old_format(tmp_path: Path) -> None:
    outputs_root = tmp_path / "outputs"
    old_path = outputs_root / "latest" / "daily_report_pre_metadata_20260608.xlsx"
    _write_latest_report(old_path)

    result = check_outputs_latest.classify_latest_file(old_path, outputs_root)

    assert result["category"] == check_outputs_latest.CATEGORY_OLD_FORMAT
    assert "pre_metadata" in result["reason"]


def test_missing_required_column_fails_strict(tmp_path: Path) -> None:
    outputs_root = tmp_path / "outputs"
    report_path = outputs_root / "latest" / "daily_report_sales.xlsx"
    _write_latest_report(report_path, missing_column="target_status")

    result = check_outputs_latest.check_outputs_latest(outputs_root, strict=True)

    assert result["result"] == "FAIL"
    assert result["old_format_files"] == ["latest/daily_report_sales.xlsx"]
    assert result["missing_required_columns"] == {
        "latest/daily_report_sales.xlsx": ["target_status"]
    }


def test_invalid_xlsx_is_classified(tmp_path: Path) -> None:
    outputs_root = tmp_path / "outputs"
    invalid_path = outputs_root / "latest" / "broken.xlsx"
    invalid_path.parent.mkdir(parents=True, exist_ok=True)
    invalid_path.write_text("not a workbook", encoding="utf-8")

    result = check_outputs_latest.check_outputs_latest(outputs_root, strict=True)

    assert result["result"] == "FAIL"
    assert result["invalid_files"] == ["latest/broken.xlsx"]


def test_organize_moves_old_format_to_archive_old_format(tmp_path: Path) -> None:
    outputs_root = tmp_path / "outputs"
    old_path = outputs_root / "latest" / "daily_report_pre_metadata.xlsx"
    _write_old_format_report(old_path)

    result = check_outputs_latest.check_outputs_latest(outputs_root, organize=True)

    assert result["result"] == "PASS"
    assert not old_path.exists()
    assert (outputs_root / "archive_old_format" / "daily_report_pre_metadata.xlsx").is_file()
    assert result["moved_to_archive_old_format"] == [
        "archive_old_format/daily_report_pre_metadata.xlsx"
    ]


def test_organize_adds_suffix_when_archive_filename_exists(tmp_path: Path) -> None:
    outputs_root = tmp_path / "outputs"
    old_path = outputs_root / "latest" / "daily_report_pre_metadata.xlsx"
    existing_archive = outputs_root / "archive_old_format" / old_path.name
    _write_old_format_report(old_path)
    _write_old_format_report(existing_archive)

    result = check_outputs_latest.check_outputs_latest(outputs_root, organize=True)

    assert existing_archive.is_file()
    assert (outputs_root / "archive_old_format" / "daily_report_pre_metadata_1.xlsx").is_file()
    assert result["moved_to_archive_old_format"] == [
        "archive_old_format/daily_report_pre_metadata_1.xlsx"
    ]

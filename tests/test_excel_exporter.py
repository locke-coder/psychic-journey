from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from src.excel_exporter import (
    CLOSECYCLE_REQUIRED_CUMULATIVE_COLUMNS,
    EXPORT_VERSION,
    SCENARIO_GRID_REQUIRED_COLUMNS,
    SHEET_NAMES,
    export_daily_report,
    prepare_scenario_grid_export_frame,
)


def _output_path() -> Path:
    return Path(__file__).resolve().parents[1] / "outputs" / "test_report.xlsx"


def _versioned_output_path(path: Path) -> Path:
    return path.with_name(f"{path.stem}_{EXPORT_VERSION}{path.suffix}")


@pytest.fixture
def output_path() -> Path:
    path = _output_path()
    for candidate in (path, _versioned_output_path(path)):
        if candidate.exists():
            candidate.unlink()
    yield path
    # The generated test workbook is removed so outputs stays clean after pytest.
    for candidate in (path, _versioned_output_path(path)):
        if candidate.exists():
            candidate.unlink()


def _scenario_df() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "scenario_id": [f"F{forecast}_P{provision}" for forecast in range(1, 4) for provision in range(1, 4)],
            "forecast_model": ["F1", "F1", "F1", "F2", "F2", "F2", "F3", "F3", "F3"],
            "provision_strategy": ["P1", "P2", "P3"] * 3,
            "forecast_amount": [95.0, 96.1, 97.2, 94.0, 95.1, 96.2, 99.0, 100.1, 101.2],
            "forecast_rate": [0.95, 0.961, 0.972, 0.94, 0.951, 0.962, 0.99, 1.001, 1.012],
            "status": ["OK"] * 9,
        }
    )


def _revised_targets_df() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "date": pd.to_datetime(["2026-06-08", "2026-06-09"]),
            "revised_target": [20.5, 21.5],
        }
    )


def _close_cycle_df() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "date": pd.to_datetime(
                ["2026-06-01", "2026-06-04", "2026-06-05", "2026-06-08"]
            ),
            "day_name": ["Mon", "Thu", "Fri", "Mon"],
            "business_day_no": [1, 2, 3, 4],
            "is_close_day": [False, False, True, False],
            "close_type": ["", "", "manual_close", ""],
            "sales_target_daily": [10.0, 20.0, 30.0, 40.0],
            "recognized_target_daily": [5.0, 10.0, 15.0, 20.0],
            "sales_actual_cum": [8.0, 27.0, 58.0, 95.0],
            "recognized_actual_cum": [4.0, 14.0, 33.0, 51.0],
        }
    )


def _backtest_summary_df() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "forecast_model": ["F1", "F2", "F3"],
            "sample_count": [3, 3, 3],
            "error_rate": [0.05, 0.08, 0.03],
            "bias": [1.0, -2.0, 0.5],
        }
    )


def _export(
    output_path: Path,
    *,
    backtest_summary_df: pd.DataFrame | None = None,
    overwrite: bool = True,
) -> Path:
    return export_daily_report(
        output_path,
        {"metric": "sales", "monthly_target": 100.0, "forecast_rate": 0.95},
        _scenario_df(),
        _revised_targets_df(),
        _close_cycle_df(),
        {"errors": [], "warnings": ["sample warning"]},
        "Daily report body",
        forecast_history_df=pd.DataFrame(),
        final_actuals_df=pd.DataFrame(),
        backtest_summary_df=backtest_summary_df
        if backtest_summary_df is not None
        else pd.DataFrame(),
        model_weights_df=pd.DataFrame(),
        confidence_band_df=pd.DataFrame(),
        insights_df=pd.DataFrame(),
        overwrite=overwrite,
    )


def test_export_daily_report_creates_outputs_file(output_path: Path) -> None:
    saved_path = _export(output_path)

    assert saved_path == _versioned_output_path(output_path)
    assert EXPORT_VERSION in saved_path.stem
    assert saved_path.is_file()


def test_export_daily_report_contains_required_sheets(output_path: Path) -> None:
    saved_path = _export(output_path)

    workbook = load_workbook(saved_path)

    assert workbook.sheetnames == list(SHEET_NAMES)
    assert workbook.sheetnames[:6] == [
        "Summary",
        "ScenarioGrid",
        "DailyRevisedTargets",
        "CloseCycle",
        "Validation",
        "ReportText",
    ]
    assert {
        "ForecastHistory",
        "FinalActuals",
        "BacktestSummary",
        "ModelWeights",
        "ConfidenceBand",
        "Insights",
    }.issubset(workbook.sheetnames)


def test_report_text_sheet_contains_report_body(output_path: Path) -> None:
    saved_path = _export(output_path)

    workbook = load_workbook(saved_path)

    assert workbook["ReportText"]["A2"].value == "Daily report body"


def test_scenario_grid_has_nine_data_rows(output_path: Path) -> None:
    saved_path = _export(output_path)

    workbook = load_workbook(saved_path)
    scenario_sheet = workbook["ScenarioGrid"]

    assert scenario_sheet.max_row - 1 == 9


def test_scenario_grid_contains_target_and_strategy_columns(output_path: Path) -> None:
    saved_path = _export(output_path)

    workbook = load_workbook(saved_path)
    scenario_sheet = workbook["ScenarioGrid"]
    headers = [cell.value for cell in scenario_sheet[1]]

    assert set(SCENARIO_GRID_REQUIRED_COLUMNS).issubset(headers)
    assert headers[: len(SCENARIO_GRID_REQUIRED_COLUMNS)] == list(
        SCENARIO_GRID_REQUIRED_COLUMNS
    )


def test_prepare_scenario_grid_export_frame_adds_display_columns() -> None:
    prepared = prepare_scenario_grid_export_frame(_scenario_df())

    assert prepared.shape[0] == 9
    assert set(SCENARIO_GRID_REQUIRED_COLUMNS).issubset(prepared.columns)
    assert prepared["scenario"].tolist()[0] == "F1_P1"
    assert prepared["strategy_code"].tolist()[:3] == ["P1", "P2", "P3"]
    assert prepared["strategy_label"].tolist()[:3] == [
        "잔여목표 균등 배분",
        "마감일 집중 보정",
        "비마감일 분산 보정",
    ]
    assert prepared["strategy_group"].unique().tolist() == ["목표 보정"]


def test_prepare_scenario_grid_export_frame_keeps_o_strategy_labels_distinct() -> None:
    overachievement = pd.DataFrame(
        {
            "scenario_id": ["F1_O1", "F1_O2", "F1_O3"],
            "forecast_model": ["F1"] * 3,
            "provision_strategy": [
                "O1_TARGET_HOLD_BUFFER",
                "O2_STRETCH_TARGET_CAPTURE",
                "O3_QUALITY_GUARD_RELIEF",
            ],
            "target_status": ["OVER_TARGET"] * 3,
            "target_variance": [12.0] * 3,
            "surplus_to_target": [12.0] * 3,
            "forecast_after_provision": [112.0] * 3,
        }
    )

    prepared = prepare_scenario_grid_export_frame(overachievement)

    assert prepared["strategy_label"].tolist() == ["버퍼 유지", "Stretch 전환", "품질 방어"]
    assert prepared["strategy_group"].unique().tolist() == ["초과달성 운영"]


def test_close_cycle_sheet_contains_cumulative_columns(output_path: Path) -> None:
    saved_path = _export(output_path)

    workbook = load_workbook(saved_path)
    headers = [cell.value for cell in workbook["CloseCycle"][1]]

    assert set(CLOSECYCLE_REQUIRED_CUMULATIVE_COLUMNS).issubset(headers)
    assert {
        "business_day_no",
        "date",
        "day_name",
        "is_close_day",
        "close_type",
        "close_cycle_no",
        "close_cycle_label",
        "cycle_sales_target",
        "cycle_recognized_target",
        "cycle_sales_actual",
        "cycle_recognized_actual",
    }.issubset(headers)


def test_close_cycle_cumulative_values_are_calculated_from_input_rows(
    output_path: Path,
) -> None:
    saved_path = _export(output_path)

    workbook = load_workbook(saved_path)
    sheet = workbook["CloseCycle"]
    headers = [cell.value for cell in sheet[1]]
    rows = [
        {
            header: sheet.cell(row=row_index, column=headers.index(header) + 1).value
            for header in headers
        }
        for row_index in range(2, sheet.max_row + 1)
    ]

    assert [row["sales_target_cum"] for row in rows] == pytest.approx(
        [10.0, 30.0, 60.0, 100.0]
    )
    assert [row["recognized_target_cum"] for row in rows] == pytest.approx(
        [5.0, 15.0, 30.0, 50.0]
    )
    assert [row["sales_actual_cum"] for row in rows] == pytest.approx(
        [8.0, 27.0, 58.0, 95.0]
    )
    assert [row["recognized_actual_cum"] for row in rows] == pytest.approx(
        [4.0, 14.0, 33.0, 51.0]
    )
    assert [row["sales_gap_to_plan_cum"] for row in rows] == pytest.approx(
        [-2.0, -3.0, -2.0, -5.0]
    )
    assert [row["recognized_gap_to_plan_cum"] for row in rows] == pytest.approx(
        [-1.0, -1.0, 3.0, 1.0]
    )
    assert [row["sales_attainment_rate_cum"] for row in rows] == pytest.approx(
        [0.8, 0.9, 58.0 / 60.0, 0.95]
    )
    assert [row["recognized_attainment_rate_cum"] for row in rows] == pytest.approx(
        [0.8, 14.0 / 15.0, 1.1, 1.02]
    )


def test_close_cycle_numbering_uses_is_close_day_not_day_name(
    output_path: Path,
) -> None:
    saved_path = _export(output_path)

    workbook = load_workbook(saved_path)
    sheet = workbook["CloseCycle"]
    headers = [cell.value for cell in sheet[1]]
    cycle_column = headers.index("close_cycle_no") + 1

    assert [
        sheet.cell(row=row_index, column=cycle_column).value
        for row_index in range(2, sheet.max_row + 1)
    ] == [1, 1, 1, 2]


def test_existing_file_can_be_protected_by_overwrite_flag(output_path: Path) -> None:
    _export(output_path)

    with pytest.raises(FileExistsError):
        _export(output_path, overwrite=False)


def test_history_export_sheets_are_safe_when_history_is_empty(output_path: Path) -> None:
    saved_path = _export(output_path)

    workbook = load_workbook(saved_path)

    assert workbook["ForecastHistory"]["A1"].value == "message"
    assert "forecast_history" in workbook["ForecastHistory"]["A2"].value
    assert workbook["FinalActuals"]["A1"].value == "message"
    assert "final_actuals" in workbook["FinalActuals"]["A2"].value


def test_backtest_summary_sheet_contains_required_columns(output_path: Path) -> None:
    saved_path = _export(output_path, backtest_summary_df=_backtest_summary_df())

    workbook = load_workbook(saved_path)
    backtest_sheet = workbook["BacktestSummary"]
    headers = [cell.value for cell in backtest_sheet[1]]

    assert {"forecast_model", "error_rate", "bias", "sample_count"}.issubset(headers)
    assert backtest_sheet.max_row == 4


def test_report_text_optionally_includes_model_error_summary(output_path: Path) -> None:
    saved_path = _export(output_path, backtest_summary_df=_backtest_summary_df())

    workbook = load_workbook(saved_path)

    assert "모델 오차율 요약" in workbook["ReportText"]["A2"].value

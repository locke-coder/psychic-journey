"""Excel workbook export helpers for daily closing reports."""

from __future__ import annotations

import json
import re
from collections.abc import Mapping
from datetime import date, datetime
from math import isfinite
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.display_labels import (
    get_forecast_model_label,
    get_strategy_code,
    get_strategy_group,
    get_strategy_label,
)
from src.report_builder import append_model_error_summary_to_report


SHEET_NAMES = (
    "Summary",
    "ScenarioGrid",
    "DailyRevisedTargets",
    "CloseCycle",
    "Validation",
    "ReportText",
    "ForecastHistory",
    "FinalActuals",
    "BacktestSummary",
    "ModelWeights",
    "ConfidenceBand",
    "Insights",
)
EXPORT_VERSION = "v2"

AMOUNT_NUMBER_FORMAT = '#,##0.0'
RATE_NUMBER_FORMAT = '0.0%'
DATE_NUMBER_FORMAT = 'yyyy-mm-dd'

_HEADER_FILL = PatternFill(fill_type="solid", fgColor="44546A")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_BORDER = Border(bottom=Side(style="thin", color="B7C9D6"))
_REPO_ROOT = Path(__file__).resolve().parents[1]

_AMOUNT_NAME_HINTS = (
    "amount",
    "target",
    "actual",
    "forecast",
    "gap",
    "uplift",
    "sales",
    "recognized",
    "required",
    "allocated",
    "unallocated",
    "remaining",
    "sum",
)
_RATE_NAME_HINTS = ("rate", "ratio", "pct", "percent", "achievement")
_COUNT_NAME_HINTS = ("count", "cycle_id", "day_no", "business_day_no")
_PERCENT_BASIS_COLUMNS = {"achievement_rate"}
_VERSION_TOKEN_PATTERN = re.compile(
    r"(^|[_-])(v\d+|version\d+|generated_at[_-]?\d{8,})([_-]|$)",
    re.IGNORECASE,
)
FORECAST_HISTORY_MESSAGE = "No forecast_history data is available yet."
FINAL_ACTUALS_MESSAGE = "No final_actuals data is available yet."
BACKTEST_SUMMARY_MESSAGE = (
    "No BacktestSummary data is available. Save forecast_history and final_actuals first."
)
MODEL_WEIGHTS_MESSAGE = "No ModelWeights data is available until model error rates exist."
CONFIDENCE_BAND_MESSAGE = "No ConfidenceBand data is available yet."
INSIGHTS_MESSAGE = "No Insights data is available yet."
BACKTEST_SUMMARY_COLUMNS = (
    "forecast_model",
    "sample_count",
    "error_rate",
    "bias",
    "mean_abs_error",
    "mean_error_rate",
    "median_error_rate",
    "best_model_by_error_rate",
)
MODEL_WEIGHTS_COLUMNS = (
    "forecast_model",
    "sample_count",
    "error_rate",
    "bias",
    "model_weight",
)
CONFIDENCE_BAND_COLUMNS = (
    "target_month",
    "as_of_date",
    "metric",
    "forecast_model",
    "forecast_amount",
    "confidence_lower",
    "confidence_upper",
)
INSIGHTS_COLUMNS = ("insight",)
SCENARIO_GRID_REQUIRED_COLUMNS = (
    "scenario",
    "forecast_model",
    "model_name",
    "expected_month_end_amount",
    "target_status",
    "target_variance",
    "surplus_to_target",
    "strategy_type",
    "strategy_code",
    "overachievement_strategy",
    "strategy_label",
    "strategy_group",
    "stretch_uplift",
    "revised_monthly_target",
    "remaining_surplus_buffer",
    "minimum_remaining_to_hit_target",
    "relief_amount",
    "recommended_action",
    "risk_note",
)
SCENARIO_GRID_EXPORT_COLUMN_ORDER = (
    *SCENARIO_GRID_REQUIRED_COLUMNS,
    "scenario_id",
    "provision_strategy",
    "metric",
    "as_of_date",
    "monthly_target",
    "current_actual_cum",
    "current_target_cum",
    "remaining_target",
    "forecast_amount",
    "forecast_rate",
    "gap_to_target",
    "required_uplift",
    "allocated_uplift",
    "unallocated_uplift",
    "revised_remaining_target",
    "forecast_after_provision",
    "gap_after_provision",
    "next_close_date",
    "next_close_required",
    "risk_level",
    "status",
    "comment",
    "warnings",
)
CLOSECYCLE_REQUIRED_CUMULATIVE_COLUMNS = (
    "sales_target_cum",
    "recognized_target_cum",
    "sales_actual_cum",
    "recognized_actual_cum",
    "sales_gap_to_plan_cum",
    "recognized_gap_to_plan_cum",
    "sales_attainment_rate_cum",
    "recognized_attainment_rate_cum",
)
CLOSECYCLE_BASE_EXPORT_COLUMNS = (
    "business_day_no",
    "date",
    "day_name",
    "is_close_day",
    "close_type",
    "close_cycle_no",
    "close_cycle_label",
    "cycle_sales_target",
    "cycle_recognized_target",
    "cycle_sales_actual",
    "cycle_recognized_actual",
)
CONFIDENCE_BAND_COLUMN_PAIRS = (
    ("confidence_lower", "confidence_upper"),
    ("forecast_lower", "forecast_upper"),
    ("lower_bound", "upper_bound"),
)


def export_daily_report(
    output_path: str | Path,
    summary_dict: Mapping[str, Any],
    scenario_df: pd.DataFrame,
    revised_targets_df: pd.DataFrame,
    close_cycle_df: pd.DataFrame,
    validation_result: Mapping[str, Any] | Any,
    report_text: str,
    *,
    forecast_history_df: pd.DataFrame | Any | None = None,
    final_actuals_df: pd.DataFrame | Any | None = None,
    backtest_summary_df: pd.DataFrame | Any | None = None,
    model_weights_df: pd.DataFrame | Any | None = None,
    confidence_band_df: pd.DataFrame | Any | None = None,
    insights_df: pd.DataFrame | Any | None = None,
    overwrite: bool = True,
) -> Path:
    """Export the daily report workbook and return the saved path."""
    resolved_output_path = _resolve_output_path(output_path)
    resolved_output_path.parent.mkdir(parents=True, exist_ok=True)

    if resolved_output_path.exists() and not overwrite:
        raise FileExistsError(
            f"{resolved_output_path} already exists. Set overwrite=True to replace it."
        )

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"

    history_tables = _build_history_export_tables(
        forecast_history_df,
        final_actuals_df,
        backtest_summary_df,
        model_weights_df,
        confidence_band_df,
        insights_df,
    )
    report_text = append_model_error_summary_to_report(
        report_text,
        history_tables["BacktestSummary"],
    )

    _write_mapping_sheet(summary_sheet, summary_dict)
    _write_dataframe_sheet(
        workbook.create_sheet("ScenarioGrid"),
        prepare_scenario_grid_export_frame(scenario_df),
    )
    _write_dataframe_sheet(workbook.create_sheet("DailyRevisedTargets"), revised_targets_df)
    _write_dataframe_sheet(
        workbook.create_sheet("CloseCycle"),
        _prepare_close_cycle_export_frame(close_cycle_df),
    )
    _write_mapping_sheet(workbook.create_sheet("Validation"), validation_result)
    _write_report_text_sheet(workbook.create_sheet("ReportText"), report_text)
    _write_optional_dataframe_sheet(
        workbook.create_sheet("ForecastHistory"),
        history_tables["ForecastHistory"],
        FORECAST_HISTORY_MESSAGE,
    )
    _write_optional_dataframe_sheet(
        workbook.create_sheet("FinalActuals"),
        history_tables["FinalActuals"],
        FINAL_ACTUALS_MESSAGE,
    )
    _write_optional_dataframe_sheet(
        workbook.create_sheet("BacktestSummary"),
        history_tables["BacktestSummary"],
        BACKTEST_SUMMARY_MESSAGE,
        BACKTEST_SUMMARY_COLUMNS,
    )
    _write_optional_dataframe_sheet(
        workbook.create_sheet("ModelWeights"),
        history_tables["ModelWeights"],
        MODEL_WEIGHTS_MESSAGE,
        MODEL_WEIGHTS_COLUMNS,
    )
    _write_optional_dataframe_sheet(
        workbook.create_sheet("ConfidenceBand"),
        history_tables["ConfidenceBand"],
        CONFIDENCE_BAND_MESSAGE,
        CONFIDENCE_BAND_COLUMNS,
    )
    _write_optional_dataframe_sheet(
        workbook.create_sheet("Insights"),
        history_tables["Insights"],
        INSIGHTS_MESSAGE,
        INSIGHTS_COLUMNS,
    )

    workbook.save(resolved_output_path)
    return resolved_output_path


def _build_history_export_tables(
    forecast_history_df: pd.DataFrame | Any | None,
    final_actuals_df: pd.DataFrame | Any | None,
    backtest_summary_df: pd.DataFrame | Any | None,
    model_weights_df: pd.DataFrame | Any | None,
    confidence_band_df: pd.DataFrame | Any | None,
    insights_df: pd.DataFrame | Any | None,
) -> dict[str, pd.DataFrame]:
    forecast_history = (
        _load_default_forecast_history()
        if forecast_history_df is None
        else _as_dataframe(forecast_history_df)
    )
    final_actuals = (
        _load_default_final_actuals()
        if final_actuals_df is None
        else _as_dataframe(final_actuals_df)
    )
    backtest_summary = (
        _build_default_backtest_summary(forecast_history, final_actuals)
        if backtest_summary_df is None
        else _as_dataframe(backtest_summary_df)
    )
    backtest_summary = _normalize_backtest_summary(backtest_summary)

    model_weights = (
        _build_model_weights(backtest_summary)
        if model_weights_df is None
        else _as_dataframe(model_weights_df)
    )
    confidence_band = (
        _build_confidence_band_frame(forecast_history)
        if confidence_band_df is None
        else _as_dataframe(confidence_band_df)
    )
    insights = (
        _build_insights_frame(forecast_history, final_actuals, backtest_summary)
        if insights_df is None
        else _as_dataframe(insights_df)
    )

    return {
        "ForecastHistory": forecast_history,
        "FinalActuals": final_actuals,
        "BacktestSummary": backtest_summary,
        "ModelWeights": model_weights,
        "ConfidenceBand": confidence_band,
        "Insights": insights,
    }


def _load_default_forecast_history() -> pd.DataFrame:
    try:
        from src import history_schema
        from src.history_store import load_forecast_history

        path = history_schema.get_storage_paths(repo_root=_REPO_ROOT)[
            history_schema.FORECAST_HISTORY
        ]
        return load_forecast_history(path)
    except Exception:  # noqa: BLE001 - export should remain safe without history.
        return pd.DataFrame()


def _load_default_final_actuals() -> pd.DataFrame:
    try:
        from src import history_schema
        from src.final_actual_store import load_final_actuals

        path = history_schema.get_storage_paths(repo_root=_REPO_ROOT)[
            history_schema.FINAL_ACTUALS
        ]
        return load_final_actuals(path)
    except Exception:  # noqa: BLE001 - export should remain safe without final actuals.
        return pd.DataFrame()


def _build_default_backtest_summary(
    forecast_history: pd.DataFrame,
    final_actuals: pd.DataFrame,
) -> pd.DataFrame:
    try:
        from src.backtest_engine import build_backtest_dataset, summarize_by_forecast_model

        backtest_df = build_backtest_dataset(forecast_history, final_actuals)
        return summarize_by_forecast_model(backtest_df)
    except Exception:  # noqa: BLE001 - empty summary is safer than blocking export.
        return pd.DataFrame(columns=BACKTEST_SUMMARY_COLUMNS)


def _normalize_backtest_summary(backtest_summary: pd.DataFrame) -> pd.DataFrame:
    summary = _as_dataframe(backtest_summary)
    if summary.empty and len(summary.columns) == 0:
        return pd.DataFrame(columns=BACKTEST_SUMMARY_COLUMNS)
    if "error_rate" not in summary.columns and "mean_error_rate" in summary.columns:
        summary = summary.copy()
        summary["error_rate"] = summary["mean_error_rate"]

    for column in BACKTEST_SUMMARY_COLUMNS:
        if column not in summary.columns:
            summary[column] = None
    ordered_columns = [
        *BACKTEST_SUMMARY_COLUMNS,
        *(column for column in summary.columns if column not in BACKTEST_SUMMARY_COLUMNS),
    ]
    return summary.loc[:, ordered_columns]


def _build_model_weights(backtest_summary: pd.DataFrame) -> pd.DataFrame:
    if backtest_summary.empty:
        return pd.DataFrame(columns=MODEL_WEIGHTS_COLUMNS)
    if "forecast_model" not in backtest_summary.columns or "error_rate" not in backtest_summary.columns:
        return pd.DataFrame(columns=MODEL_WEIGHTS_COLUMNS)

    weights = backtest_summary.copy()
    weights["error_rate"] = pd.to_numeric(weights["error_rate"], errors="coerce")
    finite_rows = weights.loc[weights["error_rate"].map(_is_finite_number)].copy()
    if finite_rows.empty:
        return pd.DataFrame(columns=MODEL_WEIGHTS_COLUMNS)

    zero_error = finite_rows.loc[finite_rows["error_rate"] == 0]
    if not zero_error.empty:
        finite_rows["model_weight"] = 0.0
        finite_rows.loc[zero_error.index, "model_weight"] = 1.0 / len(zero_error)
    else:
        finite_rows = finite_rows.loc[finite_rows["error_rate"] > 0].copy()
        if finite_rows.empty:
            return pd.DataFrame(columns=MODEL_WEIGHTS_COLUMNS)
        inverse_error = 1.0 / finite_rows["error_rate"]
        finite_rows["model_weight"] = inverse_error / inverse_error.sum()

    for column in MODEL_WEIGHTS_COLUMNS:
        if column not in finite_rows.columns:
            finite_rows[column] = None
    return finite_rows.loc[:, list(MODEL_WEIGHTS_COLUMNS)].reset_index(drop=True)


def _build_confidence_band_frame(forecast_history: pd.DataFrame) -> pd.DataFrame:
    history = _as_dataframe(forecast_history)
    if history.empty:
        return pd.DataFrame(columns=CONFIDENCE_BAND_COLUMNS)

    band_pair = _confidence_band_pair(history)
    if band_pair is None:
        return pd.DataFrame(columns=CONFIDENCE_BAND_COLUMNS)

    lower_column, upper_column = band_pair
    result = history.copy()
    if lower_column != "confidence_lower":
        result["confidence_lower"] = result[lower_column]
    if upper_column != "confidence_upper":
        result["confidence_upper"] = result[upper_column]

    for column in CONFIDENCE_BAND_COLUMNS:
        if column not in result.columns:
            result[column] = None
    return result.loc[:, list(CONFIDENCE_BAND_COLUMNS)].dropna(
        how="all",
        subset=["confidence_lower", "confidence_upper"],
    )


def _confidence_band_pair(df: pd.DataFrame) -> tuple[str, str] | None:
    columns = set(df.columns)
    for lower_column, upper_column in CONFIDENCE_BAND_COLUMN_PAIRS:
        if lower_column in columns and upper_column in columns:
            return lower_column, upper_column
    return None


def _build_insights_frame(
    forecast_history: pd.DataFrame,
    final_actuals: pd.DataFrame,
    backtest_summary: pd.DataFrame,
) -> pd.DataFrame:
    insights: list[str] = []
    if forecast_history.empty:
        insights.append(FORECAST_HISTORY_MESSAGE)
    if final_actuals.empty:
        insights.append(FINAL_ACTUALS_MESSAGE)
    if backtest_summary.empty:
        insights.append(BACKTEST_SUMMARY_MESSAGE)
    else:
        ranked = backtest_summary.copy()
        ranked["error_rate"] = pd.to_numeric(ranked["error_rate"], errors="coerce")
        ranked = ranked.loc[ranked["error_rate"].map(_is_finite_number)].sort_values(
            ["error_rate", "forecast_model"],
            ascending=[True, True],
            kind="mergesort",
        )
        if not ranked.empty:
            best = ranked.iloc[0]
            insights.append(
                "Best model by Backtest error_rate: "
                f"{best.get('forecast_model')} ({best.get('error_rate'):.1%})."
            )
    if not insights:
        insights.append("History export data is available.")
    return pd.DataFrame({"insight": insights})


def _resolve_output_path(output_path: str | Path) -> Path:
    path = Path(output_path)
    if path.suffix.lower() != ".xlsx":
        path = path.with_suffix(".xlsx")
    path = _ensure_versioned_path(path)

    if path.is_absolute():
        return path

    if path.parts and path.parts[0].lower() == "outputs":
        return _REPO_ROOT / path

    return _REPO_ROOT / "outputs" / path.name


def _ensure_versioned_path(path: Path) -> Path:
    if _VERSION_TOKEN_PATTERN.search(path.stem):
        return path
    return path.with_name(f"{path.stem}_{EXPORT_VERSION}{path.suffix}")


def _write_dataframe_sheet(ws: Worksheet, data: pd.DataFrame | Any) -> None:
    df = _as_dataframe(data)

    if df.empty and len(df.columns) == 0:
        ws.append(["message"])
        ws.append(["No data"])
        _apply_common_style(ws)
        return

    headers = [str(column) for column in df.columns]
    ws.append(headers)
    for row in df.itertuples(index=False, name=None):
        ws.append([_to_excel_value(value) for value in row])

    _apply_common_style(ws)
    _apply_column_formats(ws, headers)


def _write_optional_dataframe_sheet(
    ws: Worksheet,
    data: pd.DataFrame | Any,
    empty_message: str,
    fallback_columns: tuple[str, ...] = (),
) -> None:
    df = _as_dataframe(data)
    if df.empty:
        headers = list(df.columns) if len(df.columns) > 0 else list(fallback_columns)
        if not headers:
            headers = ["message"]
        if "message" not in headers:
            headers.append("message")
        ws.append([str(column) for column in headers])
        row = [None] * len(headers)
        row[headers.index("message")] = empty_message
        ws.append(row)
        _apply_common_style(ws)
        _apply_column_formats(ws, [str(column) for column in headers])
        return

    _write_dataframe_sheet(ws, df)


def _ensure_dataframe_columns(
    data: pd.DataFrame | Any,
    required_columns: tuple[str, ...],
) -> pd.DataFrame:
    df = _as_dataframe(data)
    for column in required_columns:
        if column not in df.columns:
            df[column] = None
    return df


def prepare_scenario_grid_export_frame(data: pd.DataFrame | Any) -> pd.DataFrame:
    """Add D03 display/export columns to ScenarioGrid without changing rows."""
    df = _as_dataframe(data)
    if df.empty and len(df.columns) == 0:
        return pd.DataFrame(columns=SCENARIO_GRID_EXPORT_COLUMN_ORDER)

    result = df.copy()
    result["scenario"] = _first_series(result, ("scenario", "scenario_id"))
    result["forecast_model"] = _first_series(result, ("forecast_model",))
    result["model_name"] = [
        get_forecast_model_label(_forecast_model_for_row(row))
        for _, row in result.iterrows()
    ]
    result["strategy_code"] = [
        get_strategy_code(_strategy_source_for_row(row))
        for _, row in result.iterrows()
    ]
    if "strategy_type" not in result.columns:
        result["strategy_type"] = None
    result["strategy_type"] = [
        _strategy_type_for_code(row.get("strategy_code"), row.get("strategy_type"))
        for _, row in result.iterrows()
    ]
    result["strategy_label"] = result["strategy_code"].map(get_strategy_label)
    result["strategy_group"] = result["strategy_code"].map(get_strategy_group)
    result["expected_month_end_amount"] = [
        _first_value(row, ("forecast_after_provision", "forecast_amount"))
        for _, row in result.iterrows()
    ]
    result["risk_note"] = [_risk_note_for_row(row) for _, row in result.iterrows()]

    for column in SCENARIO_GRID_REQUIRED_COLUMNS:
        if column not in result.columns:
            result[column] = None

    ordered_columns = [
        *[column for column in SCENARIO_GRID_EXPORT_COLUMN_ORDER if column in result.columns],
        *[
            column
            for column in result.columns
            if column not in set(SCENARIO_GRID_EXPORT_COLUMN_ORDER)
        ],
    ]
    return result.loc[:, ordered_columns]


def _first_series(df: pd.DataFrame, columns: tuple[str, ...]) -> pd.Series:
    for column in columns:
        if column in df.columns:
            return df[column]
    return pd.Series([None] * len(df), index=df.index)


def _forecast_model_for_row(row: Mapping[str, Any]) -> object:
    value = row.get("forecast_model")
    if not _is_missing_value(value):
        return value
    scenario_id = str(row.get("scenario_id") or row.get("scenario") or "")
    prefix = scenario_id.split("_", maxsplit=1)[0]
    return prefix if prefix in {"F1", "F2", "F3"} else value


def _strategy_source_for_row(row: Mapping[str, Any]) -> object:
    for key in (
        "overachievement_strategy",
        "provision_strategy",
        "neutral_strategy",
        "strategy_id",
        "strategy_code",
        "scenario",
        "scenario_id",
    ):
        value = row.get(key)
        if not _is_missing_value(value) and str(value):
            return value
    return ""


def _strategy_type_for_code(strategy_code: object, existing: object) -> object:
    if not _is_missing_value(existing) and str(existing):
        return existing
    code = get_strategy_code(strategy_code)
    if code in {"P1", "P2", "P3"}:
        return "PROVISION"
    if code in {"O1", "O2", "O3"}:
        return "OVERACHIEVEMENT"
    return "NEUTRAL"


def _first_value(row: Mapping[str, Any], columns: tuple[str, ...]) -> object:
    for column in columns:
        value = row.get(column)
        if not _is_missing_value(value):
            return value
    return None


def _risk_note_for_row(row: Mapping[str, Any]) -> str:
    for key in ("risk_note", "comment", "warnings", "recommended_action"):
        value = row.get(key)
        if _is_missing_value(value):
            continue
        if isinstance(value, (list, tuple, set)):
            text = ", ".join(str(item) for item in value if str(item))
        else:
            text = str(value)
        if text.strip():
            return text.strip()
    return "특이 리스크 없음"


def _is_missing_value(value: object) -> bool:
    if value is None:
        return True
    try:
        missing = pd.isna(value)
    except (TypeError, ValueError):
        return False
    if missing is pd.NA:
        return True
    if isinstance(missing, bool):
        return missing
    return False


def _prepare_close_cycle_export_frame(data: pd.DataFrame | Any) -> pd.DataFrame:
    df = _as_dataframe(data)
    if df.empty and len(df.columns) == 0:
        return df

    if _has_columns(df, _daily_close_cycle_input_columns()):
        return _build_daily_close_cycle_export_frame(df)

    return _ensure_close_cycle_cumulative_columns(df)


def _daily_close_cycle_input_columns() -> tuple[str, ...]:
    return (
        "business_day_no",
        "date",
        "is_close_day",
        "sales_target_daily",
        "recognized_target_daily",
        "sales_actual_cum",
        "recognized_actual_cum",
    )


def _build_daily_close_cycle_export_frame(df: pd.DataFrame) -> pd.DataFrame:
    working = df.copy()
    working["_business_day_sort"] = pd.to_numeric(
        working["business_day_no"],
        errors="raise",
    )
    working = working.sort_values(
        "_business_day_sort",
        kind="mergesort",
    ).reset_index(drop=True)

    is_close_day = _coerce_close_day_series(working["is_close_day"])
    close_cycle_no = _close_cycle_numbers(is_close_day)
    working["close_cycle_no"] = close_cycle_no
    working["close_cycle_label"] = [f"C{cycle_no:02d}" for cycle_no in close_cycle_no]

    for column in ("day_name", "close_type"):
        if column not in working.columns:
            working[column] = None

    for metric in ("sales", "recognized"):
        target_daily_column = f"{metric}_target_daily"
        actual_cum_column = f"{metric}_actual_cum"
        target_cum_column = f"{metric}_target_cum"
        gap_column = f"{metric}_gap_to_plan_cum"
        rate_column = f"{metric}_attainment_rate_cum"
        cycle_target_column = f"cycle_{metric}_target"
        cycle_actual_column = f"cycle_{metric}_actual"

        target_daily = pd.to_numeric(
            working[target_daily_column],
            errors="raise",
        ).fillna(0.0)
        actual_cum = pd.to_numeric(
            working[actual_cum_column],
            errors="coerce",
        )

        working[target_cum_column] = target_daily.cumsum()
        working[actual_cum_column] = actual_cum
        working[gap_column] = actual_cum - working[target_cum_column]
        working[rate_column] = actual_cum / working[target_cum_column].where(
            working[target_cum_column] != 0
        )

        working[cycle_target_column] = target_daily.groupby(
            working["close_cycle_no"],
            sort=False,
        ).transform("sum")
        working[cycle_actual_column] = _daily_actual_from_cumulative(actual_cum).groupby(
            working["close_cycle_no"],
            sort=False,
        ).transform(lambda values: values.sum(min_count=1))

    ordered_columns = [
        *[column for column in CLOSECYCLE_BASE_EXPORT_COLUMNS if column in working.columns],
        "sales_target_daily",
        "recognized_target_daily",
        *CLOSECYCLE_REQUIRED_CUMULATIVE_COLUMNS,
        *[
            column
            for column in working.columns
            if column
            not in {
                "_business_day_sort",
                *CLOSECYCLE_BASE_EXPORT_COLUMNS,
                "sales_target_daily",
                "recognized_target_daily",
                *CLOSECYCLE_REQUIRED_CUMULATIVE_COLUMNS,
            }
        ],
    ]
    return working.loc[:, ordered_columns]


def _ensure_close_cycle_cumulative_columns(df: pd.DataFrame) -> pd.DataFrame:
    result = df.copy()

    if "date" not in result.columns and "cycle_end_date" in result.columns:
        result["date"] = result["cycle_end_date"]
    if "close_cycle_no" not in result.columns and "cycle_id" in result.columns:
        result["close_cycle_no"] = result["cycle_id"]
    if "close_cycle_label" not in result.columns and "close_cycle_no" in result.columns:
        result["close_cycle_label"] = [
            f"C{int(cycle_no):02d}" if _is_finite_number(cycle_no) else None
            for cycle_no in result["close_cycle_no"]
        ]

    _ensure_metric_close_cycle_columns(
        result,
        "sales",
        target_source=_first_existing_column(
            result,
            ("cycle_sales_target", "sales_target_daily", "sales_target", "target_sum"),
        ),
        actual_source=_first_existing_column(
            result,
            ("cycle_sales_actual", "sales_actual_daily", "sales_actual", "actual_sum"),
        ),
    )
    _ensure_metric_close_cycle_columns(
        result,
        "recognized",
        target_source=_first_existing_column(
            result,
            (
                "cycle_recognized_target",
                "recognized_target_daily",
                "recognized_target",
            ),
        ),
        actual_source=_first_existing_column(
            result,
            (
                "cycle_recognized_actual",
                "recognized_actual_daily",
                "recognized_actual",
            ),
        ),
    )

    for column in CLOSECYCLE_REQUIRED_CUMULATIVE_COLUMNS:
        if column not in result.columns:
            result[column] = None

    preferred_columns = [
        *[column for column in CLOSECYCLE_BASE_EXPORT_COLUMNS if column in result.columns],
        *CLOSECYCLE_REQUIRED_CUMULATIVE_COLUMNS,
    ]
    return result.loc[
        :,
        [
            *preferred_columns,
            *[column for column in result.columns if column not in set(preferred_columns)],
        ],
    ]


def _ensure_metric_close_cycle_columns(
    df: pd.DataFrame,
    metric: str,
    *,
    target_source: str | None,
    actual_source: str | None,
) -> None:
    target_cum_column = f"{metric}_target_cum"
    actual_cum_column = f"{metric}_actual_cum"
    gap_column = f"{metric}_gap_to_plan_cum"
    rate_column = f"{metric}_attainment_rate_cum"

    if target_cum_column not in df.columns and target_source is not None:
        df[target_cum_column] = pd.to_numeric(
            df[target_source],
            errors="coerce",
        ).fillna(0.0).cumsum()

    if actual_cum_column not in df.columns and actual_source is not None:
        df[actual_cum_column] = pd.to_numeric(
            df[actual_source],
            errors="coerce",
        ).cumsum()

    if gap_column not in df.columns and {target_cum_column, actual_cum_column} <= set(df.columns):
        df[gap_column] = pd.to_numeric(
            df[actual_cum_column],
            errors="coerce",
        ) - pd.to_numeric(
            df[target_cum_column],
            errors="coerce",
        )

    if rate_column not in df.columns and {target_cum_column, actual_cum_column} <= set(df.columns):
        target_cum = pd.to_numeric(df[target_cum_column], errors="coerce")
        actual_cum = pd.to_numeric(df[actual_cum_column], errors="coerce")
        df[rate_column] = actual_cum / target_cum.where(target_cum != 0)


def _daily_actual_from_cumulative(actual_cum: pd.Series) -> pd.Series:
    daily = actual_cum.diff()
    if not daily.empty:
        daily.iloc[0] = actual_cum.iloc[0]
    daily.loc[actual_cum.isna()] = pd.NA
    return daily


def _close_cycle_numbers(is_close_day: pd.Series) -> list[int]:
    cycle_numbers: list[int] = []
    current_cycle_no = 1
    for is_close in is_close_day:
        cycle_numbers.append(current_cycle_no)
        if is_close:
            current_cycle_no += 1
    return cycle_numbers


def _coerce_close_day_series(values: pd.Series) -> pd.Series:
    true_tokens = {"Y", "YES", "TRUE", "1"}
    false_tokens = {"N", "NO", "FALSE", "0", ""}
    coerced: list[bool] = []
    for value in values:
        try:
            if pd.isna(value):
                coerced.append(False)
                continue
        except TypeError:
            pass

        if isinstance(value, bool):
            coerced.append(value)
            continue

        if isinstance(value, str):
            token = value.strip().upper()
            if token in true_tokens:
                coerced.append(True)
                continue
            if token in false_tokens:
                coerced.append(False)
                continue

        if _is_number(value) and value in (0, 1):
            coerced.append(bool(value))
            continue

        raise ValueError(f"Unsupported is_close_day value: {value!r}")

    return pd.Series(coerced, index=values.index, dtype=bool)


def _has_columns(df: pd.DataFrame, columns: tuple[str, ...]) -> bool:
    return set(columns) <= set(df.columns)


def _first_existing_column(df: pd.DataFrame, candidates: tuple[str, ...]) -> str | None:
    for column in candidates:
        if column in df.columns:
            return column
    return None


def _write_mapping_sheet(ws: Worksheet, data: Mapping[str, Any] | Any) -> None:
    ws.append(["item", "value"])

    if isinstance(data, Mapping):
        items = data.items()
    else:
        items = [("value", data)]

    for key, value in items:
        ws.append([str(key), _to_excel_value(value)])

    _apply_common_style(ws)
    _apply_key_value_formats(ws)


def _write_report_text_sheet(ws: Worksheet, report_text: str) -> None:
    ws.append(["report_text"])
    ws.append([str(report_text)])
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 100
    ws.row_dimensions[2].height = 90
    _apply_common_style(ws)


def _as_dataframe(data: pd.DataFrame | Any) -> pd.DataFrame:
    if isinstance(data, pd.DataFrame):
        return data.copy(deep=False)
    if data is None:
        return pd.DataFrame()
    return pd.DataFrame(data)


def _to_excel_value(value: object) -> object:
    if isinstance(value, pd.Timestamp):
        timestamp = value.to_pydatetime()
        return timestamp.replace(tzinfo=None) if timestamp.tzinfo else timestamp

    if isinstance(value, datetime):
        return value.replace(tzinfo=None) if value.tzinfo else value

    if isinstance(value, date):
        return value

    if isinstance(value, Mapping):
        return json.dumps(value, ensure_ascii=False, default=str)

    if isinstance(value, (list, tuple, set)):
        return json.dumps(list(value), ensure_ascii=False, default=str)

    try:
        missing = pd.isna(value)
    except (TypeError, ValueError):
        missing = False

    if isinstance(missing, bool) and missing:
        return None

    if isinstance(value, float) and not isfinite(value):
        return None

    if hasattr(value, "item"):
        try:
            return value.item()
        except ValueError:
            return value

    return value


def _apply_common_style(ws: Worksheet) -> None:
    _style_header(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = _used_range_ref(ws)
    _resize_columns(ws)


def _style_header(ws: Worksheet) -> None:
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _HEADER_BORDER
        cell.alignment = Alignment(horizontal="center")


def _used_range_ref(ws: Worksheet) -> str:
    return f"A1:{get_column_letter(ws.max_column)}{max(ws.max_row, 1)}"


def _resize_columns(ws: Worksheet) -> None:
    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        if ws.title == "ReportText" and column_letter == "A":
            continue

        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 42)


def _apply_column_formats(ws: Worksheet, headers: list[str]) -> None:
    for column_index, header in enumerate(headers, start=1):
        for row_index in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_index, column=column_index)
            _apply_cell_format(cell, header)


def _apply_key_value_formats(ws: Worksheet) -> None:
    for row_index in range(2, ws.max_row + 1):
        label = str(ws.cell(row=row_index, column=1).value or "")
        cell = ws.cell(row=row_index, column=2)
        _apply_cell_format(cell, label)


def _apply_cell_format(cell: Any, label: str) -> None:
    if not _is_number(cell.value):
        if _is_date_label(label) and cell.value is not None:
            cell.number_format = DATE_NUMBER_FORMAT
        return

    if _is_rate_label(label):
        if _is_percent_basis_label(label) and abs(cell.value) > 1:
            cell.value = cell.value / 100
        cell.number_format = RATE_NUMBER_FORMAT
        return

    if _is_amount_label(label):
        cell.number_format = AMOUNT_NUMBER_FORMAT


def _is_number(value: object) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def _is_finite_number(value: object) -> bool:
    try:
        return isfinite(float(value))
    except (TypeError, ValueError):
        return False


def _is_rate_label(label: str) -> bool:
    normalized = label.lower()
    return any(hint in normalized for hint in _RATE_NAME_HINTS)


def _is_percent_basis_label(label: str) -> bool:
    return label.lower() in _PERCENT_BASIS_COLUMNS


def _is_amount_label(label: str) -> bool:
    normalized = label.lower()
    if any(hint in normalized for hint in _COUNT_NAME_HINTS):
        return False
    if _is_rate_label(normalized) or _is_date_label(normalized):
        return False
    return any(hint in normalized for hint in _AMOUNT_NAME_HINTS)


def _is_date_label(label: str) -> bool:
    return "date" in label.lower()

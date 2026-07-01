"""Scan and organize generated output workbooks."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, Sequence
import fnmatch

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


REPO_ROOT = Path(__file__).resolve().parents[1]
OUTPUTS_ROOT = REPO_ROOT / "outputs"

STATUS_LATEST = "latest"
STATUS_OLD_FORMAT = "old_format"
STATUS_INVALID = "invalid"

LATEST_DIR_NAME = "latest"
ARCHIVE_OLD_FORMAT_DIR_NAME = "archive_old_format"
ARCHIVE_INVALID_DIR_NAME = "archive_invalid"
INPUT_TEMPLATE_NAME = "month_close_forecast_input_template.xlsx"

REQUIRED_LATEST_SHEETS = {
    "Summary",
    "ScenarioGrid",
    "DailyRevisedTargets",
    "CloseCycle",
    "Validation",
    "ReportText",
}
ADVANCED_LATEST_SHEETS = {
    "ForecastHistory",
    "FinalActuals",
    "BacktestSummary",
    "ModelWeights",
    "ConfidenceBand",
    "Insights",
}
LEGACY_OUTPUT_NAME_PATTERNS = (
    "daily_report_*_20260602.xlsx",
    "daily_report_sales_20260604.xlsx",
)
INPUT_TEMPLATE_HEADERS = [
    "date",
    "day_name",
    "business_day_no",
    "is_close_day",
    "close_type",
    "sales_target_daily",
    "recognized_target_daily",
    "sales_actual_cum",
    "recognized_actual_cum",
    "memo",
]


def scan_outputs(outputs_root: Path | str | None = None) -> dict[str, Any]:
    """Classify all xlsx files below outputs."""
    root = _outputs_root(outputs_root)
    files = [classify_workbook(path, root) for path in iter_xlsx_files(root)]
    return {
        "outputs_root": str(root),
        "counts": _status_counts(files),
        "files": files,
    }


def organize_outputs(
    outputs_root: Path | str | None = None,
    *,
    dry_run: bool = False,
    create_template: bool = True,
) -> dict[str, Any]:
    """Move output workbooks into latest/archive folders."""
    root = _outputs_root(outputs_root)
    scan = scan_outputs(root)
    moves: list[dict[str, str]] = []

    if not dry_run:
        ensure_output_dirs(root)

    for item in scan["files"]:
        source = root / item["relative_path"]
        target_dir = target_dir_for_status(root, item["status"])
        desired_target = target_dir / source.name
        if source.resolve() == desired_target.resolve():
            continue
        target = unique_destination(desired_target)

        move_entry = {
            "source": _relative_to_root(source, root),
            "destination": _relative_to_root(target, root),
            "status": item["status"],
        }
        moves.append(move_entry)
        if not dry_run:
            target.parent.mkdir(parents=True, exist_ok=True)
            source.rename(target)

    created_template: str | None = None
    if create_template and not dry_run:
        template_path = root / LATEST_DIR_NAME / INPUT_TEMPLATE_NAME
        if not is_valid_input_template_file(template_path):
            create_input_template(template_path)
            created_template = _relative_to_root(template_path, root)

    return {
        "status": "DRY_RUN" if dry_run else "ORGANIZED",
        "outputs_root": str(root),
        "counts": scan["counts"],
        "moves": moves,
        "created_template": created_template,
    }


def classify_workbook(path: Path | str, outputs_root: Path | str | None = None) -> dict[str, Any]:
    """Classify one workbook by readability and required sheet/header structure."""
    root = _outputs_root(outputs_root)
    workbook_path = Path(path)
    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - invalid workbooks are an expected category.
        return {
            "path": str(workbook_path),
            "relative_path": _relative_to_root(workbook_path, root),
            "status": STATUS_INVALID,
            "reason": f"{type(exc).__name__}: {exc}",
            "sheets": [],
            "advanced_sheet_count": 0,
        }

    try:
        sheet_names = set(workbook.sheetnames)
        advanced_sheet_count = len(sheet_names & ADVANCED_LATEST_SHEETS)
        if is_legacy_output_name(workbook_path.name):
            status = STATUS_OLD_FORMAT
            reason = "legacy output filename pattern"
        elif REQUIRED_LATEST_SHEETS.issubset(sheet_names):
            status = STATUS_LATEST
            reason = "required latest report sheets present"
        elif is_valid_input_template_workbook(workbook):
            status = STATUS_LATEST
            reason = "valid input template headers present"
        else:
            status = STATUS_OLD_FORMAT
            missing = sorted(REQUIRED_LATEST_SHEETS - sheet_names)
            reason = "missing latest report sheets: " + ", ".join(missing)

        return {
            "path": str(workbook_path),
            "relative_path": _relative_to_root(workbook_path, root),
            "status": status,
            "reason": reason,
            "sheets": workbook.sheetnames,
            "advanced_sheet_count": advanced_sheet_count,
        }
    finally:
        workbook.close()


def iter_xlsx_files(outputs_root: Path | str | None = None) -> list[Path]:
    """Return xlsx files under outputs, sorted for stable reports."""
    root = _outputs_root(outputs_root)
    if not root.exists():
        return []
    return sorted(path for path in root.rglob("*.xlsx") if path.is_file())


def is_legacy_output_name(filename: str) -> bool:
    """Return True for explicitly identified old output snapshots."""
    return any(
        fnmatch.fnmatch(filename, pattern)
        for pattern in LEGACY_OUTPUT_NAME_PATTERNS
    )


def ensure_output_dirs(outputs_root: Path | str | None = None) -> None:
    """Create managed output directories and keep empty dirs visible."""
    root = _outputs_root(outputs_root)
    for dirname in (
        LATEST_DIR_NAME,
        ARCHIVE_OLD_FORMAT_DIR_NAME,
        ARCHIVE_INVALID_DIR_NAME,
    ):
        directory = root / dirname
        directory.mkdir(parents=True, exist_ok=True)
        gitkeep = directory / ".gitkeep"
        if not gitkeep.exists():
            gitkeep.write_text("", encoding="utf-8")


def target_dir_for_status(outputs_root: Path, status: str) -> Path:
    """Return the managed folder for a classification status."""
    if status == STATUS_LATEST:
        return outputs_root / LATEST_DIR_NAME
    if status == STATUS_OLD_FORMAT:
        return outputs_root / ARCHIVE_OLD_FORMAT_DIR_NAME
    if status == STATUS_INVALID:
        return outputs_root / ARCHIVE_INVALID_DIR_NAME
    raise ValueError(f"Unknown output status: {status}")


def unique_destination(path: Path) -> Path:
    """Return a non-overwriting destination path."""
    if not path.exists():
        return path

    for index in range(1, 1000):
        candidate = path.with_name(f"{path.stem}_{index}{path.suffix}")
        if not candidate.exists():
            return candidate
    raise FileExistsError(f"Could not find a free destination for {path}")


def create_input_template(path: Path | str) -> Path:
    """Create a valid, non-sensitive input template workbook."""
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "InputTemplate"
    worksheet.append(INPUT_TEMPLATE_HEADERS)
    worksheet.append(
        [
            "YYYY-MM-DD",
            "display_only",
            1,
            False,
            "",
            None,
            None,
            None,
            None,
            "",
        ]
    )
    header_fill = PatternFill(fill_type="solid", fgColor="44546A")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    worksheet.freeze_panes = "A2"

    workbook.save(output_path)
    workbook.close()
    return output_path


def is_valid_input_template_file(path: Path | str) -> bool:
    """Return True when the workbook opens and has required input headers."""
    template_path = Path(path)
    if not template_path.is_file():
        return False
    try:
        workbook = load_workbook(template_path, read_only=True, data_only=True)
    except Exception:  # noqa: BLE001 - invalid files should be replaced.
        return False
    try:
        return is_valid_input_template_workbook(workbook)
    finally:
        workbook.close()


def is_valid_input_template_workbook(workbook: Any) -> bool:
    """Return True when the active sheet has the required input headers."""
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(min_row=1, max_row=1, values_only=True)
    first_row = next(rows, ())
    headers = {str(value).strip() for value in first_row if value is not None}
    return set(INPUT_TEMPLATE_HEADERS).issubset(headers)


def format_scan_markdown(scan: dict[str, Any]) -> str:
    """Format a scan result as a compact Markdown table."""
    lines = [
        "# Outputs Scan Report",
        "",
        "| status | file | reason |",
        "| --- | --- | --- |",
    ]
    for item in scan["files"]:
        lines.append(
            f"| {item['status']} | {item['relative_path']} | {item['reason']} |"
        )
    if not scan["files"]:
        lines.append("| none | - | no xlsx files found |")
    return "\n".join(lines) + "\n"


def main(argv: Sequence[str] | None = None) -> int:
    """CLI entry point."""
    parser = argparse.ArgumentParser(description="Scan and organize outputs/*.xlsx files.")
    subparsers = parser.add_subparsers(dest="command", required=True)

    scan_parser = subparsers.add_parser("scan", help="Print output classification JSON.")
    scan_parser.add_argument(
        "--format",
        choices=("json", "markdown"),
        default="json",
        help="Output format.",
    )

    organize_parser = subparsers.add_parser("organize", help="Move outputs into managed folders.")
    organize_parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Show planned moves without creating folders or moving files.",
    )
    organize_parser.add_argument(
        "--format",
        choices=("json", "markdown"),
        default="json",
        help="Output format.",
    )
    organize_parser.add_argument(
        "--no-template",
        action="store_true",
        help="Do not create the replacement input template.",
    )

    report_parser = subparsers.add_parser("report", help="Print a Markdown output report.")
    report_parser.add_argument(
        "--format",
        choices=("markdown", "json"),
        default="markdown",
        help="Output format.",
    )

    args = parser.parse_args(argv)
    if args.command == "organize":
        result = organize_outputs(
            dry_run=args.dry_run,
            create_template=not args.no_template,
        )
        print(_format_result(result, args.format))
        return 0

    scan = scan_outputs()
    output_format = args.format
    print(_format_result(scan, output_format))
    return 0


def _format_result(result: dict[str, Any], output_format: str) -> str:
    if output_format == "markdown":
        if "files" in result:
            return format_scan_markdown(result)
        lines = [
            "# Outputs Organize Report",
            "",
            f"- status: {result['status']}",
            f"- created_template: {result['created_template'] or 'none'}",
            "",
            "| status | source | destination |",
            "| --- | --- | --- |",
        ]
        for move in result["moves"]:
            lines.append(
                f"| {move['status']} | {move['source']} | {move['destination']} |"
            )
        if not result["moves"]:
            lines.append("| none | - | - |")
        return "\n".join(lines) + "\n"
    return json.dumps(result, ensure_ascii=False, indent=2)


def _status_counts(files: list[dict[str, Any]]) -> dict[str, int]:
    counts = {
        STATUS_LATEST: 0,
        STATUS_OLD_FORMAT: 0,
        STATUS_INVALID: 0,
    }
    for item in files:
        counts[item["status"]] += 1
    return counts


def _outputs_root(outputs_root: Path | str | None = None) -> Path:
    return Path(outputs_root).resolve() if outputs_root is not None else OUTPUTS_ROOT


def _relative_to_root(path: Path, root: Path) -> str:
    try:
        return path.resolve().relative_to(root.resolve()).as_posix()
    except ValueError:
        return path.as_posix()


if __name__ == "__main__":
    raise SystemExit(main())

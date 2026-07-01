"""Strict checker for shareable files in outputs/latest."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, Sequence

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
OUTPUTS_ROOT = REPO_ROOT / "outputs"
LATEST_DIR_NAME = "latest"
ARCHIVE_OLD_FORMAT_DIR_NAME = "archive_old_format"
ARCHIVE_INVALID_DIR_NAME = "archive_invalid"

CATEGORY_LATEST_REPORT = "latest_report"
CATEGORY_INPUT_TEMPLATE = "input_template"
CATEGORY_OLD_FORMAT = "old_format"
CATEGORY_INVALID = "invalid"
CATEGORY_UNKNOWN = "unknown"
CATEGORY_ALLOWED_MARKER = "allowed_marker"

REQUIRED_REPORT_COLUMNS = [
    "target_status",
    "target_variance",
    "surplus_to_target",
    "strategy_type",
    "overachievement_strategy",
    "recommended_action",
]
ADVANCED_SHEETS = [
    "ForecastHistory",
    "FinalActuals",
    "BacktestSummary",
    "ModelWeights",
    "ConfidenceBand",
    "Insights",
]
INPUT_TEMPLATE_HEADERS = [
    "date",
    "day_name",
    "business_day_no",
    "is_close_day",
    "close_type",
    "sales_target_daily",
    "recognized_target_daily",
    "sales_actual_cum",
    "recognized_actual_cum",
    "memo",
]

REPORT_NAME_MARKERS = ("report", "daily_report", "e2e_report")


def check_outputs_latest(
    outputs_root: Path | str | None = None,
    *,
    strict: bool = False,
    organize: bool = False,
) -> dict[str, Any]:
    """Inspect outputs/latest and optionally archive disallowed old or invalid files."""
    root = _outputs_root(outputs_root)
    latest_dir = root / LATEST_DIR_NAME
    initial_results = scan_latest(latest_dir, root)
    moves: list[dict[str, str]] = []

    if organize:
        moves = organize_latest_files(initial_results["file_results"], root)

    final_results = scan_latest(latest_dir, root)
    result = _summary_result(
        final_results["file_results"],
        root,
        latest_dir,
        strict=strict or organize,
    )
    result["initial_checked_files"] = initial_results["checked_files"]
    result["moved_to_archive_old_format"] = [
        move["destination"]
        for move in moves
        if move["category"] == CATEGORY_OLD_FORMAT
    ]
    result["moved_to_archive_invalid"] = [
        move["destination"]
        for move in moves
        if move["category"] == CATEGORY_INVALID
    ]
    result["moves"] = moves
    return result


def scan_latest(latest_dir: Path | str, outputs_root: Path | str | None = None) -> dict[str, Any]:
    """Return per-file classifications for outputs/latest."""
    root = _outputs_root(outputs_root)
    directory = Path(latest_dir)
    file_results = [
        classify_latest_file(path, root)
        for path in _iter_latest_files(directory)
    ]
    return {
        "outputs_root": str(root),
        "latest_dir": str(directory),
        "checked_files": [item["relative_path"] for item in file_results],
        "file_results": file_results,
    }


def classify_latest_file(path: Path | str, outputs_root: Path | str | None = None) -> dict[str, Any]:
    """Classify one file by the strict outputs/latest sharing policy."""
    root = _outputs_root(outputs_root)
    workbook_path = Path(path)
    relative_path = _relative_to_root(workbook_path, root)
    base: dict[str, Any] = {
        "path": str(workbook_path),
        "relative_path": relative_path,
        "filename": workbook_path.name,
        "category": CATEGORY_UNKNOWN,
        "reason": "",
        "sheets": [],
        "missing_required_columns": [],
        "missing_advanced_sheets": [],
    }

    if workbook_path.name == ".gitkeep":
        base["category"] = CATEGORY_ALLOWED_MARKER
        base["reason"] = "allowed latest directory marker"
        return base

    if workbook_path.suffix.lower() != ".xlsx":
        base["reason"] = "non-xlsx file is not allowed in outputs/latest"
        return base

    if "pre_metadata" in workbook_path.name.lower():
        base["category"] = CATEGORY_OLD_FORMAT
        base["reason"] = "pre_metadata workbook belongs in archive_old_format"
        return base

    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - invalid workbooks are expected findings.
        base["category"] = CATEGORY_INVALID
        base["reason"] = f"workbook could not be opened: {type(exc).__name__}"
        return base

    try:
        base["sheets"] = list(workbook.sheetnames)
        if _is_input_template_workbook(workbook):
            base["category"] = CATEGORY_INPUT_TEMPLATE
            base["reason"] = "input template headers present"
            return base

        if "ScenarioGrid" not in workbook.sheetnames:
            if _looks_like_report_name(workbook_path.name):
                base["category"] = CATEGORY_OLD_FORMAT
                base["reason"] = "report workbook is missing ScenarioGrid sheet"
            else:
                base["reason"] = "xlsx is neither a latest report nor an input template"
            return base

        headers = _worksheet_headers(workbook["ScenarioGrid"])
        missing_columns = [
            column
            for column in REQUIRED_REPORT_COLUMNS
            if column not in headers
        ]
        base["missing_required_columns"] = missing_columns
        base["missing_advanced_sheets"] = [
            sheet
            for sheet in ADVANCED_SHEETS
            if sheet not in workbook.sheetnames
        ]
        if missing_columns:
            base["category"] = CATEGORY_OLD_FORMAT
            base["reason"] = "ScenarioGrid is missing required latest columns"
            return base

        base["category"] = CATEGORY_LATEST_REPORT
        base["reason"] = "ScenarioGrid required latest columns present"
        return base
    finally:
        workbook.close()


def organize_latest_files(
    file_results: Sequence[dict[str, Any]],
    outputs_root: Path | str | None = None,
) -> list[dict[str, str]]:
    """Move old-format and invalid files out of outputs/latest without overwriting."""
    root = _outputs_root(outputs_root)
    archive_old = root / ARCHIVE_OLD_FORMAT_DIR_NAME
    archive_invalid = root / ARCHIVE_INVALID_DIR_NAME
    archive_old.mkdir(parents=True, exist_ok=True)
    archive_invalid.mkdir(parents=True, exist_ok=True)
    _ensure_gitkeep(archive_old)
    _ensure_gitkeep(archive_invalid)

    moves: list[dict[str, str]] = []
    for item in file_results:
        category = item["category"]
        if category not in {CATEGORY_OLD_FORMAT, CATEGORY_INVALID}:
            continue

        source = Path(item["path"])
        if not source.is_file():
            continue

        target_dir = archive_invalid if category == CATEGORY_INVALID else archive_old
        target = unique_destination(target_dir / source.name)
        target.parent.mkdir(parents=True, exist_ok=True)
        source.rename(target)
        moves.append(
            {
                "source": _relative_to_root(source, root),
                "destination": _relative_to_root(target, root),
                "category": category,
            }
        )
    return moves


def unique_destination(path: Path) -> Path:
    """Return a destination path that does not overwrite an existing file."""
    if not path.exists():
        return path

    for index in range(1, 1000):
        candidate = path.with_name(f"{path.stem}_{index}{path.suffix}")
        if not candidate.exists():
            return candidate
    raise FileExistsError(f"Could not find a free destination for {path}")


def main(argv: Sequence[str] | None = None) -> int:
    """CLI entry point."""
    parser = argparse.ArgumentParser(
        description="Check that outputs/latest contains only shareable current files."
    )
    parser.add_argument("--json", action="store_true", help="Print JSON output.")
    parser.add_argument("--strict", action="store_true", help="Fail on disallowed files.")
    parser.add_argument(
        "--organize",
        action="store_true",
        help="Archive old-format and invalid files out of outputs/latest.",
    )
    args = parser.parse_args(argv)

    result = check_outputs_latest(strict=args.strict, organize=args.organize)
    if args.json:
        print(json.dumps(result, ensure_ascii=False, indent=2))
    else:
        print(format_text_result(result))
    return 0 if result["result"] == "PASS" else 1


def format_text_result(result: dict[str, Any]) -> str:
    """Return a compact text report without workbook cell values."""
    lines = [
        f"outputs_latest_strict: {result['result']}",
        f"checked_files: {len(result['checked_files'])}",
        f"latest_reports: {len(result['latest_reports'])}",
        f"input_templates: {len(result['input_templates'])}",
        f"old_format_files: {len(result['old_format_files'])}",
        f"invalid_files: {len(result['invalid_files'])}",
        f"unknown_files: {len(result['unknown_files'])}",
    ]
    if result["old_format_files"]:
        lines.append("old_format_files:")
        lines.extend(f"  - {path}" for path in result["old_format_files"])
    if result["invalid_files"]:
        lines.append("invalid_files:")
        lines.extend(f"  - {path}" for path in result["invalid_files"])
    if result["unknown_files"]:
        lines.append("unknown_files:")
        lines.extend(f"  - {path}" for path in result["unknown_files"])
    if result["moved_to_archive_old_format"]:
        lines.append("moved_to_archive_old_format:")
        lines.extend(f"  - {path}" for path in result["moved_to_archive_old_format"])
    if result["moved_to_archive_invalid"]:
        lines.append("moved_to_archive_invalid:")
        lines.extend(f"  - {path}" for path in result["moved_to_archive_invalid"])
    return "\n".join(lines)


def _summary_result(
    file_results: Sequence[dict[str, Any]],
    outputs_root: Path,
    latest_dir: Path,
    *,
    strict: bool,
) -> dict[str, Any]:
    latest_reports = _paths_for_category(file_results, CATEGORY_LATEST_REPORT)
    input_templates = _paths_for_category(file_results, CATEGORY_INPUT_TEMPLATE)
    old_format_files = _paths_for_category(file_results, CATEGORY_OLD_FORMAT)
    invalid_files = _paths_for_category(file_results, CATEGORY_INVALID)
    unknown_files = _paths_for_category(file_results, CATEGORY_UNKNOWN)
    fail_files = old_format_files + invalid_files + unknown_files
    return {
        "outputs_root": str(outputs_root),
        "latest_dir": str(latest_dir),
        "checked_files": [item["relative_path"] for item in file_results],
        "latest_reports": latest_reports,
        "input_templates": input_templates,
        "old_format_files": old_format_files,
        "invalid_files": invalid_files,
        "unknown_files": unknown_files,
        "missing_required_columns": {
            item["relative_path"]: item["missing_required_columns"]
            for item in file_results
            if item["missing_required_columns"]
        },
        "missing_advanced_sheets": {
            item["relative_path"]: item["missing_advanced_sheets"]
            for item in file_results
            if item["category"] == CATEGORY_LATEST_REPORT
            and item["missing_advanced_sheets"]
        },
        "file_results": list(file_results),
        "result": "FAIL" if strict and fail_files else "PASS",
    }


def _paths_for_category(
    file_results: Sequence[dict[str, Any]],
    category: str,
) -> list[str]:
    return [
        item["relative_path"]
        for item in file_results
        if item["category"] == category
    ]


def _iter_latest_files(latest_dir: Path) -> list[Path]:
    if not latest_dir.exists():
        return []
    return sorted(path for path in latest_dir.iterdir() if path.is_file())


def _worksheet_headers(worksheet: Any) -> set[str]:
    row = next(
        worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
        (),
    )
    return {str(value).strip() for value in row if value is not None}


def _is_input_template_workbook(workbook: Any) -> bool:
    if not workbook.sheetnames:
        return False
    headers = _worksheet_headers(workbook[workbook.sheetnames[0]])
    return set(INPUT_TEMPLATE_HEADERS).issubset(headers)


def _looks_like_report_name(filename: str) -> bool:
    lowered = filename.lower()
    return any(marker in lowered for marker in REPORT_NAME_MARKERS)


def _ensure_gitkeep(directory: Path) -> None:
    gitkeep = directory / ".gitkeep"
    if not gitkeep.exists():
        gitkeep.write_text("", encoding="utf-8")


def _outputs_root(outputs_root: Path | str | None = None) -> Path:
    return Path(outputs_root).resolve() if outputs_root is not None else OUTPUTS_ROOT


def _relative_to_root(path: Path, root: Path) -> str:
    try:
        return path.resolve().relative_to(root.resolve()).as_posix()
    except ValueError:
        return path.as_posix()


if __name__ == "__main__":
    raise SystemExit(main())
